import pandas as pd
import subprocess
import sys
import os
import warnings
import unicodedata
import re
import io
import base64
import numpy as np
import matplotlib.pyplot as plt
from IPython.display import display, HTML

# Potlač zbytečná varování z openpyxl (neznámé Excel extensions, conditional formatting)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- Kontrola závislostí ---
try:
    import openpyxl
except ImportError:
    print("Modul 'openpyxl' není nainstalován. Instalace...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

try:
    import shapely
except ImportError:
    print("Modul 'shapely' není nainstalován. Instalace...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "shapely"])
    import shapely

try:
    import pyproj
except ImportError:
    print("Modul 'pyproj' není nainstalován. Instalace...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pyproj"])
    import pyproj


# =============================================================================
# KONFIGURACE
# =============================================================================

# ── Datum exportu DBS ─────────────────────────────────────────────────────────
# Změň toto datum při každém novém exportu DBS.
# Musí odpovídat hodnotě EFFECTIVE_DATE v souboru dbs_branch_network_epb.csv.
# Populace poboček zůstane uzamčena dle branch_snapshot.csv bez ohledu na datum.
DBS_DATE = "2026-02-10"

# Definice souborů
soubory = {
    "profitabilita": {
        "path": '../vypocet_ir_2026/zdroje/Pobockova_profitabilita_4Q2025.xlsx',
        "validity": "EOY2025",
        "source_desc": "Hlavní report pobočkové profitability z controllingu.",
        "extra_params": {"header": 2, "usecols": "A:AB"}
    },
    "revenues": {
        "path": "../in/tables/YEARLY_VYNOSY_NOVYCH_OBCHODU_BASED_ON_LTV.csv",
        "validity": "EOY2025",
        "source_desc": "Soubor s výnosy z OKO vynásobené LTV values 2H2024.",
        "extra_params": {}
    },
    "dbs": {
        "path": "../in/tables/dbs_branch_network_epb.csv",
        "validity": DBS_DATE,
        "source_desc": f"Data z databáze síť — export ke dni {DBS_DATE}.",
        "extra_params": {"filter_date": DBS_DATE}
    },
    "old_ir": {
        "path": "../vypocet_ir_2026/zdroje/IR_2024_v06.xlsx",
        "validity": "EOY2024",
        "source_desc": "Interní rating pro rok 2025 zpracovaný k EOY2024.",
        "extra_params": {}
    },
    "rent_and_ownership": {
        "path": "export_najmy_2026.pkl",
        "validity": "EOY2025",
        "source_desc": "Nájemné a vlastnictví poboček — export ze smluv, zdroj Edita Mudrová.",
        "extra_params": {},
        "hj_merge": True          # join přes CONCATENATED_HJ (explode)
    },
    "invest": {
        "path": "export_investice_2026.pkl",
        "validity": "EOY2025",
        "source_desc": "Přehled provedených investic do poboček, zdroj Adéla Grusová.",
        "extra_params": {},
        "hj_merge": True          # join přes CONCATENATED_HJ (explode)
    },
    "kpis": {
        "path": "kpis_grouped_2026.pkl",
        "validity": "EOY2025",
        "source_desc": "KPI metriky návštěvnosti a schůzek agregované na pobočku za celý rok.",
        "extra_params": {},
        "merge_on": {"left_on": "BRANCH_CODE", "right_on": "POBOCKA_ID"}
    },
    "parties": {
        "path": "parties_2026.pkl",
        "validity": "EOY2025",
        "source_desc": "BNS parties, data předávaná o klientech do CleverMaps.",
        "extra_params": {},
        "merge_on": {"left_on": "BRANCH_CODE", "right_on": "DBS_HOME_BRANCH_CODE"},
        "drop_key": "DBS_HOME_BRANCH_CODE",
    },
    "obchody_detail": {
        "path": "export_obchodu_detail_2026.pkl",
        "validity": "EOY2025",
        "source_desc": "Data s obchody z OKO za rok 2024 a 2025 podle produktové skupiny.",
        "extra_params": {},
        "merge_on": {"left_on": "BRANCH_CODE", "right_on": "BRANCH_CODE"},
    },
    "footprint_bns": {
        "path": "../vypocet_ir_2026/zdroje/footprint_bns.xlsx",
        "validity": "2026-02-26",
        "source_desc": "Data se strategiemi BNS. Workshopy, výstupy regionu k 26.2.2026",
        "extra_params": {},
        "merge_on": {"left_on": "BRANCH_CODE", "right_on": "BRANCH_ID"},
    },
    "hotovostni_trn": {
        "path": '../vypocet_ir_2026/zdroje/prevoditelnost_cash_trn_2025_10_02_2026_12.xlsx',
        "validity": "EOY2025",
        "source_desc": "Převoditelné transakce na pokladnách z Generátoru, celý rok 2025.",
        "extra_params": {"sheet_name": "only_prevoditelnost_cash_trn"}
    },
    "ar": {
        "path": "export_ar_2026.pkl",
        "validity": "EOY2025",
        "source_desc": "Data z alokačního reportu pro rok 2025.",
        "extra_params": {},
        "hj_merge": True
    },
    "specialiste": {
        "path": "export_specialiste.pkl",
        "validity": "2026-03-01",
        "source_desc": "Obsazení pozic specialistů na pobočkách — počty pracovníků dle pracovní pozice.",
        "extra_params": {},
        "no_merge": True   # načítáme zvlášť, nemergeujeme do df
    },
    "kapacita_atm": {
        "path": '../vypocet_ir_2026/export_kapacita_atm.pkl',
        "validity": "EOY2025",
        "source_desc": "Kapacita ATM vypočítaná Honzou Boreckým na datech za rok 2025.",
        "extra_params": {},
        "no_merge": True
    },
    "export_atm": {
        "path": '../vypocet_ir_2026/export_atm.pkl',
        "validity": "2026-09-03",
        "source_desc": "Report atm 9.3.2026",
        "extra_params": {},
        "no_merge": True
    },
    "visits": {
        "path": '../in/tables/VISITS_2025.csv',
        "validity": "EOY2025",
        "source_desc": "Návštěvní data poboček za celý rok 2025 — datum, čas, typ návštěvy, segment klienta.",
        "extra_params": {},
        "no_merge": True
    },
    "parties_full": {
        "path": '../in/tables/BNS_PARTIES_2026.csv',
        "validity": "EOY2025",
        "source_desc": "Úplný BNS parties s GPS adresou klienta (LOC_GEO_LATITUDE/LONGITUDE) a CP_SUBSEGMENT_ID.",
        "extra_params": {"usecols": ["PT_UNIFIED_KEY", "LOC_GEO_LATITUDE", "LOC_GEO_LONGITUDE", "CP_SUBSEGMENT_ID", "DBS_HOME_BRANCH_CODE"]},
        "no_merge": True
    },
    "spadovky": {
        "path": '../vypocet_ir_2026/zdroje/ČS - návštěvnost poboček ve spádovkách.xlsx',
        "validity": "31.12.2024",
        "source_desc": "Spádové pobočky — kam přejdou klienti v případě uzavření pobočky (top 3 alternativy dle analýzy návštěvnosti).",
        "extra_params": {},
        "no_merge": True
    }
}

# Konfigurace sloupců pro mergeování
MERGE_COLS = {
    "profitabilita": {
        "cols": ["ID_POBOCKY", "FTE", "PRIMARNI_KLIENTI", "AKTIVNI_KLIENTI", "POCET_KLIENTU",
                 "VYNOSY", "OSOBNI_NAKLADY", "PRIME_NAKLADY", "PRIME_NAKLADY/VYNOSY"],
        "left_on": "BRANCH_CODE",
        "right_on": "ID_POBOCKY",
        "how": "left",
        "suffixes": ("", ""),
        "drop_key": "ID_POBOCKY",
    },
    "revenues": {
        "cols": ["BRANCH_CODE", "POCET_PRODEJU", "OBJEM_VYNOSU_CZK"],
        "on": "BRANCH_CODE",
        "how": "left",
    },
    "old_ir": {
        "cols": ["BRANCH_ID", "INTERNI_RATING_2023", "INTERNI_RATING_2024",
                 "FORMAT_2023", "FORMAT_2024_(FIX_FS)", "NEW_BUSINESS_2024_-_OBJEM_VYNOSU",
                 "NEW_BUSINESS_2024_-_POCTY_KS", "C/I_RATIO_(YTD_2024)", "REVENUES_2024",
                 "REVENUES_2023", "REVENUES_2022", "REVENUES_2021",
                 "C/I_RATIO_(YTD_2021)", "C/I_RATIO_(YTD_2022)", "C/I_RATIO_(YTD_2023)",
                 "INTERNI_RATING_2024_-_KVINTILY", "INTERNI_RATING_2024_-_PERCENTILY",
                 "HODNOTA_INDEXU_EXPOZICE",
                 "POCET_KLIENTU_EOY_2021", "POCET_KLIENTU_EOY_2022",
                 "POCET_KLIENTU_EOY_2023", "POCET_KLIENTU_EOY_2024"],
        "left_on": "BRANCH_CODE",
        "right_on": "BRANCH_ID",
        "how": "left",
        "suffixes": ("", "_OLD_2024"),
        "drop_key": "BRANCH_ID",
    },
    "dbs": {
        "cols": ["BRANCH_CODE", "BRANCH_TYPE_NAME", "BRANCH_BUILDING_NF_SF",
                 "NEW_FORMAT_SINCE", "CASHLESS", "CASHLESS_SINCE",
                 "GPS_X", "GPS_Y",
                 "CITY", "CITY_DISTRICT", "STREET", "CISLO_POPISNE",
                 "CISLO_ORIENTACNI", "RUIAN_ID"],
        "on": "BRANCH_CODE",
        "how": "left",
    },
    "footprint_bns": {
        "cols": ["BRANCH_ID", "FOOTPRINT_STRATEGIE_(2030)", "PLAN_CLOSE_(ROK)",
                 "CLOSE_STATUS", "INVESTICE_NF_(ROK)", "INVESTICE_FHC_(ROK)",
                 "INVESTICE_NF_OR_FHC_(ROK)",
                 "PLAN_NEW_CASHIERLESS_(ROK)", "NEW_CASHIERLESS_STATUS"],
        "left_on": "BRANCH_CODE",
        "right_on": "BRANCH_ID",
        "how": "left",
        "drop_key": "BRANCH_ID",
    },
    # rent_and_ownership se merguje přes explode_hj_merge z pkl — není zde
}

# Sloupce pro zobrazení a jejich přejmenování (sdíleno oběma reporty)
# Pořadí odpovídá pořadí skupin v COL_GROUPS
NOVE_NAZVY = {
    # ── Identita ──────────────────────────────────────────────────
    "BRANCH_CODE": "ID Pobočky",
    "BRANCH_NAME": "Název pobočky",
    "REGION_NAME": "Region",
    "OBLAST":      "Oblast",

    # ── 📍 Adresa ─────────────────────────────────────────────────
    "CITY":              "Město",
    "CITY_DISTRICT":     "Obvod / část",
    "STREET":            "Ulice",
    "CISLO_POPISNE":     "Č. popisné",
    "CISLO_ORIENTACNI":  "Č. orientační",
    "RUIAN_ID":          "RUIAN ID",
    "ORP_NAZEV":         "ORP",
    "ORP_KOD":           "ORP kód",

    # ── 👥 Klienti ────────────────────────────────────────────────
    "POCET_KLIENTU_EOY_2021": "Klienti EOY 2021",
    "POCET_KLIENTU_EOY_2022": "Klienti EOY 2022",
    "POCET_KLIENTU_EOY_2023": "Klienti EOY 2023",
    "POCET_KLIENTU_EOY_2024": "Klienti EOY 2024",
    "POCET_KLIENTU":          "Klienti EOY 2025",
    "KLIENTI_TREND_HTML":     "Trend klientů 21–25",
    "PRIMARNI_KLIENTI":   "Primární klienti",
    "AKTIVNI_KLIENTI":    "Aktivní klienti",
    "PRIM_RATIO":         "Primární / celkem",
    "AKTIVNI_RATIO":      "Aktivní / celkem",
    "KLIENTI_RATIO_HTML": "Poměr klientů",
    "CIZINCI":            "Cizinci celkem",
    "CIZINCI_SLOVENSKO":  "Cizinci SK",
    "CIZINCI_UKRAJINA":   "Cizinci UA",

    # ── ⭐ Interní rating ─────────────────────────────────────────
    "INTERNI_RATING_2023":   "Rating 23",
    "INTERNI_RATING_2024":   "Rating 24",
    "IR":                    "Rating 25",
    "IR_Q_HTML":             "Rating 25 kvintil",
    "IR_CHANGE_HTML":        "Změna ratingu 25/24",
    "IR_PERC_CHANGE_HTML":   "Změna ratingu perc 25/24",

    # ── 📊 C/I ────────────────────────────────────────────────────
    "C/I_RATIO_(YTD_2021)":  "C/I ratio 21",
    "C/I_RATIO_(YTD_2022)":  "C/I ratio 22",
    "C/I_RATIO_(YTD_2023)":  "C/I ratio 23",
    "C/I_RATIO_(YTD_2024)":  "C/I ratio 24",
    "PRIME_NAKLADY/VYNOSY":  "C/I ratio 25",
    "CI_RATIO_STATUS":       "C/I kvintil",
    "CI_CHANGE_HTML":        "Změna C/I 25/24",

    # ── 💰 Výnosy ─────────────────────────────────────────────────
    "REVENUES_2021":                     "Výnosy 21",
    "REVENUES_2022":                     "Výnosy 22",
    "REVENUES_2023":                     "Výnosy 23",
    "REVENUES_2024":                     "Výnosy 24",
    "VYNOSY":                            "Výnosy 25",
    "REV_SPARKLINE":                     "Trend výnosů 21–25",
    "NEW_BUSINESS_2024_-_OBJEM_VYNOSU":  "Nové výnosy 24",
    "OBJEM_VYNOSU_CZK":                  "Nové výnosy 25",
    "NEW_BUSINESS_VOL_STATUS":           "Nové výnosy kvintil",
    "REV_CHANGE_HTML":                   "Změna nových výnosů 25/24",

    # ── 📈 Business rating ────────────────────────────────────────
    "CAP_REV_FTE":           "Výnos/FTE",
    "CAP_REV_FTE_Q_HTML":    "Výnos/FTE kvintil",
    "CAP_NEWCLI_FTE":        "Noví klienti/FTE",
    "CAP_NEWCLI_FTE_Q_HTML": "Noví klienti/FTE kvintil",
    "CAP_PRIMCLI_FTE":       "Primární klienti/FTE",
    "CAP_PRIMCLI_FTE_Q_HTML":"Prim. klienti/FTE kvintil",
    "CAP_SCHUZKY_FTE":       "Schůzky/FTE",
    "CAP_SCHUZKY_FTE_Q_HTML":"Schůzky/FTE kvintil",
    "CAP_PEREX":             "PEREX",
    "CAP_PEREX_Q_HTML":      "PEREX kvintil",
    "CAP_RNK":               "Business pořadí",
    "CAP_Q":                 "Business rating Q",
    "CAP_Q_HTML":            "Business rating kvintil",

    # ── 🏦 Hotovostní rating ──────────────────────────────────────
    "TRN_RNK":       "Cash rating pořadí",
    "TRN_STRATEGIE": "Vypočítaná strategie",
    "TRN_Q_HTML":    "Cash rating kvintil",
    "TRN_Q":         "Cash rating Q",
    "TRN_POCET_CELKEM": "Trn. počet celkem",
    "TRN_CASTKA_CELKEM":"Trn. objem celkem",
    "TRN_DENNI_POCET":  "Denní trn. průměr",

    # ── 🏠 Nájemné ────────────────────────────────────────────────
    "VLASTNICTVI":                 "Vlastnictví",
    "ROCNI_SPLATKY_S_DPH_CZK":    "Roční nájemné",
    "ROCNI_SPLATKY_S_DPH_CZK_STATUS": "Nájemné kvintil",
    "ZACATEK_DATUM":               "Začátek smlouvy",
    "KONEC_DATUM":                 "Konec smlouvy",
    "VYPOVEDNI_DOBA":              "Výpovědní lhůta",

    # ── 🔨 Investice ──────────────────────────────────────────────
    "DESC_LAST_INV":     "Poslední investice",
    "DATUM_LAST_INV":    "Datum posl. inv.",
    "LAST_INV":          "Výše posl. inv.",
    "DESC_BIGGEST_INV":  "Nejvyšší investice",
    "DATUM_BIGGEST_INV": "Datum nejv. inv.",
    "BIGGEST_INV":       "Výše nejv. inv.",
    "TOTAL_INV":         "Celkem zainvestováno",

    # ── 🚶 Návštěvy ───────────────────────────────────────────────
    "POCET_SCHUZEK_ONLINE":  "Online schůzky",
    "POCET_SCHUZEK_FYZICKY": "Fyzické schůzky",
    "POCET_BEZHOT_WALK_IN":  "Bezhot. walkin",
    "POCET_HOT_WALK_IN":     "Hot. walkin",
    "POCET_NAVSTEV_CELKEM":  "Celkové návštěvy",
    "NAVSTEVY_RATIO_HTML":   "Poměr návštěv",

    # ── 🗺️ Strategie BNS ─────────────────────────────────────────
    "FOOTPRINT_STRATEGIE_(2030)":  "Strategie BNS",
    "PLAN_CLOSE_(ROK)":            "Rok uzavření",
    "CLOSE_STATUS":                "Status uzavření",
    "INVESTICE_NF_(ROK)":          "Rok investice",
    "INVESTICE_FHC_(ROK)":         "Rok FHC investice",
    "INVESTICE_NF_OR_FHC_(ROK)":   "Rok plánované investice",

    # ── 💵 Strategie hotovosti ────────────────────────────────────
    "PLAN_NEW_CASHIERLESS_(ROK)":  "Rok cashless přechodu",
    "NEW_CASHIERLESS_STATUS":      "Status cashless",
    "CASHLESS":                    "Bezhotovostní",
    "CASHLESS_SINCE":              "Datum bezhotovostní",

    # ── 📐 Alokační report ────────────────────────────────────────
    "ONLY_D5_FLAG":                 "Pouze D5",
    "WPL_MAX_2026":                 "WPL max 2026",
    "WPL_IN_USE_2026":              "WPL využito 2026",
    "CELK_PLOCHA_POBOCKY_2026":     "Celk. plocha pobočky",
    "PLOCHA_POUZE_D5":              "Plocha pouze D5",
    "PLOCHA_NAD_OPTIMALNI_POBOCKU": "Plocha nad optimál",

    # ── 🏢 Budova ─────────────────────────────────────────────────
    "BRANCH_TYPE_NAME":      "Typologie",
    "BRANCH_FORMAT":          "Formát pobočky (celk. FTE)",
    "BRANCH_FORMAT_OBCHODNI": "Formát pobočky (obch. FTE)",
    "FORMAT_2024_(FIX_FS)":  "Realizovaný formát",
    "BRANCH_BUILDING_NF_SF": "Formát NF/SF",
    "NEW_FORMAT_SINCE":      "Datum NF",

    # ── 💵 Transakční data ────────────────────────────────────────
    "CASH_PREVODITELNOST_PCT":        "% převoditelnosti",
    "CASH_IN_PREVODITELNA_CASTKA":    "Cash IN převod. Kč",
    "CASH_IN_NEPREVODITELNA_CASTKA":  "Cash IN nepřevod. Kč",
    "CASH_IN_PREVODITELNA_POCET":     "Cash IN převod. počet",
    "CASH_IN_NEPREVODITELNA_POCET":   "Cash IN nepřevod. počet",
    "CASH_OUT_PREVODITELNA_CASTKA":   "Cash OUT převod. Kč",
    "CASH_OUT_NEPREVODITELNA_CASTKA": "Cash OUT nepřevod. Kč",
    "CASH_OUT_PREVODITELNA_POCET":    "Cash OUT převod. počet",
    "CASH_OUT_NEPREVODITELNA_POCET":  "Cash OUT nepřevod. počet",

    # ── 🔀 Spádová oblast ─────────────────────────────────────────
    "SP_NAZEV_1":  "Spádová pobočka 1",
    "SP_VISITS_1": "Spád. návštěvy 1",
    "SP_PCT_1":    "Spád. podíl 1",
    "SP_NAZEV_2":  "Spádová pobočka 2",
    "SP_VISITS_2": "Spád. návštěvy 2",
    "SP_PCT_2":    "Spád. podíl 2",
    "SP_NAZEV_3":  "Spádová pobočka 3",
    "SP_VISITS_3": "Spád. návštěvy 3",
    "SP_PCT_3":    "Spád. podíl 3",

    # ── 👷 FTE ────────────────────────────────────────────────────
    "FTE":               "FTE",
    "OBCHODNI_FTE":      "Obchodní FTE",
    "BANKERS_COUNT":     "Počet bankéřů",
    "BANKERS_NEEDED":    "Potřeba bankéřů",
    "BANKERS_DIFF_HTML": "Bankéři — rozdíl",
    "_total_spec":       "Všechny pozice celkem",

    # ── 🛍️ Prodeje (kvintily per produkt + celkový počet) ────────
    "POCET_PRODEJU_CELKEM_2025":         "Prodeje celkem 2025",
    "POCET_PRODEJU_ACCOUNTS_2025":       "Prodeje Účty 2025",
    "POCET_PRODEJU_ACCOUNTS_2025_Q_HTML":"Účty kvintil",
    "POCET_PRODEJU_HYPO_2025":           "Prodeje Hypotéky 2025",
    "POCET_PRODEJU_HYPO_2025_Q_HTML":    "Hypotéky kvintil",
    "POCET_PRODEJU_INS_LIFE_2025":       "Prodeje Poj. život 2025",
    "POCET_PRODEJU_INS_LIFE_2025_Q_HTML":"Poj. život kvintil",
    "POCET_PRODEJU_INS_NONLIFE_2025":    "Prodeje Poj. neživot 2025",
    "POCET_PRODEJU_INS_NONLIFE_2025_Q_HTML":"Poj. neživot kvintil",
    "POCET_PRODEJU_INV_REGULAR_2025":    "Prodeje Inv. pravid. 2025",
    "POCET_PRODEJU_INV_REGULAR_2025_Q_HTML":"Inv. pravid. kvintil",
    "POCET_PRODEJU_INV_SIMPLE_2025":     "Prodeje Inv. jednor. 2025",
    "POCET_PRODEJU_INV_SIMPLE_2025_Q_HTML":"Inv. jednor. kvintil",
    "POCET_PRODEJU_LOANREV_2025":        "Prodeje Revol. úvěry 2025",
    "POCET_PRODEJU_LOANREV_2025_Q_HTML": "Revol. úvěry kvintil",
    "POCET_PRODEJU_LOANS_2025":          "Prodeje Úvěry 2025",
    "POCET_PRODEJU_LOANS_2025_Q_HTML":   "Úvěry kvintil",
    "POCET_PRODEJU_PENSION_2025":        "Prodeje Penze 2025",
    "POCET_PRODEJU_PENSION_2025_Q_HTML": "Penze kvintil",

    # ── 📡 Index expozice ─────────────────────────────────────────
    "HODNOTA_INDEXU_EXPOZICE": "Index expozice",
    "IE_Q_HTML":               "Index expozice kvintil",

    # ── 🏧 ATM ────────────────────────────────────────────────────
    "ATM_SUMMARY_HTML":  "ATM přehled",
    "ATM_VOLNA_KAPACITA": "ATM volná kap. %",
    "ATM_VYBERY_MESIC":  "ATM výběry průměr/měs.",
    "ATM_VKLADY_MESIC":  "ATM vklady průměr/měs.",
    "ATM_VYBERY_PO_PREV": "ATM výběry po převodu/měs.",
    "ATM_VKLADY_PO_PREV": "ATM vklady po převodu/měs.",
}

COLS_TO_FIX = [
    'INTERNI_RATING_2023', 'INTERNI_RATING_2024', 'IR',
    'NEW_BUSINESS_2024_-_POCTY_KS', 'POCET_PRODEJU',
    'NEW_BUSINESS_2024_-_OBJEM_VYNOSU', 'OBJEM_VYNOSU_CZK',
    'C/I_RATIO_(YTD_2021)', 'C/I_RATIO_(YTD_2022)', 'C/I_RATIO_(YTD_2023)',
    'C/I_RATIO_(YTD_2024)', 'PRIME_NAKLADY/VYNOSY',
    'ROCNI_SPLATKY_S_DPH_CZK', 'VYNOSY', 'REVENUES_2024',
    "DESC_LAST_INV", "DATUM_LAST_INV", "LAST_INV",
    "DESC_BIGGEST_INV", "DATUM_BIGGEST_INV", "BIGGEST_INV", "TOTAL_INV",
    'POCET_SCHUZEK_ONLINE', 'POCET_SCHUZEK_FYZICKY',
    'POCET_BEZHOT_WALK_IN', 'POCET_HOT_WALK_IN',
    'POCET_NAVSTEV_CELKEM', 'IR_DIFF_NUM',
    'CIZINCI', 'CIZINCI_SLOVENSKO', 'CIZINCI_UKRAJINA',
    'POCET_KLIENTU_EOY_2021', 'POCET_KLIENTU_EOY_2022',
    'POCET_KLIENTU_EOY_2023', 'POCET_KLIENTU_EOY_2024',
    'HODNOTA_INDEXU_EXPOZICE', 'OBCHODNI_FTE',
    'POCET_PRODEJU_CELKEM_2025',
    # hotovostní transakce
    'CASH_IN_PREVODITELNA_CASTKA',   'CASH_IN_NEPREVODITELNA_CASTKA',
    'CASH_IN_PREVODITELNA_POCET',    'CASH_IN_NEPREVODITELNA_POCET',
    'CASH_OUT_PREVODITELNA_CASTKA',  'CASH_OUT_NEPREVODITELNA_CASTKA',
    'CASH_OUT_PREVODITELNA_POCET',   'CASH_OUT_NEPREVODITELNA_POCET',
    # cash rating
    'TRN_POCET_CELKEM', 'TRN_CASTKA_CELKEM', 'TRN_DENNI_POCET',
    'TRN_SCORE', 'TRN_Q', 'TRN_RNK', 'TRN_POCET_Q', 'TRN_CASTKA_Q',
    # business rating vstupní sloupce
    'FTE', 'PRIMARNI_KLIENTI', 'AKTIVNI_KLIENTI', 'POCET_KLIENTU',
    'OSOBNI_NAKLADY', 'PRIME_NAKLADY', 'NR_NEW_ARRIVALS',
    # business rating výstup
    'WPL_MAX_2026', 'WPL_IN_USE_2026', 'CELK_PLOCHA_POBOCKY_2026',
    'PLOCHA_POUZE_D5', 'PLOCHA_NAD_OPTIMALNI_POBOCKU',
    'CAP_RNK', 'CAP_Q',
    'CAP_REV_FTE', 'CAP_NEWCLI_FTE', 'CAP_PRIMCLI_FTE', 'CAP_SCHUZKY_FTE', 'CAP_PEREX',
    'CAP_REV_FTE_Q', 'CAP_NEWCLI_FTE_Q', 'CAP_PRIMCLI_FTE_Q', 'CAP_SCHUZKY_FTE_Q', 'CAP_PEREX_Q',
]

# Definice věkových skupin (centrálně pro snadnou úpravu)
AGE_GROUPS = ['1-15', '16-25', '26-45', '46-65', '65+']
AGE_COLORS = ['#2770ef', '#0bb440', '#00a3a5', '#fd6230', '#9b59b6']

# Produktové skupiny pro obchody_detail
PRODUCT_GROUPS = [
    ("ACCOUNTS",    "Účty"),
    ("HYPO",        "Hypotéky"),
    ("INS_LIFE",    "Poj. život"),
    ("INS_NONLIFE", "Poj. neživot"),
    ("INV_REGULAR", "Inv. pravid."),
    ("INV_SIMPLE",  "Inv. jednor."),
    ("LOANREV",     "Revol. úvěry"),
    ("LOANS",       "Úvěry"),
    ("PENSION",     "Penze"),
]

# Sloupce s datumy pro zobrazení
# *** OPRAVA: přidány ZACATEK_DATUM a KONEC_DATUM ***
DATE_COLS_DISPLAY = [
    'DATUM_LAST_INV', 'DATUM_BIGGEST_INV',
    'NEW_FORMAT_SINCE', 'CASHLESS_SINCE',
    'ZACATEK_DATUM', 'KONEC_DATUM',
]

# Hotovostní sloupce — skupiny pro přehled
CASH_CASTKA_COLS = [
    'CASH_IN_PREVODITELNA_CASTKA',
    'CASH_IN_NEPREVODITELNA_CASTKA',
    'CASH_OUT_PREVODITELNA_CASTKA',
    'CASH_OUT_NEPREVODITELNA_CASTKA',
]
CASH_POCET_COLS = [
    'CASH_IN_PREVODITELNA_POCET',
    'CASH_IN_NEPREVODITELNA_POCET',
    'CASH_OUT_PREVODITELNA_POCET',
    'CASH_OUT_NEPREVODITELNA_POCET',
]

# =============================================================================
# POMOCNÉ FUNKCE
# =============================================================================

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    def clean(col):
        if not isinstance(col, str): col = str(col)
        col = ''.join(c for c in unicodedata.normalize('NFD', col) if unicodedata.category(c) != 'Mn')
        col = col.upper()
        col = re.sub(r'\s+', '_', col.strip())
        return col
    df.columns = [clean(col) for col in df.columns]
    return df


def format_date(val):
    """Formátuje datum do dd.mm.yyyy, prázdné/NaN vrátí prázdný řetězec."""
    if val == "" or val is None:
        return ""
    if isinstance(val, float) and np.isnan(val):
        return ""
    try:
        parsed = pd.to_datetime(val, dayfirst=True)
        if pd.isna(parsed):
            return ""
        return parsed.strftime("%d.%m.%Y")
    except:
        return ""


def format_money(val):
    if val == "" or pd.isna(val): return "0 Kč"
    try: return f"{int(float(val)):,}".replace(",", " ") + " Kč"
    except: return "0 Kč"


def format_pct(val):
    try: return f"{float(val) * 100:.2f}%".replace(".", ",")
    except: return "0,00%"


def format_num_round(val):
    try: return f"{int(round(float(val))):,}".replace(",", "\u00a0")
    except: return "0"


def format_int_sep(val):
    """Celé číslo s mezerou jako oddělovačem tisíců."""
    try: return f"{int(float(val)):,}".replace(",", "\u00a0")
    except: return "0"


dot_scale = {1: '#0ab33f', 2: '#42db72', 3: '#f0b55d', 4: '#f0845d', 5: '#e64343'}

# Sjednocená funkce pro všechny kvintily — tečka + šedé číslo kvintilu
def get_quintile_dot(val, tooltip_prefix="Kvintil"):
    """
    Barevná tečka + šedé číslo kvintilu.
    1 = nejlepší (zelená), 5 = nejhorší (červená).
    Používá se pro všechna kvintilová zobrazení v reportu.
    """
    try:
        v = int(val)
        color = dot_scale.get(v, "gray")
        label = {1: "top 20 %", 2: "20–40 %", 3: "40–60 %", 4: "60–80 %", 5: "bottom 20 %"}.get(v, "")
        return (f'<span title="{tooltip_prefix} {v}: {label}" '
                f'style="color:{color}; font-size:22px; line-height:1; cursor:default;">●</span>'
                f'<span style="font-size:0.72rem;color:#888;margin-left:3px;">{v}</span>')
    except:
        return ""

# Aliasy pro zpětnou kompatibilitu (interně i pro obchody tabulku)
def get_dot(val):
    return get_quintile_dot(val)

def get_ir_q_dot(val):
    return get_quintile_dot(val, tooltip_prefix="Rating kvintil")


def get_change_html(val, mode='num'):
    """
    Generuje HTML pro barevné zobrazení změny.
    mode='num':   kladné = zelené (vyšší je lepší, např. počet prodejů, výnosy)
    mode='money': kladné = zelené (vyšší je lepší)
    mode='pct':   kladné = červené (vyšší C/I ratio je horší)
    mode='rating': kladné = červené (vyšší pořadí ratingu je horší)
    """
    if mode == 'money':
        color = "#0ab33f" if val > 0 else ("#e64343" if val < 0 else "gray")
        txt = format_money(val)
    elif mode == 'pct':
        color = "#0ab33f" if val < 0 else ("#e64343" if val > 0 else "gray")
        txt = format_pct(val)
    elif mode == 'rating':
        color = "#0ab33f" if val < 0 else ("#e64343" if val > 0 else "gray")
        txt = f"{int(val)}" if val == int(val) else f"{val:.1f}"
    else:
        color = "#0ab33f" if val > 0 else ("#e64343" if val < 0 else "gray")
        txt = f"{int(val)}" if val == int(val) else f"{val:.1f}"
    sign = "+" if val > 0 else ""
    return f'<b style="color:{color}">{sign}{txt}</b>'


def color_strategy_html(val):
    """Obarví hodnotu FOOTPRINT_STRATEGIE — keep=zelená, close=červená."""
    v = str(val).lower().strip()
    if not v or v in ('nan', 'none', ''):
        return ""
    if 'keep' in v:
        return (f"<span style='background:#d4edda;color:#155724;"
                f"padding:2px 8px;border-radius:4px;font-weight:bold;"
                f"display:inline-block;'>{val}</span>")
    elif 'close' in v:
        return (f"<span style='background:#f8d7da;color:#721c24;"
                f"padding:2px 8px;border-radius:4px;font-weight:bold;"
                f"display:inline-block;'>{val}</span>")
    return str(val)


# =============================================================================
# PIVOT HOTOVOSTNÍCH TRANSAKCÍ
# =============================================================================

def pivot_hotovostni_trn(df_trn: pd.DataFrame) -> pd.DataFrame:
    """
    Převede hotovostni_trn do wide formátu na úrovni pobočky.

    Vstupní sloupce (po normalize_columns):
      POBOCKA, POHYB (CASH_IN/CASH_OUT/OTHER), PREVODITELNOST (Y/N),
      SUM_CASTKA_TRANSAKCE, POCET_OPERACI

    Výstup (jeden řádek per pobočka):
      CASH_IN_PREVODITELNA_CASTKA / _NEPREVODITELNA_CASTKA / _POCET (×2)
      CASH_OUT_PREVODITELNA_CASTKA / _NEPREVODITELNA_CASTKA / _POCET (×2)
    """
    df = df_trn.copy()

    # --- Detekce sloupce s ID pobočky ---
    id_col_candidates = ['POBOCKA', 'BRANCH_CODE', 'POBOCKA_ID', 'ID_POBOCKY']
    id_col = next((c for c in id_col_candidates if c in df.columns), None)
    if id_col is None:
        raise ValueError(f"Nelze najít sloupec s ID pobočky. Dostupné: {list(df.columns)}")

    # --- Detekce ostatních sloupců ---
    pohyb_col  = next((c for c in ['POHYB', 'TYP_POHYBU'] if c in df.columns), None)
    prev_col   = next((c for c in ['PREVODITELNOST', 'PREV'] if c in df.columns), None)
    castka_col = next((c for c in ['SUM_CASTKA_TRANSAKCE', 'CASTKA', 'SUM_CASTKA'] if c in df.columns), None)
    pocet_col  = next((c for c in ['POCET_OPERACI', 'POCET'] if c in df.columns), None)

    for name, val in [('POHYB', pohyb_col), ('PREVODITELNOST', prev_col),
                      ('SUM_CASTKA_TRANSAKCE', castka_col), ('POCET_OPERACI', pocet_col)]:
        if val is None:
            raise ValueError(f"Nelze najít sloupec {name}. Dostupné: {list(df.columns)}")

    # --- Normalizace ---
    df[id_col]    = pd.to_numeric(df[id_col], errors='coerce')
    df[pohyb_col] = df[pohyb_col].astype(str).str.strip().str.upper()
    df[prev_col]  = df[prev_col].astype(str).str.strip().str.upper()
    df[castka_col] = pd.to_numeric(df[castka_col], errors='coerce').fillna(0)
    df[pocet_col]  = pd.to_numeric(df[pocet_col],  errors='coerce').fillna(0)

    # --- Filtr: jen CASH_IN a CASH_OUT ---
    df = df[df[pohyb_col].isin(['CASH_IN', 'CASH_OUT'])].copy()

    # --- Mapování Y/N → PREVODITELNA/NEPREVODITELNA ---
    df['PREV_LABEL'] = df[prev_col].map({'Y': 'PREVODITELNA', 'N': 'NEPREVODITELNA'})
    df = df[df['PREV_LABEL'].notna()].copy()

    if df.empty:
        empty = pd.DataFrame(columns=['BRANCH_CODE'] + CASH_CASTKA_COLS + CASH_POCET_COLS)
        return empty

    # --- Pivot částka ---
    pivot_castka = df.pivot_table(
        index=id_col,
        columns=[pohyb_col, 'PREV_LABEL'],
        values=castka_col,
        aggfunc='sum',
        fill_value=0
    )
    pivot_castka.columns = [f"{p}_{l}_CASTKA" for p, l in pivot_castka.columns]

    # --- Pivot počet ---
    pivot_pocet = df.pivot_table(
        index=id_col,
        columns=[pohyb_col, 'PREV_LABEL'],
        values=pocet_col,
        aggfunc='sum',
        fill_value=0
    )
    pivot_pocet.columns = [f"{p}_{l}_POCET" for p, l in pivot_pocet.columns]

    result = pivot_castka.join(pivot_pocet, how='outer').fillna(0).reset_index()

    # Přejmenovat ID sloupec na BRANCH_CODE a zajistit typ int (stejný jako dbs_final)
    result = result.rename(columns={id_col: 'BRANCH_CODE'})
    result['BRANCH_CODE'] = result['BRANCH_CODE'].astype(int)

    # Doplnit chybějící sloupce nulami
    for col in CASH_CASTKA_COLS + CASH_POCET_COLS:
        if col not in result.columns:
            result[col] = 0

    display(HTML(f"""
        <div style="background:#e8f8f0;padding:8px 12px;border-left:4px solid #27ae60;
                    font-size:0.85rem;margin-top:6px;">
            ✅ <b>pivot_hotovostni_trn OK</b> — {len(result)} poboček,
            vzorová data: CASH_IN převoditelná průměr =
            <b>{int(result['CASH_IN_PREVODITELNA_CASTKA'].mean()):,} Kč</b>
        </div>
    """))

    return result[['BRANCH_CODE'] + CASH_CASTKA_COLS + CASH_POCET_COLS]


# Počet pracovních dní v roce (standard pro výpočet denních operací)
WORKING_DAYS_PER_YEAR = 252

# Nové sloupce pro cash rating
CASH_RATING_COLS = [
    'TRN_POCET_CELKEM', 'TRN_CASTKA_CELKEM',
    'TRN_DENNI_POCET', 'TRN_SCORE',
    'TRN_POCET_Q', 'TRN_CASTKA_Q', 'TRN_Q',
]

def pivot_transakce_trn(df_trn: pd.DataFrame) -> pd.DataFrame:
    """
    Agreguje transakce_trn na úroveň pobočky — všechny kategorie (CASH_IN, CASH_OUT, OTHER).

    Vstup:  POBOCKA (nebo BRANCH_CODE), POHYB, SUM_CASTKA_TRANSAKCE, POCET_OPERACI
    Výstup: BRANCH_CODE, TRN_POCET_CELKEM, TRN_CASTKA_CELKEM, TRN_DENNI_POCET
    """
    df = df_trn.copy()

    id_col     = next((c for c in ['POBOCKA', 'BRANCH_CODE', 'POBOCKA_ID', 'ID_POBOCKY',
                                      'BRANCH_ID', 'KOD_POBOCKY', 'KOD', 'BRANCH'] if c in df.columns), None)
    castka_col = next((c for c in ['SUM_CASTKA_TRANSAKCE', 'CASTKA', 'SUM_CASTKA',
                                    'CASTKA_TRANSAKCE', 'OBJEM', 'SUM_OBJEM'] if c in df.columns), None)
    pocet_col  = next((c for c in ['POCET_OPERACI', 'POCET', 'POCET_TRANSAKCI',
                                    'COUNT', 'POCET_TRN', 'CNT'] if c in df.columns), None)

    if id_col is None:
        raise ValueError(f"Nelze najít ID pobočky. Sloupce: {list(df.columns)}")

    display(HTML(f"""
        <div style="background:#fff3cd;padding:6px 12px;border-left:4px solid #f39c12;font-size:0.82rem;">
            🔍 <b>pivot_transakce_trn detekce sloupců:</b>
            ID=<b>{id_col}</b> | Částka=<b>{castka_col}</b> | Počet=<b>{pocet_col}</b>
        </div>
    """))
    if castka_col is None:
        raise ValueError(f"Nelze najít sloupec s částkou. Sloupce: {list(df.columns)}")
    if pocet_col is None:
        raise ValueError(f"Nelze najít sloupec s počtem. Sloupce: {list(df.columns)}")

    df[id_col]     = pd.to_numeric(df[id_col],     errors='coerce')
    df[castka_col] = pd.to_numeric(df[castka_col], errors='coerce').fillna(0)
    df[pocet_col]  = pd.to_numeric(df[pocet_col],  errors='coerce').fillna(0)
    df = df.dropna(subset=[id_col])

    # Sečti CASH_IN + CASH_OUT + OTHER dohromady — žádný filtr na POHYB
    result = df.groupby(id_col).agg(
        TRN_POCET_CELKEM =(pocet_col,  'sum'),
        TRN_CASTKA_CELKEM=(castka_col, 'sum'),
    ).reset_index().rename(columns={id_col: 'BRANCH_CODE'})

    result['BRANCH_CODE']    = result['BRANCH_CODE'].astype(int)
    result['TRN_DENNI_POCET'] = (result['TRN_POCET_CELKEM'] / WORKING_DAYS_PER_YEAR).round(1)

    display(HTML(f"""
        <div style="background:#e8f8f0;padding:8px 12px;border-left:4px solid #27ae60;font-size:0.85rem;margin-top:6px;">
            ✅ <b>pivot_transakce_trn OK</b> — {len(result)} poboček |
            Celkový počet transakcí: <b>{int(result['TRN_POCET_CELKEM'].sum()):,}</b> |
            Průměrné denní operace/pobočka: <b>{result['TRN_DENNI_POCET'].mean():.1f}</b>
        </div>
    """))

    return result[['BRANCH_CODE', 'TRN_POCET_CELKEM', 'TRN_CASTKA_CELKEM', 'TRN_DENNI_POCET']]


def compute_capacity_rating(df: pd.DataFrame) -> pd.DataFrame:
    """
    Kapacitní rating pobočky — jak efektivně pobočka využívá svůj personál.

    Složky (všechny počítány na FTE):
      0.15 × výnos z nových obchodů na FTE          (OBJEM_VYNOSU_CZK / FTE)
      0.20 × počet nových klientů na FTE             (NR_NEW_ARRIVALS / FTE)
      0.15 × počet primárních klientů na FTE         (PRIMARNI_KLIENTI / FTE)
      0.10 × počet schůzek na FTE                    ((POCET_SCHUZEK_FYZICKY + POCET_SCHUZEK_ONLINE) / FTE)
      0.40 × PEREX nových výnosů na FTE              (( OSOBNI_NAKLADY/FTE ) / ( OBJEM_VYNOSU_CZK/FTE ))

    PEREX = osobní náklady na FTE / výnos z nových obchodů na FTE
    Nižší PEREX = efektivnější pobočka → rank ascending

    Výstup: CAP_RNK (pořadí 1 = nejlepší), CAP_Q (kvintil 1=nejlepší), CAP_Q_HTML
    """
    d = df.copy()

    required = ['FTE', 'OBJEM_VYNOSU_CZK', 'OSOBNI_NAKLADY', 'PRIMARNI_KLIENTI',
                'POCET_SCHUZEK_FYZICKY', 'POCET_SCHUZEK_ONLINE', 'NR_NEW_ARRIVALS']
    missing = [c for c in required if c not in d.columns]
    if missing:
        display(HTML(f"<b style='color:orange;'>⚠️ Business rating: chybí sloupce {missing} — přeskakuji.</b>"))
        for col in ['CAP_RNK', 'CAP_Q', 'CAP_Q_HTML']:
            d[col] = np.nan if col != 'CAP_Q_HTML' else ''
        return d

    # OSOBNI_NAKLADY jsou záporné (účetní konvence) — pracujeme s absolutní hodnotou
    d['_OSOBNI_NAKLADY_ABS'] = d['OSOBNI_NAKLADY'].abs()

    # Maska: pobočky s FTE > 0, nenulovými výnosy a nenulovými náklady
    mask = (
        (d['FTE'].notna()) & (d['FTE'] > 0) &
        (d['OBJEM_VYNOSU_CZK'].notna()) & (d['OBJEM_VYNOSU_CZK'] > 0) &
        (d['_OSOBNI_NAKLADY_ABS'].notna()) & (d['_OSOBNI_NAKLADY_ABS'] > 0)
    )

    if mask.sum() < 5:
        for col in ['CAP_RNK', 'CAP_Q', 'CAP_Q_HTML']:
            d[col] = np.nan if col != 'CAP_Q_HTML' else ''
        return d

    fte  = d.loc[mask, 'FTE']

    # ── Dílčí metriky na FTE ──────────────────────────────────────────
    rev_fte      = d.loc[mask, 'OBJEM_VYNOSU_CZK'] / fte           # výnos na FTE
    new_cli_fte  = d.loc[mask, 'NR_NEW_ARRIVALS'].fillna(0) / fte   # nových klientů na FTE
    prim_cli_fte = d.loc[mask, 'PRIMARNI_KLIENTI'].fillna(0) / fte  # primárních klientů na FTE
    schuzky_fte  = (d.loc[mask, 'POCET_SCHUZEK_FYZICKY'].fillna(0) +
                    d.loc[mask, 'POCET_SCHUZEK_ONLINE'].fillna(0)) / fte  # schůzky na FTE

    # PEREX = (osobní náklady / FTE) / (nové výnosy / FTE)
    # = osobní náklady / nové výnosy  — FTE se zkrátí
    # OSOBNI_NAKLADY jsou záporné (účetní konvence) → abs()
    # Nižší = lepší (méně nákladů na každou Kč výnosu)
    perex = (d.loc[mask, '_OSOBNI_NAKLADY_ABS'] / fte) / rev_fte

    # ── Ranky každé složky ────────────────────────────────────────────
    # U všech metrik vyšší = lepší → ascending=False → rank 1 = nejvyšší hodnota
    # Výjimka PEREX: nižší = lepší → ascending=True
    rnk_rev      = rev_fte.rank(ascending=False, method='first')
    rnk_new_cli  = new_cli_fte.rank(ascending=False, method='first')
    rnk_prim_cli = prim_cli_fte.rank(ascending=False, method='first')
    rnk_schuzky  = schuzky_fte.rank(ascending=False, method='first')
    rnk_perex    = perex.rank(ascending=True,  method='first')  # nižší PEREX = lepší

    # ── Vážený skóre ─────────────────────────────────────────────────
    score = (
        rnk_rev      * 0.15 +
        rnk_new_cli  * 0.20 +
        rnk_prim_cli * 0.15 +
        rnk_schuzky  * 0.10 +
        rnk_perex    * 0.40
    )

    # ── Dílčí hodnoty na FTE — uložit pro zobrazení ─────────────────
    d.loc[mask, 'CAP_REV_FTE']      = rev_fte.round(0)
    d.loc[mask, 'CAP_NEWCLI_FTE']   = new_cli_fte.round(2)
    d.loc[mask, 'CAP_PRIMCLI_FTE']  = prim_cli_fte.round(1)
    d.loc[mask, 'CAP_SCHUZKY_FTE']  = schuzky_fte.round(2)
    d.loc[mask, 'CAP_PEREX']        = perex.round(4)

    # ── Kvintily dílčích parametrů (Q1 = nejlepší) ───────────────────
    def _q_asc(series):   # vyšší = lepší
        return 6 - (pd.qcut(series.rank(ascending=False, method='first'), q=5, labels=False, duplicates='drop') + 1)
    def _q_desc(series):  # nižší = lepší (PEREX)
        return pd.qcut(series.rank(ascending=True, method='first'), q=5, labels=False, duplicates='drop') + 1

    d.loc[mask, 'CAP_REV_FTE_Q']     = _q_asc(rev_fte)
    d.loc[mask, 'CAP_NEWCLI_FTE_Q']  = _q_asc(new_cli_fte)
    d.loc[mask, 'CAP_PRIMCLI_FTE_Q'] = _q_asc(prim_cli_fte)
    d.loc[mask, 'CAP_SCHUZKY_FTE_Q'] = _q_asc(schuzky_fte)
    d.loc[mask, 'CAP_PEREX_Q']       = _q_desc(perex)

    for qcol in ['CAP_REV_FTE_Q','CAP_NEWCLI_FTE_Q','CAP_PRIMCLI_FTE_Q','CAP_SCHUZKY_FTE_Q','CAP_PEREX_Q']:
        d[f'{qcol}_HTML'] = d[qcol].map(
            lambda v, _c=qcol: get_quintile_dot(int(v), tooltip_prefix=_c.replace('_Q','').replace('CAP_','').replace('_',' ') + ' kvintil')
            if pd.notna(v) and v > 0 else ''
        )

    # ── Finální pořadí: nižší skóre = lepší = pořadí 1 ──────────────
    d.loc[mask, 'CAP_SCORE'] = score
    d.loc[mask, 'CAP_RNK']   = score.rank(ascending=True, method='first').astype(int)

    # ── Celkový kvintil: Q1 = top 20 % (nejlepší) ────────────────────
    d.loc[mask, 'CAP_Q'] = (
        pd.qcut(d.loc[mask, 'CAP_RNK'], q=5, labels=False, duplicates='drop') + 1
    )

    # ── HTML tečka celkového kvintilu ────────────────────────────────
    d['CAP_Q_HTML'] = d['CAP_Q'].map(
        lambda v: get_quintile_dot(int(v), tooltip_prefix="Business rating kvintil")
        if pd.notna(v) and v > 0 else ''
    )

    d = d.drop(columns=[c for c in ['CAP_SCORE', '_OSOBNI_NAKLADY_ABS'] if c in d.columns])
    return d


def compute_cash_rating(df: pd.DataFrame) -> pd.DataFrame:
    """
    Vypočítá hotovostní rating na základě počtu (45 %) a objemu (55 %) transakcí.
    Vrátí df s novými sloupci: TRN_SCORE, TRN_POCET_Q, TRN_CASTKA_Q, TRN_Q, TRN_Q_HTML.
    Kvintil 1 = nejvyšší aktivita (nejlepší), 5 = nejnižší.
    """
    d = df.copy()
    # Pouze pobočky s nenulovými transakcemi
    mask = (d['TRN_POCET_CELKEM'] > 0) & (d['TRN_CASTKA_CELKEM'] > 0)

    if mask.sum() < 5:
        d['TRN_POCET_Q']  = np.nan
        d['TRN_CASTKA_Q'] = np.nan
        d['TRN_SCORE']    = np.nan
        d['TRN_Q']        = np.nan
        d['TRN_Q_HTML']   = ''
        return d

    # ── Rank v celé populaci (všechny pobočky s hotovostí) ──────────────
    # ascending=False: největší počet/objem = rank 1
    d.loc[mask, 'TRN_POCET_RNK']  = d.loc[mask, 'TRN_POCET_CELKEM'].rank(ascending=False, method='first')
    d.loc[mask, 'TRN_CASTKA_RNK'] = d.loc[mask, 'TRN_CASTKA_CELKEM'].rank(ascending=False, method='first')

    # Vážený skóre: 45 % počet + 55 % objem
    d.loc[mask, 'TRN_SCORE'] = (
        d.loc[mask, 'TRN_POCET_RNK']  * 0.45 +
        d.loc[mask, 'TRN_CASTKA_RNK'] * 0.55
    )

    # Finální pořadí v celé populaci: nižší skóre = více transakcí = pořadí 1
    d.loc[mask, 'TRN_RNK'] = d.loc[mask, 'TRN_SCORE'].rank(ascending=True, method='first').astype(int)

    # Kvintily z celkového pořadí: Q1 = top 20 % (nejvyšší aktivita)
    d.loc[mask, 'TRN_Q'] = (
        pd.qcut(d.loc[mask, 'TRN_RNK'], q=5, labels=False, duplicates='drop') + 1
    )

    # HTML tečka s kvintily (Q1 zelená = nejlepší)
    d['TRN_Q_HTML'] = d['TRN_Q'].map(
        lambda v: get_quintile_dot(int(v), tooltip_prefix="Cash rating kvintil") if pd.notna(v) and v > 0 else ''
    )

    # ── Cashless rank — pořadí jen mezi cashless pobočkami ───────────
    # Cashless pobočky mají méně hotovostních transakcí než ne-cashless,
    # proto je jejich pořadí v celkové populaci horší — cashless rank
    # umožňuje srovnání cashless poboček navzájem.
    is_cashless = pd.Series(False, index=d.index)
    if 'CASHLESS' in d.columns:
        is_cashless = d['CASHLESS'].astype(str).str.strip().str.upper().isin(['Y', 'YES', 'ANO', '1', 'TRUE'])
    cash_mask = mask & is_cashless

    if cash_mask.sum() > 0:
        # Ranky počtu a objemu jen v rámci cashless podmnožiny
        d.loc[cash_mask, '_CL_POCET_RNK']  = d.loc[cash_mask, 'TRN_POCET_CELKEM'].rank(ascending=False, method='first')
        d.loc[cash_mask, '_CL_CASTKA_RNK'] = d.loc[cash_mask, 'TRN_CASTKA_CELKEM'].rank(ascending=False, method='first')
        d.loc[cash_mask, '_CL_SCORE'] = (
            d.loc[cash_mask, '_CL_POCET_RNK']  * 0.45 +
            d.loc[cash_mask, '_CL_CASTKA_RNK'] * 0.55
        )
        d.loc[cash_mask, 'TRN_CASHLESS_RNK'] = (
            d.loc[cash_mask, '_CL_SCORE'].rank(ascending=True, method='first').astype(int)
        )
        d = d.drop(columns=['_CL_POCET_RNK', '_CL_CASTKA_RNK', '_CL_SCORE'])

    # ── Vypočítaná strategie — prvních 20/35/50 podle CELKOVÉHO pořadí ─
    def _calc_strategie(rnk):
        try:
            r = int(rnk)
            if r <= 20: return "VAR. 20 poboček s hotovostí"
            if r <= 35: return "VAR. 35 poboček s hotovostí"
            if r <= 50: return "VAR. 50 poboček s hotovostí"
        except:
            pass
        return ""
    d.loc[mask, 'TRN_STRATEGIE'] = d.loc[mask, 'TRN_RNK'].map(_calc_strategie)

    d = d.drop(columns=[c for c in ['TRN_POCET_RNK', 'TRN_CASTKA_RNK'] if c in d.columns])
    return d


# =============================================================================
# GENERÁTORY GRAFŮ A TABULEK (beze změny)
# =============================================================================

def generate_apex_chart_html(df_reg, chart_id):
    """Vrátí HTML snippet s animovaným sloupcovým grafem (ApexCharts)."""
    labels = ['Online schůzky', 'Fyzické schůzky', 'Bezhot. walkin', 'Hot. walkin']
    colors = ['#2770ef', '#0bb440', '#00a3a5', '#fd6230']
    values = [
        int(df_reg['POCET_SCHUZEK_ONLINE'].sum()),
        int(df_reg['POCET_SCHUZEK_FYZICKY'].sum()),
        int(df_reg['POCET_BEZHOT_WALK_IN'].sum()),
        int(df_reg['POCET_HOT_WALK_IN'].sum())
    ]
    if sum(values) == 0:
        return "<p style='color:#aaa;'>Žádná data o návštěvách.</p>"

    labels_js = str(labels).replace("'", '"')
    colors_js = str(colors).replace("'", '"')
    values_js = str(values)

    return f"""
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3"></script>
<div id="{chart_id}" style="min-height:260px;"></div>
<script>
(function() {{
  var options = {{
    chart: {{
      type: "bar",
      height: 260,
      toolbar: {{ show: false }},
      animations: {{
        enabled: true,
        easing: "easeinout",
        speed: 600,
        animateGradually: {{ enabled: true, delay: 100 }},
        dynamicAnimation: {{ enabled: true, speed: 350 }}
      }},
      background: "transparent"
    }},
    series: [{{ name: "Po\u010det", data: {values_js} }}],
    xaxis: {{
      categories: {labels_js},
      labels: {{ style: {{ fontSize: "12px", colors: "#555" }} }}
    }},
    yaxis: {{
      labels: {{
        formatter: function(v) {{ return v.toLocaleString("cs-CZ"); }},
        style: {{ fontSize: "11px", colors: "#888" }}
      }}
    }},
    colors: {colors_js},
    plotOptions: {{
      bar: {{
        distributed: true,
        borderRadius: 5,
        columnWidth: "55%",
        dataLabels: {{ position: "top" }}
      }}
    }},
    dataLabels: {{
      enabled: true,
      formatter: function(v) {{ return v.toLocaleString("cs-CZ"); }},
      offsetY: -22,
      style: {{ fontSize: "11px", colors: ["#333"], fontWeight: "bold" }}
    }},
    legend: {{ show: false }},
    grid: {{
      borderColor: "#e0e0e0",
      strokeDashArray: 4,
      yaxis: {{ lines: {{ show: true }} }}
    }},
    tooltip: {{
      y: {{ formatter: function(v) {{ return v.toLocaleString("cs-CZ"); }} }}
    }}
  }};
  new ApexCharts(document.getElementById("{chart_id}"), options).render();
}})();
</script>
"""


def generate_age_segment_table(target_df, table_id="age-tbl"):
    count_cols = [f'{g}_POCET_KLIENTU' for g in AGE_GROUPS]
    rev_cols   = [f'{g}_PRUMERNY_VYNOS' for g in AGE_GROUPS]

    available_count = [c for c in count_cols if c in target_df.columns]
    available_rev   = [c for c in rev_cols   if c in target_df.columns]

    if not available_count:
        return "<p style='color:#aaa;font-style:italic;'>Žádná data o věkových skupinách.</p>"

    CIZINCI_COLS = [
        ('CIZINCI',           'Cizinci celkem', '#6c5ce7'),
        ('CIZINCI_SLOVENSKO', 'Cizinci SK',     '#00b894'),
        ('CIZINCI_UKRAJINA',  'Cizinci UA',     '#e17055'),
    ]
    available_cizinci = [(col, lbl, clr) for col, lbl, clr in CIZINCI_COLS if col in target_df.columns]

    base_cols = ['BRANCH_CODE', 'BRANCH_NAME']
    extra_ciz = [col for col, _, _ in available_cizinci]
    d = target_df[base_cols + available_count + available_rev + extra_ciz].copy()

    group_headers = "".join(
        f"<th colspan='2' style='"
        f"background:{AGE_COLORS[i]};"
        f"color:white;text-align:center;"
        f"border:1px solid rgba(0,0,0,0.15);"
        f"padding:6px 10px;white-space:nowrap;'>"
        f"{g}</th>"
        for i, g in enumerate(AGE_GROUPS) if f'{g}_POCET_KLIENTU' in available_count
    )
    cizinci_th1 = "".join(
        f"<th style='background:{clr};color:white;text-align:center;"
        f"border:1px solid rgba(0,0,0,0.15);padding:6px 10px;white-space:nowrap;'>"
        f"{lbl}</th>"
        for _, lbl, clr in available_cizinci
    )

    sub_headers = "".join(
        f"<th style='background:{AGE_COLORS[i]};opacity:0.85;color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);"
        f"padding:5px 8px;font-size:0.78rem;font-weight:500;'>Počet kl.</th>"
        f"<th style='background:{AGE_COLORS[i]};opacity:0.75;color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);"
        f"padding:5px 8px;font-size:0.78rem;font-weight:500;'>Prům. výnos</th>"
        for i, g in enumerate(AGE_GROUPS) if f'{g}_POCET_KLIENTU' in available_count
    )
    cizinci_th2 = "".join(
        f"<th style='background:{clr};opacity:0.8;color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);"
        f"padding:5px 8px;font-size:0.78rem;font-weight:500;'>počet</th>"
        for _, lbl, clr in available_cizinci
    )

    rows_html = ""
    for i, (_, row) in enumerate(d.iterrows()):
        bg = "#ffffff" if i % 2 == 0 else "#eef4ff"
        td_base = f"border:1px solid #dee2e6;padding:5px 8px;background:{bg};"

        cells = (
            f"<td data-sort='{int(row['BRANCH_CODE'])}' style='{td_base}text-align:center;font-weight:bold;white-space:nowrap;'>"
            f"{int(row['BRANCH_CODE'])}</td>"
            f"<td style='{td_base}text-align:left;white-space:nowrap;'>"
            f"{row['BRANCH_NAME']}</td>"
        )
        for g in AGE_GROUPS:
            cnt_col = f'{g}_POCET_KLIENTU'
            rev_col = f'{g}_PRUMERNY_VYNOS'
            if cnt_col not in available_count:
                continue
            cnt = int(row[cnt_col]) if pd.notna(row.get(cnt_col)) else 0
            rev_raw = float(row[rev_col]) if rev_col in available_rev and pd.notna(row.get(rev_col)) else 0.0
            rev = format_money(rev_raw)
            cnt_fmt = f"{cnt:,}".replace(",", "\u00a0")
            cells += (
                f"<td data-sort='{cnt}' style='{td_base}text-align:center;'>{cnt_fmt}</td>"
                f"<td data-sort='{rev_raw:.2f}' style='{td_base}text-align:center;font-size:0.8rem;color:#555;'>{rev}</td>"
            )
        for col, lbl, clr in available_cizinci:
            val = row.get(col, 0)
            try:
                v = int(float(val)) if pd.notna(val) else 0
            except:
                v = 0
            fmt = f"{v:,}".replace(",", "\u00a0")
            cells += (
                f"<td data-sort='{v}' style='{td_base}text-align:center;font-weight:500;'>{fmt}</td>"
            )
        rows_html += f"<tr>{cells}</tr>"

    tbl_id = table_id
    return f"""
<div class="table-responsive" style="max-height:480px;overflow-y:auto;border-radius:6px;box-shadow:0 2px 6px rgba(0,0,0,0.07);">
  <table id="{tbl_id}" class="table table-bordered table-sm mb-0"
         style="font-size:0.82rem;background:white;">
    <thead style="position:sticky;top:0;z-index:10;">
      <tr>
        <th rowspan='2' data-col='0' style='background:#2770f0;color:white;text-align:center;
            border:1px solid #1e5bc9;padding:6px 10px;vertical-align:middle;cursor:pointer;'
            title='Klikni pro řazení'>
          ID Pobočky <span class='age-sort-arrow' data-arrow='0' style='font-size:0.7rem;opacity:0.4;'>&#x2195;</span></th>
        <th rowspan='2' data-col='1' style='background:#2770f0;color:white;text-align:left;
            border:1px solid #1e5bc9;padding:6px 10px;vertical-align:middle;cursor:pointer;'
            title='Klikni pro řazení'>
          Název pobočky <span class='age-sort-arrow' data-arrow='1' style='font-size:0.7rem;opacity:0.4;'>&#x2195;</span></th>
        {group_headers}
        {cizinci_th1}
      </tr>
      <tr>{sub_headers}{cizinci_th2}</tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
</div>
<script>
(function() {{
  var tbl = document.getElementById("{tbl_id}");
  if (!tbl) return;
  var _ss = {{col: -1, asc: true}};
  // Add sort arrows and click to sub-header row cells (row 2)
  var allTh = Array.from(tbl.querySelectorAll("thead th:not([rowspan='2'])"));
  allTh.forEach(function(th, i) {{
    var ci = i + 2; // offset by 2 fixed cols
    th.setAttribute("data-col", ci);
    th.style.cursor = "pointer";
    th.title = "Klikni pro řazení";
    var arr = document.createElement("span");
    arr.className = "age-sort-arrow";
    arr.setAttribute("data-arrow", ci);
    arr.style.cssText = "font-size:0.7rem;opacity:0.4;pointer-events:none;";
    arr.innerHTML = "&#x2195;";
    th.appendChild(arr);
  }});
  function doSort(colIdx) {{
    var asc = (_ss.col === colIdx) ? !_ss.asc : true;
    _ss = {{col: colIdx, asc: asc}};
    var tbody = tbl.querySelector("tbody");
    var allRows = Array.from(tbody.querySelectorAll("tr"));
    allRows.sort(function(a, b) {{
      var ca = a.cells[colIdx], cb = b.cells[colIdx];
      if (!ca || !cb) return 0;
      var va = ca.getAttribute("data-sort") || ca.textContent.trim();
      var vb = cb.getAttribute("data-sort") || cb.textContent.trim();
      var na = parseFloat(va), nb = parseFloat(vb);
      if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
      return asc ? va.localeCompare(vb,"cs") : vb.localeCompare(va,"cs");
    }});
    allRows.forEach(function(row, i) {{
      var bg = (i%2===0)?"#ffffff":"#eef4ff";
      Array.from(row.cells).forEach(function(td) {{ td.style.background = bg; }});
      tbody.appendChild(row);
    }});
    // Arrows + orange highlight
    tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
      var ci = parseInt(th.getAttribute("data-col"));
      var arr = th.querySelector(".age-sort-arrow");
      if (ci === colIdx) {{
        th.style.setProperty("background","#f39c12","important");
        if (arr) arr.innerHTML = asc ? "&#x2191;" : "&#x2193;";
        if (arr) arr.style.opacity = "1";
      }} else {{
        th.style.removeProperty("background");
        if (arr) arr.innerHTML = "&#x2195;";
        if (arr) arr.style.opacity = "0.4";
      }}
    }});
    // Orange column cells
    var ncols = allRows[0] ? allRows[0].cells.length : 0;
    allRows.forEach(function(row) {{
      Array.from(row.cells).forEach(function(td, ci) {{
        if (ci === colIdx) td.style.background = "rgba(243,156,18,0.15)";
      }});
    }});
  }}
  tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
    th.addEventListener("click", function() {{
      doSort(parseInt(th.getAttribute("data-col")));
    }});
  }});
}})();
</script>"""


def generate_age_apex_chart(df_reg, chart_id):
    labels = AGE_GROUPS
    colors = AGE_COLORS
    values = [
        int(df_reg[f'{g}_POCET_KLIENTU'].sum()) if f'{g}_POCET_KLIENTU' in df_reg.columns else 0
        for g in AGE_GROUPS
    ]

    if sum(values) == 0:
        return "<p style='color:#aaa;font-style:italic;'>Žádná data o věkových skupinách.</p>"

    import json
    labels_js = json.dumps(labels)
    colors_js = json.dumps(colors)
    values_js = str(values)

    return f"""
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3"></script>
<div id="{chart_id}" style="min-height:260px;"></div>
<script>
(function() {{
  var options = {{
    chart: {{
      type: "bar",
      height: 260,
      toolbar: {{ show: false }},
      animations: {{
        enabled: true,
        easing: "easeinout",
        speed: 600,
        animateGradually: {{ enabled: true, delay: 100 }},
        dynamicAnimation: {{ enabled: true, speed: 350 }}
      }},
      background: "transparent"
    }},
    series: [{{ name: "Po\u010det klient\u016f", data: {values_js} }}],
    xaxis: {{
      categories: {labels_js},
      labels: {{ style: {{ fontSize: "12px", colors: "#555" }} }}
    }},
    yaxis: {{
      labels: {{
        formatter: function(v) {{ return v.toLocaleString("cs-CZ"); }},
        style: {{ fontSize: "11px", colors: "#888" }}
      }}
    }},
    colors: {colors_js},
    plotOptions: {{
      bar: {{
        distributed: true,
        borderRadius: 5,
        columnWidth: "55%",
        dataLabels: {{ position: "top" }}
      }}
    }},
    dataLabels: {{
      enabled: true,
      formatter: function(v) {{ return v.toLocaleString("cs-CZ"); }},
      offsetY: -22,
      style: {{ fontSize: "11px", colors: ["#333"], fontWeight: "bold" }}
    }},
    legend: {{ show: false }},
    grid: {{
      borderColor: "#e0e0e0",
      strokeDashArray: 4,
      yaxis: {{ lines: {{ show: true }} }}
    }},
    tooltip: {{
      y: {{ formatter: function(v) {{ return v.toLocaleString("cs-CZ") + " klient\u016f"; }} }}
    }}
  }};
  new ApexCharts(document.getElementById("{chart_id}"), options).render();
}})();
</script>"""


def generate_revenue_curve_chart(df_reg, chart_id):
    """Spojnicový graf průměrného výnosu klienta dle věkové skupiny (agregát za region)."""
    import json as _json

    rev_vals = []
    for g in AGE_GROUPS:
        col = f"{g}_PRUMERNY_VYNOS"
        if col in df_reg.columns:
            val = df_reg[col].mean()
            rev_vals.append(round(float(val), 0) if not __import__("math").isnan(val) else 0)
        else:
            rev_vals.append(0)

    if all(v == 0 for v in rev_vals):
        return "<p style='color:#aaa;font-style:italic;'>Žádná data o výnosech dle věku.</p>"

    labels_js = _json.dumps(AGE_GROUPS)
    values_js = str(rev_vals)

    return f"""
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3"></script>
<div id="{chart_id}" style="min-height:280px;"></div>
<script>
(function() {{
  var options = {{
    chart: {{
      type: "area",
      height: 280,
      toolbar: {{ show: false }},
      animations: {{ enabled: true, speed: 700 }},
      background: "transparent"
    }},
    series: [{{ name: "Pr\u016fm. v\u00fdnos", data: {values_js} }}],
    xaxis: {{
      categories: {labels_js},
      title: {{ text: "V\u011bkov\u00e1 skupina", style: {{ fontSize: "11px", color: "#888" }} }},
      labels: {{ style: {{ fontSize: "12px", colors: "#555" }} }}
    }},
    yaxis: {{
      title: {{ text: "K\u010d / klient (pr\u016fm.)", style: {{ fontSize: "11px", color: "#888" }} }},
      labels: {{
        formatter: function(v) {{
          return (v/1000).toFixed(0) + "\u00a0K\u010d";
        }},
        style: {{ fontSize: "11px", colors: "#888" }}
      }}
    }},
    colors: ["#2770ef"],
    fill: {{
      type: "gradient",
      gradient: {{
        shadeIntensity: 1, opacityFrom: 0.35, opacityTo: 0.05,
        stops: [0, 100]
      }}
    }},
    stroke: {{ curve: "smooth", width: 3 }},
    markers: {{
      size: 6, colors: ["#2770ef"],
      strokeColors: "#fff", strokeWidth: 2, hover: {{ size: 8 }}
    }},
    dataLabels: {{
      enabled: true,
      formatter: function(v) {{ return (v/1000).toFixed(0) + "\u00a0K"; }},
      style: {{ fontSize: "10px", colors: ["#2770ef"], fontWeight: "600" }},
      offsetY: -8,
      background: {{ enabled: false }}
    }},
    grid: {{
      borderColor: "#e0e0e0", strokeDashArray: 4,
      yaxis: {{ lines: {{ show: true }} }}
    }},
    tooltip: {{
      y: {{ formatter: function(v) {{
        return new Intl.NumberFormat("cs-CZ").format(Math.round(v)) + "\u00a0K\u010d";
      }} }}
    }}
  }};
  new ApexCharts(document.getElementById("{chart_id}"), options).render();
}})();
</script>"""


def generate_obchody_benchmark_chart(target_df, chart_id="obch-bench", region_df=None):
    """
    Interaktivní graf produktových skupin:
    - Autocomplete našeptávač pro výběr pobočky
    - ApexCharts sloupcový graf: vybraná pobočka vs průměr regionu vs průměr stejného formátu
    - Přepínač: Počet prodejů 2025 / Objem výnosů 2025
    """
    import json as _json

    d = target_df.copy()
    # Benchmark region = target_df (regionální nebo celá síť)
    d_bench = (region_df if region_df is not None else target_df).copy()

    available_groups = [
        (key, label) for key, label in PRODUCT_GROUPS
        if f'POCET_PRODEJU_{key}_2025' in d.columns
    ]
    if not available_groups:
        return ""

    for key, _ in available_groups:
        for col in [f'POCET_PRODEJU_{key}_2025', f'OBJEM_VYNOSU_CZK_{key}_2025']:
            for df_ in [d, d_bench]:
                if col in df_.columns:
                    df_[col] = pd.to_numeric(df_[col], errors='coerce').fillna(0)

    group_keys   = [g[0] for g in available_groups]
    group_labels = [g[1] for g in available_groups]

    # ── Průměr regionu ─────────────────────────────────────────────
    reg_cnt_avg = {k: float(d_bench[f'POCET_PRODEJU_{k}_2025'].mean())
                   if f'POCET_PRODEJU_{k}_2025' in d_bench.columns else 0.0
                   for k in group_keys}
    reg_obj_avg = {k: float(d_bench[f'OBJEM_VYNOSU_CZK_{k}_2025'].mean())
                   if f'OBJEM_VYNOSU_CZK_{k}_2025' in d_bench.columns else 0.0
                   for k in group_keys}

    # ── Průměry dle formátu pobočky ────────────────────────────────
    fmt_col = next((c for c in ['BRANCH_FORMAT', 'BRANCH_TYPE_NAME'] if c in d.columns), None)
    fmt_cnt_avgs = {}
    fmt_obj_avgs = {}
    if fmt_col and fmt_col in d_bench.columns:
        for fmt_val, grp in d_bench.groupby(fmt_col, observed=True):
            fv = str(fmt_val)
            fmt_cnt_avgs[fv] = {k: float(grp[f'POCET_PRODEJU_{k}_2025'].mean())
                                 if f'POCET_PRODEJU_{k}_2025' in grp.columns else 0.0
                                 for k in group_keys}
            fmt_obj_avgs[fv] = {k: float(grp[f'OBJEM_VYNOSU_CZK_{k}_2025'].mean())
                                 if f'OBJEM_VYNOSU_CZK_{k}_2025' in grp.columns else 0.0
                                 for k in group_keys}

    # ── Data pro každou pobočku ────────────────────────────────────
    branches_data = {}
    for _, row in d.iterrows():
        bc  = int(row.get('BRANCH_CODE', 0))
        bn  = str(row.get('BRANCH_NAME', bc))
        fmt = str(row.get(fmt_col, '')) if fmt_col else ''
        branches_data[bc] = {
            'name': bn, 'format': fmt,
            'cnt': {k: float(row.get(f'POCET_PRODEJU_{k}_2025', 0) or 0) for k in group_keys},
            'obj': {k: float(row.get(f'OBJEM_VYNOSU_CZK_{k}_2025', 0) or 0) for k in group_keys},
        }

    branches_list = sorted(branches_data.items(), key=lambda x: x[1]['name'])  # (bc, {...})

    data_json     = _json.dumps(branches_data)
    reg_cnt_json  = _json.dumps(reg_cnt_avg)
    reg_obj_json  = _json.dumps(reg_obj_avg)
    fmt_cnt_json  = _json.dumps(fmt_cnt_avgs)
    fmt_obj_json  = _json.dumps(fmt_obj_avgs)
    keys_json     = _json.dumps(group_keys)
    labels_json   = _json.dumps(group_labels)

    # Datalist options
    dl_options = "".join(
        f'<option value="{row["name"]} ({bc})" data-bc="{bc}">'
        for bc, row in branches_list
    )

    return f"""
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3"></script>
<div id="obch-bench-wrap-{chart_id}"
     style="background:#fff;border-radius:10px;padding:16px 18px;
            border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);margin-top:16px;">
  <div style="font-size:0.72rem;color:#2770f0;font-weight:700;text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:8px;">
    📊 Interaktivní benchmark — produktové skupiny
  </div>
  <div style="font-size:0.75rem;color:#666;margin-bottom:12px;">
    Vyber pobočku pro srovnání výkonu dle produktové skupiny vůči průměru regionu
    a průměru poboček stejného formátu.
  </div>

  <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:12px;">
    <input id="obch-search-{chart_id}" type="text" list="obch-dl-{chart_id}"
           placeholder="🔍 Vyhledej pobočku…"
           style="padding:7px 12px;border:1px solid #ccc;border-radius:6px;
                  font-size:0.88rem;width:300px;outline:none;">
    <datalist id="obch-dl-{chart_id}">{dl_options}</datalist>

    <div style="display:flex;gap:6px;">
      <button id="obch-metric-cnt-{chart_id}"
              style="padding:5px 14px;border:2px solid #2770f0;border-radius:6px;
                     background:#2770f0;color:white;font-size:0.82rem;font-weight:600;cursor:pointer;">
        Počet prodejů
      </button>
      <button id="obch-metric-obj-{chart_id}"
              style="padding:5px 14px;border:2px solid #ccc;border-radius:6px;
                     background:#f5f5f5;color:#555;font-size:0.82rem;font-weight:600;cursor:pointer;">
        Objem výnosů
      </button>
    </div>
    <span id="obch-fmt-info-{chart_id}"
          style="font-size:0.78rem;color:#888;font-style:italic;"></span>
  </div>

  <div id="obch-chart-{chart_id}" style="min-height:340px;"></div>
  <div id="obch-nosel-{chart_id}"
       style="text-align:center;color:#aaa;font-style:italic;padding:60px 0;font-size:0.9rem;">
    ← Vyber pobočku pro zobrazení grafu
  </div>
</div>

<script>
(function() {{
  var branchData  = {data_json};
  var regCntAvg   = {reg_cnt_json};
  var regObjAvg   = {reg_obj_json};
  var fmtCntAvgs  = {fmt_cnt_json};
  var fmtObjAvgs  = {fmt_obj_json};
  var gKeys       = {keys_json};
  var gLabels     = {labels_json};

  var searchEl    = document.getElementById("obch-search-{chart_id}");
  var chartEl     = document.getElementById("obch-chart-{chart_id}");
  var noSelEl     = document.getElementById("obch-nosel-{chart_id}");
  var fmtInfoEl   = document.getElementById("obch-fmt-info-{chart_id}");
  var btnCnt      = document.getElementById("obch-metric-cnt-{chart_id}");
  var btnObj      = document.getElementById("obch-metric-obj-{chart_id}");
  if (!searchEl || !chartEl) return;

  var currentMetric = "cnt";  // "cnt" | "obj"
  var currentBc     = null;
  var apexChart     = null;

  function fmtMoney(v) {{
    if (v >= 1e9) return (v/1e9).toFixed(1) + " mld Kč";
    if (v >= 1e6) return (v/1e6).toFixed(1) + " M Kč";
    if (v >= 1e3) return (v/1e3).toFixed(0) + " K Kč";
    return Math.round(v) + " Kč";
  }}

  function buildSeries(bc) {{
    var b    = branchData[bc];
    var avg  = currentMetric === "cnt" ? regCntAvg  : regObjAvg;
    var fAvg = currentMetric === "cnt" ? fmtCntAvgs : fmtObjAvgs;
    var bData  = currentMetric === "cnt" ? b.cnt  : b.obj;
    var hasFmt = b.format && fAvg[b.format];
    var branchVals = gKeys.map(function(k) {{ return Math.round(bData[k] || 0); }});
    var regVals    = gKeys.map(function(k) {{ return Math.round(avg[k]  || 0); }});
    var fmtVals    = hasFmt ? gKeys.map(function(k) {{ return Math.round(fAvg[b.format][k] || 0); }}) : null;

    var series = [
      {{ name: b.name,          data: branchVals, color: "#2770f0" }},
      {{ name: "⌀ Průměr regionu", data: regVals,    color: "#94a3b8" }},
    ];
    if (fmtVals) {{
      series.push({{ name: "⌀ Průměr formátu (" + b.format + ")", data: fmtVals, color: "#22c55e" }});
    }}
    return series;
  }}

  function renderChart(bc) {{
    var b = branchData[bc];
    if (!b) return;
    var series = buildSeries(bc);
    var isObj  = currentMetric === "obj";
    var opts = {{
      chart:  {{ type: "bar", height: 340, toolbar: {{ show: false }},
                 animations: {{ enabled: true, speed: 280 }} }},
      series: series,
      xaxis:  {{ categories: gLabels,
                 labels:     {{ style: {{ fontSize: "11px" }}, rotate: -25 }} }},
      yaxis:  {{ labels: {{ formatter: isObj
                  ? function(v) {{ return fmtMoney(v); }}
                  : function(v) {{ return Math.round(v).toLocaleString("cs-CZ"); }} }} }},
      plotOptions: {{ bar: {{ borderRadius: 4, columnWidth: "68%",
                              dataLabels: {{ position: "top" }} }} }},
      dataLabels: {{
        enabled: true,
        formatter: isObj
          ? function(v) {{ return v >= 1e6 ? (v/1e6).toFixed(1)+"M" : v >= 1e3 ? (v/1e3).toFixed(0)+"k" : Math.round(v); }}
          : function(v) {{ return Math.round(v).toLocaleString("cs-CZ"); }},
        style: {{ fontSize: "10px", colors: ["#334"] }},
        offsetY: -4,
      }},
      legend:  {{ position: "top" }},
      tooltip: {{ y: {{ formatter: isObj
                  ? function(v) {{ return fmtMoney(v); }}
                  : function(v) {{ return Math.round(v).toLocaleString("cs-CZ") + " prodejů"; }} }} }},
      grid:    {{ borderColor: "#f0f0f0" }},
    }};

    noSelEl.style.display = "none";
    chartEl.style.display = "";

    if (apexChart) {{
      apexChart.updateOptions(opts, true, true);
    }} else {{
      apexChart = new ApexCharts(chartEl, opts);
      apexChart.render();
    }}

    var hasFmt = b.format && (currentMetric === "cnt" ? fmtCntAvgs : fmtObjAvgs)[b.format];
    fmtInfoEl.textContent = b.format
      ? (hasFmt ? "Formát: " + b.format : "Formát: " + b.format + " (benchmark nedostupný)")
      : "";
  }}

  function parseBc(val) {{
    var m = val.match(/\((\d+)\)\s*$/);
    return m ? parseInt(m[1]) : null;
  }}

  chartEl.style.display = "none";

  searchEl.addEventListener("input", function() {{
    var bc = parseBc(searchEl.value);
    if (bc && branchData[bc]) {{ currentBc = bc; renderChart(bc); }}
  }});
  searchEl.addEventListener("change", function() {{
    var bc = parseBc(searchEl.value);
    if (bc && branchData[bc]) {{ currentBc = bc; renderChart(bc); }}
  }});

  function setMetric(m) {{
    currentMetric = m;
    if (m === "cnt") {{
      btnCnt.style.background = "#2770f0"; btnCnt.style.color = "white"; btnCnt.style.borderColor = "#2770f0";
      btnObj.style.background = "#f5f5f5"; btnObj.style.color = "#555";  btnObj.style.borderColor = "#ccc";
    }} else {{
      btnObj.style.background = "#2770f0"; btnObj.style.color = "white"; btnObj.style.borderColor = "#2770f0";
      btnCnt.style.background = "#f5f5f5"; btnCnt.style.color = "#555";  btnCnt.style.borderColor = "#ccc";
    }}
    if (currentBc) renderChart(currentBc);
  }}
  btnCnt.addEventListener("click", function() {{ setMetric("cnt"); }});
  btnObj.addEventListener("click", function() {{ setMetric("obj"); }});
}})();
</script>"""


def generate_obchody_table(target_df, table_id="obchody-tbl"):
    import json as _json

    all_cnt_cols = [f'POCET_PRODEJU_{k}_2024' for k, _ in PRODUCT_GROUPS] + \
                   [f'POCET_PRODEJU_{k}_2025' for k, _ in PRODUCT_GROUPS]
    available_cnt = [c for c in all_cnt_cols if c in target_df.columns]
    if not available_cnt:
        return "<p style='color:#aaa;font-style:italic;'>Žádná data o obchodech.</p>"

    d = target_df.copy()
    for key, _ in PRODUCT_GROUPS:
        for col in [f'POCET_PRODEJU_{key}_2024', f'POCET_PRODEJU_{key}_2025',
                    f'OBJEM_VYNOSU_CZK_{key}_2024', f'OBJEM_VYNOSU_CZK_{key}_2025']:
            if col in d.columns:
                d[col] = pd.to_numeric(d[col], errors='coerce').fillna(0)

    for key, _ in PRODUCT_GROUPS:
        col25 = f'POCET_PRODEJU_{key}_2025'
        qcol  = f'POCET_PRODEJU_{key}_2025_Q'
        if col25 in d.columns and d[col25].sum() > 0:
            try:
                d[qcol] = 6 - (pd.qcut(d[col25], q=5, labels=False, duplicates='drop') + 1)
            except Exception:
                d[qcol] = np.nan
        else:
            d[qcol] = np.nan

    prod_group_colors = [
        '#4a90d9', '#45b065', '#9b6bbf', '#d4682a',
        '#2aacbd', '#b89a00', '#a83250', '#4a8c4a', '#7a5c3a'
    ]
    prod_group_pastel = [
        '#dceeff', '#d4f5e0', '#eeddf7', '#fce8d8',
        '#d4f0f5', '#faf0cc', '#fad8e0', '#d8edd8', '#ede0d4'
    ]

    available_groups = [
        (i, key, label, prod_group_colors[i % len(prod_group_colors)],
         prod_group_pastel[i % len(prod_group_pastel)])
        for i, (key, label) in enumerate(PRODUCT_GROUPS)
        if f'POCET_PRODEJU_{key}_2024' in d.columns
    ]

    th_groups = "".join(
        f"<th colspan='7' data-group='{key}' style='background:{color};color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.2);padding:5px 8px;white-space:nowrap;'>"
        f"{label}</th>"
        for i, key, label, color, pastel in available_groups
    )

    th_sub = "".join(
        f"<th data-group='{key}' style='background:{color};filter:brightness(1.15);color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);padding:4px 5px;font-size:0.74rem;font-weight:500;"
        f"white-space:nowrap;'>Poč. 24</th>"
        f"<th data-group='{key}' style='background:{color};filter:brightness(1.15);color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);padding:4px 5px;font-size:0.74rem;font-weight:500;"
        f"white-space:nowrap;'>Poč. 25</th>"
        f"<th data-group='{key}' style='background:{color};filter:brightness(1.3);color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);padding:4px 5px;font-size:0.74rem;font-weight:600;"
        f"white-space:nowrap;'>Δ Poč.</th>"
        f"<th data-group='{key}' style='background:{color};filter:brightness(1.4);color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);padding:4px 5px;font-size:0.74rem;font-weight:600;"
        f"white-space:nowrap;'>Kvint.</th>"
        f"<th data-group='{key}' style='background:{color};filter:brightness(1.15);color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);padding:4px 5px;font-size:0.74rem;font-weight:500;"
        f"white-space:nowrap;'>Obj. 24</th>"
        f"<th data-group='{key}' style='background:{color};filter:brightness(1.15);color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);padding:4px 5px;font-size:0.74rem;font-weight:500;"
        f"white-space:nowrap;'>Obj. 25</th>"
        f"<th data-group='{key}' style='background:{color};filter:brightness(1.3);color:white;"
        f"text-align:center;border:1px solid rgba(0,0,0,0.15);padding:4px 5px;font-size:0.74rem;font-weight:600;"
        f"white-space:nowrap;'>Δ Obj.</th>"
        for i, key, label, color, pastel in available_groups
    )

    def diff_html_num(val):
        try:
            v = int(round(float(val)))
        except (ValueError, TypeError):
            return "—"
        if v == 0:
            return "<span style='color:#aaa;'>0</span>"
        color = "#0ab33f" if v > 0 else "#e64343"
        sign  = "+" if v > 0 else ""
        return f"<b style='color:{color};'>{sign}{v:,}</b>"

    def diff_html_money(val):
        try:
            v = float(val)
        except (ValueError, TypeError):
            return "—"
        if v == 0:
            return "<span style='color:#aaa;'>0\u00a0Kč</span>"
        color = "#0ab33f" if v > 0 else "#e64343"
        sign  = "+" if v > 0 else ""
        txt   = f"{int(v):,}".replace(",", "\u00a0") + "\u00a0Kč"
        return f"<b style='color:{color};'>{sign}{txt}</b>"

    rows_html = ""
    for i, (_, row) in enumerate(d.iterrows()):
        bg = "#ffffff" if i % 2 == 0 else "#eef4ff"
        td = f"border:1px solid #dee2e6;padding:4px 6px;background:{bg};"

        cells = (
            f"<td data-group='__id__' data-sort='{int(row['BRANCH_CODE'])}' style='{td}text-align:center;font-weight:bold;'>{int(row['BRANCH_CODE'])}</td>"
            f"<td data-group='__id__' style='{td}text-align:left;white-space:nowrap;'>{row['BRANCH_NAME']}</td>"
        )
        for gi, key, label, color, pastel in available_groups:
            cnt24_col = f'POCET_PRODEJU_{key}_2024'
            cnt25_col = f'POCET_PRODEJU_{key}_2025'
            obj24_col = f'OBJEM_VYNOSU_CZK_{key}_2024'
            obj25_col = f'OBJEM_VYNOSU_CZK_{key}_2025'
            qcol      = f'POCET_PRODEJU_{key}_2025_Q'

            cnt24 = row.get(cnt24_col, 0) or 0
            cnt25 = row.get(cnt25_col, 0) or 0
            obj24 = row.get(obj24_col, 0) or 0
            obj25 = row.get(obj25_col, 0) or 0
            q_val = row.get(qcol, np.nan)

            cnt24_fmt = f"{int(cnt24):,}".replace(",", "\u00a0")
            cnt25_fmt = f"{int(cnt25):,}".replace(",", "\u00a0")
            obj24_fmt = f"{int(obj24):,}".replace(",", "\u00a0") + "\u00a0Kč"
            obj25_fmt = f"{int(obj25):,}".replace(",", "\u00a0") + "\u00a0Kč"
            dot_html  = get_quintile_dot(q_val, tooltip_prefix="Prodeje kvintil") if pd.notna(q_val) else "<span style='color:#ccc;'>●</span>"
            dcnt = cnt25 - cnt24
            dobj = obj25 - obj24

            cells += (
                f"<td data-group='{key}' data-sort='{int(cnt24)}' style='{td}text-align:right;color:#555;font-size:0.8rem;'>{cnt24_fmt}</td>"
                f"<td data-group='{key}' data-sort='{int(cnt25)}' style='{td}text-align:right;font-size:0.8rem;'>{cnt25_fmt}</td>"
                f"<td data-group='{key}' data-sort='{dcnt}' style='{td}text-align:right;'>{diff_html_num(dcnt)}</td>"
                f"<td data-group='{key}' data-sort='{int(q_val) if pd.notna(q_val) else 99}' style='{td}text-align:center;'>{dot_html}</td>"
                f"<td data-group='{key}' data-sort='{int(obj24)}' style='{td}text-align:right;color:#555;font-size:0.78rem;'>{obj24_fmt}</td>"
                f"<td data-group='{key}' data-sort='{int(obj25)}' style='{td}text-align:right;font-size:0.78rem;'>{obj25_fmt}</td>"
                f"<td data-group='{key}' data-sort='{dobj}' style='{td}text-align:right;'>{diff_html_money(dobj)}</td>"
            )
        rows_html += f"<tr>{cells}</tr>"

    groups_json = _json.dumps([
        {"key": key, "label": label, "color": color, "pastel": pastel}
        for i, key, label, color, pastel in available_groups
    ])

    return f"""
<style>
  #outer-{table_id} tbody tr:hover td {{ background-color: #d1fae5 !important; }}
  #outer-{table_id} .prod-toggle-bar {{
    display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 10px;
    padding: 10px 12px; background: #f8faff; border: 1px solid #dde4f5; border-radius: 8px;
  }}
  #outer-{table_id} .prod-toggle-btn {{
    display: flex; align-items: center; gap: 5px;
    padding: 4px 12px; border-radius: 20px; border: 2px solid transparent;
    font-size: 0.82rem; cursor: pointer; user-select: none;
    transition: opacity 0.15s, border-color 0.15s;
    color: #333; font-weight: 600;
  }}
  #outer-{table_id} .prod-toggle-btn.inactive {{
    opacity: 0.35; border-color: #ccc !important;
  }}
  #outer-{table_id} .prod-toggle-btn input {{
    width: 13px; height: 13px; cursor: pointer; margin: 0;
  }}
  #outer-{table_id} .prod-toggle-all {{
    padding: 4px 12px; border-radius: 20px; border: 1px solid #aaa;
    background: #f0f0f0; font-size: 0.82rem; cursor: pointer; font-weight: 600; color: #333;
  }}
</style>

<div id="outer-{table_id}">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">
    <input id="ob-search-{table_id}" type="text"
           placeholder="🔍 Filtruj název nebo kód pobočky..."
           style="padding:6px 12px;border:1px solid #ccc;border-radius:6px;
                  font-size:0.88rem;width:280px;outline:none;
                  box-shadow:inset 0 1px 3px rgba(0,0,0,.08);">
    <button id="ob-search-clear-{table_id}"
            style="padding:5px 12px;border:1px solid #ccc;border-radius:6px;
                   background:#f5f5f5;font-size:0.85rem;cursor:pointer;">✖ Zrušit</button>
    <span id="ob-search-count-{table_id}" style="font-size:0.82rem;color:#888;"></span>
  </div>
  <div style="font-size:0.82rem;color:#555;font-weight:600;margin-bottom:5px;">
    📦 Zobrazit produktové skupiny:
  </div>
  <div class="prod-toggle-bar" id="prod-toggles-{table_id}">
    <button class="prod-toggle-all" id="prod-all-{table_id}">✔ Vše</button>
    <button class="prod-toggle-all" id="prod-none-{table_id}">✕ Nic</button>
  </div>

  <div class="table-responsive" style="max-height:520px;overflow:auto;border-radius:6px;
       box-shadow:0 2px 6px rgba(0,0,0,0.07);">
    <table class="table table-bordered table-sm mb-0"
           id="{table_id}" style="font-size:0.81rem;background:white;min-width:max-content;">
      <thead style="position:sticky;top:0;z-index:10;">
        <tr>
          <th rowspan='2' data-group='__id__' style='background:#2770f0;color:white;text-align:center;
              border:1px solid #1e5bc9;padding:6px 10px;vertical-align:middle;'>ID&nbsp;Pobočky</th>
          <th rowspan='2' data-group='__id__' style='background:#2770f0;color:white;text-align:left;
              border:1px solid #1e5bc9;padding:6px 10px;vertical-align:middle;'>Název pobočky</th>
          {th_groups}
        </tr>
        <tr>{th_sub}</tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
</div>

<script>
(function() {{
  var groups = {groups_json};
  var bar    = document.getElementById("prod-toggles-{table_id}");
  var tbl    = document.getElementById("{table_id}");
  var btnAll  = document.getElementById("prod-all-{table_id}");
  var btnNone = document.getElementById("prod-none-{table_id}");
  if (!tbl || !bar) return;

  var active = {{}};
  groups.forEach(function(g) {{ active[g.key] = true; }});

  groups.forEach(function(g) {{
    var btn = document.createElement("label");
    btn.className = "prod-toggle-btn";
    btn.style.background = g.pastel;
    btn.style.borderColor = g.color;
    btn.title = "Zobrazit / skrýt: " + g.label;

    var cb = document.createElement("input");
    cb.type = "checkbox";
    cb.checked = true;
    cb.style.accentColor = g.color;
    cb.addEventListener("change", function() {{
      active[g.key] = cb.checked;
      btn.className = "prod-toggle-btn" + (cb.checked ? "" : " inactive");
      applyVisibility();
    }});

    btn.appendChild(cb);
    btn.appendChild(document.createTextNode(" " + g.label));
    bar.appendChild(btn);
  }});

  function applyVisibility() {{
    tbl.querySelectorAll("thead tr:first-child th[data-group]").forEach(function(th) {{
      var g = th.getAttribute("data-group");
      if (g === "__id__") return;
      th.style.display = active[g] ? "" : "none";
    }});
    tbl.querySelectorAll("thead tr:nth-child(2) th[data-group]").forEach(function(th) {{
      var g = th.getAttribute("data-group");
      th.style.display = active[g] ? "" : "none";
    }});
    tbl.querySelectorAll("tbody tr").forEach(function(row) {{
      row.querySelectorAll("td[data-group]").forEach(function(td) {{
        var g = td.getAttribute("data-group");
        if (g === "__id__") return;
        td.style.display = active[g] ? "" : "none";
      }});
    }});
  }}

  btnAll.addEventListener("click", function() {{
    groups.forEach(function(g) {{ active[g.key] = true; }});
    bar.querySelectorAll("input[type=checkbox]").forEach(function(cb) {{ cb.checked = true; }});
    bar.querySelectorAll(".prod-toggle-btn").forEach(function(b) {{
      b.className = "prod-toggle-btn";
    }});
    applyVisibility();
  }});

  btnNone.addEventListener("click", function() {{
    groups.forEach(function(g) {{ active[g.key] = false; }});
    bar.querySelectorAll("input[type=checkbox]").forEach(function(cb) {{ cb.checked = false; }});
    bar.querySelectorAll(".prod-toggle-btn").forEach(function(b) {{
      b.className = "prod-toggle-btn inactive";
    }});
    applyVisibility();
  }});

  // ─── Řazení dle sloupce ──────────────────────────────────
  var _obSS = {{col: -1, asc: true}};

  // Přidej šipky do sub-header řádku (thead tr:nth-child(2))
  var subHeaderRow = tbl.querySelector("thead tr:nth-child(2)");
  if (subHeaderRow) {{
    Array.from(subHeaderRow.querySelectorAll("th")).forEach(function(th, i) {{
      // skutečný index sloupce v tabulce = i + 2 (ID + Název jsou rowspan=2)
      var ci = i + 2;
      th.setAttribute("data-col", ci);
      th.style.cursor = "pointer";
      th.title = "Klikni pro řazení";
      var arr = document.createElement("span");
      arr.className = "ob-sort-arrow";
      arr.setAttribute("data-arrow", ci);
      arr.style.cssText = "font-size:0.65rem;opacity:0.4;pointer-events:none;display:block;";
      arr.innerHTML = "&#x2195;";
      th.appendChild(arr);
    }});
  }}
  // Přidej sort i na první 2 hlavičky (ID, Název) z prvního řádku
  var firstHeaderRow = tbl.querySelector("thead tr:first-child");
  if (firstHeaderRow) {{
    Array.from(firstHeaderRow.querySelectorAll("th[rowspan='2']")).forEach(function(th, i) {{
      th.setAttribute("data-col", i);
      th.style.cursor = "pointer";
      th.title = "Klikni pro řazení";
      var arr = document.createElement("span");
      arr.className = "ob-sort-arrow";
      arr.setAttribute("data-arrow", i);
      arr.style.cssText = "font-size:0.65rem;opacity:0.4;pointer-events:none;display:block;";
      arr.innerHTML = "&#x2195;";
      th.appendChild(arr);
    }});
  }}

  function doObSort(colIdx) {{
    var asc = (_obSS.col === colIdx) ? !_obSS.asc : true;
    _obSS = {{col: colIdx, asc: asc}};
    var tbody = tbl.querySelector("tbody");
    var allRows = Array.from(tbody.querySelectorAll("tr"));
    allRows.sort(function(a, b) {{
      var ca = a.cells[colIdx], cb = b.cells[colIdx];
      if (!ca || !cb) return 0;
      var va = ca.getAttribute("data-sort") || ca.textContent.trim();
      var vb = cb.getAttribute("data-sort") || cb.textContent.trim();
      var na = parseFloat(va), nb = parseFloat(vb);
      if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
      return asc ? va.localeCompare(vb,"cs") : vb.localeCompare(va,"cs");
    }});
    allRows.forEach(function(row, i) {{
      var bg = (i%2===0)?"#ffffff":"#eef4ff";
      // Resetuj background, pak obarvi (respektuj group colors)
      tbody.appendChild(row);
    }});
    // Šipky v obou header řádcích
    tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
      var ci = parseInt(th.getAttribute("data-col"));
      var arr = th.querySelector(".ob-sort-arrow");
      if (ci === colIdx) {{
        th.style.setProperty("background","#f39c12","important");
        if (arr) {{ arr.innerHTML = asc ? "&#x2191;" : "&#x2193;"; arr.style.opacity = "1"; }}
      }} else {{
        th.style.removeProperty("background");
        if (arr) {{ arr.innerHTML = "&#x2195;"; arr.style.opacity = "0.4"; }}
      }}
    }});
    // Orange buňky v aktivním sloupci
    allRows.forEach(function(row) {{
      Array.from(row.cells).forEach(function(td, ci) {{
        if (ci === colIdx) td.style.setProperty("background","rgba(243,156,18,0.2)","important");
        else td.style.removeProperty("background");
      }});
    }});
  }}

  tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
    th.addEventListener("click", function() {{
      doObSort(parseInt(th.getAttribute("data-col")));
    }});
  }});
  // ──────────────────────────────────────────────────────────

  // ─── Vyhledávání dle pobočky ────────────────────────────────────────────
  var obSearchInput = document.getElementById("ob-search-{table_id}");
  var obSearchClear = document.getElementById("ob-search-clear-{table_id}");
  var obSearchCount = document.getElementById("ob-search-count-{table_id}");

  function applyObSearch() {{
    var q = obSearchInput ? obSearchInput.value.trim().toLowerCase() : '';
    var tbody = tbl.querySelector("tbody");
    var rows  = Array.from(tbody.querySelectorAll("tr"));
    var visible = 0;
    rows.forEach(function(row) {{
      if (!q) {{ row.style.display = ''; visible++; return; }}
      var cells = Array.from(row.querySelectorAll("td[data-group='__id__']"));
      var text  = cells.map(function(c) {{ return c.textContent.trim().toLowerCase(); }}).join(' ');
      var match = text.indexOf(q) !== -1;
      row.style.display = match ? '' : 'none';
      if (match) visible++;
    }});
    if (obSearchCount) {{
      obSearchCount.textContent = q ? ('Zobrazeno: ' + visible + ' z ' + rows.length) : '';
    }}
  }}

  if (obSearchInput) {{
    obSearchInput.addEventListener("input", applyObSearch);
    obSearchInput.addEventListener("focus",  function() {{ obSearchInput.style.borderColor = '#2770f0'; }});
    obSearchInput.addEventListener("blur",   function() {{ obSearchInput.style.borderColor = '#ccc'; }});
  }}
  if (obSearchClear) {{
    obSearchClear.addEventListener("click", function() {{
      if (obSearchInput) {{ obSearchInput.value = ''; }}
      applyObSearch();
    }});
  }}
  // ────────────────────────────────────────────────────────────────────────

}})();
</script>"""


def generate_branch_pie_chart(df, chart_id):
    """Vrátí HTML s dvěma koláčovými grafy: dle regionu a dle typologie."""
    import json as _json

    # ── Dle regionu ──────────────────────────────────────────────────
    reg_counts = df['REGION_NAME'].value_counts().dropna()
    reg_labels = _json.dumps(list(reg_counts.index.astype(str)))
    reg_values = _json.dumps([int(v) for v in reg_counts.values])

    # ── Dle typologie ─────────────────────────────────────────────────
    typ_counts = df['BRANCH_TYPE_NAME'].fillna('Neznámá').value_counts()
    typ_labels = _json.dumps(list(typ_counts.index.astype(str)))
    typ_values = _json.dumps([int(v) for v in typ_counts.values])

    reg_id  = f"{chart_id}-reg"
    typ_id  = f"{chart_id}-typ"

    # ── Cash / Cashless ──────────────────────────────────────────────
    cash_counts = (
        df['CASHLESS'].fillna('N')
        .map(lambda v: 'Cashless' if str(v).strip().upper() in ('Y','YES','ANO','1','TRUE') else 'Hotovostní')
        .value_counts()
    ) if 'CASHLESS' in df.columns else _json.loads('{}') or {}
    cash_labels = _json.dumps(list(cash_counts.index.astype(str)) if hasattr(cash_counts, 'index') else [])
    cash_values = _json.dumps([int(v) for v in cash_counts.values] if hasattr(cash_counts, 'values') else [])
    cash_id = f"{chart_id}-cash"

    # ── Vlastnictví ───────────────────────────────────────────────────
    vlastnictvi_id = f"{chart_id}-vlast"
    has_vlast = 'VLASTNICTVI' in df.columns
    if has_vlast:
        vlast_counts = (
            df['VLASTNICTVI'].fillna('?')
            .map(lambda v: {
                'L': 'Pronajatá', 'l': 'Pronajatá',
                'O': 'Vlastní',   'o': 'Vlastní',
            }.get(str(v).strip(), str(v).strip() if str(v).strip() not in ('','nan','None') else 'Nezadáno'))
            .value_counts()
        )
        vlast_labels = _json.dumps(list(vlast_counts.index.astype(str)))
        vlast_values = _json.dumps([int(v) for v in vlast_counts.values])
    else:
        vlast_labels, vlast_values = '[]', '[]'

    return f"""
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3"></script>
<div style="display:flex;gap:16px;flex-wrap:wrap;">
  <div style="flex:1;min-width:220px;">
    <div style="font-size:0.82rem;font-weight:600;color:#555;margin-bottom:6px;">🗺️ Dle regionu</div>
    <div id="{reg_id}"></div>
  </div>
  <div style="flex:1;min-width:220px;">
    <div style="font-size:0.82rem;font-weight:600;color:#555;margin-bottom:6px;">🏢 Dle typologie</div>
    <div id="{typ_id}"></div>
  </div>
  <div style="flex:1;min-width:220px;">
    <div style="font-size:0.82rem;font-weight:600;color:#555;margin-bottom:6px;">💵 Cash / Cashless</div>
    <div id="{cash_id}"></div>
  </div>
  <div style="flex:1;min-width:220px;">
    <div style="font-size:0.82rem;font-weight:600;color:#555;margin-bottom:6px;">🏠 Vlastnictví pobočky</div>
    <div id="{vlastnictvi_id}"></div>
  </div>
</div>
<script>
(function() {{
  var pieOpts = function(elId, labels, values, colors) {{
    return {{
      chart: {{ type: 'pie', height: 240, id: elId,
                animations: {{ enabled: true, speed: 600 }},
                toolbar: {{ show: false }} }},
      series: values,
      labels: labels,
      legend: {{ position: 'bottom', fontSize: '11px',
                 itemMargin: {{ horizontal: 4, vertical: 2 }} }},
      dataLabels: {{ enabled: true, style: {{ fontSize: '11px' }},
                    formatter: function(val, opts) {{
                      return opts.w.config.series[opts.seriesIndex];
                    }} }},
      tooltip: {{ y: {{ formatter: function(v) {{ return v + ' poboček'; }} }} }},
      colors: colors || ['#2770ef','#0bb440','#fd6230','#00a3a5','#f9a825','#7b1fa2','#e64343','#0097a7'],
    }};
  }};
  new ApexCharts(document.getElementById('{reg_id}'),  pieOpts('{reg_id}',  {reg_labels},  {reg_values})).render();
  new ApexCharts(document.getElementById('{typ_id}'),  pieOpts('{typ_id}',  {typ_labels},  {typ_values})).render();
  new ApexCharts(document.getElementById('{cash_id}'), pieOpts('{cash_id}', {cash_labels}, {cash_values}, ['#0bb440','#2770ef'])).render();
  new ApexCharts(document.getElementById('{vlastnictvi_id}'), pieOpts('{vlastnictvi_id}', {vlast_labels}, {vlast_values}, ['#e65100','#1565c0','#aaa'])).render();
}})();
</script>"""


def generate_top_investments_table(df, table_id="inv-top"):
    """Top 5 nejnovějších a top 5 největších investic."""
    import math

    needed = ['BRANCH_CODE', 'BRANCH_NAME', 'DESC_LAST_INV', 'DATUM_LAST_INV',
              'LAST_INV', 'DESC_BIGGEST_INV', 'DATUM_BIGGEST_INV', 'BIGGEST_INV']
    missing = [c for c in needed if c not in df.columns]
    if missing:
        return f"<p style='color:#aaa;font-style:italic;'>Chybí sloupce: {missing}</p>"

    def _parse_inv(df_in, date_col, val_col, label, color, n=5):
        d = df_in.copy()
        d[val_col]  = pd.to_numeric(d[val_col],  errors='coerce')
        d[date_col] = pd.to_datetime(d[date_col], errors='coerce')
        d = d.dropna(subset=[val_col]).query(f"`{val_col}` > 0")
        sort_by_date = (label == "nejnovější")
        if sort_by_date:
            d = d.dropna(subset=[date_col]).sort_values(date_col, ascending=False).head(n)
        else:
            d = d.sort_values(val_col, ascending=False).head(n)

        # Oranžové zvýraznění: datum pro nejnovější, výše pro nejdražší
        HL   = "background:rgba(243,156,18,0.20);font-weight:600;"
        HL_H = "background:#f39c12!important;color:white!important;"

        rows_html = ""
        for _, r in d.iterrows():
            date_str = r[date_col].strftime("%d.%m.%Y") if pd.notna(r[date_col]) else "—"
            val_str  = format_money(r[val_col]) if r[val_col] > 0 else "—"
            desc     = str(r.get("DESC_LAST_INV" if sort_by_date else "DESC_BIGGEST_INV", "")).strip() or "—"
            td_date = f"padding:4px 8px;border:1px solid #dee2e6;white-space:nowrap;{HL if sort_by_date else ''}"
            td_val  = f"padding:4px 8px;border:1px solid #dee2e6;text-align:right;color:{color};white-space:nowrap;{HL if not sort_by_date else 'font-weight:600;'}"
            rows_html += f"""<tr>
              <td style='padding:4px 8px;border:1px solid #dee2e6;font-weight:600;white-space:nowrap;'>{int(r['BRANCH_CODE'])}</td>
              <td style='padding:4px 8px;border:1px solid #dee2e6;white-space:nowrap;'>{r['BRANCH_NAME']}</td>
              <td style='padding:4px 8px;border:1px solid #dee2e6;color:#555;font-size:0.82rem;'>{desc}</td>
              <td style='{td_date}'>{date_str}</td>
              <td style='{td_val}'>{val_str}</td>
            </tr>"""
        header_bg = "#1a5fb4" if sort_by_date else "#7b3f9e"
        th_date = f"color:white;padding:5px 8px;border:1px solid rgba(255,255,255,0.2);{HL_H if sort_by_date else ''}"
        th_val  = f"color:white;padding:5px 8px;border:1px solid rgba(255,255,255,0.2);text-align:right;{HL_H if not sort_by_date else ''}"
        arrow_date = " ↓" if sort_by_date else ""
        arrow_val  = " ↓" if not sort_by_date else ""
        return f"""
<div style="margin-bottom:16px;">
  <div style="font-size:0.83rem;font-weight:700;color:{header_bg};margin-bottom:6px;">
    {"🕐" if sort_by_date else "💰"} Top 5: {label.capitalize()} investice
  </div>
  <div class="table-responsive" style="border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,0.07);">
    <table class="table table-bordered table-sm mb-0" style="font-size:0.83rem;background:white;">
      <thead>
        <tr style="background:{header_bg};">
          <th style="color:white;padding:5px 8px;border:1px solid rgba(255,255,255,0.2);">ID</th>
          <th style="color:white;padding:5px 8px;border:1px solid rgba(255,255,255,0.2);">Pobočka</th>
          <th style="color:white;padding:5px 8px;border:1px solid rgba(255,255,255,0.2);">Popis</th>
          <th style="{th_date}">Datum{arrow_date}</th>
          <th style="{th_val}">Výše inv.{arrow_val}</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
</div>"""

    t1 = _parse_inv(df, 'DATUM_LAST_INV',    'LAST_INV',    'nejnovější', '#1a5fb4')
    t2 = _parse_inv(df, 'DATUM_BIGGEST_INV', 'BIGGEST_INV', 'nejdražší',  '#7b3f9e')
    return t1 + t2


def render_top5_cards(df, mode, title, subtitle="", n=5):
    """
    Vykreslí top N poboček ve stylu spádových karet.

    mode:
      'improve'  — největší zlepšení ratingu (IR_DIFF_NUM ascending)
      'worsen'   — největší zhoršení ratingu (IR_DIFF_NUM descending)
      'inv_new'  — nejnovější investice (DATUM_LAST_INV descending)
      'inv_big'  — největší investice (BIGGEST_INV descending)
    """
    # ── Barvy dle modu ────────────────────────────────────────────────
    _RANK_COLS = {
        1: '#2770f0', 2: '#4d8ef0', 3: '#74aaf5', 4: '#9cc4f8', 5: '#c2dafc'
    }
    _HDR_COL = {
        'improve': '#0bb440', 'worsen': '#eb4d79',
        'inv_new': '#2770f0', 'inv_big': '#7b3f9e',
        'revenue': '#0bb440',
    }
    hdr_col = _HDR_COL.get(mode, '#2770f0')

    # ── Seřazení a výběr ──────────────────────────────────────────────
    d = df.copy()
    if mode == 'improve':
        if 'IR_DIFF_NUM' not in d.columns: return ""
        d = d[d['IR_DIFF_NUM'].notna() & (d['IR_DIFF_NUM'] < 0)]
        d = d.sort_values('IR_DIFF_NUM', ascending=True).head(n)
    elif mode == 'worsen':
        if 'IR_DIFF_NUM' not in d.columns: return ""
        d = d[d['IR_DIFF_NUM'].notna() & (d['IR_DIFF_NUM'] > 0)]
        d = d.sort_values('IR_DIFF_NUM', ascending=False).head(n)
    elif mode == 'inv_new':
        if 'DATUM_LAST_INV' not in d.columns: return ""
        d['DATUM_LAST_INV'] = pd.to_datetime(d['DATUM_LAST_INV'], errors='coerce')
        d['LAST_INV'] = pd.to_numeric(d['LAST_INV'], errors='coerce')
        d = d.dropna(subset=['DATUM_LAST_INV']).query('LAST_INV > 0')
        d = d.sort_values('DATUM_LAST_INV', ascending=False).head(n)
    elif mode == 'inv_big':
        if 'BIGGEST_INV' not in d.columns: return ""
        d['BIGGEST_INV'] = pd.to_numeric(d['BIGGEST_INV'], errors='coerce')
        d['DATUM_BIGGEST_INV'] = pd.to_datetime(d.get('DATUM_BIGGEST_INV', pd.NaT), errors='coerce')
        d = d.dropna(subset=['BIGGEST_INV']).query('BIGGEST_INV > 0')
        d = d.sort_values('BIGGEST_INV', ascending=False).head(n)
    elif mode == 'revenue':
        if 'OBJEM_VYNOSU_CZK' not in d.columns: return ""
        d['OBJEM_VYNOSU_CZK'] = pd.to_numeric(d['OBJEM_VYNOSU_CZK'], errors='coerce')
        d = d.dropna(subset=['OBJEM_VYNOSU_CZK']).query('OBJEM_VYNOSU_CZK > 0')
        d = d.sort_values('OBJEM_VYNOSU_CZK', ascending=False).head(n)

    if d.empty:
        return f"<p style='color:#aaa;font-style:italic;padding:8px 0;'>Žádná data.</p>"

    # ── Pomocné formátovací funkce ─────────────────────────────────────
    def _val_line(mode, row):
        """Vrátí hlavní metriku karty."""
        if mode == 'improve':
            diff = int(row.get('IR_DIFF_NUM', 0))
            ir25 = row.get('IR', '—')
            ir24 = row.get('INTERNI_RATING_2024', '—')
            try: ir25 = int(float(ir25))
            except: pass
            try: ir24 = int(float(ir24))
            except: pass
            return (f"<span style='color:#0bb440;font-size:1.1rem;font-weight:900;'>▲ {abs(diff)}</span>"
                    f"<span style='font-size:0.75rem;color:#888;margin-left:6px;'>míst nahoru</span>"
                    f"<span style='font-size:0.75rem;color:#aaa;margin-left:10px;'>{ir24} → {ir25}</span>")
        elif mode == 'worsen':
            diff = int(row.get('IR_DIFF_NUM', 0))
            ir25 = row.get('IR', '—')
            ir24 = row.get('INTERNI_RATING_2024', '—')
            try: ir25 = int(float(ir25))
            except: pass
            try: ir24 = int(float(ir24))
            except: pass
            return (f"<span style='color:#eb4d79;font-size:1.1rem;font-weight:900;'>▼ {abs(diff)}</span>"
                    f"<span style='font-size:0.75rem;color:#888;margin-left:6px;'>míst dolů</span>"
                    f"<span style='font-size:0.75rem;color:#aaa;margin-left:10px;'>{ir24} → {ir25}</span>")
        elif mode == 'inv_new':
            dt = row.get('DATUM_LAST_INV')
            val = row.get('LAST_INV', 0)
            dt_s = dt.strftime('%d.%m.%Y') if pd.notna(dt) else '—'
            return (f"<span style='color:#2770f0;font-size:0.92rem;font-weight:800;'>{dt_s}</span>"
                    f"<span style='font-size:0.75rem;color:#888;margin-left:8px;'>{format_money(val)}</span>")
        elif mode == 'inv_big':
            val = row.get('BIGGEST_INV', 0)
            dt  = row.get('DATUM_BIGGEST_INV')
            dt_s = dt.strftime('%d.%m.%Y') if pd.notna(dt) else '—'
            return (f"<span style='color:#7b3f9e;font-size:0.92rem;font-weight:800;'>{format_money(val)}</span>"
                    f"<span style='font-size:0.75rem;color:#888;margin-left:8px;'>{dt_s}</span>")
        elif mode == 'revenue':
            rev = row.get('OBJEM_VYNOSU_CZK', 0)
            rev24 = row.get('NEW_BUSINESS_2024_-_OBJEM_VYNOSU')
            try:
                diff_pct = (float(rev) / float(rev24) - 1) * 100 if rev24 and float(rev24) > 0 else None
            except: diff_pct = None
            diff_html = ""
            if diff_pct is not None:
                sign = "▲" if diff_pct >= 0 else "▼"
                clr  = "#0bb440" if diff_pct >= 0 else "#eb4d79"
                diff_html = f"<span style='color:{clr};font-size:0.75rem;margin-left:8px;font-weight:700;'>{sign} {abs(diff_pct):.1f} %</span>"
            return (f"<span style='color:#0bb440;font-size:0.95rem;font-weight:800;'>{format_money(rev)}</span>"
                    + diff_html)
        return ""

    def _sub_line(mode, row):
        """Vrátí doplňkovou info řádku."""
        if mode in ('improve','worsen'):
            ci = row.get('PRIME_NAKLADY/VYNOSY')
            rev = row.get('OBJEM_VYNOSU_CZK')
            parts = []
            if pd.notna(ci): parts.append(f"C/I {ci*100:.1f} %")
            if pd.notna(rev): parts.append(f"Výnosy {format_money(rev)}")
            return " · ".join(parts)
        elif mode == 'inv_new':
            desc = str(row.get('DESC_LAST_INV', '') or '')
            return desc[:60] + ('…' if len(desc) > 60 else '') if desc else '—'
        elif mode == 'inv_big':
            desc = str(row.get('DESC_BIGGEST_INV', '') or '')
            return desc[:60] + ('…' if len(desc) > 60 else '') if desc else '—'
        elif mode == 'revenue':
            cnt  = row.get('POCET_PRODEJU')
            ci   = row.get('PRIME_NAKLADY/VYNOSY')
            ir   = row.get('IR')
            parts = []
            try: parts.append(f"Rating {int(float(ir))}")
            except: pass
            try: parts.append(f"{int(float(cnt)):,} prodejů")
            except: pass
            try: parts.append(f"C/I {float(ci)*100:.1f} %")
            except: pass
            return " · ".join(parts)
        return ""

    # ── Karty ──────────────────────────────────────────────────────────
    cards = ""
    for rank, (_, row) in enumerate(d.iterrows(), 1):
        rcol   = _RANK_COLS.get(rank, '#c2dafc')
        bcode  = int(row.get('BRANCH_CODE', 0))
        bname  = str(row.get('BRANCH_NAME', str(bcode)))
        region = str(row.get('REGION_NAME', ''))
        val_html = _val_line(mode, row)
        sub_html = _sub_line(mode, row)

        cards += f"""
<div style="display:flex;align-items:stretch;gap:0;border-radius:8px;overflow:hidden;
            border:1px solid #e8edf5;margin-bottom:8px;background:#fff;
            box-shadow:0 1px 4px rgba(0,0,0,0.05);">
  <!-- rank badge -->
  <div style="background:{rcol};color:white;font-size:1rem;font-weight:900;
              min-width:40px;display:flex;align-items:center;justify-content:center;
              flex-shrink:0;">{rank}</div>
  <!-- obsah -->
  <div style="padding:9px 14px;flex:1;min-width:0;">
    <div style="display:flex;align-items:baseline;gap:8px;flex-wrap:wrap;">
      <span style="font-size:0.85rem;font-weight:700;color:#0f172a;">{bname}</span>
      <span style="font-size:0.72rem;color:#aaa;">ID {bcode}</span>
      <span style="font-size:0.72rem;color:#888;margin-left:auto;">{region}</span>
    </div>
    <div style="margin-top:5px;display:flex;align-items:center;flex-wrap:wrap;gap:4px;">
      {val_html}
    </div>
    <div style="font-size:0.72rem;color:#888;margin-top:3px;white-space:nowrap;
                overflow:hidden;text-overflow:ellipsis;">{sub_html}</div>
  </div>
</div>"""

    return f"""
<div style="background:#fff;border-radius:10px;padding:14px 16px;
            border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">
  <div style="font-size:0.72rem;font-weight:700;color:{hdr_col};text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:4px;">{title}</div>
  {"" if not subtitle else f'<div style="font-size:0.72rem;color:#aaa;margin-bottom:10px;">{subtitle}</div>'}
  {cards}
</div>"""


def generate_danger_zone_table(df):
    """
    Pobočky v nebezpečné zóně — pět kritérií:
      A) Vysoké nájemné + nízké výnosy
      B) Vysoké C/I + podprůměrné výnosy + málo FTE
      C) Zhoršení ratingu 2 roky za sebou
      D) Nízká primární klientela (PRIM_RATIO pod Q25)
      E) Výnosy v dolním kvintilu + žádný růst prodejů (POCET_PRODEJU_CELKEM_2025 <= 2024)
    Štítky jsou zaklikávatelné — slouží k filtrování zobrazených řádků.
    """
    import json as _json

    d = df.copy()
    for c in ['OBJEM_VYNOSU_CZK', 'ROCNI_SPLATKY_S_DPH_CZK', 'PRIME_NAKLADY/VYNOSY', 'FTE',
              'IR', 'INTERNI_RATING_2024', 'INTERNI_RATING_2023', 'IR_DIFF_NUM',
              'PRIM_RATIO', 'POCET_PRODEJU_CELKEM_2025', 'VYNOSY']:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors='coerce')

    # Vyřaď pobočky bez vypočítaného interního ratingu
    if 'IR_FLAG' in d.columns:
        d = d[d['IR_FLAG'] == 'Y'].copy()
    elif 'IR' in d.columns:
        d = d[d['IR'].notna() & (d['IR'] > 0)].copy()

    _rev_med  = d['OBJEM_VYNOSU_CZK'].median()               if 'OBJEM_VYNOSU_CZK' in d.columns else None
    _rent_q75 = d['ROCNI_SPLATKY_S_DPH_CZK'].quantile(0.75)  if 'ROCNI_SPLATKY_S_DPH_CZK' in d.columns else None
    _ci_q75   = d['PRIME_NAKLADY/VYNOSY'].quantile(0.75)      if 'PRIME_NAKLADY/VYNOSY' in d.columns else None
    _fte_q25  = d['FTE'].quantile(0.25)                       if 'FTE' in d.columns else None
    _rev_q25  = d['OBJEM_VYNOSU_CZK'].quantile(0.25)          if 'OBJEM_VYNOSU_CZK' in d.columns else None
    _rev_q20  = d['OBJEM_VYNOSU_CZK'].quantile(0.20)          if 'OBJEM_VYNOSU_CZK' in d.columns else None
    _prim_q25 = d['PRIM_RATIO'].quantile(0.25)                if 'PRIM_RATIO' in d.columns else None

    flags = []
    _seen = set()

    # A) Vysoké nájemné + nízké výnosy
    if _rent_q75 is not None and _rev_med is not None:
        mask_a = (d['ROCNI_SPLATKY_S_DPH_CZK'] > _rent_q75) & (d['OBJEM_VYNOSU_CZK'] < _rev_med)
        for _, r in d[mask_a].iterrows():
            key = (int(r['BRANCH_CODE']), 'A')
            if key not in _seen:
                _seen.add(key)
                flags.append({'row': r, 'důvod': 'A', 'popis': 'Vysoké nájemné + nízké výnosy',
                              'col': '#eb4d79', 'icon': '🏠'})

    # B) Vysoké C/I + podprůměrné výnosy + málo FTE
    if _ci_q75 is not None and _rev_q25 is not None and _fte_q25 is not None:
        mask_b = ((d['PRIME_NAKLADY/VYNOSY'] > _ci_q75) &
                  (d['OBJEM_VYNOSU_CZK'] < _rev_q25) &
                  (d['FTE'] <= _fte_q25))
        for _, r in d[mask_b].iterrows():
            key = (int(r['BRANCH_CODE']), 'B')
            if key not in _seen:
                _seen.add(key)
                flags.append({'row': r, 'důvod': 'B', 'popis': 'Vysoké C/I + slabé výnosy + málo FTE',
                              'col': '#f59e0b', 'icon': '📉'})

    # C) Zhoršení ratingu 2 roky za sebou (IR roste = horší)
    if all(c in d.columns for c in ['IR', 'INTERNI_RATING_2024', 'INTERNI_RATING_2023', 'IR_DIFF_NUM']):
        mask_c = (
            d['IR_DIFF_NUM'] > 0
        ) & (
            d['INTERNI_RATING_2024'].notna() & d['INTERNI_RATING_2023'].notna()
        ) & (
            d['INTERNI_RATING_2024'] > d['INTERNI_RATING_2023']
        )
        for _, r in d[mask_c].iterrows():
            key = (int(r['BRANCH_CODE']), 'C')
            if key not in _seen:
                _seen.add(key)
                flags.append({'row': r, 'důvod': 'C', 'popis': 'Rating se zhoršuje 2 roky za sebou',
                              'col': '#7c3aed', 'icon': '📛'})

    # D) Nízká primární klientela (PRIM_RATIO pod Q25)
    if _prim_q25 is not None and 'PRIM_RATIO' in d.columns:
        mask_d = d['PRIM_RATIO'] < _prim_q25
        for _, r in d[mask_d].iterrows():
            key = (int(r['BRANCH_CODE']), 'D')
            if key not in _seen:
                _seen.add(key)
                flags.append({'row': r, 'důvod': 'D', 'popis': 'Nízká primární klientela (pod Q25)',
                              'col': '#0891b2', 'icon': '👤'})

    # E) Výnosy v dolním kvintilu + stagnace/pokles prodejů 2025 vs 2024
    if _rev_q20 is not None and 'POCET_PRODEJU_CELKEM_2025' in d.columns:
        _prod_2024_cols = [c for c in d.columns if c.startswith('POCET_PRODEJU_') and c.endswith('_2024')]
        if _prod_2024_cols:
            d['_prod_sum_2024'] = d[_prod_2024_cols].sum(axis=1)
        else:
            d['_prod_sum_2024'] = 0
        mask_e = (
            (d['OBJEM_VYNOSU_CZK'] < _rev_q20) &
            (d['POCET_PRODEJU_CELKEM_2025'] <= d['_prod_sum_2024'])
        )
        for _, r in d[mask_e].iterrows():
            key = (int(r['BRANCH_CODE']), 'E')
            if key not in _seen:
                _seen.add(key)
                flags.append({'row': r, 'důvod': 'E', 'popis': 'Nízké výnosy + stagnace prodejů',
                              'col': '#059669', 'icon': '📦'})

    if not flags:
        return "<p style='color:#aaa;font-style:italic;'>Žádné pobočky nesplňují kritéria nebezpečné zóny.</p>"

    _order = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4}
    flags.sort(key=lambda x: (_order.get(x['důvod'], 9), x['row'].get('OBJEM_VYNOSU_CZK', 0) or 0))

    # Metadata kritérií pro JS
    criteria_meta = [
        {'key': 'A', 'col': '#eb4d79', 'icon': '🏠',
         'desc': 'Vysoké nájemné + nízké výnosy (nájemné &gt; Q75, výnosy &lt; medián)'},
        {'key': 'B', 'col': '#f59e0b', 'icon': '📉',
         'desc': 'Vysoké C/I + slabé výnosy + málo FTE (nad/pod Q75/Q25)'},
        {'key': 'C', 'col': '#7c3aed', 'icon': '📛',
         'desc': 'Rating se zhoršuje 2 roky za sebou (IR↑ v 2024 i 2025)'},
        {'key': 'D', 'col': '#0891b2', 'icon': '👤',
         'desc': 'Nízká primární klientela (PRIM_RATIO &lt; Q25)'},
        {'key': 'E', 'col': '#059669', 'icon': '📦',
         'desc': 'Nízké výnosy (Q20) + stagnace nebo pokles počtu prodejů'},
    ]
    # Pouze kritéria, která mají alespoň jeden záznam
    active_keys = {f['důvod'] for f in flags}
    criteria_meta = [c for c in criteria_meta if c['key'] in active_keys]
    criteria_json = _json.dumps([{'key': c['key'], 'col': c['col']} for c in criteria_meta])

    rows_html = ""
    for f in flags:
        r    = f['row']
        col  = f['col']
        bc   = int(r.get('BRANCH_CODE', 0))
        bn   = str(r.get('BRANCH_NAME', ''))
        reg  = str(r.get('REGION_NAME', ''))
        rev  = r.get('OBJEM_VYNOSU_CZK')
        rent = r.get('ROCNI_SPLATKY_S_DPH_CZK')
        ci   = r.get('PRIME_NAKLADY/VYNOSY')
        fte  = r.get('FTE')
        prim = r.get('PRIM_RATIO')
        ir25 = r.get('IR'); ir24 = r.get('INTERNI_RATING_2024'); ir23 = r.get('INTERNI_RATING_2023')
        try: ir25 = int(float(ir25))
        except: ir25 = '—'
        try: ir24 = int(float(ir24))
        except: ir24 = '—'
        try: ir23 = int(float(ir23))
        except: ir23 = '—'

        _metrics = []
        if pd.notna(rev):   _metrics.append(f"Výnosy {format_money(rev)}")
        if pd.notna(rent):  _metrics.append(f"Nájemné {format_money(rent)}/rok")
        if pd.notna(ci):    _metrics.append(f"C/I {ci*100:.1f} %")
        if pd.notna(fte):   _metrics.append(f"FTE {fte:.1f}")
        if pd.notna(prim):  _metrics.append(f"Prim. kl. {prim*100:.1f} %")
        _rating_str = f"Rating: {ir23} → {ir24} → {ir25}"

        rows_html += f"""
<div class="dz-row" data-dz-reason="{f['důvod']}"
     style="display:flex;align-items:stretch;gap:0;border-radius:8px;overflow:hidden;
            border:1px solid #e8edf5;margin-bottom:8px;background:#fff;
            box-shadow:0 1px 4px rgba(0,0,0,0.05);">
  <div style="background:{col};color:white;font-size:0.78rem;font-weight:900;
              min-width:36px;display:flex;align-items:center;justify-content:center;
              flex-shrink:0;writing-mode:vertical-rl;text-orientation:mixed;
              padding:8px 0;letter-spacing:1px;">{f['důvod']}</div>
  <div style="padding:9px 14px;flex:1;min-width:0;">
    <div style="display:flex;align-items:baseline;gap:8px;flex-wrap:wrap;">
      <span style="font-size:0.85rem;font-weight:700;color:#0f172a;">{bn}</span>
      <span style="font-size:0.72rem;color:#aaa;">ID {bc}</span>
      <span style="font-size:0.72rem;color:#888;">{reg}</span>
      <span style="margin-left:auto;background:{col}18;color:{col};font-size:0.72rem;
                   font-weight:700;padding:1px 8px;border-radius:10px;border:1px solid {col}44;">
        {f['icon']} {f['popis']}</span>
    </div>
    <div style="font-size:0.75rem;color:#555;margin-top:4px;">{' · '.join(_metrics)}</div>
    <div style="font-size:0.72rem;color:#aaa;margin-top:2px;">{_rating_str}</div>
  </div>
</div>"""

    # Legenda — zaklikávatelné štítky (toggle)
    legend_items = "".join(
        f"""<button class="dz-toggle-btn" data-dz-key="{c['key']}"
              style="background:{c['col']}18;color:{c['col']};border:2px solid {c['col']}55;
                     border-radius:6px;padding:4px 12px;font-size:0.78rem;font-weight:700;
                     cursor:pointer;user-select:none;transition:opacity .15s,border-color .15s;"
              title="Klikni pro zobrazení/skrytí">
          {c['icon']} {c['key']} — {c['desc']}
        </button>"""
        for c in criteria_meta
    )

    total = len(flags)
    return f"""
<div id="dz-wrap" style="background:#fff;border-radius:10px;padding:16px 18px;
     border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">
  <div style="font-size:0.72rem;color:#eb4d79;font-weight:700;text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:4px;">⚠️ Pobočky v nebezpečné zóně</div>
  <div style="font-size:0.72rem;color:#aaa;margin-bottom:10px;">
    <span id="dz-count">{total}</span> z {total} poboček · klikni na štítek pro filtrování
  </div>
  <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:14px;"
       id="dz-legend">{legend_items}</div>
  <div id="dz-rows">{rows_html}</div>
</div>

<script>
(function() {{
  var meta    = {criteria_json};
  var active  = {{}};
  meta.forEach(function(m) {{ active[m.key] = true; }});

  var legend  = document.getElementById("dz-legend");
  var countEl = document.getElementById("dz-count");
  if (!legend) return;

  function applyDzFilter() {{
    var rows    = document.querySelectorAll("#dz-rows .dz-row");
    var visible = 0;
    rows.forEach(function(row) {{
      var k = row.getAttribute("data-dz-reason");
      var show = active[k] !== false;
      row.style.display = show ? "" : "none";
      if (show) visible++;
    }});
    if (countEl) countEl.textContent = visible;
  }}

  legend.querySelectorAll(".dz-toggle-btn").forEach(function(btn) {{
    btn.addEventListener("click", function() {{
      var k = btn.getAttribute("data-dz-key");
      active[k] = !active[k];
      btn.style.opacity = active[k] ? "1" : "0.3";
      btn.style.borderStyle = active[k] ? "solid" : "dashed";
      applyDzFilter();
    }});
  }});
}})();
</script>"""


def generate_age_pyramid(df):
    """
    Věková pyramida celé sítě — dva horizontální bar charty vedle sebe:
    levý = počet klientů, pravý = průměrný výnos. Čistý SVG, bez závislostí.
    """
    count_cols = [f'{g}_POCET_KLIENTU'  for g in AGE_GROUPS]
    rev_cols   = [f'{g}_PRUMERNY_VYNOS' for g in AGE_GROUPS]

    avail_c = [c for c in count_cols if c in df.columns]
    if not avail_c:
        return "<p style='color:#aaa;font-style:italic;'>Žádná věková data.</p>"

    totals_count, totals_rev = [], []
    for g in AGE_GROUPS:
        cc = f'{g}_POCET_KLIENTU'
        rc = f'{g}_PRUMERNY_VYNOS'
        cnt = float(df[cc].sum()) if cc in df.columns else 0.0
        if rc in df.columns and cc in df.columns:
            mask = df[cc].notna() & df[rc].notna() & (df[cc] > 0)
            wav  = float((df.loc[mask, rc] * df.loc[mask, cc]).sum() / df.loc[mask, cc].sum()) if mask.any() else 0.0
        else:
            wav = 0.0
        totals_count.append(cnt)
        totals_rev.append(wav)

    max_c = max(totals_count) or 1
    max_r = max(totals_rev)   or 1

    BAR_H    = 32
    GAP      = 10
    LABEL_W  = 58
    BAR_AREA = 260
    TOTAL_H  = len(AGE_GROUPS) * (BAR_H + GAP) + 48

    def _bar_row(i, g, val, max_val, color, fmt_fn, side):
        y     = 40 + i * (BAR_H + GAP)
        bw    = max(int(val / max_val * BAR_AREA), 2) if val > 0 else 0
        label = fmt_fn(val)
        if side == 'left':
            return (
                f'<rect x="{LABEL_W}" y="{y}" width="{bw}" height="{BAR_H}" '
                f'rx="4" fill="{color}" opacity="0.88"/>'
                f'<text x="{LABEL_W + bw + 5}" y="{y + BAR_H//2 + 5}" '
                f'font-size="11" fill="#444" font-family="Segoe UI,Arial,sans-serif">{label}</text>'
                f'<text x="{LABEL_W - 6}" y="{y + BAR_H//2 + 5}" '
                f'font-size="12" fill="#333" text-anchor="end" font-weight="600" '
                f'font-family="Segoe UI,Arial,sans-serif">{g}</text>'
            )
        else:
            return (
                f'<rect x="{LABEL_W}" y="{y}" width="{bw}" height="{BAR_H}" '
                f'rx="4" fill="{color}" opacity="0.88"/>'
                f'<text x="{LABEL_W + bw + 5}" y="{y + BAR_H//2 + 5}" '
                f'font-size="11" fill="#444" font-family="Segoe UI,Arial,sans-serif">{label}</text>'
                f'<text x="{LABEL_W - 6}" y="{y + BAR_H//2 + 5}" '
                f'font-size="12" fill="#333" text-anchor="end" font-weight="600" '
                f'font-family="Segoe UI,Arial,sans-serif">{g}</text>'
            )

    def _fmt_cnt(v):
        if v >= 1_000_000: return f"{v/1_000_000:.1f}M"
        if v >= 1_000:     return f"{v/1_000:.0f}k"
        return str(int(v))

    def _fmt_rev(v):
        if v >= 1_000_000: return f"{v/1_000_000:.2f}M Kč"
        if v >= 1_000:     return f"{v/1_000:.0f}k Kč"
        return f"{int(v)} Kč"

    W = LABEL_W + BAR_AREA + 90

    # Levý chart — počet klientů
    rows_c = "".join(
        _bar_row(i, g, totals_count[i], max_c, AGE_COLORS[i], _fmt_cnt, 'left')
        for i, g in enumerate(AGE_GROUPS)
    )
    svg_cnt = (
        f'<svg viewBox="0 0 {W} {TOTAL_H}" xmlns="http://www.w3.org/2000/svg" '
        f'style="width:100%;max-width:{W}px;display:block;">'
        f'<text x="{LABEL_W}" y="22" font-size="12" font-weight="700" fill="#2770f0" '
        f'font-family="Segoe UI,Arial,sans-serif">👥 Počet klientů</text>'
        f'{rows_c}</svg>'
    )

    # Pravý chart — průměrný výnos
    rows_r = "".join(
        _bar_row(i, g, totals_rev[i], max_r, AGE_COLORS[i], _fmt_rev, 'right')
        for i, g in enumerate(AGE_GROUPS)
    )
    svg_rev = (
        f'<svg viewBox="0 0 {W} {TOTAL_H}" xmlns="http://www.w3.org/2000/svg" '
        f'style="width:100%;max-width:{W}px;display:block;">'
        f'<text x="{LABEL_W}" y="22" font-size="12" font-weight="700" fill="#0bb440" '
        f'font-family="Segoe UI,Arial,sans-serif">💰 Průměrný výnos na klienta</text>'
        f'{rows_r}</svg>'
    )

    # Celkové součty pod grafy
    total_klienti = int(sum(totals_count))
    avg_vynos_all = sum(c * r for c, r in zip(totals_count, totals_rev)) / sum(totals_count) if sum(totals_count) else 0

    return f"""
<div style="background:#fff;border-radius:10px;padding:16px 18px;
            border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">
  <div style="font-size:0.72rem;color:#2770f0;font-weight:700;text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:4px;">👥 Věková pyramida klientů — celá síť</div>
  <div style="font-size:0.72rem;color:#aaa;margin-bottom:14px;">
    Agregát všech poboček · celkem {total_klienti:,} klientů · průměrný výnos {_fmt_rev(avg_vynos_all)}
  </div>
  <div style="display:flex;gap:28px;flex-wrap:wrap;">
    <div style="flex:1;min-width:280px;">{svg_cnt}</div>
    <div style="flex:1;min-width:280px;">{svg_rev}</div>
  </div>
</div>"""


def generate_revenue_client_heatmap(df, parties_full_df=None):
    """
    Mapa kde bydlí klienti s nejvyššími výnosy (top 20 % dle LTV).
    Používá parties_full s LOC_GEO_LATITUDE/LONGITUDE filtrovaný
    na pobočky z top 20 % výnosového kvintilu.
    """
    import json as _jh

    if parties_full_df is None or parties_full_df.empty:
        return "<p style='color:#aaa;font-style:italic;'>parties_full není k dispozici.</p>"

    # Top 20 % poboček dle výnosů
    _d = df.copy()
    _d['OBJEM_VYNOSU_CZK'] = pd.to_numeric(_d['OBJEM_VYNOSU_CZK'], errors='coerce')
    _thresh = _d['OBJEM_VYNOSU_CZK'].quantile(0.80)
    _top_codes = set(_d[_d['OBJEM_VYNOSU_CZK'] >= _thresh]['BRANCH_CODE'].astype(int).tolist())

    _pf = parties_full_df.copy()
    _pf.columns = [c.upper() for c in _pf.columns]

    # Filtr na klienty top poboček — HNED na začátku před dalším zpracováním
    _bid_col = next((c for c in ['DBS_HOME_BRANCH_CODE','BRANCH_ID','BRANCH_CODE'] if c in _pf.columns), None)
    if not _bid_col:
        return "<p style='color:#aaa;font-style:italic;'>parties_full nemá sloupec s ID pobočky.</p>"

    _pf[_bid_col] = pd.to_numeric(_pf[_bid_col], errors='coerce')
    _pf_top = _pf[_pf[_bid_col].isin(_top_codes)].copy()

    # Vzorkování — max 50 000 řádků pro výkon
    if len(_pf_top) > 50_000:
        _pf_top = _pf_top.sample(50_000, random_state=42)

    if 'LOC_GEO_LATITUDE' not in _pf_top.columns or 'LOC_GEO_LONGITUDE' not in _pf_top.columns:
        return "<p style='color:#aaa;font-style:italic;'>GPS souřadnice nejsou k dispozici v parties_full.</p>"

    _pf_top['_lat'] = pd.to_numeric(_pf_top['LOC_GEO_LATITUDE'],  errors='coerce')
    _pf_top['_lon'] = pd.to_numeric(_pf_top['LOC_GEO_LONGITUDE'], errors='coerce')
    _pf_top = _pf_top.dropna(subset=['_lat','_lon'])
    _pf_top = _pf_top[(_pf_top['_lat'].between(48.5, 51.2)) & (_pf_top['_lon'].between(12.0, 18.8))]

    if _pf_top.empty:
        return "<p style='color:#aaa;font-style:italic;'>Žádné GPS záznamy pro top 20 % pobočky.</p>"

    # Agregace na mřížku 0.02°
    _pf_top['_glat'] = (_pf_top['_lat'] / 0.02).round() * 0.02
    _pf_top['_glon'] = (_pf_top['_lon'] / 0.02).round() * 0.02
    _grid = _pf_top.groupby(['_glat','_glon']).size().reset_index(name='cnt')

    import plotly.graph_objects as _pgo2
    _fig = _pgo2.Figure()
    _fig.add_trace(_pgo2.Densitymapbox(
        lat=_grid['_glat'], lon=_grid['_glon'], z=_grid['cnt'],
        radius=22,
        colorscale=[
            [0.0,  'rgba(39,112,240,0)'],
            [0.2,  'rgba(39,112,240,0.5)'],
            [0.5,  'rgba(11,180,64,0.75)'],
            [0.75, 'rgba(245,158,11,0.88)'],
            [1.0,  'rgba(235,77,121,1.0)'],
        ],
        showscale=True,
        colorbar=dict(title=dict(text='Klientů'), thickness=12, len=0.6,
                      tickvals=[], ticktext=[]),
        hovertemplate='Klientů v oblasti: %{z}<extra></extra>',
    ))
    # Pobočky top 20 % jako markery
    _top_df = _d[_d['OBJEM_VYNOSU_CZK'] >= _thresh].copy()
    if 'GPS_X' in _top_df.columns and 'GPS_Y' in _top_df.columns:
        _top_df['GPS_X'] = pd.to_numeric(_top_df['GPS_X'], errors='coerce')
        _top_df['GPS_Y'] = pd.to_numeric(_top_df['GPS_Y'], errors='coerce')
        _top_df = _top_df.dropna(subset=['GPS_X','GPS_Y'])
        _fig.add_trace(_pgo2.Scattermapbox(
            lat=_top_df['GPS_X'], lon=_top_df['GPS_Y'],
            mode='markers',
            marker=dict(size=9, color='#eb4d79', opacity=0.9),
            text=_top_df.apply(lambda r: f"{r.get('BRANCH_NAME','')} ({format_money(r['OBJEM_VYNOSU_CZK'])})", axis=1),
            hoverinfo='text', showlegend=False,
        ))

    _clat = float(_pf_top['_lat'].median())
    _clon = float(_pf_top['_lon'].median())
    _fig.update_layout(
        mapbox=_mapbox_layout(center_lat=_clat, center_lon=_clon, zoom=6),
        height=480, margin=dict(l=0, r=0, t=0, b=0),
        paper_bgcolor='white',
    )
    _map_div = _fig.to_html(
        full_html=False, include_plotlyjs='cdn',
        div_id="revenue-heatmap-net",
        config={'displayModeBar': False, 'scrollZoom': True},
    )

    _cnt_clients = len(_pf_top)
    _n_branches  = len(_top_codes)
    return f"""
<div style="background:#fff;border-radius:10px;padding:16px 18px;
            border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">
  <div style="font-size:0.72rem;color:#0bb440;font-weight:700;text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:4px;">💎 Kde bydlí klienti s nejvyššími výnosy</div>
  <div style="font-size:0.72rem;color:#aaa;margin-bottom:10px;">
    Top 20 % poboček dle LTV výnosů ({_n_branches} poboček, práh {format_money(_thresh)}) ·
    {_cnt_clients:,} klientů s GPS adresou · růžové tečky = pobočky top 20 %
  </div>
  <div style="border-radius:10px;overflow:hidden;border:1px solid #dde4f5;">
    {_map_div}
  </div>
</div>"""


def generate_orp_coverage_html(df, geojson_path=None):
    """Přehled pokrytí ORP s přepínačem Všechny pobočky / Pouze Keep."""
    import json as _json
    import os as _os
    _fname = '../vypocet_ir_2026/zdroje/csu_geodb_sde_CISORP_byty_etl_20210326.geojson'
    _path  = geojson_path if (geojson_path and _os.path.isfile(geojson_path)) else (
             _fname if _os.path.isfile(_fname) else None)
    if _path is None:
        return (f"<p style='color:#e64343;font-style:italic;'>"
                f"GeoJSON ORP nenalezen: <b>{_fname}</b> (CWD: {_os.getcwd()})</p>")
    try:
        with open(_path, 'r', encoding='utf-8') as f:
            geojson = _json.load(f)
    except Exception as _e:
        return f"<p style='color:#aaa;font-style:italic;'>Chyba při čtení GeoJSON: {_e}</p>"

    if 'ORP_NAZEV' not in df.columns or df['ORP_NAZEV'].eq('').all():
        return "<p style='color:#aaa;font-style:italic;'>ORP data nejsou dostupná.</p>"

    all_orp = {str(f['properties']['kod']): f['properties']['nazev'] for f in geojson['features']}
    n_total = len(all_orp)

    _strat_col = 'FOOTPRINT_STRATEGIE_(2030)'
    if _strat_col in df.columns:
        df_keep = df[~df[_strat_col].astype(str).str.lower().str.contains('close', na=False)].copy()
    else:
        df_keep = df.copy()

    def _card(val, lbl, col):
        return (f'<div style="background:#f8fbff;border:1px solid #e8edf5;border-radius:8px;'
                f'padding:10px 18px;text-align:center;min-width:110px;">'
                f'<div style="font-size:1.4rem;font-weight:800;color:{col};">{val}</div>'
                f'<div style="font-size:0.68rem;color:#888;text-transform:uppercase;margin-top:2px;">{lbl}</div>'
                f'</div>')

    def _build_panel(dff, uid):
        orp_counts = (
            dff[dff['ORP_NAZEV'].notna() & (dff['ORP_NAZEV'] != '')]
            .groupby(['ORP_KOD', 'ORP_NAZEV'], observed=True)
            .size().reset_index(name='pocet')
            .sort_values('pocet', ascending=False)
        )
        covered = set(orp_counts['ORP_KOD'].astype(str))
        missing = {k: v for k, v in all_orp.items() if k not in covered}
        n_cov   = len(covered)
        n_miss  = len(missing)
        pct     = n_cov / n_total * 100 if n_total else 0

        cards = (f'<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:16px;">'
                 + _card(str(n_total), 'ORP celkem v ČR', '#0f172a')
                 + _card(str(n_cov),   'ORP s pobočkou',  '#0bb440')
                 + _card(str(n_miss),  'ORP bez pobočky', '#eb4d79')
                 + _card(f'{pct:.0f}\u202f%', 'pokrytí ORP', '#2770f0')
                 + '</div>')

        max_c = int(orp_counts['pocet'].max()) if not orp_counts.empty else 1
        with_rows = ''
        for i, (_, r) in enumerate(orp_counts.iterrows()):
            bg  = '#fff' if i % 2 == 0 else '#f8fbff'
            cnt = int(r['pocet'])
            bar = min(cnt / max_c * 100, 100)
            col = '#0bb440' if cnt >= 2 else '#2770f0'
            with_rows += (f'<tr style="background:{bg};">'
                          f'<td style="padding:4px 10px;font-size:0.75rem;color:#aaa;">{r["ORP_KOD"]}</td>'
                          f'<td style="padding:4px 10px;font-size:0.8rem;font-weight:600;">{r["ORP_NAZEV"]}</td>'
                          f'<td style="padding:4px 10px;"><div style="display:flex;align-items:center;gap:6px;">'
                          f'<div style="width:55px;background:#eef4ff;border-radius:3px;height:6px;">'
                          f'<div style="width:{bar:.0f}%;background:{col};height:100%;border-radius:3px;"></div></div>'
                          f'<span style="font-size:0.8rem;font-weight:700;color:{col};">{cnt}</span>'
                          f'</div></td></tr>')
        with_tbl = (f'<div style="flex:1;min-width:280px;">'
                    f'<div style="font-size:0.73rem;font-weight:700;color:#0bb440;text-transform:uppercase;'
                    f'letter-spacing:.4px;margin-bottom:7px;">&#x2705; ORP s pobočkou ({n_cov})</div>'
                    f'<div style="max-height:400px;overflow-y:auto;border-radius:6px;border:1px solid #e8edf5;">'
                    f'<table style="width:100%;border-collapse:collapse;">'
                    f'<thead><tr style="background:#eefaf0;position:sticky;top:0;z-index:2;">'
                    f'<th style="padding:5px 10px;color:#888;text-align:left;font-size:0.7rem;">Kód</th>'
                    f'<th style="padding:5px 10px;color:#888;text-align:left;font-size:0.7rem;">ORP</th>'
                    f'<th style="padding:5px 10px;color:#888;text-align:left;font-size:0.7rem;">Poboček</th>'
                    f'</tr></thead><tbody>{with_rows}</tbody></table></div></div>')

        without_rows = ''
        for i, (kod, nazev) in enumerate(sorted(missing.items(), key=lambda x: x[1])):
            bg = '#fff' if i % 2 == 0 else '#fff8f8'
            without_rows += (f'<tr style="background:{bg};">'
                             f'<td style="padding:4px 10px;font-size:0.75rem;color:#aaa;">{kod}</td>'
                             f'<td style="padding:4px 10px;font-size:0.8rem;color:#eb4d79;font-weight:500;">{nazev}</td>'
                             f'</tr>')
        without_tbl = (f'<div style="flex:1;min-width:220px;">'
                       f'<div style="font-size:0.73rem;font-weight:700;color:#eb4d79;text-transform:uppercase;'
                       f'letter-spacing:.4px;margin-bottom:7px;">&#x274C; ORP bez pobočky ({n_miss})</div>'
                       f'<div style="max-height:400px;overflow-y:auto;border-radius:6px;border:1px solid #fde8e8;">'
                       f'<table style="width:100%;border-collapse:collapse;">'
                       f'<thead><tr style="background:#fdecea;position:sticky;top:0;z-index:2;">'
                       f'<th style="padding:5px 10px;color:#888;text-align:left;font-size:0.7rem;">Kód</th>'
                       f'<th style="padding:5px 10px;color:#888;text-align:left;font-size:0.7rem;">ORP</th>'
                       f'</tr></thead><tbody>{without_rows}</tbody></table></div></div>')

        return (f'<div id="orp-panel-{uid}" style="display:none;">'
                f'{cards}'
                f'<div style="display:flex;gap:24px;flex-wrap:wrap;align-items:flex-start;">'
                f'{with_tbl}{without_tbl}'
                f'</div></div>')

    panel_all  = _build_panel(df,      'all')
    panel_keep = _build_panel(df_keep, 'keep')
    n_keep = len(df_keep)
    n_all  = len(df)

    btn_style_active   = "padding:6px 16px;border-radius:6px;border:1px solid #2770f0;background:#2770f0;color:#fff;font-size:0.82rem;cursor:pointer;font-weight:600;"
    btn_style_inactive = "padding:6px 16px;border-radius:6px;border:1px solid #ccc;background:#f0f4f8;color:#555;font-size:0.82rem;cursor:pointer;font-weight:600;"

    toggle_js = (
        "document.getElementById('orp-panel-all').style.display='';"
        "document.getElementById('orp-panel-keep').style.display='none';"
        "document.getElementById('orp-btn-all').setAttribute('style','" + btn_style_active + "');"
        "document.getElementById('orp-btn-keep').setAttribute('style','" + btn_style_inactive + "');"
    )
    toggle_js_keep = (
        "document.getElementById('orp-panel-keep').style.display='';"
        "document.getElementById('orp-panel-all').style.display='none';"
        "document.getElementById('orp-btn-keep').setAttribute('style','" + btn_style_active + "');"
        "document.getElementById('orp-btn-all').setAttribute('style','" + btn_style_inactive + "');"
    )

    return (f'<div style="background:#fff;border-radius:10px;padding:16px 18px;'
            f'border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">'
            f'<div style="font-size:0.72rem;color:#2770f0;font-weight:700;text-transform:uppercase;'
            f'letter-spacing:.5px;margin-bottom:10px;">&#x1F5FA; Pokryt&#237; ORP pobo&#269;kovou s&#237;t&#237;</div>'
            f'<div style="display:flex;gap:8px;margin-bottom:14px;">'
            f'<button id="orp-btn-all" onclick="{toggle_js}" style="{btn_style_active}">Všechny pobočky ({n_all})</button>'
            f'<button id="orp-btn-keep" onclick="{toggle_js_keep}" style="{btn_style_inactive}">Pouze Keep ({n_keep})</button>'
            f'</div>'
            f'{panel_all}{panel_keep}'
            f'<script>(function(){{document.getElementById("orp-panel-all").style.display=""}})()</script>'
            f'</div>')


def generate_ie_top5_html(df, title="Top 5 — Index expozice", n=5):
    """Top N poboček s nejlepším indexem expozice — design jako render_top5_cards."""
    _ie = 'HODNOTA_INDEXU_EXPOZICE'
    if _ie not in df.columns:
        return ''
    d = df.copy()
    d[_ie] = pd.to_numeric(d[_ie], errors='coerce')
    d = d[d[_ie].notna()].nlargest(n, _ie)
    if d.empty:
        return ''

    _RANK_COLS = {1:'#0ab33f', 2:'#42db72', 3:'#74aaf5', 4:'#9cc4f8', 5:'#c2dafc'}
    _ie_dot_scale = {1:'#0ab33f', 2:'#42db72', 3:'#f0b55d', 4:'#f0845d', 5:'#eb4d79'}

    cards = ''
    for rank, (_, row) in enumerate(d.iterrows(), 1):
        rcol = _RANK_COLS.get(rank, '#c2dafc')
        bn   = str(row.get('BRANCH_NAME', int(row.get('BRANCH_CODE', 0))))
        reg  = str(row.get('REGION_NAME', ''))
        ie   = float(row[_ie])
        ir   = int(float(row['IR'])) if pd.notna(row.get('IR')) else '—'
        qi   = int(float(row['IE_Q'])) if 'IE_Q' in row and pd.notna(row.get('IE_Q')) else None
        fmt  = str(row.get('BRANCH_FORMAT', '') or '').lower().strip()
        fmt_lbl = fmt.title() if fmt else '—'
        fmt_col = {'small':'#f59e0b','medium economy':'#fb923c',
                   'medium':'#2770f0','flagship':'#0ab33f'}.get(fmt, '#aaa')

        qi_html = (
            f'<span style="color:{_ie_dot_scale.get(qi,"#aaa")};font-size:16px;">&#9679;</span>'
            f'<span style="font-size:0.7rem;color:#888;margin-left:2px;">Q{qi}</span>'
        ) if qi else ''

        cards += (
            f'<div style="display:flex;align-items:stretch;gap:0;border-radius:8px;overflow:hidden;'
            f'border:1px solid #e8edf5;margin-bottom:8px;background:#fff;'
            f'box-shadow:0 1px 4px rgba(0,0,0,0.05);">'
            f'<div style="background:{rcol};color:white;font-size:1rem;font-weight:900;'
            f'min-width:40px;display:flex;align-items:center;justify-content:center;'
            f'flex-shrink:0;">{rank}</div>'
            f'<div style="padding:9px 14px;flex:1;min-width:0;">'
            f'<div style="display:flex;align-items:baseline;gap:8px;flex-wrap:wrap;">'
            f'<span style="font-weight:700;font-size:0.9rem;color:#0f172a;">{bn}</span>'
            f'<span style="font-size:0.72rem;color:#aaa;">{reg}</span>'
            f'</div>'
            f'<div style="margin-top:4px;display:flex;align-items:center;gap:10px;flex-wrap:wrap;">'
            f'<span style="color:#0ab33f;font-size:1.1rem;font-weight:900;">{ie:.2f}</span>'
            f'<span style="font-size:0.72rem;color:#888;">index expozice</span>'
            f'{qi_html}'
            f'</div>'
            f'<div style="margin-top:3px;font-size:0.72rem;color:#aaa;">'
            f'Rating {ir} &nbsp;·&nbsp; '
            f'<span style="color:{fmt_col};">{fmt_lbl}</span>'
            f'</div>'
            f'</div></div>'
        )

    return (
        f'<div style="background:#fafbff;border-radius:8px;padding:14px 16px;'
        f'border:1px solid #e8edf5;flex:1;min-width:260px;">'
        f'<div style="font-size:0.75rem;font-weight:700;color:#2e7d32;'
        f'text-transform:uppercase;letter-spacing:.4px;margin-bottom:10px;">{title}</div>'
        f'{cards}'
        f'</div>'
    )


def generate_atm_region_html(df_reg, export_atm=None, kapacita_atm=None):
    """
    Přehled ATM pro region:
    - Souhrnné počty bankomatů
    - Top 10 nejhůře kapacitně zajištěných poboček
    - Top 10 nejvytíženějších bankomatů (dle průměrného počtu výběrů + vkladů / měsíc)
    """
    if export_atm is None or kapacita_atm is None:
        return ""
    if export_atm.empty or kapacita_atm.empty:
        return ""

    try:
        # Připrav branch_codes regionu
        reg_codes = set(pd.to_numeric(df_reg['BRANCH_CODE'], errors='coerce').dropna().astype(int))
        if not reg_codes:
            return ""

        # ── Propoj kapacita_atm s regionem ───────────────────────
        _kid = next((c for c in ['BRANCH_ID','BRANCH_CODE'] if c in kapacita_atm.columns), None)
        _aid = next((c for c in ['BRANCH_ID','BRANCH_CODE','POBOCKA'] if c in export_atm.columns), None)
        if not _kid or not _aid:
            return ""

        kap = kapacita_atm.copy()
        kap[_kid] = pd.to_numeric(kap[_kid], errors='coerce')
        kap_reg = kap[kap[_kid].isin(reg_codes)].copy()

        atm = export_atm.copy()
        atm[_aid] = pd.to_numeric(atm[_aid], errors='coerce')
        atm_reg = atm[atm[_aid].isin(reg_codes)].copy()

        if kap_reg.empty and atm_reg.empty:
            return ""

        # Přidej BRANCH_NAME do kap_reg
        _bn_map = dict(zip(
            pd.to_numeric(df_reg['BRANCH_CODE'], errors='coerce'),
            df_reg['BRANCH_NAME']
        ))

        # ── 1. Souhrnné statistiky ────────────────────────────────
        n_atm_total = len(atm_reg)
        n_branches_with_atm = atm_reg[_aid].nunique() if n_atm_total > 0 else 0

        def _pct(col):
            if col not in kap_reg.columns: return None
            return pd.to_numeric(kap_reg[col], errors='coerce') * 100

        def _num_sum(col):
            if col not in kap_reg.columns: return None
            s = pd.to_numeric(kap_reg[col], errors='coerce')
            return float(s.sum()) if not s.isna().all() else None

        _avg_akt  = float(_pct('AKTUALNI_VYTIZENI_STROJU').mean())  if _pct('AKTUALNI_VYTIZENI_STROJU')  is not None and not _pct('AKTUALNI_VYTIZENI_STROJU').isna().all()  else None
        _avg_prev = float(_pct('VYTIZENI_VSECH_ATM_PO_PREVODU').mean()) if _pct('VYTIZENI_VSECH_ATM_PO_PREVODU') is not None and not _pct('VYTIZENI_VSECH_ATM_PO_PREVODU').isna().all() else None
        _sum_vybery_prumer = _num_sum('PRUMER._POCET_VYBERU_ATM_/_MESIC')
        _sum_vklady_prumer = _num_sum('PRUMER._POCET_VKLADU_ATM_/_MESIC')
        _sum_vybery_prev   = _num_sum('POCET_ATM_VYBERU_PO_PREVODU_/_M')
        _sum_vklady_prev   = _num_sum('POCET_ATM_VKLADU_PO_PREVODU_/_M')

        def _fmt_pct(v):
            if v is None: return '—'
            c = '#eb4d79' if v > 100 else ('#f59e0b' if v > 80 else '#0bb440')
            return f'<b style="color:{c};">{v:.0f} %{"⚠️" if v>100 else ""}</b>'

        def _fmt_num(v):
            if v is None: return '—'
            return f'{int(v):,}'.replace(',', '\u00a0')

        summary_html = (
            f'<div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;">'
            + ''.join(
                f'<div style="background:#f8fbff;border:1px solid #e8edf5;border-radius:8px;'
                f'padding:10px 16px;text-align:center;min-width:110px;">'
                f'<div style="font-size:1.3rem;font-weight:800;color:{vc};">{vv}</div>'
                f'<div style="font-size:0.68rem;color:#888;text-transform:uppercase;">{vl}</div>'
                f'</div>'
                for vv, vl, vc in [
                    (str(n_atm_total),             'bankomatů celkem',       '#2770f0'),
                    (_fmt_pct(_avg_akt),            'průměr. akt. vytíž.',    '#e07020'),
                    (_fmt_pct(_avg_prev),           'průměr. po převodu',     '#2770f0'),
                    (_fmt_num(_sum_vybery_prumer),  'výběry průměr/měs.',     '#e07020'),
                    (_fmt_num(_sum_vklady_prumer),  'vklady průměr/měs.',     '#2770f0'),
                    (_fmt_num(_sum_vybery_prev),    'výběry po převodu/měs.', '#e07020'),
                    (_fmt_num(_sum_vklady_prev),    'vklady po převodu/měs.', '#2770f0'),
                ]
            )
            + '</div>'
        )

        # ── 2. Top 10 nejhůře kapacitně zajištěných poboček ───────
        # Řadíme dle ZBYVAJICI_VOLNA_KAPACITA vzestupně — záporná = přetíženo
        worst_html = ''
        if 'ZBYVAJICI_VOLNA_KAPACITA' in kap_reg.columns:
            kap_reg['_volno_pct'] = pd.to_numeric(
                kap_reg['ZBYVAJICI_VOLNA_KAPACITA'], errors='coerce') * 100
            worst5 = (kap_reg.dropna(subset=['_volno_pct'])
                      .sort_values('_volno_pct', ascending=True)   # nejméně volná = nahoře
                      .head(10))
            if not worst5.empty:
                rows = ''
                for i, (_, r) in enumerate(worst5.iterrows()):
                    bc   = int(r[_kid])
                    bnam = _bn_map.get(bc, str(bc))
                    akt  = float(r.get('AKTUALNI_VYTIZENI_STROJU', 0) or 0) * 100
                    prev = float(r.get('VYTIZENI_VSECH_ATM_PO_PREVODU', 0) or 0) * 100
                    vol  = float(r['_volno_pct'])
                    n_atm_b = len(atm_reg[atm_reg[_aid] == bc])
                    bg = '#fff' if i % 2 == 0 else '#fafbff'
                    _pc = lambda v: (
                        f'<span style="font-weight:700;color:{"#eb4d79" if v>100 else "#f59e0b" if v>80 else "#0bb440"};">'
                        f'{v:.0f}%{"⚠️" if v>100 else ""}</span>'
                    )
                    # Volno: záporné = červeně (přetíženo), kladné = zeleně
                    _vol_col = '#eb4d79' if vol < 0 else '#0bb440'
                    _vol_pfx = '⚠️ ' if vol < 0 else ''
                    rows += (
                        f'<tr style="background:{bg};">'
                        f'<td style="padding:5px 10px;font-weight:600;color:#0f172a;">{bnam}</td>'
                        f'<td style="padding:5px 10px;text-align:center;color:#888;">{n_atm_b}</td>'
                        f'<td style="padding:5px 10px;text-align:center;">{_pc(akt)}</td>'
                        f'<td style="padding:5px 10px;text-align:center;">{_pc(prev)}</td>'
                        f'<td style="padding:5px 10px;text-align:center;color:{_vol_col};font-weight:700;">'
                        f'{_vol_pfx}{vol:.0f}%</td>'
                        f'</tr>'
                    )
                worst_html = (
                    f'<div style="flex:1;min-width:300px;">'
                    f'<div style="font-size:0.75rem;font-weight:700;color:#e07020;'
                    f'text-transform:uppercase;letter-spacing:.4px;margin-bottom:8px;">'
                    f'⚠️ Top 10 — nejhůře kapacitně zajištěné pobočky</div>'
                    f'<table style="width:100%;border-collapse:collapse;font-size:0.78rem;">'
                    f'<thead><tr style="background:#fff5ef;">'
                    f'<th style="padding:5px 10px;text-align:left;color:#888;">Pobočka</th>'
                    f'<th style="padding:5px 10px;text-align:center;color:#888;">ATM</th>'
                    f'<th style="padding:5px 10px;text-align:center;color:#888;">Akt.</th>'
                    f'<th style="padding:5px 10px;text-align:center;color:#888;">Po přev.</th>'
                    f'<th style="padding:5px 10px;text-align:center;color:#888;">Volno</th>'
                    f'</tr></thead>'
                    f'<tbody>{rows}</tbody>'
                    f'</table></div>'
                )

        # ── 3. Top 10 nejvytíženějších bankomatů ──────────────────
        # Vytíženost = průměr výběrů + vkladů / měsíc z nových sloupců
        busy_html = ''
        _vyb_col = 'PRUMER._POCET_VYBERU_ATM_/_MESIC'
        _vkl_col = 'PRUMER._POCET_VKLADU_ATM_/_MESIC'

        # Propojíme export_atm s kapacita_atm přes BRANCH_CODE
        atm_kap = atm_reg.merge(
            kap_reg[[_kid, _vyb_col, _vkl_col]].rename(columns={_kid: _aid}),
            on=_aid, how='left'
        ) if (_vyb_col in kap_reg.columns or _vkl_col in kap_reg.columns) else atm_reg.copy()

        # Pokud jsou sloupce přímo v export_atm, použij je odtud
        if _vyb_col not in atm_kap.columns and _vyb_col in atm_reg.columns:
            atm_kap = atm_reg.copy()

        if _vyb_col in atm_kap.columns or _vkl_col in atm_kap.columns:
            atm_kap['_busy'] = (
                pd.to_numeric(atm_kap.get(_vyb_col, 0), errors='coerce').fillna(0)
                + pd.to_numeric(atm_kap.get(_vkl_col, 0), errors='coerce').fillna(0)
            )
            busy5 = atm_kap.sort_values('_busy', ascending=False).head(10)
            if not busy5.empty:
                rows = ''
                for i, (_, r) in enumerate(busy5.iterrows()):
                    bc   = int(r[_aid]) if pd.notna(r.get(_aid)) else 0
                    bnam = _bn_map.get(bc, str(bc))
                    atm_id   = str(r.get('ATM_ID', '—'))
                    atm_type = str(r.get('ATM_TYPE', '—'))
                    loc      = str(r.get('BRANCH_LOCATION', '—'))
                    vyb  = float(r.get(_vyb_col, 0) or 0)
                    vkl  = float(r.get(_vkl_col, 0) or 0)
                    busy = vyb + vkl
                    bg = '#fff' if i % 2 == 0 else '#fafbff'
                    rows += (
                        f'<tr style="background:{bg};">'
                        f'<td style="padding:5px 10px;font-weight:600;color:#2770f0;">{atm_id}</td>'
                        f'<td style="padding:5px 10px;color:#555;font-size:0.75rem;">{atm_type}</td>'
                        f'<td style="padding:5px 10px;color:#555;font-size:0.75rem;">{loc}</td>'
                        f'<td style="padding:5px 10px;color:#888;font-size:0.75rem;">{bnam}</td>'
                        f'<td style="padding:5px 10px;text-align:right;font-weight:700;color:#e07020;">{int(vyb):,}</td>'
                        f'<td style="padding:5px 10px;text-align:right;font-weight:700;color:#2770f0;">{int(vkl):,}</td>'
                        f'<td style="padding:5px 10px;text-align:right;font-weight:800;color:#0f172a;">{int(busy):,}</td>'
                        f'</tr>'
                    )
                busy_html = (
                    f'<div style="flex:1;min-width:320px;">'
                    f'<div style="font-size:0.75rem;font-weight:700;color:#2770f0;'
                    f'text-transform:uppercase;letter-spacing:.4px;margin-bottom:8px;">'
                    f'📈 Top 10 — nejvytíženější bankomaty (výběry + vklady / měsíc)</div>'
                    f'<table style="width:100%;border-collapse:collapse;font-size:0.78rem;">'
                    f'<thead><tr style="background:#eef4ff;">'
                    f'<th style="padding:5px 10px;text-align:left;color:#888;">ATM ID</th>'
                    f'<th style="padding:5px 10px;text-align:left;color:#888;">Typ</th>'
                    f'<th style="padding:5px 10px;text-align:left;color:#888;">Umístění</th>'
                    f'<th style="padding:5px 10px;text-align:left;color:#888;">Pobočka</th>'
                    f'<th style="padding:5px 10px;text-align:right;color:#e07020;">⬆️ Výběry</th>'
                    f'<th style="padding:5px 10px;text-align:right;color:#2770f0;">⬇️ Vklady</th>'
                    f'<th style="padding:5px 10px;text-align:right;color:#0f172a;">Celkem</th>'
                    f'</tr></thead>'
                    f'<tbody>{rows}</tbody>'
                    f'</table></div>'
                )

        if not worst_html and not busy_html:
            return ""

        return (
            f'<div style="background:#fff;border-radius:10px;padding:16px 18px;'
            f'border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">'
            f'<div style="font-size:0.72rem;color:#2770f0;font-weight:700;'
            f'text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px;">'
            f'🏧 Bankomaty — přehled regionu</div>'
            f'{summary_html}'
            f'<div style="display:flex;gap:20px;flex-wrap:wrap;align-items:flex-start;">'
            f'{worst_html}{busy_html}'
            f'</div></div>'
        )

    except Exception as e:
        return f"<p style='color:#aaa;font-style:italic;'>ATM přehled nedostupný: {e}</p>"


def generate_cohort_analysis(df):
    """
    Kohortová analýza výnosů — pobočky seskupené dle různých dimenzí.
    Zahrnuje pouze pobočky s vypočteným interním ratingem (IR_FLAG == Y).
    """
    # ── Filtr: pouze pobočky s IR ratingem ────────────────────────
    if 'IR_FLAG' in df.columns:
        d = df[df['IR_FLAG'] == 'Y'].copy()
    elif 'IR' in df.columns:
        d = df[df['IR'].notna() & (df['IR'] > 0)].copy()
    else:
        d = df.copy()

    for c in ['OBJEM_VYNOSU_CZK', 'FTE', 'VYNOSY', 'ROCNI_SPLATKY_S_DPH_CZK',
              'PLOCHA_POUZE_D5', 'CAP_Q', 'CAP_SCHUZKY_FTE_Q']:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors='coerce')

    sections = []

    # ── helper: jedna sekce z groupby ─────────────────────────────
    # Barevná škála pro pořadí dle průměrného výnosu (1.=nejlepší=zelená, poslední=červená)
    _rank_colors  = ['#0ab33f', '#42db72', '#74aaf5', '#fb923c', '#eb4d79',
                     '#aaa', '#ccc', '#e0e0e0']
    _cat_color    = '#2770f0'   # jednotná modrá pro kategorické sekce

    def _make_section(groups, label_col, label_map, default_col='#888', is_quintile=False):
        """Karet styl — každá skupina = karta s barevným pruhem.
        is_quintile=True  → barvy dle kvintilového pořadí (zelená→červená)
        is_quintile=False → jednotná modrá (kategorická proměnná)
        """
        rows = ''
        for rank, (_, r) in enumerate(groups.iterrows()):
            key  = r[label_col]
            lbl, _col = label_map.get(key, (str(key) if pd.notna(key) else '—', default_col))
            bar_col = (_rank_colors[rank] if rank < len(_rank_colors) else '#aaa') if is_quintile else _cat_color
            pct     = float(r['rev_mean']) / (float(groups['rev_mean'].max()) or 1) * 100
            fte_html = (
                f'<span style="font-size:0.72rem;color:#0bb440;">'
                f'FTE: {format_money(float(r["rev_fte"]))}</span>'
                if pd.notna(r.get('rev_fte')) and float(r.get('rev_fte', 0)) > 0 else ''
            )
            rows += (
                f'<div style="display:flex;align-items:stretch;gap:0;border-radius:8px;'
                f'overflow:hidden;border:1px solid #e8edf5;margin-bottom:8px;background:#fff;'
                f'box-shadow:0 1px 4px rgba(0,0,0,0.05);">'
                f'<div style="background:{bar_col};min-width:8px;flex-shrink:0;"></div>'
                f'<div style="padding:9px 14px;flex:1;min-width:0;">'
                f'<div style="display:flex;align-items:baseline;gap:8px;flex-wrap:wrap;">'
                f'<span style="font-weight:700;font-size:0.88rem;color:#0f172a;">{lbl}</span>'
                f'<span style="font-size:0.72rem;color:#aaa;">{int(r["pocet"])} pb.</span>'
                f'</div>'
                f'<div style="margin-top:5px;background:#f0f0f0;border-radius:3px;'
                f'height:5px;overflow:hidden;">'
                f'<div style="width:{pct:.0f}%;background:{bar_col};height:100%;'
                f'border-radius:3px;"></div></div>'
                f'<div style="margin-top:5px;display:flex;align-items:center;'
                f'gap:10px;flex-wrap:wrap;">'
                f'<span style="font-size:0.9rem;font-weight:800;color:{bar_col};">'
                f'⌀ {format_money(float(r["rev_mean"]))}</span>'
                f'<span style="font-size:0.72rem;color:#888;">'
                f'med. {format_money(float(r["rev_med"]))}</span>'
                f'{fte_html}'
                f'</div>'
                f'</div></div>'
            )
        return rows

    def _agg(grp_col):
        return d.groupby(grp_col, observed=True).agg(
            pocet     = ('BRANCH_CODE', 'count'),
            rev_mean  = ('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med   = ('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte   = ('OBJEM_VYNOSU_CZK',
                         lambda x: (x / d.loc[x.index, 'FTE']).mean()
                         if 'FTE' in d.columns else float('nan')),
        ).reset_index().sort_values('rev_mean', ascending=False)

    # ── Sekce 1: BRANCH_FORMAT (small/medium/flagship) ────────────
    if 'BRANCH_FORMAT' in d.columns and d['BRANCH_FORMAT'].notna().any():
        _fmt_map = {
            'small':          ('Small',          '#f59e0b'),
            'medium economy': ('Medium Economy', '#fb923c'),
            'medium':         ('Medium',         '#2770f0'),
            'flagship':       ('Flagship',       '#0bb440'),
        }
        d['_fmt_clean'] = d['BRANCH_FORMAT'].astype(str).str.lower().str.strip()
        g = _agg('_fmt_clean')
        g['_fmt_clean'] = g['_fmt_clean'].str.lower().str.strip()
        # Seřaď dle logiky formátu
        _fmt_order = ['small', 'medium economy', 'medium', 'flagship']
        g['_ord'] = g['_fmt_clean'].map({v: i for i, v in enumerate(_fmt_order)}).fillna(99)
        g = g.sort_values('_ord')
        rows = _make_section(g, '_fmt_clean', _fmt_map)
        if rows:
            sections.append(('🏢 Formát pobočky', rows))

    # ── Sekce 2: Formát NF/SF/klasika (BRANCH_BUILDING_NF_SF) ─────
    _nfsf_col = next((c for c in ['BRANCH_BUILDING_NF_SF'] if c in d.columns), None)
    if _nfsf_col:
        g = _agg(_nfsf_col)
        _nfsf_colors = ['#1a5fb4', '#2770f0', '#4a90d9', '#7ab4f5', '#aecef9', '#ccc']
        _nfsf_map = {
            str(v): (str(v), _nfsf_colors[i % len(_nfsf_colors)])
            for i, v in enumerate(g[_nfsf_col].dropna().unique())
        }
        rows = _make_section(g, _nfsf_col, _nfsf_map)
        if rows:
            sections.append(('🏗️ Formát NF/SF', rows))

    # ── Sekce 3: Velikost pobočky dle FTE ─────────────────────────
    if 'FTE' in d.columns and d['FTE'].notna().any():
        _q33 = d['FTE'].quantile(0.33)
        _q67 = d['FTE'].quantile(0.67)
        def _size_label(fte):
            if pd.isna(fte): return None
            if fte <= _q33:  return f'Malá (≤{_q33:.0f} FTE)'
            if fte <= _q67:  return f'Střední ({_q33:.0f}–{_q67:.0f} FTE)'
            return f'Velká (>{_q67:.0f} FTE)'
        d['_size'] = d['FTE'].apply(_size_label)
        d_sz = d[d['_size'].notna()]
        _size_order = [f'Malá (≤{_q33:.0f} FTE)', f'Střední ({_q33:.0f}–{_q67:.0f} FTE)', f'Velká (>{_q67:.0f} FTE)']
        _size_map = {
            _size_order[0]: (_size_order[0], '#a5b4fc'),
            _size_order[1]: (_size_order[1], '#2770f0'),
            _size_order[2]: (_size_order[2], '#1a4db5'),
        }
        g = d_sz.groupby('_size', observed=True).agg(
            pocet    = ('BRANCH_CODE', 'count'),
            rev_mean = ('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med  = ('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte  = ('OBJEM_VYNOSU_CZK',
                        lambda x: (x / d_sz.loc[x.index, 'FTE']).mean()),
        ).reset_index()
        g['_ord'] = g['_size'].map({v: i for i, v in enumerate(_size_order)}).fillna(99)
        g = g.sort_values('_ord')
        rows = _make_section(g, '_size', _size_map)
        if rows:
            sections.append(('📐 Velikost dle FTE', rows))

    # ── Sekce 4: Business rating kvintil ──────────────────────────
    if 'CAP_Q' in d.columns and d['CAP_Q'].notna().any():
        _bq_map = {
            1: ('BQ1 — Business elite', '#0bb440'),
            2: ('BQ2 — Nadprůměr',      '#4ade80'),
            3: ('BQ3 — Průměr',          '#f59e0b'),
            4: ('BQ4 — Podprůměr',       '#fb923c'),
            5: ('BQ5 — Slabé',           '#eb4d79'),
        }
        d['_bq'] = pd.to_numeric(d['CAP_Q'], errors='coerce')
        d_bq = d[d['_bq'].notna() & d['_bq'].between(1, 5)].copy()
        d_bq['_bq'] = d_bq['_bq'].astype(int)
        g = d_bq.groupby('_bq', observed=True).agg(
            pocet    = ('BRANCH_CODE', 'count'),
            rev_mean = ('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med  = ('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte  = ('OBJEM_VYNOSU_CZK',
                        lambda x: (x / d_bq.loc[x.index, 'FTE']).mean()
                        if 'FTE' in d_bq.columns else float('nan')),
        ).reset_index().sort_values('_bq')
        rows = _make_section(g, '_bq', _bq_map, is_quintile=True)
        if rows:
            sections.append(('📈 Business rating kvintil', rows))

    # ── Sekce 5: Schůzky/FTE kvintil ──────────────────────────────
    if 'CAP_SCHUZKY_FTE_Q' in d.columns and d['CAP_SCHUZKY_FTE_Q'].notna().any():
        _sq_map = {
            1: ('SQ1 — Nejvíce schůzek/FTE', '#0bb440'),
            2: ('SQ2 — Nadprůměr',            '#4ade80'),
            3: ('SQ3 — Průměr',                '#f59e0b'),
            4: ('SQ4 — Podprůměr',             '#fb923c'),
            5: ('SQ5 — Nejméně schůzek/FTE',  '#eb4d79'),
        }
        d['_sq'] = pd.to_numeric(d['CAP_SCHUZKY_FTE_Q'], errors='coerce')
        d_sq = d[d['_sq'].notna() & d['_sq'].between(1, 5)].copy()
        d_sq['_sq'] = d_sq['_sq'].astype(int)
        g = d_sq.groupby('_sq', observed=True).agg(
            pocet             = ('BRANCH_CODE', 'count'),
            rev_mean          = ('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med           = ('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte           = ('OBJEM_VYNOSU_CZK',
                                 lambda x: (x / d_sq.loc[x.index, 'FTE']).mean()
                                 if 'FTE' in d_sq.columns else float('nan')),
        ).reset_index().sort_values('_sq')
        rows = _make_section(g, '_sq', _sq_map, is_quintile=True)
        if rows:
            sections.append(('🤝 Schůzky/FTE kvintil', rows))

    # ── Sekce 6: Kvintil nájemného ────────────────────────────────
    _rent_q_col = 'ROCNI_SPLATKY_S_DPH_CZK_Q'
    if _rent_q_col in d.columns and d[_rent_q_col].notna().any():
        d[_rent_q_col] = pd.to_numeric(d[_rent_q_col], errors='coerce')
        _rq_map = {
            1: ('NQ1 — Nejlevnější nájemné', '#0bb440'),
            2: ('NQ2 — Nižší nájemné',       '#4ade80'),
            3: ('NQ3 — Průměrné nájemné',    '#f59e0b'),
            4: ('NQ4 — Vyšší nájemné',       '#fb923c'),
            5: ('NQ5 — Nejdražší nájemné',   '#eb4d79'),
        }
        d_rq = d[d[_rent_q_col].notna() & d[_rent_q_col].between(1, 5)].copy()
        d_rq[_rent_q_col] = d_rq[_rent_q_col].astype(int)
        g = d_rq.groupby(_rent_q_col, observed=True).agg(
            pocet     = ('BRANCH_CODE', 'count'),
            rev_mean  = ('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med   = ('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte   = ('OBJEM_VYNOSU_CZK',
                         lambda x: (x / d_rq.loc[x.index, 'FTE']).mean()
                         if 'FTE' in d_rq.columns else float('nan')),
            rent_mean = ('ROCNI_SPLATKY_S_DPH_CZK', 'mean'),
        ).reset_index().sort_values(_rent_q_col)
        rows = _make_section(g, _rent_q_col, _rq_map, is_quintile=True)
        if rows:
            sections.append(('🏠 Kvintil nájemného', rows))

    # ── Sekce 7: Plocha D5 kvintil ────────────────────────────────
    _d5_col = 'PLOCHA_POUZE_D5'
    if _d5_col in d.columns and d[_d5_col].notna().any() and (d[_d5_col] > 0).any():
        d_d5 = d[d[_d5_col] > 0].copy()
        try:
            d_d5['_d5q'] = pd.qcut(d_d5[_d5_col], q=5, labels=False, duplicates='drop') + 1
            d_d5['_d5q'] = d_d5['_d5q'].astype(int)
            _d5q_map = {
                1: ('D5Q1 — Nejmenší plocha D5', '#0bb440'),
                2: ('D5Q2',                       '#4ade80'),
                3: ('D5Q3 — Střední plocha D5',   '#f59e0b'),
                4: ('D5Q4',                       '#fb923c'),
                5: ('D5Q5 — Největší plocha D5',  '#eb4d79'),
            }
            g = d_d5.groupby('_d5q', observed=True).agg(
                pocet    = ('BRANCH_CODE', 'count'),
                rev_mean = ('OBJEM_VYNOSU_CZK', 'mean'),
                rev_med  = ('OBJEM_VYNOSU_CZK', 'median'),
                rev_fte  = ('OBJEM_VYNOSU_CZK',
                            lambda x: (x / d_d5.loc[x.index, 'FTE']).mean()
                            if 'FTE' in d_d5.columns else float('nan')),
                d5_mean  = (_d5_col, 'mean'),
            ).reset_index().sort_values('_d5q')
            rows = _make_section(g, '_d5q', _d5q_map, is_quintile=True)
            if rows:
                sections.append(('📐 Plocha D5 kvintil', rows))
        except Exception:
            pass

    # ── Sekce 8: Typologie pobočky (BRANCH_TYPE_NAME) ─────────────
    if 'BRANCH_TYPE_NAME' in d.columns and d['BRANCH_TYPE_NAME'].notna().any():
        _typ_colors = ['#2770f0', '#4d8ef0', '#74aaf5', '#9cc4f8', '#c2dafc',
                       '#a5b4fc', '#818cf8', '#6366f1']
        g = _agg('BRANCH_TYPE_NAME')
        _typ_map = {
            str(v): (str(v), _typ_colors[i % len(_typ_colors)])
            for i, v in enumerate(g['BRANCH_TYPE_NAME'].dropna().unique())
        }
        rows = _make_section(g, 'BRANCH_TYPE_NAME', _typ_map)
        if rows:
            sections.append(('🏢 Typologie pobočky', rows))

    # ── Sekce 9: Kvintil indexu expozice (IE_Q) ───────────────────
    _ie_q_col = 'IE_Q'
    if _ie_q_col in d.columns and d[_ie_q_col].notna().any():
        d[_ie_q_col] = pd.to_numeric(d[_ie_q_col], errors='coerce')
        _ieq_map = {
            1: ('IEQ1 — Nejlepší index expozice', '#0bb440'),
            2: ('IEQ2 — Nadprůměrná expozice',    '#4ade80'),
            3: ('IEQ3 — Průměrná expozice',        '#f59e0b'),
            4: ('IEQ4 — Podprůměrná expozice',     '#fb923c'),
            5: ('IEQ5 — Nejhorší index expozice',  '#eb4d79'),
        }
        d_ie = d[d[_ie_q_col].notna() & d[_ie_q_col].between(1, 5)].copy()
        d_ie[_ie_q_col] = d_ie[_ie_q_col].astype(int)
        g = d_ie.groupby(_ie_q_col, observed=True).agg(
            pocet    = ('BRANCH_CODE', 'count'),
            rev_mean = ('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med  = ('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte  = ('OBJEM_VYNOSU_CZK',
                        lambda x: (x / d_ie.loc[x.index, 'FTE']).mean()
                        if 'FTE' in d_ie.columns else float('nan')),
        ).reset_index().sort_values(_ie_q_col)
        rows = _make_section(g, _ie_q_col, _ieq_map, is_quintile=True)
        if rows:
            sections.append(('📡 Kvintil indexu expozice', rows))

    # ── Sekce 10: PEREX kvintil (CAP_PEREX_Q) ─────────────────────
    _perex_q_col = 'CAP_PEREX_Q'
    if _perex_q_col in d.columns and d[_perex_q_col].notna().any():
        d[_perex_q_col] = pd.to_numeric(d[_perex_q_col], errors='coerce')
        _pq_map = {
            1: ('PQ1 — Nejefektivnější PEREX',  '#0bb440'),
            2: ('PQ2 — Nadprůměrný PEREX',      '#4ade80'),
            3: ('PQ3 — Průměrný PEREX',          '#f59e0b'),
            4: ('PQ4 — Podprůměrný PEREX',       '#fb923c'),
            5: ('PQ5 — Nejméně efektivní PEREX', '#eb4d79'),
        }
        d_pq = d[d[_perex_q_col].notna() & d[_perex_q_col].between(1, 5)].copy()
        d_pq[_perex_q_col] = d_pq[_perex_q_col].astype(int)
        g = d_pq.groupby(_perex_q_col, observed=True).agg(
            pocet    = ('BRANCH_CODE', 'count'),
            rev_mean = ('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med  = ('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte  = ('OBJEM_VYNOSU_CZK',
                        lambda x: (x / d_pq.loc[x.index, 'FTE']).mean()
                        if 'FTE' in d_pq.columns else float('nan')),
        ).reset_index().sort_values(_perex_q_col)
        rows = _make_section(g, _perex_q_col, _pq_map, is_quintile=True)
        if rows:
            sections.append(('⚡ PEREX kvintil', rows))

    # ── Sekce 11: Primární klienti kvintil (CAP_PRIMCLI_FTE_Q) ────
    _prim_q_col = 'CAP_PRIMCLI_FTE_Q'
    if _prim_q_col in d.columns and d[_prim_q_col].notna().any():
        d[_prim_q_col] = pd.to_numeric(d[_prim_q_col], errors='coerce')
        _primq_map = {
            1: ('PKQ1 — Nejvíce prim. klientů/FTE', '#0bb440'),
            2: ('PKQ2 — Nadprůměr',                  '#4ade80'),
            3: ('PKQ3 — Průměr',                      '#f59e0b'),
            4: ('PKQ4 — Podprůměr',                   '#fb923c'),
            5: ('PKQ5 — Nejméně prim. klientů/FTE',  '#eb4d79'),
        }
        d_primq = d[d[_prim_q_col].notna() & d[_prim_q_col].between(1, 5)].copy()
        d_primq[_prim_q_col] = d_primq[_prim_q_col].astype(int)
        g = d_primq.groupby(_prim_q_col, observed=True).agg(
            pocet    = ('BRANCH_CODE', 'count'),
            rev_mean = ('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med  = ('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte  = ('OBJEM_VYNOSU_CZK',
                        lambda x: (x / d_primq.loc[x.index, 'FTE']).mean()
                        if 'FTE' in d_primq.columns else float('nan')),
        ).reset_index().sort_values(_prim_q_col)
        rows = _make_section(g, _prim_q_col, _primq_map, is_quintile=True)
        if rows:
            sections.append(('👤 Primární klienti/FTE kvintil', rows))

    # ── Sekce 12: Počet bankéřů (BANKERS_COUNT) ───────────────────
    if 'BANKERS_COUNT' in d.columns and d['BANKERS_COUNT'].notna().any():
        d['BANKERS_COUNT'] = pd.to_numeric(d['BANKERS_COUNT'], errors='coerce')
        d_bnk = d[d['BANKERS_COUNT'].notna()].copy()
        d_bnk['BANKERS_COUNT'] = d_bnk['BANKERS_COUNT'].astype(int)
        # Seskup do skupin: 0, 1, 2, 3, 4-5, 6+
        def _bnk_grp(n):
            if n == 0: return '0 — bez bankéřů'
            if n == 1: return '1 bankéř'
            if n == 2: return '2 bankéři'
            if n == 3: return '3 bankéři'
            if n <= 5: return '4–5 bankéřů'
            return '6+ bankéřů'
        d_bnk['_bnk_grp'] = d_bnk['BANKERS_COUNT'].apply(_bnk_grp)
        _bnk_order = ['0 — bez bankéřů', '1 bankéř', '2 bankéři', '3 bankéři', '4–5 bankéřů', '6+ bankéřů']
        _bnk_colors = ['#eb4d79', '#fb923c', '#f59e0b', '#4ade80', '#0bb440', '#047857']
        _bnk_map = {v: (v, _bnk_colors[i % len(_bnk_colors)]) for i, v in enumerate(_bnk_order)}
        g = d_bnk.groupby('_bnk_grp', observed=True).agg(
            pocet    = ('BRANCH_CODE', 'count'),
            rev_mean = ('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med  = ('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte  = ('OBJEM_VYNOSU_CZK',
                        lambda x: (x / d_bnk.loc[x.index, 'FTE']).mean()
                        if 'FTE' in d_bnk.columns else float('nan')),
        ).reset_index()
        g['_ord'] = g['_bnk_grp'].map({v: i for i, v in enumerate(_bnk_order)}).fillna(99)
        g = g.sort_values('_ord')
        rows = _make_section(g, '_bnk_grp', _bnk_map)
        if rows:
            sections.append(('🏦 Počet bankéřů', rows))

    if not sections:
        return "<p style='color:#aaa;font-style:italic;'>Nedostatek dat pro kohortovou analýzu.</p>"

    _panels = ""
    for _title, _content in sections:
        _panels += (
            f'<div style="flex:1;min-width:280px;background:#fafbff;border-radius:8px;'
            f'padding:14px 16px;border:1px solid #e8edf5;">'
            f'<div style="font-size:0.75rem;font-weight:700;color:#444;text-transform:uppercase;'
            f'letter-spacing:.4px;margin-bottom:12px;">{_title}</div>'
            f'{_content}'
            f'</div>'
        )

    return (
        f'<div style="background:#fff;border-radius:10px;padding:16px 18px;'
        f'border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">'
        f'<div style="font-size:0.72rem;color:#2770f0;font-weight:700;text-transform:uppercase;'
        f'letter-spacing:.5px;margin-bottom:4px;">📊 Kohortová analýza výnosů</div>'
        f'<div style="font-size:0.72rem;color:#aaa;margin-bottom:14px;">'
        f'Pouze pobočky s vypočteným interním ratingem · '
        f'průměrné nové výnosy a výnos/FTE dle různých dimenzí</div>'
        f'<div style="display:flex;gap:20px;flex-wrap:wrap;align-items:flex-start;">'
        f'{_panels}'
        f'</div></div>'
    )



    # Kohorty dle formátu pobočky
    _fmt_col = next((c for c in ['BRANCH_BUILDING_NF_SF', 'BRANCH_FORMAT', 'FORMAT_2024_(FIX_FS)'] if c in d.columns), None)
    _has_fmt = _fmt_col is not None

    # Kohorty dle FTE velikosti (malá / střední / velká)
    _has_fte = 'FTE' in d.columns and d['FTE'].notna().any()

    rows_html = ""
    sections  = []

    # ── Sekce 1: Formát pobočky ─────────────────────────────────────
    if _has_fmt:
        _fmt_groups = d.groupby(_fmt_col, observed=True).agg(
            pocet=('BRANCH_CODE', 'count'),
            rev_mean=('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med=('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte_mean=('OBJEM_VYNOSU_CZK', lambda x: (
                (x / d.loc[x.index, 'FTE']).mean() if 'FTE' in d.columns else float('nan')
            )),
        ).reset_index().sort_values('rev_mean', ascending=False)
        _fmt_max = float(_fmt_groups['rev_mean'].max() or 1)

        _fmt_rows = ""
        for _, r in _fmt_groups.iterrows():
            _pct = float(r['rev_mean']) / _fmt_max * 100
            _fmt_rows += f"""
<div style="display:flex;align-items:center;gap:10px;margin-bottom:9px;flex-wrap:wrap;">
  <span style="font-size:0.8rem;font-weight:700;color:#0f172a;min-width:90px;
               white-space:nowrap;">{r[_fmt_col] or '—'}</span>
  <span style="font-size:0.72rem;color:#aaa;min-width:38px;">{int(r['pocet'])} pb.</span>
  <div style="flex:1;min-width:100px;background:#eef4ff;border-radius:4px;height:10px;overflow:hidden;">
    <div style="width:{_pct:.1f}%;background:#2770f0;height:100%;border-radius:4px;"></div>
  </div>
  <span style="font-size:0.8rem;font-weight:700;color:#2770f0;min-width:90px;text-align:right;">
    ⌀ {format_money(float(r['rev_mean']))}</span>
  <span style="font-size:0.72rem;color:#888;min-width:80px;">
    med. {format_money(float(r['rev_med']))}</span>
  {"" if pd.isna(r['rev_fte_mean']) else
    f'<span style="font-size:0.72rem;color:#0bb440;min-width:90px;">FTE: {format_money(float(r["rev_fte_mean"]))}</span>'}
</div>"""
        sections.append(('🏢 Formát pobočky', _fmt_rows))

    # ── Sekce 2: Velikost pobočky dle FTE ──────────────────────────
    if _has_fte:
        _q33 = d['FTE'].quantile(0.33)
        _q67 = d['FTE'].quantile(0.67)
        def _size_label(fte):
            if pd.isna(fte): return 'Neznámá'
            if fte <= _q33:  return f'Malá (≤ {_q33:.1f} FTE)'
            if fte <= _q67:  return f'Střední ({_q33:.1f}–{_q67:.1f} FTE)'
            return f'Velká (> {_q67:.1f} FTE)'
        d['_size'] = d['FTE'].apply(_size_label)
        _size_order = [f'Malá (≤ {_q33:.1f} FTE)', f'Střední ({_q33:.1f}–{_q67:.1f} FTE)', f'Velká (> {_q67:.1f} FTE)']
        _size_cols  = ['#a5b4fc', '#2770f0', '#1a4db5']

        _sz_groups = d.groupby('_size', observed=True).agg(
            pocet=('BRANCH_CODE', 'count'),
            rev_mean=('OBJEM_VYNOSU_CZK', 'mean'),
            rev_med=('OBJEM_VYNOSU_CZK', 'median'),
            rev_fte=('OBJEM_VYNOSU_CZK', lambda x: (x / d.loc[x.index, 'FTE']).mean()),
        ).reset_index()
        _sz_max = float(_sz_groups['rev_mean'].max() or 1)

        _sz_rows = ""
        for _si, _slbl in enumerate(_size_order):
            _sr = _sz_groups[_sz_groups['_size'] == _slbl]
            if _sr.empty: continue
            _sr = _sr.iloc[0]
            _pct = float(_sr['rev_mean']) / _sz_max * 100
            _scol = _size_cols[_si]
            _sz_rows += f"""
<div style="display:flex;align-items:center;gap:10px;margin-bottom:9px;flex-wrap:wrap;">
  <span style="font-size:0.8rem;font-weight:700;color:#0f172a;min-width:160px;white-space:nowrap;">{_slbl}</span>
  <span style="font-size:0.72rem;color:#aaa;min-width:38px;">{int(_sr['pocet'])} pb.</span>
  <div style="flex:1;min-width:100px;background:#eef4ff;border-radius:4px;height:10px;overflow:hidden;">
    <div style="width:{_pct:.1f}%;background:{_scol};height:100%;border-radius:4px;"></div>
  </div>
  <span style="font-size:0.8rem;font-weight:700;color:{_scol};min-width:90px;text-align:right;">
    ⌀ {format_money(float(_sr['rev_mean']))}</span>
  <span style="font-size:0.72rem;color:#888;min-width:80px;">med. {format_money(float(_sr['rev_med']))}</span>
  <span style="font-size:0.72rem;color:#0bb440;min-width:90px;">FTE: {format_money(float(_sr['rev_fte']))}</span>
</div>"""
        sections.append(('📐 Velikost pobočky (FTE)', _sz_rows))

    # ── Sekce 3: Business rating kvintil ───────────────────────────
    if 'CAP_Q' in d.columns:
        d['CAP_Q'] = pd.to_numeric(d['CAP_Q'], errors='coerce')
        _bq_labels = {1: ('BQ1 — Business elite', '#0bb440'),
                      2: ('BQ2 — Nadprůměr',      '#4ade80'),
                      3: ('BQ3 — Průměr',          '#f59e0b'),
                      4: ('BQ4 — Podprůměr',       '#fb923c'),
                      5: ('BQ5 — Slabé',           '#eb4d79')}
        _bq_groups = d.groupby('CAP_Q', observed=True).agg(
            pocet=('BRANCH_CODE', 'count'),
            rev_mean=('OBJEM_VYNOSU_CZK', 'mean'),
            rev_fte=('OBJEM_VYNOSU_CZK', lambda x: (x / d.loc[x.index, 'FTE']).mean() if 'FTE' in d.columns else float('nan')),
        ).reset_index()
        _bq_max = float(_bq_groups['rev_mean'].max() or 1)

        _bq_rows = ""
        for _, bqr in _bq_groups.sort_values('CAP_Q').iterrows():
            _bqi = int(bqr['CAP_Q']) if pd.notna(bqr['CAP_Q']) else 0
            _bqlbl, _bqcol = _bq_labels.get(_bqi, (f'BQ{_bqi}', '#888'))
            _pct = float(bqr['rev_mean']) / _bq_max * 100
            _bq_rows += f"""
<div style="display:flex;align-items:center;gap:10px;margin-bottom:9px;flex-wrap:wrap;">
  <span style="font-size:0.8rem;font-weight:700;color:{_bqcol};min-width:150px;white-space:nowrap;">{_bqlbl}</span>
  <span style="font-size:0.72rem;color:#aaa;min-width:38px;">{int(bqr['pocet'])} pb.</span>
  <div style="flex:1;min-width:100px;background:#eef4ff;border-radius:4px;height:10px;overflow:hidden;">
    <div style="width:{_pct:.1f}%;background:{_bqcol};height:100%;border-radius:4px;"></div>
  </div>
  <span style="font-size:0.8rem;font-weight:700;color:{_bqcol};min-width:90px;text-align:right;">
    ⌀ {format_money(float(bqr['rev_mean']))}</span>
  {"" if pd.isna(bqr['rev_fte']) else f'<span style="font-size:0.72rem;color:#0bb440;min-width:90px;">FTE: {format_money(float(bqr["rev_fte"]))}</span>'}
</div>"""
        sections.append(('📈 Business rating kvintil', _bq_rows))

    # ── Sekce 4: Schůzky/FTE kvintil ───────────────────────────────
    if 'CAP_SCHUZKY_FTE_Q' in d.columns:
        d['CAP_SCHUZKY_FTE_Q'] = pd.to_numeric(d['CAP_SCHUZKY_FTE_Q'], errors='coerce')
        _sq_labels = {1: ('SQ1 — Nejvíce schůzek/FTE', '#0bb440'),
                      2: ('SQ2 — Nadprůměr',            '#4ade80'),
                      3: ('SQ3 — Průměr',                '#f59e0b'),
                      4: ('SQ4 — Podprůměr',             '#fb923c'),
                      5: ('SQ5 — Nejméně schůzek/FTE',  '#eb4d79')}
        _sq_groups = d.groupby('CAP_SCHUZKY_FTE_Q', observed=True).agg(
            pocet=('BRANCH_CODE', 'count'),
            schuzky_fte_mean=('CAP_SCHUZKY_FTE', 'mean'),
            rev_mean=('OBJEM_VYNOSU_CZK', 'mean'),
        ).reset_index()
        _sq_rev_max = float(_sq_groups['rev_mean'].max() or 1)

        _sq_rows = ""
        for _, sqr in _sq_groups.sort_values('CAP_SCHUZKY_FTE_Q').iterrows():
            _sqi = int(sqr['CAP_SCHUZKY_FTE_Q']) if pd.notna(sqr['CAP_SCHUZKY_FTE_Q']) else 0
            _sqlbl, _sqcol = _sq_labels.get(_sqi, (f'SQ{_sqi}', '#888'))
            _pct = float(sqr['rev_mean']) / _sq_rev_max * 100
            _s_val = f"{float(sqr['schuzky_fte_mean']):.2f}" if pd.notna(sqr['schuzky_fte_mean']) else "—"
            _sq_rows += f"""
<div style="display:flex;align-items:center;gap:10px;margin-bottom:9px;flex-wrap:wrap;">
  <span style="font-size:0.8rem;font-weight:700;color:{_sqcol};min-width:180px;white-space:nowrap;">{_sqlbl}</span>
  <span style="font-size:0.72rem;color:#aaa;min-width:38px;">{int(sqr['pocet'])} pb.</span>
  <div style="flex:1;min-width:100px;background:#eef4ff;border-radius:4px;height:10px;overflow:hidden;">
    <div style="width:{_pct:.1f}%;background:{_sqcol};height:100%;border-radius:4px;"></div>
  </div>
  <span style="font-size:0.8rem;font-weight:700;color:{_sqcol};min-width:90px;text-align:right;">
    ⌀ {format_money(float(sqr['rev_mean']))}</span>
  <span style="font-size:0.72rem;color:#888;min-width:80px;">sch./FTE: {_s_val}</span>
</div>"""
        sections.append(('🤝 Schůzky/FTE kvintil', _sq_rows))

    if not sections:
        return "<p style='color:#aaa;font-style:italic;'>Nedostatek dat pro kohortovou analýzu.</p>"

    # Sestav HTML
    _panels = ""
    for _title, _content in sections:
        _panels += f"""
<div style="flex:1;min-width:280px;background:#fafbff;border-radius:8px;
            padding:14px 16px;border:1px solid #e8edf5;">
  <div style="font-size:0.75rem;font-weight:700;color:#444;text-transform:uppercase;
              letter-spacing:.4px;margin-bottom:12px;">{_title}</div>
  {_content}
</div>"""

    return f"""
<div style="background:#fff;border-radius:10px;padding:16px 18px;
            border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">
  <div style="font-size:0.72rem;color:#2770f0;font-weight:700;text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:4px;">📊 Kohortová analýza výnosů</div>
  <div style="font-size:0.72rem;color:#aaa;margin-bottom:14px;">
    Průměrné nové výnosy a výnos/FTE dle formátu, velikosti a ratingového kvintilu
  </div>
  <div style="display:flex;gap:20px;flex-wrap:wrap;align-items:flex-start;">
    {_panels}
  </div>
</div>"""


def generate_network_simulation_200(df, spadovky_df=None, target_n=250):
    """
    Simulace sítě 250 poboček s výhledem do roku 2030.

    Logika:
    1. Vyřaď všechny pobočky označené ke close (jakýkoliv rok)
    2. Simuluj přesun klientů ze zavřených poboček dle spádovosti
       — kapacita dle formátu: small 10 %, medium-economy 25 %, medium 40 %, flagship 70 %
       — zbytek → online
    3. Simuluj interní rating 2030: C/I trend × 40 % + výnosový trend × 60 %
       (CAGR z REVENUES_2022→2024 projektován na +5 let)
    4. Sestav tabulku seřazenou dle simulovaného ratingu
    5. Zobraz kapacitu osobních bankéřů (1 bankéř = 4 schůzky/den × 248 pracovních dnů)
    """
    import math as _math

    BANKER_POSITIONS = [
        "osobní bankéř - junior", "osobní bankéř - medior",
        "osobní bankéř - senior", "osobní bankéř - master",
    ]
    # Kapacita formátu: kolik % navíc klientů pobočka zvládne
    FORMAT_CAP = {'small': 0.10, 'medium economy': 0.25, 'medium': 0.40, 'flagship': 0.70}
    # ── Kapacita bankéře ────────────────────────────────────────────
    # 5 schůzek/den × 248 pracovních dní = 1 240 schůzek/rok na bankéře
    # Skutečná zátěž = fyzické schůzky + online schůzky
    #                 + 30 % bezhotovostního walk-in (trvá ~20 min → přepočet na 45 min)
    #                   → 30 % × bezhot_walkin × (20/45) ekvivalent plné schůzky
    # HOT walk-in NEBERE se v potaz (hotovostní obsluha není bankéřská schůzka)
    MEETINGS_PER_DAY   = 5
    WORKING_DAYS       = 248
    BANKER_CAP_YEAR    = MEETINGS_PER_DAY * WORKING_DAYS   # 1 240 schůzek/bankéř/rok
    BEZHOT_TO_SCH_RATE = 0.30   # 30 % bezhot walk-in → schůzka
    BEZHOT_SCH_EQUIV   = 20/45  # bezhot walk-in trvá ~20 min vs. 45 min schůzka
    PORTFOLIO_PER_BANKER = 1500  # každý bankéř zvládne max 1 500 klientů v portfoliu

    d = df.copy()
    for c in ['OBJEM_VYNOSU_CZK','VYNOSY','PRIME_NAKLADY','PRIME_NAKLADY/VYNOSY',
              'POCET_NAVSTEV_CELKEM','FTE','IR','BRANCH_CODE','POCET_KLIENTU',
              'POCET_SCHUZEK_FYZICKY','POCET_SCHUZEK_ONLINE','POCET_BEZHOT_WALK_IN',
              'REVENUES_2022','REVENUES_2023','REVENUES_2024']:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors='coerce')

    if 'BRANCH_FORMAT' not in d.columns:
        d['BRANCH_FORMAT'] = 'medium'

    # ── 1. Rozděl na close a zbývající ───────────────────────────
    _is_close = (
        d['FOOTPRINT_STRATEGIE_(2030)'].astype(str).str.lower()
         .str.contains('close', na=False)
        if 'FOOTPRINT_STRATEGIE_(2030)' in d.columns
        else pd.Series(False, index=d.index)
    )
    d_close  = d[_is_close].copy()
    d_remain = d[~_is_close].copy()

    n_close  = len(d_close)
    n_remain = len(d_remain)

    d_remain = d_remain.set_index('BRANCH_CODE', drop=False)

    # ── 2. Kapacita dle formátu ───────────────────────────────────
    d_remain['_fmt_cap'] = (
        d_remain['BRANCH_FORMAT'].str.lower().str.strip()
        .map(lambda f: FORMAT_CAP.get(f, FORMAT_CAP['medium']))
    )
    d_remain['_max_extra_cli'] = d_remain['POCET_KLIENTU'].fillna(0) * d_remain['_fmt_cap']

    # ── Výnos / klient pro přepočet výnosů ───────────────────────
    d_remain['_rev_per_cli'] = np.where(
        d_remain['POCET_KLIENTU'].fillna(0) > 0,
        d_remain['OBJEM_VYNOSU_CZK'].fillna(0) / d_remain['POCET_KLIENTU'],
        0
    )

    # ── 3. Přesun klientů ze zavřených poboček ────────────────────
    d_remain['_delta_cli']           = 0.0
    d_remain['_delta_cli_online']    = 0.0
    d_remain['_delta_rev']           = 0.0
    # Kumulované "vážené výnosy spádových klientů" pro výpočet výnosu/klient ze spádu
    # = součet (cli_ze_zavřené × rev_per_cli_zavřené) — vlastní váha každé zavřené pobočky
    d_remain['_delta_spado_rev_sum'] = 0.0   # součet výnosů od spádových klientů
    d_remain['_delta_spado_cli_sum'] = 0.0   # počet fyzicky přijatých spádových klientů
    # Zdroje: {target_bc: [(close_name, close_bc, cli_phys, cli_online, redirected)]}
    _sources = {bc: [] for bc in d_remain.index}

    if spadovky_df is not None and not spadovky_df.empty:
        _sp = spadovky_df.copy()
        _sp.columns = [c.strip().upper() for c in _sp.columns]
        _sp['BRANCH_ID'] = pd.to_numeric(_sp['BRANCH_ID'], errors='coerce')

        for _, crow in d_close.iterrows():
            bc      = int(crow['BRANCH_CODE'])
            bc_cli  = float(crow.get('POCET_KLIENTU') or 0)
            sp_r    = _sp[_sp['BRANCH_ID'] == bc]
            if sp_r.empty or bc_cli <= 0:
                continue
            sp_r = sp_r.iloc[0]

            # Celkové spádové návštěvy → váhy
            total_vis = sum(
                float(sp_r.get(f'TOP_POBOCKA_VISITS_{i}', 0) or 0)
                for i in [1, 2, 3]
            )
            if total_vis <= 0:
                continue

            candidates = []
            for _i in [1, 2, 3]:
                _tid  = sp_r.get(f'TOP_POBOCKA_ID_{_i}')
                _tvis = sp_r.get(f'TOP_POBOCKA_VISITS_{_i}', 0)
                if pd.isna(_tid) or pd.isna(_tvis): continue
                try: _tid = int(float(_tid)); _tvis = float(_tvis)
                except: continue
                if _tid not in d_remain.index or _tvis <= 0: continue
                _weight = _tvis / total_vis
                candidates.append((_tid, _weight))

            if not candidates:
                continue

            # Přerozděl klienty dle spádové váhy — BEZ kapacitního re-routingu
            # (re-routing způsoboval duplicitní labely: stejná pobočka dostala
            #  záznam jako cíl i jako přesměrování jiného kandidáta)
            for _orig_tid, _weight in candidates:
                _cli_wave   = bc_cli * _weight
                _cap_rem    = max(0, d_remain.at[_orig_tid, '_max_extra_cli']
                                     - d_remain.at[_orig_tid, '_delta_cli'])
                _cli_phys   = min(_cli_wave, _cap_rem)
                _cli_online = _cli_wave - _cli_phys
                # Výnos/klient zavřené pobočky (ne cílové!)
                _close_rpv  = float(crow.get('OBJEM_VYNOSU_CZK') or 0)
                _close_cli  = float(crow.get('POCET_KLIENTU') or 1)
                _rpv_close  = _close_rpv / _close_cli if _close_cli > 0 else 0.0

                d_remain.at[_orig_tid, '_delta_cli']           += _cli_phys
                d_remain.at[_orig_tid, '_delta_cli_online']    += _cli_online
                d_remain.at[_orig_tid, '_delta_rev']           += (
                    _cli_phys * _rpv_close + _cli_online * _rpv_close * 0.60
                )
                # Track spádové výnosy zvlášť pro výpočet rev/klient 2030
                d_remain.at[_orig_tid, '_delta_spado_rev_sum'] += _cli_phys * _rpv_close
                d_remain.at[_orig_tid, '_delta_spado_cli_sum'] += _cli_phys
                # Přidej záznam pouze jednou za zavřenou pobočku (ne za kandidáta)
                _sources[_orig_tid].append((
                    str(crow.get('BRANCH_NAME', bc)), bc,
                    _cli_phys, _cli_online, False
                ))

    # ── 4. Simulované hodnoty 2025 (po přesunu) ───────────────────
    d_remain['SIM_KLIENTI']     = d_remain['POCET_KLIENTU'].fillna(0) + d_remain['_delta_cli']
    d_remain['SIM_KLIENTI_ONL'] = d_remain['_delta_cli_online']
    d_remain['SIM_REV']         = d_remain['OBJEM_VYNOSU_CZK'].fillna(0) + d_remain['_delta_rev']

    # ── 5. Výnosový trend → simulovaný výnos a rating 2030 ────────
    # Metodika (postavená na klientech a jejich průměrném výnosu):
    #
    # Základ: průměrný nový výnos na klienta = OBJEM_VYNOSU_CZK / POCET_KLIENTU
    # Trend:  porovnání průměrného nového výnosu 2024 vs 2025
    #         + porovnání celkových revenues 22→23→24→25 (jako korekce)
    #         váhy: 70% nové výnosy, 30% revenues
    # Výsledek: průměrný výnos/klient 2030 = výnos/klient_2025 × (1+r)^5
    # SIM výnosy 2030 = SIM_KLIENTI × průměrný výnos/klient 2030

    for _rc in ['REVENUES_2022', 'REVENUES_2023', 'REVENUES_2024', 'VYNOSY',
                'NEW_BUSINESS_2024_-_OBJEM_VYNOSU', 'OBJEM_VYNOSU_CZK',
                'NEW_BUSINESS_2024_-_POCTY_KS', 'POCET_PRODEJU', 'POCET_KLIENTU']:
        if _rc in d_remain.columns:
            d_remain[_rc] = pd.to_numeric(d_remain[_rc], errors='coerce')

    _cli    = d_remain['POCET_KLIENTU'].fillna(0)
    _nb25   = d_remain['OBJEM_VYNOSU_CZK'].fillna(0)     # nové výnosy 2025
    _nb24   = d_remain.get('NEW_BUSINESS_2024_-_OBJEM_VYNOSU',
                            pd.Series(0.0, index=d_remain.index)).fillna(0)
    _pr25   = d_remain.get('POCET_PRODEJU',
                            pd.Series(0.0, index=d_remain.index)).fillna(0)
    _pr24   = d_remain.get('NEW_BUSINESS_2024_-_POCTY_KS',
                            pd.Series(0.0, index=d_remain.index)).fillna(0)

    # Průměrný nový výnos na klienta v roce 2025
    _rev_per_cli_now = np.where(_cli > 0, _nb25 / _cli, 0.0)
    d_remain['_rev_per_cli_now'] = _rev_per_cli_now

    # Trend nového výnosu na klienta 2024 → 2025
    _rpc24 = np.where(_cli > 0, _nb24 / _cli, 0.0)
    _trend_new_per_cli = np.where(
        (_rpc24 > 0) & (_rev_per_cli_now > 0),
        np.clip(_rev_per_cli_now / _rpc24 - 1, -0.30, 0.50),
        0.0
    )

    # Trend celkových revenues (roční průměr 22→23→24→25)
    _r22 = d_remain['REVENUES_2022'].fillna(0)
    _r23 = d_remain['REVENUES_2023'].fillna(0)
    _r24 = d_remain['REVENUES_2024'].fillna(0)
    _r25 = d_remain['VYNOSY'].fillna(0)

    def _avg_rev_growth(r22, r23, r24, r25):
        pts = []
        for a, b in [(r22, r23), (r23, r24), (r24, r25)]:
            if a > 0 and b > 0:
                pts.append(b / a - 1)
        return float(np.clip(np.mean(pts), -0.20, 0.30)) if pts else 0.0

    _trend_rev = np.vectorize(_avg_rev_growth)(_r22, _r23, _r24, _r25)

    # Kombinovaná roční změna: 60% revenues + 40% trend výnosu/klienta
    _annual_change = np.clip(
        0.40 * _trend_new_per_cli + 0.60 * _trend_rev,
        -0.15, 0.25
    )
    d_remain['_cagr'] = _annual_change

    # Průměrný výnos/klient 2030 pro stávající klienty
    _rev_per_cli_2030 = _rev_per_cli_now * (1 + _annual_change) ** 5
    d_remain['_rev_per_cli_2030'] = _rev_per_cli_2030

    # Průměrný výnos/klient spádových klientů (jejich původní pobočka)
    # Trend pro spádové = stejná roční změna jako pro cílovou pobočku
    # (předpokládáme že po přechodu klient vygeneruje podobný výnos s trendem cílové pobočky)
    _spado_rev_sum = d_remain['_delta_spado_rev_sum'].fillna(0)
    _spado_cli_sum = d_remain['_delta_spado_cli_sum'].fillna(0)
    _rev_per_spado_cli = np.where(
        _spado_cli_sum > 0,
        _spado_rev_sum / _spado_cli_sum,
        _rev_per_cli_now   # fallback: pokud nejsou spádoví, použij vlastní průměr
    )
    _rev_per_spado_cli_2030 = _rev_per_spado_cli * (1 + _annual_change) ** 5
    d_remain['_rev_per_spado_cli_2030'] = _rev_per_spado_cli_2030

    # SIM výnosy 2030 = stávající klienti × jejich výnos/klient 2030
    #                  + spádoví klienti × jejich výnos/klient 2030
    _existing_cli = d_remain['POCET_KLIENTU'].fillna(0)
    d_remain['SIM_REV_2030'] = (
        _existing_cli              * _rev_per_cli_2030
        + _spado_cli_sum           * _rev_per_spado_cli_2030
    )

    # C/I simulace 2030 — čistě z historické řady, BEZ výnosového efektu
    # Vezmi dostupné hodnoty z let 2021-2025 a extrapoluj lineárním trendem
    # max dovolená změna za celých 6 let (2025→2030): ±15 % (absolutně, ne relativně)
    CI_MAX_CHANGE_6Y = 0.15   # max 15 procentních bodů změny za 6 let

    for _ci_col in ['C/I_RATIO_(YTD_2021)', 'C/I_RATIO_(YTD_2022)',
                    'C/I_RATIO_(YTD_2023)', 'C/I_RATIO_(YTD_2024)']:
        if _ci_col in d_remain.columns:
            d_remain[_ci_col] = pd.to_numeric(d_remain[_ci_col], errors='coerce')

    _ci_cols_years = [
        ('C/I_RATIO_(YTD_2021)', 2021),
        ('C/I_RATIO_(YTD_2022)', 2022),
        ('C/I_RATIO_(YTD_2023)', 2023),
        ('C/I_RATIO_(YTD_2024)', 2024),
        ('PRIME_NAKLADY/VYNOSY',  2025),
    ]
    _ci_now = d_remain['PRIME_NAKLADY/VYNOSY'].fillna(np.nan).abs()

    def _ci_trend_2030(row):
        """Lineární trend z dostupné historické řady C/I → extrapolace na 2030."""
        pts = []
        for col, yr in _ci_cols_years:
            if col in row.index:
                v = row[col]
                try:
                    fv = float(v)
                    if 0.01 < fv < 5.0:   # rozumný rozsah (1 % – 500 %)
                        pts.append((yr, fv))
                except (TypeError, ValueError):
                    pass

        ci_base = pts[-1][1] if pts else np.nan
        if np.isnan(ci_base) or ci_base <= 0:
            return np.nan

        if len(pts) >= 2:
            # Prostá lineární regrese OLS
            n   = len(pts)
            sx  = sum(p[0] for p in pts)
            sy  = sum(p[1] for p in pts)
            sxy = sum(p[0]*p[1] for p in pts)
            sxx = sum(p[0]**2 for p in pts)
            denom = n * sxx - sx * sx
            slope = (n * sxy - sx * sy) / denom if denom != 0 else 0.0
            years_ahead = 2030 - pts[-1][0]

            if slope <= 0:
                # Klesající nebo plochý trend → C/I se zlepšovalo nebo stagnuje.
                # V roce 2030 počítáme s mírným nárůstem +1–3 pp (realisticky
                # se vždy najdou tlaky na náklady). Absolutní přírůstek závisí
                # na aktuální hladině: nižší C/I → menší nárůst.
                drift = 0.01 + 0.02 * ci_base   # ~1 pp u 0 %, ~3 pp u 100 %
                ci_proj = ci_base + drift
            else:
                # Rostoucí trend → extrapoluj, ale clipni na max +15 pp
                ci_proj = ci_base + slope * years_ahead
        else:
            # Jediný bod → konzervativní nárůst +2 pp
            ci_proj = ci_base + 0.02

        # Omez maximální nárůst na 15 pp; nikdy nepovolíme pokles pod aktuální hodnotu
        ci_proj = float(np.clip(ci_proj, ci_base, ci_base + CI_MAX_CHANGE_6Y))
        return float(np.clip(ci_proj, 0.01, 3.0))

    d_remain['SIM_CI_2030'] = d_remain.apply(_ci_trend_2030, axis=1)
    # Fallback pro NaN
    d_remain['SIM_CI_2030'] = d_remain['SIM_CI_2030'].fillna(_ci_now)

    # Simulovaný rating 2030 (stejná metodika jako IR)
    _mask = d_remain['IR'].notna() & (d_remain['IR'] > 0)
    d_sim = d_remain[_mask].copy()

    d_sim['SIM_REV_RNK_2030']  = d_sim['SIM_REV_2030'].rank(ascending=False, method='first').astype(int)
    d_sim['SIM_CI_PERC_2030']  = (
        d_sim['SIM_CI_2030'].rank(method='first', pct=True, ascending=True) * 100
    ).apply(lambda x: max(1, int(round(x))))
    d_sim['SIM_IR_SCORE_2030'] = (
        d_sim['SIM_CI_PERC_2030'] * 0.4 + d_sim['SIM_REV_RNK_2030'] * 0.6
    )
    d_sim['SIM_IR_2030'] = (
        d_sim['SIM_IR_SCORE_2030'].rank(ascending=True, method='first').astype(int)
    )
    d_sim['SIM_IR_Q_2030'] = pd.qcut(d_sim['SIM_IR_2030'], q=5, labels=False) + 1

    # ── 6. Výběr top 250 dle simulovaného ratingu 2030 ────────────
    top  = d_sim.nsmallest(target_n, 'SIM_IR_2030').copy()
    rest = d_sim[~d_sim.index.isin(top.index)].sort_values('SIM_IR_2030').copy()

    # ── 7. Bankéřská kapacita ─────────────────────────────────────
    # Předpočítej průměrný počet schůzek na klienta pro celou síť
    # = (fyzické + online schůzky) / počet klientů  → použijeme jako koeficient
    _net_sch_total = (
        d_remain['POCET_SCHUZEK_FYZICKY'].fillna(0).sum()
        + d_remain['POCET_SCHUZEK_ONLINE'].fillna(0).sum()
        if 'POCET_SCHUZEK_FYZICKY' in d_remain.columns else 0
    )
    _net_cli_total = d_remain['POCET_KLIENTU'].fillna(0).sum()
    SCH_PER_CLIENT = (_net_sch_total / _net_cli_total) if _net_cli_total > 0 else 1.0
    # Bezpečný clip: realisticky 0.3–3 schůzky/klient/rok
    SCH_PER_CLIENT = float(np.clip(SCH_PER_CLIENT, 0.3, 3.0))

    def _banker_cap(row, bankers):
        """
        Vrátí (n_bankers, cap_sch, actual_sch, diff, html).
        Kapacita = bankers × 5 sch./den × 248 dní = bankers × 1 240
        Skutečná zátěž = fyzické + online + bezhot walk-in equiv + spádoví klienti
        """
        b = float(bankers)
        if b <= 0: return 0, 0, 0, 0, '—'

        cap   = b * BANKER_CAP_YEAR

        sch_fyz   = float(row.get('POCET_SCHUZEK_FYZICKY', 0) or 0)
        sch_onl   = float(row.get('POCET_SCHUZEK_ONLINE',  0) or 0)
        bezhot    = float(row.get('POCET_BEZHOT_WALK_IN',  0) or 0)
        delta_cli = float(row.get('_delta_cli', 0) or 0)
        sch_spado = delta_cli * SCH_PER_CLIENT
        actual    = (sch_fyz + sch_onl
                     + bezhot * BEZHOT_TO_SCH_RATE * BEZHOT_SCH_EQUIV
                     + sch_spado)

        diff  = cap - actual
        pct   = actual / cap * 100 if cap > 0 else 0
        clr   = '#0bb440' if diff >= 0 else '#eb4d79'
        lbl   = f"{'▲ +' if diff > 0 else '▼ '}{abs(int(diff)):,} sch."
        bar_w = min(int(pct), 100)
        bar_c = '#0bb440' if pct < 80 else ('#f59e0b' if pct < 100 else '#eb4d79')
        html  = (
            f"<div style='display:flex;align-items:center;gap:4px;'>"
            f"<div style='background:#f0f0f0;border-radius:3px;height:6px;"
            f"width:50px;overflow:hidden;flex-shrink:0;'>"
            f"<div style='background:{bar_c};width:{bar_w}%;height:100%;border-radius:3px;'></div></div>"
            f"<span style='font-size:0.68rem;font-weight:700;color:{clr};white-space:nowrap;'>{lbl}</span>"
            f"</div>"
        )
        return int(b), int(cap), int(actual), int(diff), html

    # Načti data specialistů
    _spec_g    = globals().get('df_specialiste')
    _spec_cols = globals().get('SPEC_POZICE_COLS', [])
    _bnk_cols  = []
    _spec_map  = {}  # branch_id → počet bankéřů
    _spec_total_map = {}  # branch_id → všechny pozice
    if _spec_g is not None and not _spec_g.empty and _spec_cols:
        _bnk_cols = [c for c in _spec_cols if c.lower().strip() in BANKER_POSITIONS]
        if not _bnk_cols:
            _bnk_cols = [c for c in _spec_cols if 'bankéř' in c.lower()]
        _sg = _spec_g.copy()
        _sg['branch_id'] = pd.to_numeric(_sg['branch_id'], errors='coerce')
        for c in _bnk_cols + _spec_cols:
            _sg[c] = pd.to_numeric(_sg[c], errors='coerce').fillna(0)
        if _bnk_cols:
            _sg['_bankers'] = _sg[_bnk_cols].sum(axis=1)
        else:
            _sg['_bankers'] = 0
        _sg['_all_pos'] = _sg[_spec_cols].sum(axis=1)
        _spec_map       = dict(zip(_sg['branch_id'].astype(int), _sg['_bankers']))
        _spec_total_map = dict(zip(_sg['branch_id'].astype(int), _sg['_all_pos']))

    # ── 8. Formátovací helpers ────────────────────────────────────
    _Q_COL = {1:'#0bb440', 2:'#4ade80', 3:'#f59e0b', 4:'#fb923c', 5:'#eb4d79'}

    def _fmtm(v):
        try: return format_money(float(v))
        except: return '—'

    def _fmti(v):
        try: return f"{int(float(v)):,}"
        except: return '—'

    def _fmtpct(v):
        try: return f"{float(v)*100:.1f} %"
        except: return '—'

    def _ir_diff_html(orig, sim):
        try:
            d2 = int(float(sim)) - int(float(orig))
            if d2 == 0: return "<span style='color:#aaa;'>±0</span>"
            clr = '#0bb440' if d2 < 0 else '#eb4d79'
            arr = '▲' if d2 < 0 else '▼'
            return f"<span style='color:{clr};font-weight:700;'>{arr}{abs(d2)}</span>"
        except: return '—'

    def _cagr_html(v):
        try:
            pct = float(v) * 100
            clr = '#0bb440' if pct >= 0 else '#eb4d79'
            arr = '▲' if pct >= 0 else '▼'
            return f"<span style='color:{clr};font-size:0.78rem;'>{arr}{abs(pct):.1f}%/r</span>"
        except: return '—'

    # ── 9. Stavba řádků tabulky ───────────────────────────────────
    rows_html = ""
    for rank, (_, row) in enumerate(top.sort_values('SIM_IR_2030').iterrows(), 1):
        bc       = int(row['BRANCH_CODE'])
        bn       = str(row.get('BRANCH_NAME', bc))
        reg      = str(row.get('REGION_NAME', ''))
        fmt_val  = str(row.get('BRANCH_FORMAT', '—') or '—').lower().strip()
        fmt_cap_pct = int(FORMAT_CAP.get(fmt_val, 0.40) * 100)
        fmt_lbl  = fmt_val.title() if fmt_val != '—' else '—'
        fmt_colors = {'small':'#f59e0b','medium economy':'#fb923c','medium':'#2770f0','flagship':'#0bb440'}
        fmt_col  = fmt_colors.get(fmt_val, '#aaa')

        # Rating hodnoty
        ir25      = row.get('IR', float('nan'))
        ir30      = row.get('SIM_IR_2030', float('nan'))
        q30       = int(row['SIM_IR_Q_2030']) if pd.notna(row.get('SIM_IR_Q_2030')) else 0
        qcol      = _Q_COL.get(q30, '#aaa')
        ir_diff   = _ir_diff_html(ir25, ir30)
        ir25_s    = f"{int(float(ir25))}" if pd.notna(ir25) else '—'
        ir30_s    = f"{int(float(ir30))}" if pd.notna(ir30) else '—'

        # Klienti
        cli_now   = row.get('POCET_KLIENTU', 0) or 0
        cli_sim   = row.get('SIM_KLIENTI', 0) or 0
        cli_onl   = row.get('SIM_KLIENTI_ONL', 0) or 0
        cli_delta = cli_sim - float(cli_now)
        cap_used  = cli_sim / (float(cli_now) * (1 + FORMAT_CAP.get(fmt_val, 0.40))) * 100 if cli_now else 0
        cap_used  = min(cap_used, 999)
        cap_col   = '#0bb440' if cap_used < 80 else ('#f59e0b' if cap_used < 100 else '#eb4d79')

        # Výnosy
        rev25     = row.get('OBJEM_VYNOSU_CZK', 0) or 0
        rev_sim   = row.get('SIM_REV', 0) or 0
        rev30     = row.get('SIM_REV_2030', 0) or 0
        cagr_v    = row.get('_cagr', 0) or 0

        # C/I
        ci25      = row.get('PRIME_NAKLADY/VYNOSY', float('nan'))
        ci30      = row.get('SIM_CI_2030', float('nan'))

        # Bankéři
        n_bankers = int(_spec_map.get(bc, 0))
        n_all_pos = int(_spec_total_map.get(bc, 0))
        _nb, _cap_sch, _act_sch, _diff_sch, _bnk_html = _banker_cap(row, n_bankers)

        # Potřební bankéři dle schůzkové kapacity
        import math as _m
        _needed_bankers = _m.ceil(_act_sch / BANKER_CAP_YEAR) if _act_sch > 0 else 0
        _banker_diff    = _needed_bankers - n_bankers
        _bnd_col        = '#0bb440' if _banker_diff <= 0 else '#eb4d79'
        _bnd_lbl        = (f"▲ +{_banker_diff} třeba doplnit" if _banker_diff > 0
                           else (f"▼ {abs(_banker_diff)} přebytek" if _banker_diff < 0
                                 else "✓ přesně"))
        _bnd_html = (
            f"<div style='font-size:0.68rem;font-weight:700;color:{_bnd_col};"
            f"margin-top:2px;'>{_bnd_lbl}</div>"
        )

        # Schůzky detaily
        _sch_fyz  = int(float(row.get('POCET_SCHUZEK_FYZICKY', 0) or 0))
        _sch_onl  = int(float(row.get('POCET_SCHUZEK_ONLINE',  0) or 0))
        _bezhot   = int(float(row.get('POCET_BEZHOT_WALK_IN',  0) or 0))
        _bezhot_e = round(_bezhot * BEZHOT_TO_SCH_RATE * BEZHOT_SCH_EQUIV)
        _delta_c  = float(row.get('_delta_cli', 0) or 0)
        _sch_spado = round(_delta_c * SCH_PER_CLIENT)

        # Zdrojové pobočky — všechny, ale max 4 na řádek (flex wrap)
        src_list  = _sources.get(bc, [])
        src_chips = ''
        if src_list:
            sorted_src = sorted(src_list, key=lambda x: -x[2])
            chips_html = ''.join(
                f"<span style='background:#f0f4ff;color:#2770f0;"
                f"border:1px solid #c7d9fb;"
                f"border-radius:6px;padding:1px 6px;font-size:0.65rem;"
                f"white-space:nowrap;margin:1px 2px 1px 0;display:inline-block;'>"
                f"{sn} <b>+{int(cp):,}</b>"
                f"</span>"
                for sn, _, cp, co, r in sorted_src
            )
            src_chips = (
                f'<div style="margin-top:3px;display:flex;flex-wrap:wrap;gap:2px;'
                f'max-width:220px;">{chips_html}</div>'
            )

        row_bg = '#eafaf1' if rank <= 10 else ('#fff' if rank % 2 == 0 else '#fafbff')

        rows_html += f"""
<tr style="background:{row_bg};">
  <td style="padding:5px 8px;border:1px solid #e8edf5;text-align:center;
             font-weight:700;color:#2770f0;font-size:0.8rem;">{rank}</td>
  <td style="padding:5px 8px;border:1px solid #e8edf5;white-space:nowrap;">
    <div style="font-weight:700;font-size:0.82rem;color:#0f172a;">{bn}</div>
    <div style="font-size:0.7rem;color:#aaa;">{reg} · #{bc}</div>
    <span style="background:{fmt_col}18;color:{fmt_col};border:1px solid {fmt_col}44;
                 border-radius:5px;padding:0 5px;font-size:0.65rem;font-weight:700;">
      {fmt_lbl} +{fmt_cap_pct}%</span>
    {src_chips}
  </td>
  <td style="padding:5px 8px;border:1px solid #e8edf5;text-align:center;">
    <span style="background:{qcol}18;color:{qcol};border:1px solid {qcol}44;
                 border-radius:8px;padding:1px 8px;font-weight:700;font-size:0.8rem;">Q{q30}</span>
  </td>
  <td style="padding:5px 8px;border:1px solid #e8edf5;text-align:center;font-size:0.8rem;">
    {ir25_s} → <b>{ir30_s}</b> {ir_diff}
  </td>
  <td style="padding:5px 8px;border:1px solid #e8edf5;font-size:0.78rem;">
    <div><b>{_fmti(cli_sim)}</b> <span style="color:#aaa;font-size:0.7rem;">(bylo {_fmti(cli_now)})</span></div>
    <div style="display:flex;align-items:center;gap:4px;margin-top:2px;">
      <div style="background:#f0f0f0;border-radius:3px;height:5px;width:60px;overflow:hidden;">
        <div style="background:{cap_col};width:{min(cap_used,100):.0f}%;height:100%;border-radius:3px;"></div>
      </div>
      <span style="font-size:0.65rem;color:{cap_col};">{cap_used:.0f}% kap.</span>
    </div>
    {f"<div style='font-size:0.68rem;color:#6b7280;margin-top:1px;'>🌐 {int(cli_onl):,} online</div>" if cli_onl > 1 else ''}
  </td>
  <td style="padding:5px 8px;border:1px solid #e8edf5;font-size:0.78rem;">
    <div><b>{_fmtm(rev_sim)}</b> <span style="color:#aaa;font-size:0.7rem;">roční výnosy 2025</span></div>
    <div style="color:#2770f0;font-size:0.78rem;margin-top:2px;font-weight:700;">
      {_fmtm(rev30)} <span style="font-weight:400;font-size:0.7rem;color:#2770f0;">roční 2030</span>
      {_cagr_html(cagr_v)}
    </div>
  </td>
  <td style="padding:5px 8px;border:1px solid #e8edf5;text-align:center;font-size:0.78rem;">
    <div>{_fmtpct(ci25)} → <b style="color:#2770f0;">{_fmtpct(ci30)}</b></div>
  </td>
  <td style="padding:5px 8px;border:1px solid #e8edf5;font-size:0.78rem;">
    <div style="font-weight:700;">{n_bankers} bankéřů / potřeba {_needed_bankers}</div>
    <div style="font-size:0.68rem;color:#aaa;">{n_all_pos} celkem pozic</div>
    <div style="font-size:0.68rem;color:#888;margin-top:2px;">
      skut. {_fmti(_act_sch)} sch. z {_fmti(_cap_sch)}
      <span style="color:#aaa;">({_fmti(_sch_fyz)}fyz+{_fmti(_sch_onl)}onl+{_fmti(_bezhot_e)}bw{('+'+_fmti(_sch_spado)+'sp') if _sch_spado > 0 else ''})</span>
    </div>
    {_bnk_html}
    {_bnd_html}
  </td>
</tr>"""

    # ── Rest rows — pobočky mimo top 250 ─────────────────────────
    rest_rows_html = ""
    for rest_rank, (_, rrow) in enumerate(rest.sort_values('SIM_IR_2030').iterrows(), 1):
        rbc   = int(rrow['BRANCH_CODE'])
        rbn   = str(rrow.get('BRANCH_NAME', rbc))
        rreg  = str(rrow.get('REGION_NAME', ''))
        rir30 = int(rrow['SIM_IR_2030']) if pd.notna(rrow.get('SIM_IR_2030')) else '—'
        rq30  = int(rrow['SIM_IR_Q_2030']) if pd.notna(rrow.get('SIM_IR_Q_2030')) else 0
        rqcol = _Q_COL.get(rq30, '#aaa')
        rir25 = int(float(rrow.get('IR', 0))) if pd.notna(rrow.get('IR')) else '—'
        rfmt  = str(rrow.get('BRANCH_FORMAT', '—') or '—').lower().strip()
        rfmt_col = {'small':'#f59e0b','medium economy':'#fb923c','medium':'#2770f0','flagship':'#0bb440'}.get(rfmt,'#aaa')
        rfmt_lbl = rfmt.title() if rfmt != '—' else '—'
        rest_rows_html += f"""
<tr class="sim-rest-row-{target_n}" style="background:#fafafa;opacity:0.75;">
  <td style="padding:4px 8px;border:1px solid #eee;text-align:center;
             font-weight:600;color:#aaa;font-size:0.78rem;">{target_n + rest_rank}</td>
  <td style="padding:4px 8px;border:1px solid #eee;white-space:nowrap;">
    <div style="font-weight:600;font-size:0.8rem;color:#555;">{rbn}</div>
    <div style="font-size:0.68rem;color:#bbb;">{rreg} · #{rbc}</div>
    <span style="background:{rfmt_col}18;color:{rfmt_col};border:1px solid {rfmt_col}44;
                 border-radius:5px;padding:0 5px;font-size:0.63rem;font-weight:600;">{rfmt_lbl}</span>
  </td>
  <td style="padding:4px 8px;border:1px solid #eee;text-align:center;">
    <span style="background:{rqcol}18;color:{rqcol};border:1px solid {rqcol}44;
                 border-radius:8px;padding:1px 7px;font-weight:700;font-size:0.78rem;">Q{rq30}</span>
  </td>
  <td style="padding:4px 8px;border:1px solid #eee;text-align:center;font-size:0.78rem;color:#888;">
    {rir25} → <b>{rir30}</b>
  </td>
  <td colspan="4" style="padding:4px 8px;border:1px solid #eee;font-size:0.72rem;color:#bbb;
                          text-align:center;font-style:italic;">
    mimo top {target_n}
  </td>
</tr>"""

    # ── 10. Souhrnné statistiky ───────────────────────────────────
    def _avg(col, dff):
        return float(dff[col].mean()) if col in dff.columns and dff[col].notna().any() else float('nan')

    _total_close_cli   = float(d_close['POCET_KLIENTU'].fillna(0).sum())
    _total_online_cli  = float(top['SIM_KLIENTI_ONL'].fillna(0).sum())
    _total_sim_cli     = float(top['SIM_KLIENTI'].fillna(0).sum())
    _total_bankers     = sum(_spec_map.get(int(bc), 0) for bc in top['BRANCH_CODE'].astype(int))
    _total_cap         = _total_bankers * BANKER_CAP_YEAR
    # Skutečná zátěž celé sítě
    def _safe(col): return float(top[col].fillna(0).sum()) if col in top.columns else 0.0
    _total_actual      = (_safe('POCET_SCHUZEK_FYZICKY') + _safe('POCET_SCHUZEK_ONLINE')
                          + _safe('POCET_BEZHOT_WALK_IN') * BEZHOT_TO_SCH_RATE * BEZHOT_SCH_EQUIV)
    _total_deficit     = _total_cap - _total_actual
    _deficit_col       = '#0bb440' if _total_deficit >= 0 else '#eb4d79'
    _deficit_lbl       = (f"+{int(_total_deficit):,} přebytek" if _total_deficit >= 0
                          else f"{int(_total_deficit):,} deficit")

    # ── Příklad výpočtu (první řádek tabulky) ─────────────────────
    _example_html = ""
    if len(top) > 0:
        _ex = top.iloc[0]
        _ex_name       = str(_ex.get('BRANCH_NAME', '?'))
        _ex_cli_now    = int(float(_ex.get('POCET_KLIENTU', 0) or 0))
        _ex_delta_cli  = int(float(_ex.get('_delta_cli', 0) or 0))
        _ex_cli_sim    = int(float(_ex.get('SIM_KLIENTI', 0) or 0))
        _ex_cli_onl    = int(float(_ex.get('SIM_KLIENTI_ONL', 0) or 0))
        _ex_fmt        = str(_ex.get('BRANCH_FORMAT', 'medium') or 'medium').lower().strip()
        _ex_fmt_lbl    = _ex_fmt.title()
        _ex_fmt_pct    = int(FORMAT_CAP.get(_ex_fmt, 0.40) * 100)
        _ex_rev_now    = _fmtm(_ex.get('OBJEM_VYNOSU_CZK', 0))
        _ex_delta_rev  = _fmtm(_ex.get('_delta_rev', 0))
        _ex_rev_sim    = _fmtm(_ex.get('SIM_REV', 0))
        _ex_cagr_pct   = float(_ex.get('_cagr', 0) or 0) * 100
        _ex_rev_2030   = _fmtm(_ex.get('SIM_REV_2030', 0))
        _ex_rpc_now    = _fmtm(_ex.get('_rev_per_cli_now', 0))
        _ex_rpc_2030   = _fmtm(_ex.get('_rev_per_cli_2030', 0))
        _ex_ci_now_pct = float(_ex.get('PRIME_NAKLADY/VYNOSY', 0) or 0) * 100
        _ex_ci_30_pct  = float(_ex.get('SIM_CI_2030', 0) or 0) * 100
        _ex_q30        = int(_ex.get('SIM_IR_Q_2030', 0) or 0)
        _example_html  = (
            f'<div style="background:#f0f7ff;border-radius:8px;padding:12px 16px;margin-bottom:14px;'
            f'border:1px solid #c7d9fb;font-size:0.78rem;line-height:1.9;color:#1e3a5f;">'
            f'<div style="font-weight:700;font-size:0.82rem;color:#2770f0;margin-bottom:6px;">'
            f'📐 Příklad výpočtu — {_ex_name}</div>'
            f'<b>Klienti:</b> Aktuálně <b>{_ex_cli_now:,}</b> kl.'
            f' + ze zavřených poboček <b>+{_ex_delta_cli:,}</b> kl. (dle spádového podílu) →'
            f' simulovaný stav <b>{_ex_cli_sim:,}</b> kl.'
            f' ({_ex_cli_onl:,} přesměrováno online, kapacita formátu {_ex_fmt_lbl} = +{_ex_fmt_pct}%)<br>'
            f'<b>Výnosy 2025:</b> {_ex_rev_now} (celkové) + {_ex_delta_rev} (ze spádu) = {_ex_rev_sim}<br>'
            f'<b>Výnos/klient 2025:</b> nové výnosy {_ex_rpc_now}/klient · '
            f'trend (40% změna výnos/klient 24→25 + 60% revenues trend) = {_ex_cagr_pct:.1f}%/rok<br>'
            f'<b>Výnos/klient 2030:</b> stávající {_ex_rpc_now} × (1+{_ex_cagr_pct/100:.2f})⁵ = {_ex_rpc_2030} · '
            f'spádoví klienti mají vlastní výnos/klient ze své původní pobočky, upravený stejným trendem<br>'
            f'<b>Roční výnosy 2030:</b> {_ex_cli_now:,} stáv.kl. × {_ex_rpc_2030} + spádoví × jejich výnos/kl. = <b>{_ex_rev_2030}</b><br>'
            f'<b>C/I 2030:</b> Náklady převážně fixní, výnosy rostou →'
            f' C/I₂₀₃₀ = C/I₂₀₂₅ × (výnosy₂₀₂₅ / výnosy₂₀₃₀) ='
            f' {_ex_ci_now_pct:.1f}% × ({_ex_rev_sim} / {_ex_rev_2030}) = <b>{_ex_ci_30_pct:.1f}%</b><br>'
            f'<b>Rating 2030:</b> Rank(CI_percentil × 40% + výnosové_pořadí × 60%) → kvintil <b>Q{_ex_q30}</b>'
            f'</div>'
        )

    # ── Rest section HTML ─────────────────────────────────────────
    if len(rest) == 0:
        _rest_section_html = ""
    else:
        _n_rest   = len(rest)
        _btn_show = f"▼ Zobrazit pobočky mimo top {target_n} ({_n_rest})"
        _btn_hide = f"▲ Skrýt pobočky mimo top {target_n} ({_n_rest})"
        _rest_section_html = (
            f'<div style="margin-top:10px;">'
            f'<button id="sim-rest-btn-{target_n}"'
            f' onclick="(function(btn){{'
            f'var rows=document.querySelectorAll(\'.sim-rest-row-{target_n}\');'
            f'var show=rows.length>0&&rows[0].style.display===\'none\';'
            f'rows.forEach(function(r){{r.style.display=show?\'\':\'none\';}});'
            f'btn.textContent=show?\'{_btn_hide}\':\'{_btn_show}\';'
            f'}})(this)"'
            f' style="padding:6px 14px;border:1px solid #ccc;border-radius:6px;'
            f'background:#f5f5f5;font-size:0.82rem;cursor:pointer;color:#555;">'
            f'{_btn_show}'
            f'</button></div>'
            f'<div style="overflow-x:auto;border-radius:8px;margin-top:6px;">'
            f'<table style="width:100%;border-collapse:collapse;font-size:0.82rem;">'
            f'<tbody>{rest_rows_html}</tbody>'
            f'</table></div>'
            f'<script>'
            f'setTimeout(function(){{'
            f'document.querySelectorAll(\'.sim-rest-row-{target_n}\')'
            f'.forEach(function(r){{r.style.display=\'none\';}});'
            f'}},0);</script>'
        )

    # ── 11. HTML výstup ───────────────────────────────────────────
    return f"""
<div style="background:#fff;border-radius:10px;padding:16px 18px;
            border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">

  <!-- Hlavička -->
  <div style="font-size:0.72rem;color:#2770f0;font-weight:700;text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:6px;">🏦 Simulace sítě {target_n} poboček — výhled 2030</div>

  <!-- Summary boxy -->
  <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:16px;">
    {"".join(f'''
    <div style="background:#f8fbff;border-radius:8px;padding:10px 16px;border:1px solid #e8edf5;text-align:center;min-width:120px;">
      <div style="font-size:1.4rem;font-weight:800;color:{vc};">{vv}</div>
      <div style="font-size:0.68rem;color:#888;text-transform:uppercase;letter-spacing:.3px;">{vl}</div>
    </div>''' for vv, vl, vc in [
        (f"{len(top)}", f"top {target_n} v simulaci", "#2770f0"),
        (f"{n_remain}", "celkem po close", "#6b7280"),
        (f"{n_close}", "ke close", "#eb4d79"),
        (f"{int(_total_close_cli):,}", "klientů ze zavřených", "#f59e0b"),
        (f"{int(_total_online_cli):,}", "→ online", "#6b7280"),
        (f"{_total_bankers:,}", "bankéřů celkem", "#2770f0"),
        (f"{_deficit_lbl}", "bankéřská kap.", _deficit_col),
    ])}
  </div>

  <!-- Legenda kapacity formátu -->
  <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;font-size:0.73rem;">
    <span style="color:#555;font-weight:600;">Kapacita formátu:</span>
    {"".join(f'<span style="background:{c}18;color:{c};border:1px solid {c}44;border-radius:5px;padding:1px 8px;font-weight:700;">{l} +{p}%</span>'
    for l,p,c in [('Small','10','#f59e0b'),('Medium-econ.','25','#fb923c'),('Medium','40','#2770f0'),('Flagship','70','#0bb440')])}
    <span style="color:#555;font-weight:600;margin-left:8px;">Bankéř:</span>
    <span style="background:#6b728018;color:#6b7280;border:1px solid #6b728044;border-radius:5px;padding:1px 8px;font-weight:700;">
      {MEETINGS_PER_DAY} sch./den × {WORKING_DAYS} dní = {BANKER_CAP_YEAR} sch./bankéř/rok · bezhot walk-in: {int(BEZHOT_TO_SCH_RATE*100)}% × {BEZHOT_SCH_EQUIV:.2f}× · portfolio: {PORTFOLIO_PER_BANKER} kl./bankéř</span>
  </div>

  <!-- Konkrétní příklad výpočtu -->
  {_example_html}

  <!-- Vyhledávač -->
  <div style="display:flex;gap:8px;align-items:center;margin-bottom:10px;">
    <input id="sim-search-{target_n}" type="text"
           placeholder="🔍 Filtruj název nebo kód pobočky…"
           style="padding:7px 12px;border:1px solid #ccc;border-radius:6px;
                  font-size:0.85rem;width:280px;outline:none;"
           oninput="simFilter{target_n}(this.value)">
    <button onclick="simFilter{target_n}(''); document.getElementById('sim-search-{target_n}').value='';"
            style="padding:6px 12px;border:1px solid #ccc;border-radius:6px;
                   background:#f5f5f5;font-size:0.82rem;cursor:pointer;">✖ Zrušit</button>
    <span id="sim-count-{target_n}" style="font-size:0.82rem;color:#888;"></span>
  </div>

  <!-- Tabulka -->
  <div style="overflow-x:auto;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,0.06);">
    <table id="sim-tbl-{target_n}" style="width:100%;border-collapse:collapse;font-size:0.82rem;">
      <thead style="position:sticky;top:0;z-index:5;">
        <tr style="background:#2770f0;color:white;">
          <th style="padding:7px 8px;border:1px solid #1a4db5;text-align:center;">#</th>
          <th style="padding:7px 8px;border:1px solid #1a4db5;text-align:left;min-width:180px;">
            Pobočka + přitékající klienti</th>
          <th style="padding:7px 8px;border:1px solid #1a4db5;text-align:center;">Sim. Q 2030</th>
          <th style="padding:7px 8px;border:1px solid #1a4db5;text-align:center;white-space:nowrap;">
            Rating 2025 → 2030</th>
          <th style="padding:7px 8px;border:1px solid #1a4db5;text-align:left;min-width:140px;">
            Klienti + kapacita</th>
          <th style="padding:7px 8px;border:1px solid #1a4db5;text-align:left;min-width:160px;">
            Výnosy 2025→2030</th>
          <th style="padding:7px 8px;border:1px solid #1a4db5;text-align:center;white-space:nowrap;">
            C/I 2025→2030</th>
          <th style="padding:7px 8px;border:1px solid #1a4db5;text-align:left;min-width:130px;">
            Bankéři / schůzky</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>

  <!-- Pobočky mimo top {target_n} -->
  {_rest_section_html}

  <!-- Metodika -->
  <div style="background:#f8fbff;border-radius:8px;padding:10px 14px;margin-top:14px;
              border:1px solid #e8edf5;font-size:0.73rem;color:#555;line-height:1.8;">
    <strong>Metodika simulace 2030:</strong><br>
    Klienti: % dle spádových návštěv → kapacitní limit dle formátu → zbytek online ·
    Výnosy 2030 = výnos/klient_2030 × sim.počet klientů · výnos/klient roste dle trendu (30% revenues 22-25 + 70% nové výnosy 24-25, max ±20-35%/rok) ·
    Rating 2030 = rank(CI_perc_2030 × 40% + rev_rnk_2030 × 60%) ·
    Bankéřská kap. = {MEETINGS_PER_DAY} sch./den × {WORKING_DAYS} dní = {BANKER_CAP_YEAR} sch./bankéř/rok ·
    Skutečná zátěž = fyzické + online schůzky + {int(BEZHOT_TO_SCH_RATE*100)}% bezhot.walk-in × {BEZHOT_SCH_EQUIV:.2f} + spádoví klienti × {SCH_PER_CLIENT:.2f} sch./klient ·
    C/I 2030 = lineární extrapolace z řady 2021–2025 (max ±15 pp)
  </div>
</div>
<script>
(function() {{
  function simFilter() {{
    var inp = document.getElementById('sim-search-{target_n}');
    var q   = inp ? inp.value.toLowerCase().trim() : '';
    var rows = document.querySelectorAll('#sim-tbl-{target_n} tbody tr');
    var vis = 0;
    rows.forEach(function(r) {{
      var show = !q || r.textContent.toLowerCase().indexOf(q) >= 0;
      r.style.display = show ? '' : 'none';
      if (show) vis++;
    }});
    var cnt = document.getElementById('sim-count-{target_n}');
    if (cnt) cnt.textContent = vis + '\u00a0/ {len(top)} poboček';
  }}
  // Attach oninput
  setTimeout(function() {{
    var inp = document.getElementById('sim-search-{target_n}');
    if (inp) inp.addEventListener('input', simFilter);
    // Override onclick for clear button (set via inline)
    simFilter();
  }}, 100);
  window['simFilter{target_n}'] = simFilter;
}})();
</script>"""
    """Jedna široká souhrnná tabulka: region × (typologie + cashless + NF/SF)."""
    needed = ['REGION_NAME']
    if not all(c in df.columns for c in needed):
        return "<p style='color:#aaa;font-style:italic;'>Chybí sloupec REGION_NAME.</p>"

    regions = sorted(df['REGION_NAME'].dropna().unique())

    # ── Pomocná funkce: počty dle kategorie pro každý region ─────────
    def cat_counts(col, rename_map=None):
        if col not in df.columns:
            return {}, []
        tmp = df[['REGION_NAME', col]].copy()
        tmp[col] = tmp[col].fillna('—').astype(str).str.strip()
        if rename_map:
            tmp[col] = tmp[col].map(lambda v: rename_map.get(v, v))
        pt = tmp.groupby(['REGION_NAME', col]).size().unstack(fill_value=0)
        cats = list(pt.columns)
        return pt.to_dict('index'), cats

    typ_data,  typ_cats  = cat_counts('BRANCH_TYPE_NAME')
    cash_data, cash_cats = cat_counts('CASHLESS', rename_map={
        'Y':'Cashless','y':'Cashless','YES':'Cashless','ANO':'Cashless','1':'Cashless','TRUE':'Cashless',
        'N':'Hotovostní','n':'Hotovostní','NO':'Hotovostní','NE':'Hotovostní','0':'Hotovostní','FALSE':'Hotovostní','—':'—'
    })
    nf_data,   nf_cats   = cat_counts('BRANCH_BUILDING_NF_SF')

    # Formát dle FTE — definované pořadí kategorií
    fmt_order = ['flagship', 'medium', 'medium economy', 'small', '—']
    fmt_data_raw, fmt_cats_raw = cat_counts('BRANCH_FORMAT')
    fmt_cats = [c for c in fmt_order if c in fmt_cats_raw] + [c for c in fmt_cats_raw if c not in fmt_order]
    fmt_data = fmt_data_raw

    # ── Skupiny sloupců ───────────────────────────────────────────────
    groups = [
        ('🏢 Typo.', typ_cats,  typ_data,  '#1a5fb4'),
        ('💳 Cash',  cash_cats, cash_data, '#2d8c4e'),
        ('🏗️ NF/SF', nf_cats,   nf_data,   '#7b3f9e'),
        ('📐 Formát', fmt_cats,  fmt_data,  '#b5660a'),
    ]

    # ── Záhlaví ───────────────────────────────────────────────────────
    header_groups = ""
    header_cols   = "<th style='background:#2d3748;color:white;border:1px solid rgba(255,255,255,0.15);padding:5px 10px;font-size:0.82rem;text-align:left;position:sticky;left:0;z-index:12;'>Region</th>"
    for grp_label, cats, _, grp_color in groups:
        span = len(cats) + 1  # +1 pro součet skupiny
        if span > 1:
            header_groups += (f"<th colspan='{span}' style='background:{grp_color};color:white;"
                              f"border:1px solid rgba(255,255,255,0.2);padding:5px 10px;"
                              f"text-align:center;font-size:0.82rem;'>{grp_label}</th>")
        for cat in cats:
            header_cols += (f"<th style='background:{grp_color};color:white;border:1px solid rgba(255,255,255,0.2);"
                            f"padding:4px 8px;text-align:center;white-space:nowrap;font-size:0.79rem;'>{cat}</th>")
        header_cols += (f"<th style='background:{grp_color};color:white;border:1px solid rgba(255,255,255,0.2);"
                        f"padding:4px 8px;text-align:center;font-size:0.79rem;font-style:italic;'>Σ</th>")

    # ── Řádky ─────────────────────────────────────────────────────────
    rows_html = ""
    col_totals = {}  # pro CELKEM řádek

    for i, region in enumerate(regions):
        bg = "#ffffff" if i % 2 == 0 else "#f4f7ff"
        row = (f"<td style='border:1px solid #dee2e6;padding:4px 10px;background:#f0f4ff;"
               f"font-weight:600;white-space:nowrap;font-size:0.82rem;position:sticky;left:0;z-index:5;"
               f"box-shadow:2px 0 4px rgba(0,0,0,0.07);'>{region}</td>")
        for grp_label, cats, data, grp_color in groups:
            grp_total = 0
            for cat in cats:
                val = data.get(region, {}).get(cat, 0)
                col_totals[f"{grp_label}_{cat}"] = col_totals.get(f"{grp_label}_{cat}", 0) + val
                grp_total += val
                row += (f"<td style='border:1px solid #dee2e6;padding:4px 8px;background:{bg};"
                        f"text-align:center;font-size:0.82rem;'>{val if val else '—'}</td>")
            col_totals[f"{grp_label}_Σ"] = col_totals.get(f"{grp_label}_Σ", 0) + grp_total
            row += (f"<td style='border:1px solid #dee2e6;padding:4px 8px;background:{bg};"
                    f"text-align:center;font-weight:600;font-size:0.82rem;color:#555;'>{grp_total}</td>")
        rows_html += f"<tr>{row}</tr>"

    # Celkový řádek
    total_row = (f"<td style='border:1px solid #dee2e6;padding:5px 10px;background:#e8f4ff;"
                 f"font-weight:700;font-size:0.82rem;border-top:2px solid #2770f0;"
                 f"position:sticky;left:0;z-index:5;'>CELKEM</td>")
    for grp_label, cats, _, _ in groups:
        for cat in cats:
            v = col_totals.get(f"{grp_label}_{cat}", 0)
            total_row += (f"<td style='border:1px solid #dee2e6;padding:5px 8px;background:#e8f4ff;"
                          f"font-weight:700;text-align:center;font-size:0.82rem;"
                          f"border-top:2px solid #2770f0;'>{v if v else '—'}</td>")
        v_s = col_totals.get(f"{grp_label}_Σ", 0)
        total_row += (f"<td style='border:1px solid #dee2e6;padding:5px 8px;background:#dce8ff;"
                      f"font-weight:700;text-align:center;font-size:0.82rem;"
                      f"border-top:2px solid #2770f0;'>{v_s}</td>")
    rows_html += f"<tr>{total_row}</tr>"

    html1 = f"""
<div style="margin-bottom:4px;">
  <div style="border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,0.09);max-height:520px;overflow:auto;width:100%;">
    <table class="table table-bordered table-sm mb-0" style="font-size:0.82rem;background:white;width:100%;table-layout:auto;">
      <thead style="position:sticky;top:0;z-index:10;">
        <tr>{header_groups}</tr>
        <tr>{header_cols}</tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
</div>"""

    html2 = ""
    html3 = ""

    return html1


def render_strategy_columns(df):
    """
    Tři sloupce vedle sebe: Close · Investice · Cashless.
    Badge barva = světlá odvozená od barvy záhlaví, plynule zesvětluje dle roku.
    Karta obsahuje: název, region, výnosy, rating, kvintil ratingu.
    """
    # ── Kvintil ratingu → emoji + barva ──────────────────────────────
    _Q_DOT = {1:'🟢', 2:'🔵', 3:'🟡', 4:'🟠', 5:'🔴'}
    _Q_COL = {1:'#0bb440', 2:'#2770f0', 3:'#f59e0b', 4:'#fb923c', 5:'#eb4d79'}

    def _q_badge(row):
        qv = row.get('IR_Q', '')
        try: qi = int(float(qv))
        except: return ''
        if qi not in _Q_DOT: return ''
        qcol = _Q_COL[qi]
        return (f"<span style='background:{qcol}18;color:{qcol};"
                f"border:1px solid {qcol}44;border-radius:10px;"
                f"padding:1px 7px;font-size:0.68rem;font-weight:700;'>"
                f"{_Q_DOT[qi]} Q{qi}</span>")

    # ── Světlé badge barvy odvozené od HEX záhlaví ───────────────────
    def _badge_colors(yrs, hdr_hex, n_steps=6):
        """
        Vrátí {rok: css_barva} — plynulý gradient od světlé (blízká bílé)
        po mírně sytější, vše odvozeno od hdr_hex. Nejnižší rok = nejsvětlejší.
        """
        import colorsys
        h = hdr_hex.lstrip('#')
        r, g, b = int(h[0:2],16)/255, int(h[2:4],16)/255, int(h[4:6],16)/255
        hue, sat, val = colorsys.rgb_to_hsv(r, g, b)
        result = {}
        n = max(len(yrs), 1)
        for i, yr in enumerate(sorted(yrs)):
            # světlost klesá od 0.96 (velmi světlá) po 0.82 (stále světlá)
            l = 0.96 - (i / max(n-1, 1)) * 0.14
            # sytost roste mírně od 0.18 po 0.38
            s = 0.18 + (i / max(n-1, 1)) * 0.20
            rr, gg, bb = colorsys.hsv_to_rgb(hue, s, l)
            result[yr] = (
                f"#{int(rr*255):02x}{int(gg*255):02x}{int(bb*255):02x}",
                # barva textu — vždy tmavá odvozená od záhlaví
                f"#{int(hue*255):02x}{max(int(gg*80),20):02x}{max(int(bb*60),20):02x}"
                if False else hdr_hex  # jednoduše použij barvu záhlaví pro text
            )
        return result

    # ── Karta ─────────────────────────────────────────────────────────
    def _card(row, bg_col, txt_col, yr, extra_html=''):
        bcode  = int(row.get('BRANCH_CODE', 0))
        bname  = str(row.get('BRANCH_NAME', str(bcode)))
        region = str(row.get('REGION_NAME', ''))
        ir_val = row.get('IR', '')
        try:    ir_disp = int(float(ir_val))
        except: ir_disp = '—'
        rev = row.get('OBJEM_VYNOSU_CZK')
        rev_s = format_money(float(rev)) if pd.notna(rev) and float(rev) > 0 else '—'
        q_html = _q_badge(row)

        return f"""
<div style="display:flex;align-items:stretch;gap:0;border-radius:8px;overflow:hidden;
            border:1px solid #e8edf5;margin-bottom:7px;background:#fff;
            box-shadow:0 1px 3px rgba(0,0,0,0.04);">
  <div style="background:{bg_col};color:white;font-size:0.78rem;font-weight:800;
              min-width:44px;display:flex;align-items:center;justify-content:center;
              flex-shrink:0;letter-spacing:-0.5px;">{yr}</div>
  <div style="padding:8px 12px;flex:1;min-width:0;">
    <div style="display:flex;align-items:baseline;gap:6px;flex-wrap:wrap;">
      <span style="font-size:0.84rem;font-weight:700;color:#0f172a;">{bname}</span>
      <span style="font-size:0.7rem;color:#bbb;">#{bcode}</span>
      <span style="font-size:0.7rem;color:#aaa;margin-left:auto;">{region}</span>
    </div>
    <div style="margin-top:4px;display:flex;align-items:center;flex-wrap:wrap;gap:6px;">
      <span style="font-size:0.77rem;color:#6b7280;font-weight:600;">{rev_s}</span>
      <span style="font-size:0.7rem;color:#aaa;">R {ir_disp}</span>
      {q_html}
      {extra_html}
    </div>
  </div>
</div>"""

    # ── Sloupec 1: Close ─────────────────────────────────────────────
    def _col_close(d):
        if 'FOOTPRINT_STRATEGIE_(2030)' not in d.columns: return "", 0
        sub = d[d['FOOTPRINT_STRATEGIE_(2030)'].astype(str).str.lower()
                  .str.contains('close', na=False)].copy()
        if sub.empty: return "", 0
        sub['_yr'] = pd.to_numeric(sub.get('PLAN_CLOSE_(ROK)', pd.Series(dtype=float)), errors='coerce')
        sub = sub.sort_values('_yr', na_position='last')
        yrs = sorted(sub['_yr'].dropna().unique().astype(int).tolist())
        yr_cols = _badge_colors(yrs, '#eb4d79')
        cards = ""
        for _, row in sub.iterrows():
            yr = int(row['_yr']) if pd.notna(row['_yr']) else '?'
            bg, tc = yr_cols.get(yr, ('#fce4ec', '#eb4d79'))
            status = str(row.get('CLOSE_STATUS', '') or '')
            status = '' if status in ('nan','None','') else status
            extra  = f"<span style='font-size:0.7rem;color:#eb4d79;font-weight:600;'>{status}</span>" if status else ''
            cards += _card(row, bg, tc, yr, extra)
        return cards, len(sub)

    # ── Sloupec 2: Investice ─────────────────────────────────────────
    def _col_invest(d):
        yr_col = 'INVESTICE_NF_OR_FHC_(ROK)'
        if yr_col not in d.columns: return "", 0
        sub = d[pd.to_numeric(d[yr_col], errors='coerce') > 0].copy()
        if sub.empty: return "", 0
        sub['_yr'] = pd.to_numeric(sub[yr_col], errors='coerce')
        sub = sub.sort_values('_yr', na_position='last')
        yrs = sorted(sub['_yr'].dropna().unique().astype(int).tolist())
        yr_cols = _badge_colors(yrs, '#2770f0')
        cards = ""
        for _, row in sub.iterrows():
            yr   = int(row['_yr']) if pd.notna(row['_yr']) else '?'
            bg, tc = yr_cols.get(yr, ('#e8f0fe', '#2770f0'))
            nf_yr  = pd.to_numeric(row.get('INVESTICE_NF_(ROK)',  0), errors='coerce')
            fhc_yr = pd.to_numeric(row.get('INVESTICE_FHC_(ROK)', 0), errors='coerce')
            typ = 'NF' if pd.notna(nf_yr) and nf_yr > 0 else ('FHC' if pd.notna(fhc_yr) and fhc_yr > 0 else '')
            extra  = (f"<span style='background:#eef4ff;color:#2770f0;border-radius:4px;"
                      f"padding:1px 6px;font-size:0.7rem;font-weight:700;'>{typ}</span>") if typ else ''
            cards += _card(row, bg, tc, yr, extra)
        return cards, len(sub)

    # ── Sloupec 3: Cashless ──────────────────────────────────────────
    def _col_cashless(d):
        yr_col = 'PLAN_NEW_CASHIERLESS_(ROK)'
        if yr_col not in d.columns: return "", 0
        sub = d[pd.to_numeric(d[yr_col], errors='coerce') > 0].copy()
        if sub.empty: return "", 0
        sub['_yr'] = pd.to_numeric(sub[yr_col], errors='coerce')
        sub = sub.sort_values('_yr', na_position='last')
        yrs = sorted(sub['_yr'].dropna().unique().astype(int).tolist())
        yr_cols = _badge_colors(yrs, '#0bb440')
        cards = ""
        for _, row in sub.iterrows():
            yr  = int(row['_yr']) if pd.notna(row['_yr']) else '?'
            bg, tc = yr_cols.get(yr, ('#e6f9ed', '#0bb440'))
            cashless_now = str(row.get('CASHLESS', '') or '').strip().upper()
            already = cashless_now in ('Y','YES','ANO','1','TRUE')
            if already:
                extra = ("<span style='background:#eafaf1;color:#0bb440;border:1px solid #a7f3d0;"
                         "border-radius:10px;padding:1px 7px;font-size:0.68rem;font-weight:700;'>✓ Již</span>")
            else:
                status = str(row.get('NEW_CASHIERLESS_STATUS', '') or '')
                status = '' if status in ('nan','None') else status
                extra  = f"<span style='font-size:0.7rem;color:#0bb440;font-weight:600;'>{status}</span>" if status else ''
            cards += _card(row, bg, tc, yr, extra)
        return cards, len(sub)

    c_close,    n_close    = _col_close(df)
    c_invest,   n_invest   = _col_invest(df)
    c_cashless, n_cashless = _col_cashless(df)

    def _col_wrap(title, icon, n, border_col, cards_html):
        if n == 0:
            return f"""
<div style="flex:1;min-width:260px;">
  <div style="background:{border_col};color:white;font-size:0.85rem;font-weight:700;
              border-radius:8px 8px 0 0;padding:9px 14px;">
    {icon} {title} <span style="opacity:.7;font-size:0.75rem;">(žádné)</span></div>
  <div style="border:1px solid #e8edf5;border-top:none;border-radius:0 0 8px 8px;
              padding:12px;color:#aaa;font-size:0.82rem;font-style:italic;">Žádné pobočky.</div>
</div>"""
        return f"""
<div style="flex:1;min-width:260px;">
  <div style="background:{border_col};color:white;font-size:0.85rem;font-weight:700;
              border-radius:8px 8px 0 0;padding:9px 14px;">
    {icon} {title}
    <span style="background:rgba(255,255,255,0.25);border-radius:10px;
                 padding:1px 8px;font-size:0.75rem;margin-left:6px;">{n}</span>
  </div>
  <div style="border:1px solid #e8edf5;border-top:none;border-radius:0 0 8px 8px;
              padding:10px 8px 4px;background:#fafbff;max-height:600px;overflow-y:auto;">
    {cards_html}
  </div>
</div>"""

    return f"""
<div style="display:flex;gap:16px;flex-wrap:wrap;align-items:flex-start;">
  {_col_wrap('Close', '🔴', n_close, '#eb4d79', c_close)}
  {_col_wrap('Investice', '🔨', n_invest, '#2770f0', c_invest)}
  {_col_wrap('Cashless', '💵', n_cashless, '#0bb440', c_cashless)}
</div>"""


def render_action_cards(df, mode):
    """
    Univerzální karta pro pobočky s plánovanou akcí — close / invest / cashless.
    Karty seřazeny od nejnižšího roku vzestupně, ve třech sloupcích.

    mode: 'close' | 'invest' | 'cashless'
    """
    _CFG = {
        'close': {
            'filter':   lambda d: d[d['FOOTPRINT_STRATEGIE_(2030)'].astype(str).str.lower().str.contains('close', na=False)].copy(),
            'year_col': 'PLAN_CLOSE_(ROK)',
            'sub_col':  'CLOSE_STATUS',
            'hdr_col':  '#eb4d79',
            'icon':     '🔴',
            'empty':    'Žádné pobočky se strategií CLOSE.',
            'year_lbl': 'Plán close',
        },
        'invest': {
            'filter':   lambda d: d[d['INVESTICE_NF_OR_FHC_(ROK)'].apply(
                            lambda x: pd.notna(x) and str(x).strip() not in ('', '0', 'nan') and float(x) > 0
                            if 'INVESTICE_NF_OR_FHC_(ROK)' in d.columns else False
                        )].copy() if 'INVESTICE_NF_OR_FHC_(ROK)' in d.columns else d.head(0),
            'year_col': 'INVESTICE_NF_OR_FHC_(ROK)',
            'sub_col':  None,
            'hdr_col':  '#2770f0',
            'icon':     '🔨',
            'empty':    'Žádné pobočky s plánovanou investicí.',
            'year_lbl': 'Rok investice',
        },
        'cashless': {
            'filter':   lambda d: d[d['PLAN_NEW_CASHIERLESS_(ROK)'].apply(
                            lambda x: pd.notna(x) and str(x).strip() not in ('', '0', 'nan') and float(x) > 0
                            if 'PLAN_NEW_CASHIERLESS_(ROK)' in d.columns else False
                        )].copy() if 'PLAN_NEW_CASHIERLESS_(ROK)' in d.columns else d.head(0),
            'year_col': 'PLAN_NEW_CASHIERLESS_(ROK)',
            'sub_col':  'NEW_CASHIERLESS_STATUS',
            'hdr_col':  '#0bb440',
            'icon':     '💵',
            'empty':    'Žádné pobočky s plánovaným přechodem na cashless.',
            'year_lbl': 'Rok cashless',
        },
    }

    cfg = _CFG[mode]
    try:
        d = cfg['filter'](df)
    except Exception:
        d = pd.DataFrame()

    if d.empty:
        return f"<p style='color:#aaa;font-style:italic;'>{cfg['empty']}</p>"

    year_col = cfg['year_col']
    d['_rok'] = pd.to_numeric(d[year_col], errors='coerce')
    d = d.sort_values('_rok', na_position='last')

    hdr  = cfg['hdr_col']
    icon = cfg['icon']

    # Roky jako skupiny — každá skupina = jeden rok
    years = sorted(d['_rok'].dropna().unique().astype(int).tolist())

    cols_html = ""
    for yr in years:
        yr_df = d[d['_rok'] == yr]
        cards = ""
        for _, row in yr_df.iterrows():
            bc    = int(row.get('BRANCH_CODE', 0))
            bn    = str(row.get('BRANCH_NAME', str(bc)))
            reg   = str(row.get('REGION_NAME', ''))
            strat = str(row.get('FOOTPRINT_STRATEGIE_(2030)', '') or '')
            sub   = ''
            if cfg['sub_col'] and cfg['sub_col'] in row.index:
                _sv = row.get(cfg['sub_col'], '')
                sub = str(_sv) if pd.notna(_sv) and str(_sv) not in ('nan', 'None', '') else ''
            ir_val = row.get('IR', '')
            try:    ir_disp = int(float(ir_val))
            except: ir_disp = '—'
            rev = row.get('OBJEM_VYNOSU_CZK')
            rev_s = format_money(float(rev)) if pd.notna(rev) else '—'

            # cashless — zobraz badge "Již cashless" pokud ano
            cashless_now = str(row.get('CASHLESS', '') or '')
            already = cashless_now.strip().upper() in ('Y', 'YES', 'ANO', '1', 'TRUE')
            already_html = (
                f"<span style='background:#eafaf1;color:#0bb440;border:1px solid #a7f3d0;"
                f"border-radius:10px;padding:1px 7px;font-size:0.68rem;font-weight:700;"
                f"margin-left:4px;'>✓ Již cashless</span>"
            ) if (mode == 'cashless' and already) else ''

            cards += f"""
<div style="display:flex;align-items:stretch;border-radius:7px;overflow:hidden;
            border:1px solid #e8edf5;margin-bottom:7px;background:#fff;
            box-shadow:0 1px 3px rgba(0,0,0,0.04);">
  <div style="background:{hdr};min-width:6px;flex-shrink:0;"></div>
  <div style="padding:8px 12px;flex:1;min-width:0;">
    <div style="display:flex;align-items:baseline;gap:6px;flex-wrap:wrap;">
      <span style="font-size:0.83rem;font-weight:700;color:#0f172a;">{bn}</span>
      <span style="font-size:0.7rem;color:#aaa;">#{bc}</span>
      {already_html}
    </div>
    <div style="font-size:0.72rem;color:#888;margin-top:2px;">{reg}</div>
    <div style="display:flex;gap:10px;margin-top:4px;flex-wrap:wrap;">
      <span style="font-size:0.72rem;color:#555;">Rating {ir_disp}</span>
      <span style="font-size:0.72rem;color:#555;">{rev_s}</span>
      {"" if not sub else f'<span style="font-size:0.72rem;color:{hdr};font-weight:600;">{sub}</span>'}
    </div>
  </div>
</div>"""

        cols_html += f"""
<div style="flex:1;min-width:200px;">
  <div style="background:{hdr};color:white;font-size:0.8rem;font-weight:700;
              border-radius:7px 7px 0 0;padding:6px 12px;text-align:center;">
    {icon} {yr}
    <span style="font-size:0.7rem;font-weight:400;opacity:0.85;margin-left:4px;">
      ({len(yr_df)} poboček)</span>
  </div>
  <div style="border:1px solid #e8edf5;border-top:none;border-radius:0 0 7px 7px;
              padding:8px 8px 4px 8px;background:#fafbff;">
    {cards}
  </div>
</div>"""

    # Pobočky bez roku (NaN) — přidat jako extra sloupec
    no_yr_df = d[d['_rok'].isna()]
    if not no_yr_df.empty:
        cards_ny = ""
        for _, row in no_yr_df.iterrows():
            bc  = int(row.get('BRANCH_CODE', 0))
            bn  = str(row.get('BRANCH_NAME', str(bc)))
            reg = str(row.get('REGION_NAME', ''))
            cards_ny += f"""
<div style="display:flex;align-items:stretch;border-radius:7px;overflow:hidden;
            border:1px solid #e8edf5;margin-bottom:7px;background:#fff;">
  <div style="background:#aaa;min-width:6px;flex-shrink:0;"></div>
  <div style="padding:8px 12px;flex:1;">
    <div style="font-size:0.83rem;font-weight:700;color:#0f172a;">{bn}</div>
    <div style="font-size:0.72rem;color:#888;">{reg} · #{bc}</div>
  </div>
</div>"""
        cols_html += f"""
<div style="flex:1;min-width:200px;">
  <div style="background:#aaa;color:white;font-size:0.8rem;font-weight:700;
              border-radius:7px 7px 0 0;padding:6px 12px;text-align:center;">
    Rok neznámý ({len(no_yr_df)})</div>
  <div style="border:1px solid #e8edf5;border-top:none;border-radius:0 0 7px 7px;
              padding:8px 8px 4px 8px;background:#fafbff;">{cards_ny}</div>
</div>"""

    return f"""
<div style="background:#fff;border-radius:10px;padding:14px 16px;
            border:1px solid #e8edf5;box-shadow:0 2px 8px rgba(0,0,0,0.05);">
  <div style="font-size:0.72rem;font-weight:700;color:{hdr};text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:12px;">
    {icon} {cfg['year_lbl']} — {len(d)} poboček · seřazeno dle roku
  </div>
  <div style="display:flex;gap:14px;flex-wrap:wrap;align-items:flex-start;">
    {cols_html}
  </div>
</div>"""


def generate_branch_summary_tables(df):
    """Kompaktní souhrnná tabulka: region × (typologie | cashless | NF/SF) v jedné tabulce."""

    d = df.copy()
    for c in ['REGION_NAME', 'BRANCH_TYPE_NAME', 'CASHLESS', 'BRANCH_BUILDING_NF_SF']:
        if c in d.columns:
            d[c] = d[c].fillna('—').astype(str).str.strip()

    if 'REGION_NAME' not in d.columns:
        return "<p style='color:#aaa;font-style:italic;'>Data nejsou k dispozici.</p>"

    regions = sorted([r for r in d['REGION_NAME'].unique() if r != '—'])

    def _vals(col):
        if col not in d.columns:
            return []
        return sorted([v for v in d[col].unique() if v not in ('—', '', 'nan', 'None')])

    typ_vals  = _vals('BRANCH_TYPE_NAME')
    cash_vals = _vals('CASHLESS')
    nfsf_vals = _vals('BRANCH_BUILDING_NF_SF')

    # Formát dle FTE — potřebujeme surové hodnoty (ne HTML badge)
    # Zkusíme načíst z df_specialiste přes globals nebo z BRANCH_FORMAT přímo
    # Pokud BRANCH_FORMAT obsahuje HTML, vytáhneme ze zdrojového df
    for _fc, _fc_raw in [('BRANCH_FORMAT', '_fmt_celk_raw'), ('BRANCH_FORMAT_OBCHODNI', '_fmt_obch_raw')]:
        if _fc in d.columns:
            _ser = d[_fc].astype(str)
            # Detekce HTML badge → extrahuj text
            if _ser.str.contains('<span', na=False).any():
                import re as _re
                d[_fc_raw] = _ser.apply(lambda x: _re.sub(r'<[^>]+>', '', x).strip().lower() if '<' in str(x) else str(x).lower().strip())
            else:
                d[_fc_raw] = _ser.str.lower().str.strip()
        else:
            d[_fc_raw] = '—'

    fmt_celk_vals = _vals('_fmt_celk_raw')
    fmt_obch_vals = _vals('_fmt_obch_raw')

    # Pořadí formátů: small → medium economy → medium → flagship
    _fmt_order = ['small', 'medium economy', 'medium', 'flagship']
    def _sort_fmt(vals):
        return sorted(vals, key=lambda v: _fmt_order.index(v) if v in _fmt_order else 99)
    fmt_celk_vals = _sort_fmt(fmt_celk_vals)
    fmt_obch_vals = _sort_fmt(fmt_obch_vals)

    # Skupiny sloupců: (label skupiny, barva, seznam hodnot, zdrojový sloupec)
    col_groups = []
    if typ_vals:
        col_groups.append(('Typologie', '#1a5fb4', typ_vals, 'BRANCH_TYPE_NAME'))
    if cash_vals:
        col_groups.append(('Bezhotovostní', '#2d8c4e', cash_vals, 'CASHLESS'))
    if nfsf_vals:
        col_groups.append(('Formát NF/SF', '#7b3f9e', nfsf_vals, 'BRANCH_BUILDING_NF_SF'))
    if fmt_celk_vals:
        col_groups.append(('Formát (celk. FTE)', '#e07020', fmt_celk_vals, '_fmt_celk_raw'))
    if fmt_obch_vals:
        col_groups.append(('Formát (obch. FTE)', '#9b59b6', fmt_obch_vals, '_fmt_obch_raw'))

    if not col_groups:
        return "<p style='color:#aaa;font-style:italic;'>Data nejsou k dispozici.</p>"

    # Předpočítej cross-tabulky
    cts = {}
    for grp_lbl, _color, vals, src_col in col_groups:
        cts[grp_lbl] = pd.crosstab(
            d['REGION_NAME'].fillna('—'),
            d[src_col].fillna('—'),
            margins=True, margins_name='CELKEM'
        ) if src_col in d.columns else pd.DataFrame()

    # ── Header ────────────────────────────────────────────────────
    header_top = '<tr>'
    header_top += '<th rowspan="2" style="background:#2770f0;color:#fff;border:1px solid rgba(255,255,255,0.2);padding:6px 12px;text-align:left;font-size:0.8rem;vertical-align:bottom;white-space:nowrap;">Region</th>'
    for grp_lbl, color, vals, _ in col_groups:
        header_top += (
            f'<th colspan="{len(vals)+1}" style="background:{color};color:#fff;'
            f'border:1px solid rgba(255,255,255,0.2);padding:6px 10px;'
            f'text-align:center;font-size:0.78rem;">{grp_lbl}</th>'
        )
    header_top += '</tr>'

    header_bot = '<tr>'
    for grp_lbl, color, vals, _ in col_groups:
        for v in vals:
            short = v[:12] + '…' if len(v) > 12 else v
            header_bot += (
                f'<th style="background:{color}dd;color:#fff;'
                f'border:1px solid rgba(255,255,255,0.2);padding:4px 8px;'
                f'text-align:center;font-size:0.72rem;white-space:nowrap;" title="{v}">{short}</th>'
            )
        header_bot += (
            f'<th style="background:{color}aa;color:#fff;'
            f'border:1px solid rgba(255,255,255,0.2);padding:4px 8px;'
            f'text-align:center;font-size:0.72rem;font-weight:800;">Σ</th>'
        )
    header_bot += '</tr>'

    # ── Řádky ─────────────────────────────────────────────────────
    all_rows = regions + ['CELKEM']
    rows_html = ''
    for region in all_rows:
        is_total = region == 'CELKEM'
        bg  = '#e8f4ff' if is_total else ('#fff' if all_rows.index(region) % 2 == 0 else '#f7f9ff')
        fw  = '700' if is_total else '400'
        bt  = 'border-top:2px solid #2770f0;' if is_total else ''
        row = f'<tr style="background:{bg};">'
        row += (
            f'<td style="border:1px solid #dee2e6;padding:5px 12px;font-weight:{fw};'
            f'white-space:nowrap;font-size:0.8rem;{bt}">{region}</td>'
        )
        for grp_lbl, color, vals, _ in col_groups:
            ct = cts[grp_lbl]
            for v in vals:
                val = int(ct.loc[region, v]) if (region in ct.index and v in ct.columns) else 0
                row += (
                    f'<td style="border:1px solid #dee2e6;padding:5px 8px;'
                    f'text-align:center;font-size:0.8rem;font-weight:{fw};{bt}">'
                    f'{"<b>" if val > 0 else ""}{val if val > 0 else "—"}{"</b>" if val > 0 else ""}</td>'
                )
            # Celkem skupiny
            tot = int(ct.loc[region, 'CELKEM']) if region in ct.index and 'CELKEM' in ct.columns else 0
            row += (
                f'<td style="border:1px solid #dee2e6;padding:5px 8px;'
                f'text-align:center;font-size:0.8rem;font-weight:700;{bt}'
                f'color:{color};">{tot}</td>'
            )
        row += '</tr>'
        rows_html += row

    return (
        f'<div style="overflow-x:auto;border-radius:8px;'
        f'box-shadow:0 2px 8px rgba(0,0,0,0.07);margin-bottom:20px;">'
        f'<table style="width:100%;border-collapse:collapse;font-size:0.8rem;background:#fff;">'
        f'<thead>{header_top}{header_bot}</thead>'
        f'<tbody>{rows_html}</tbody>'
        f'</table></div>'
    )


def generate_close_branches_table(df):
    if 'FOOTPRINT_STRATEGIE_(2030)' not in df.columns:
        return ("<p style='color:#aaa;font-style:italic;'>"
                "Sloupec FOOTPRINT_STRATEGIE_(2030) není k dispozici.</p>")

    close_df = df[
        df['FOOTPRINT_STRATEGIE_(2030)'].astype(str).str.lower().str.strip().str.contains('close', na=False)
    ].copy()

    if close_df.empty:
        return "<p style='color:#aaa;font-style:italic;'>Žádné pobočky se strategií CLOSE.</p>"

    close_df['_rok_sort'] = pd.to_numeric(
        close_df['PLAN_CLOSE_(ROK)'], errors='coerce'
    )
    close_df = close_df.sort_values(by='_rok_sort', na_position='last')

    rows_html = ""
    for i, (_, row) in enumerate(close_df.iterrows()):
        bg = "#ffffff" if i % 2 == 0 else "#fff5f5"
        td = f"border:1px solid #dee2e6;padding:5px 10px;background:{bg};"

        rok = _fmt_year_int(row.get('PLAN_CLOSE_(ROK)', ''))

        ir_val = row.get('IR', '')
        try:
            ir_disp = int(float(ir_val)) if ir_val not in ('', None) and not (isinstance(ir_val, float) and np.isnan(ir_val)) else "—"
        except:
            ir_disp = "—"

        strat = str(row.get('FOOTPRINT_STRATEGIE_(2030)', ''))
        strat_html = color_strategy_html(strat)

        status = str(row.get('CLOSE_STATUS', ''))
        if pd.isna(row.get('CLOSE_STATUS')) or status in ('nan', 'None'):
            status = ''

        rows_html += f"""<tr>
          <td style='{td}text-align:center;font-weight:bold;'>{int(row['BRANCH_CODE'])}</td>
          <td style='{td}text-align:left;white-space:nowrap;'>{row.get('BRANCH_NAME', '')}</td>
          <td style='{td}text-align:left;white-space:nowrap;'>{row.get('REGION_NAME', '')}</td>
          <td style='border:1px solid #dee2e6;padding:5px 10px;background:{bg};text-align:center;white-space:nowrap;'>{strat_html}</td>
          <td style='{td}text-align:center;font-weight:bold;font-size:0.82rem;color:#c0392b;'>{rok}</td>
          <td style='{td}text-align:center;'>{status}</td>
          <td style='{td}text-align:center;font-weight:bold;color:#2770f0;font-size:0.82rem;'>{ir_disp}</td>
        </tr>"""

    return f"""
<div style="background:#fff;border:1px solid #dee2e6;border-radius:8px;padding:20px;
            box-shadow:0 2px 8px rgba(0,0,0,0.06);margin-bottom:24px;">
  <p style="font-size:0.83rem;color:#666;margin-bottom:10px;">
    Celkem poboček se strategií
    <span style='background:#f8d7da;color:#721c24;padding:1px 7px;border-radius:4px;font-weight:bold;'>CLOSE</span>:
    <b>{len(close_df)}</b> — seřazeno dle plánovaného roku uzavření.
  </p>
  <div class="table-responsive" style="max-height:520px;overflow-y:auto;
       border-radius:6px;box-shadow:0 2px 6px rgba(0,0,0,0.07);">
    <table id="close-tbl-{id(close_df)}" class="table table-bordered table-sm mb-0"
           style="font-size:0.82rem;background:white;table-layout:fixed;width:100%;">
      <colgroup>
        <col style="width:90px">
        <col style="width:220px">
        <col style="width:140px">
        <col style="width:130px">
        <col style="width:100px">
        <col style="width:130px">
        <col style="width:80px">
      </colgroup>
      <thead style="position:sticky;top:0;z-index:10;">
        <tr>
          <th data-col="0" style="background:#c0392b;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">ID Pobočky
              <span class="close-arr" data-arrow="0" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="1" style="background:#c0392b;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:left;cursor:pointer;" title="Řadit">Název pobočky
              <span class="close-arr" data-arrow="1" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="2" style="background:#c0392b;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Region
              <span class="close-arr" data-arrow="2" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="3" style="background:#c0392b;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Strategie BNS
              <span class="close-arr" data-arrow="3" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="4" style="background:#c0392b;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Rok uzavření
              <span class="close-arr" data-arrow="4" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="5" style="background:#c0392b;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Status uzavření
              <span class="close-arr" data-arrow="5" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="6" style="background:#c0392b;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Rating 25
              <span class="close-arr" data-arrow="6" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
<script>
(function() {{
  var tbl = document.getElementById("close-tbl-{id(close_df)}");
  if (!tbl) return;
  var _ss = {{col: 4, asc: true}};
  function doSort(colIdx) {{
    var asc = (_ss.col === colIdx) ? !_ss.asc : true;
    _ss = {{col: colIdx, asc: asc}};
    var tbody = tbl.querySelector("tbody");
    var allRows = Array.from(tbody.querySelectorAll("tr"));
    allRows.sort(function(a, b) {{
      var ca = a.cells[colIdx], cb = b.cells[colIdx];
      if (!ca || !cb) return 0;
      var va = ca.textContent.trim(), vb = cb.textContent.trim();
      var na = parseFloat(va), nb = parseFloat(vb);
      if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
      return asc ? va.localeCompare(vb,"cs") : vb.localeCompare(va,"cs");
    }});
    allRows.forEach(function(row, i) {{
      Array.from(row.cells).forEach(function(td) {{ td.style.background = (i%2===0)?"#ffffff":"#fff5f5"; }});
      tbody.appendChild(row);
    }});
    tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
      var ci = parseInt(th.getAttribute("data-col"));
      var arr = th.querySelector(".close-arr");
      if (ci === colIdx) {{
        th.style.background = "#f39c12";
        if (arr) {{ arr.innerHTML = asc?"&#x2191;":"&#x2193;"; arr.style.opacity="1"; }}
      }} else {{
        th.style.background = "#c0392b";
        if (arr) {{ arr.innerHTML = "&#x2195;"; arr.style.opacity="0.4"; }}
      }}
    }});
    allRows.forEach(function(row) {{
      Array.from(row.cells).forEach(function(td, ci) {{
        if (ci === colIdx) td.style.background = "rgba(243,156,18,0.15)";
      }});
    }});
  }}
  tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
    th.addEventListener("click", function() {{ doSort(parseInt(th.getAttribute("data-col"))); }});
  }});
}})();
</script>
  </div>
</div>"""


def generate_invest_branches_table(df):
    nf_col  = 'INVESTICE_NF_(ROK)'
    fhc_col = 'INVESTICE_FHC_(ROK)'
    cmb_col = 'INVESTICE_NF_OR_FHC_(ROK)'

    has_nf  = nf_col  in df.columns
    has_fhc = fhc_col in df.columns
    has_cmb = cmb_col in df.columns

    if not has_nf and not has_fhc:
        return ("<p style='color:#aaa;font-style:italic;'>"
                "Sloupce INVESTICE_NF_(ROK) / INVESTICE_FHC_(ROK) nejsou k dispozici.</p>")

    inv_df = df.copy()

    def _has_year(val):
        if val is None or (isinstance(val, float) and np.isnan(val)) or val == '':
            return False
        try:
            return int(float(val)) > 0
        except:
            return False

    mask_inv = pd.Series([False] * len(inv_df), index=inv_df.index)
    if has_nf:
        mask_inv = mask_inv | inv_df[nf_col].apply(_has_year)
    if has_fhc:
        mask_inv = mask_inv | inv_df[fhc_col].apply(_has_year)

    inv_df = inv_df[mask_inv].copy()

    if inv_df.empty:
        return ("<p style='color:#aaa;font-style:italic;'>"
                "Žádné pobočky s plánovanou investicí NF nebo FHC.</p>")

    def _year_num(val):
        try:
            v = int(float(val))
            return v if v > 0 else 9999
        except:
            return 9999

    if has_cmb:
        inv_df['_rok_sort'] = inv_df[cmb_col].apply(_year_num)
    else:
        col_a = inv_df[nf_col].apply(_year_num) if has_nf else pd.Series([9999]*len(inv_df), index=inv_df.index)
        col_b = inv_df[fhc_col].apply(_year_num) if has_fhc else pd.Series([9999]*len(inv_df), index=inv_df.index)
        inv_df['_rok_sort'] = pd.concat([col_a, col_b], axis=1).min(axis=1)

    inv_df = inv_df.sort_values(by='_rok_sort', na_position='last')

    HDR = '#1a7a4a'

    rows_html = ""
    for i, (_, row) in enumerate(inv_df.iterrows()):
        bg = "#ffffff" if i % 2 == 0 else "#f0fff6"
        td = f"border:1px solid #dee2e6;padding:5px 10px;background:{bg};"

        rok_nf  = _fmt_year_int(row.get(nf_col,  '')) if has_nf  else ''
        rok_fhc = _fmt_year_int(row.get(fhc_col, '')) if has_fhc else ''
        rok_cmb = _fmt_year_int(row.get(cmb_col, '')) if has_cmb else (rok_nf or rok_fhc)

        ir_val = row.get('IR', '')
        try:
            ir_disp = int(float(ir_val)) if ir_val not in ('', None) and not (isinstance(ir_val, float) and np.isnan(ir_val)) else "—"
        except:
            ir_disp = "—"

        strat = str(row.get('FOOTPRINT_STRATEGIE_(2030)', ''))
        strat_html = color_strategy_html(strat)

        def fmt_rok_inv(val):
            if not val:
                return "<span style='color:#ccc;'>—</span>"
            return f"<b style='color:#1a7a4a;font-size:0.82rem;'>{val}</b>"

        rows_html += f"""<tr>
          <td style='{td}text-align:center;font-weight:bold;'>{int(row['BRANCH_CODE'])}</td>
          <td style='{td}text-align:left;white-space:nowrap;'>{row.get('BRANCH_NAME', '')}</td>
          <td style='{td}text-align:left;white-space:nowrap;'>{row.get('REGION_NAME', '')}</td>
          <td style='border:1px solid #dee2e6;padding:5px 10px;background:{bg};text-align:center;white-space:nowrap;'>{strat_html}</td>
          <td style='{td}text-align:center;'>{fmt_rok_inv(rok_nf)}</td>
          <td style='{td}text-align:center;'>{fmt_rok_inv(rok_fhc)}</td>
          <td style='{td}text-align:center;font-weight:bold;color:#1a7a4a;font-size:0.82rem;'>{rok_cmb if rok_cmb else '<span style="color:#ccc;">—</span>'}</td>
          <td style='{td}text-align:center;font-weight:bold;color:#2770f0;font-size:0.82rem;'>{ir_disp}</td>
        </tr>"""

    count_nf  = inv_df[nf_col].apply(_has_year).sum()  if has_nf  else 0
    count_fhc = inv_df[fhc_col].apply(_has_year).sum() if has_fhc else 0

    return f"""
<div style="background:#fff;border:1px solid #dee2e6;border-radius:8px;padding:20px;
            box-shadow:0 2px 8px rgba(0,0,0,0.06);margin-bottom:24px;">
  <p style="font-size:0.83rem;color:#666;margin-bottom:10px;">
    Celkem poboček s plánovanou investicí: <b>{len(inv_df)}</b>
    &nbsp;|&nbsp;
    <span style='background:#d4edda;color:#155724;padding:1px 7px;border-radius:4px;font-weight:bold;'>NF</span>:
    <b>{count_nf}</b>
    &nbsp;|&nbsp;
    <span style='background:#cce5ff;color:#004085;padding:1px 7px;border-radius:4px;font-weight:bold;'>FHC</span>:
    <b>{count_fhc}</b>
    &nbsp;— seřazeno dle nejbližšího plánovaného roku.
  </p>
  <div class="table-responsive" style="max-height:520px;overflow-y:auto;
       border-radius:6px;box-shadow:0 2px 6px rgba(0,0,0,0.07);">
    <table id="invest-tbl-{id(inv_df)}" class="table table-bordered table-sm mb-0"
           style="font-size:0.82rem;background:white;table-layout:fixed;width:100%;">
      <colgroup>
        <col style="width:90px">
        <col style="width:220px">
        <col style="width:140px">
        <col style="width:130px">
        <col style="width:110px">
        <col style="width:110px">
        <col style="width:120px">
        <col style="width:80px">
      </colgroup>
      <thead style="position:sticky;top:0;z-index:10;">
        <tr>
          <th data-col="0" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">ID Pobočky
              <span class="inv-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="1" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:left;cursor:pointer;" title="Řadit">Název pobočky
              <span class="inv-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="2" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Region
              <span class="inv-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="3" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Strategie BNS
              <span class="inv-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="4" style="background:#2d8c4e;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Rok inv. NF
              <span class="inv-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="5" style="background:#1a5fb4;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Rok inv. FHC
              <span class="inv-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="6" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Rok plán. inv.
              <span class="inv-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="7" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Rating 25
              <span class="inv-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
<script>
(function() {{
  var tbl = document.getElementById("invest-tbl-{id(inv_df)}");
  if (!tbl) return;
  var _ss = {{col: 6, asc: true}};
  var HDR_COLOR = "{HDR}";
  var COL_COLORS = [HDR_COLOR,HDR_COLOR,HDR_COLOR,HDR_COLOR,"#2d8c4e","#1a5fb4",HDR_COLOR,HDR_COLOR];
  function doSort(colIdx) {{
    var asc = (_ss.col === colIdx) ? !_ss.asc : true;
    _ss = {{col: colIdx, asc: asc}};
    var tbody = tbl.querySelector("tbody");
    var allRows = Array.from(tbody.querySelectorAll("tr"));
    allRows.sort(function(a, b) {{
      var ca = a.cells[colIdx], cb = b.cells[colIdx];
      if (!ca || !cb) return 0;
      var va = ca.textContent.trim(), vb = cb.textContent.trim();
      var na = parseFloat(va), nb = parseFloat(vb);
      if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
      return asc ? va.localeCompare(vb,"cs") : vb.localeCompare(va,"cs");
    }});
    allRows.forEach(function(row, i) {{
      Array.from(row.cells).forEach(function(td) {{ td.style.background = (i%2===0)?"#ffffff":"#f0fff6"; }});
      tbody.appendChild(row);
    }});
    tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
      var ci = parseInt(th.getAttribute("data-col"));
      var arr = th.querySelector(".inv-arr");
      if (ci === colIdx) {{
        th.style.background = "#f39c12";
        if (arr) {{ arr.innerHTML = asc?"&#x2191;":"&#x2193;"; arr.style.opacity="1"; }}
      }} else {{
        th.style.background = COL_COLORS[ci] || HDR_COLOR;
        if (arr) {{ arr.innerHTML = "&#x2195;"; arr.style.opacity="0.4"; }}
      }}
    }});
    allRows.forEach(function(row) {{
      Array.from(row.cells).forEach(function(td, ci) {{
        if (ci === colIdx) td.style.background = "rgba(243,156,18,0.15)";
      }});
    }});
  }}
  tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
    th.addEventListener("click", function() {{ doSort(parseInt(th.getAttribute("data-col"))); }});
  }});
}})();
</script>
  </div>
</div>"""



def generate_cashierless_table(df):
    """Tabulka poboček s plánovaným přechodem na cashless — řazeno dle roku."""
    year_col   = 'PLAN_NEW_CASHIERLESS_(ROK)'
    status_col = 'NEW_CASHIERLESS_STATUS'

    if year_col not in df.columns:
        return ("<p style='color:#aaa;font-style:italic;'>"
                f"Sloupec {year_col} není k dispozici.</p>")

    def _has_year(val):
        if val is None or (isinstance(val, float) and np.isnan(val)) or val == '':
            return False
        try:
            return int(float(val)) > 0
        except:
            return False

    cash_df = df[df[year_col].apply(_has_year)].copy()

    if cash_df.empty:
        return "<p style='color:#aaa;font-style:italic;'>Žádné pobočky s plánovaným přechodem na cashless.</p>"

    def _year_num(val):
        try:
            v = int(float(val))
            return v if v > 0 else 9999
        except:
            return 9999

    cash_df['_rok_sort'] = cash_df[year_col].apply(_year_num)
    cash_df = cash_df.sort_values(by='_rok_sort', na_position='last')

    HDR = '#1a7a6a'   # teal odstín pro cashless téma

    rows_html = ""
    for i, (_, row) in enumerate(cash_df.iterrows()):
        bg = "#ffffff" if i % 2 == 0 else "#f0faf8"
        td = f"border:1px solid #dee2e6;padding:5px 10px;background:{bg};"

        rok = _fmt_year_int(row.get(year_col, ''))
        status = str(row.get(status_col, '')) if status_col in cash_df.columns else ''
        if pd.isna(row.get(status_col)) or status in ('nan', 'None'):
            status = ''

        ir_val = row.get('IR', '')
        try:
            ir_disp = int(float(ir_val)) if ir_val not in ('', None) and not (isinstance(ir_val, float) and np.isnan(ir_val)) else "—"
        except:
            ir_disp = "—"

        strat = str(row.get('FOOTPRINT_STRATEGIE_(2030)', ''))
        strat_html = color_strategy_html(strat)

        cashless_now = str(row.get('CASHLESS', ''))
        if pd.isna(row.get('CASHLESS')) or cashless_now in ('nan', 'None'):
            cashless_now = ''
        cashless_badge = (
            "<span style='background:#d4edda;color:#155724;padding:1px 7px;"
            "border-radius:4px;font-weight:bold;font-size:0.78rem;'>✓ Již cashless</span>"
            if cashless_now.strip().upper() in ('Y', 'YES', 'ANO', '1', 'TRUE') else ''
        )

        rows_html += f"""<tr>
          <td style='{td}text-align:center;font-weight:bold;'>{int(row['BRANCH_CODE'])}</td>
          <td style='{td}text-align:left;white-space:nowrap;'>{row.get('BRANCH_NAME', '')}</td>
          <td style='{td}text-align:left;white-space:nowrap;'>{row.get('REGION_NAME', '')}</td>
          <td style='border:1px solid #dee2e6;padding:5px 10px;background:{bg};text-align:center;white-space:nowrap;'>{strat_html}</td>
          <td style='{td}text-align:center;font-weight:bold;font-size:0.82rem;color:{HDR};'>{rok}</td>
          <td style='{td}text-align:center;'>{status}</td>
          <td style='{td}text-align:center;'>{cashless_badge}</td>
          <td style='{td}text-align:center;font-weight:bold;color:#2770f0;font-size:0.82rem;'>{ir_disp}</td>
        </tr>"""

    tbl_id = f"cashierless-tbl-{{id(cash_df)}}"
    count = len(cash_df)

    return f"""
<div style="background:#fff;border:1px solid #dee2e6;border-radius:8px;padding:20px;
            box-shadow:0 2px 8px rgba(0,0,0,0.06);margin-bottom:24px;">
  <p style="font-size:0.83rem;color:#666;margin-bottom:10px;">
    Celkem poboček s plánovaným přechodem na cashless: <b>{count}</b>
    &nbsp;— seřazeno dle plánovaného roku.
  </p>
  <div class="table-responsive" style="max-height:520px;overflow-y:auto;
       border-radius:6px;box-shadow:0 2px 6px rgba(0,0,0,0.07);">
    <table id="{tbl_id}" class="table table-bordered table-sm mb-0"
           style="font-size:0.82rem;background:white;table-layout:fixed;width:100%;">
      <colgroup>
        <col style="width:90px">
        <col style="width:220px">
        <col style="width:140px">
        <col style="width:130px">
        <col style="width:120px">
        <col style="width:150px">
        <col style="width:130px">
        <col style="width:80px">
      </colgroup>
      <thead style="position:sticky;top:0;z-index:10;">
        <tr>
          <th data-col="0" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">ID Pobočky
              <span class="cash-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="1" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:left;cursor:pointer;" title="Řadit">Název pobočky
              <span class="cash-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="2" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Region
              <span class="cash-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="3" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Strategie BNS
              <span class="cash-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="4" style="background:#0d6e5e;color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Rok cashless
              <span class="cash-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="5" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Status cashless
              <span class="cash-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="6" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Stav nyní
              <span class="cash-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
          <th data-col="7" style="background:{HDR};color:white;border:1px solid rgba(0,0,0,0.2);
              padding:6px 12px;text-align:center;cursor:pointer;" title="Řadit">Rating 25
              <span class="cash-arr" style="font-size:0.7rem;opacity:0.4;">&#x2195;</span></th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
<script>
(function() {{
  var tbl = document.getElementById("{tbl_id}");
  if (!tbl) return;
  var COL_COLORS = ["{HDR}","{HDR}","{HDR}","{HDR}","#0d6e5e","{HDR}","{HDR}","{HDR}"];
  var _ss = {{col: 4, asc: true}};
  function doSort(colIdx) {{
    var asc = (_ss.col === colIdx) ? !_ss.asc : true;
    _ss = {{col: colIdx, asc: asc}};
    var tbody = tbl.querySelector("tbody");
    var allRows = Array.from(tbody.querySelectorAll("tr"));
    allRows.sort(function(a, b) {{
      var ca = a.cells[colIdx], cb = b.cells[colIdx];
      if (!ca || !cb) return 0;
      var va = ca.textContent.trim(), vb = cb.textContent.trim();
      var na = parseFloat(va), nb = parseFloat(vb);
      if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
      return asc ? va.localeCompare(vb,"cs") : vb.localeCompare(va,"cs");
    }});
    allRows.forEach(function(row, i) {{
      Array.from(row.cells).forEach(function(td) {{ td.style.background = (i%2===0)?"#ffffff":"#f0faf8"; }});
      tbody.appendChild(row);
    }});
    tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
      var ci = parseInt(th.getAttribute("data-col"));
      var arr = th.querySelector(".cash-arr");
      if (ci === colIdx) {{
        th.style.background = "#f39c12";
        if (arr) {{ arr.innerHTML = asc?"&#x2191;":"&#x2193;"; arr.style.opacity="1"; }}
      }} else {{
        th.style.background = COL_COLORS[ci] || "{HDR}";
        if (arr) {{ arr.innerHTML = "&#x2195;"; arr.style.opacity="0.4"; }}
      }}
    }});
    allRows.forEach(function(row) {{
      Array.from(row.cells).forEach(function(td, ci) {{
        if (ci === colIdx) td.style.background = "rgba(243,156,18,0.15)";
      }});
    }});
  }}
  tbl.querySelectorAll("thead th[data-col]").forEach(function(th) {{
    th.addEventListener("click", function() {{ doSort(parseInt(th.getAttribute("data-col"))); }});
  }});
}})();
</script>
  </div>
</div>"""


def merge_with_source(base_df, source_df, merge_cfg):
    """Provede merge podle konfigurace z MERGE_COLS."""
    cfg = merge_cfg.copy()
    cols = cfg.pop("cols")
    drop_key = cfg.pop("drop_key", None)
    available_cols = [c for c in cols if c in source_df.columns]

    source_df = source_df.copy()

    # Sjednoť typy merge klíčů na int64 aby nedocházelo k chybě object vs int64
    left_on  = cfg.get("left_on")  or cfg.get("on")
    right_on = cfg.get("right_on") or cfg.get("on")
    for _df, _col in [(base_df, left_on), (source_df, right_on)]:
        if _col and _col in _df.columns:
            try:
                _df[_col] = pd.to_numeric(_df[_col], errors='coerce').astype('Int64')
            except Exception:
                pass

    result = base_df.merge(source_df[available_cols], **cfg)
    if drop_key and drop_key in result.columns:
        result = result.drop(columns=[drop_key])
    return result


SPECIAL_BRANCH_CODES = {902, 904, 905, 906, 909, 919}

def prepare_rating_status(df):
    """Vrátí pobočky s IR_FLAG==Y + speciální pobočky (bez ratingu, ale s daty)."""
    df_ir   = df[df['IR_FLAG'] == 'Y'].copy()
    df_spec = df[df['BRANCH_CODE'].isin(SPECIAL_BRANCH_CODES)].copy()
    # Přidej pouze ty speciální, které ještě nejsou v df_ir
    df_spec = df_spec[~df_spec['BRANCH_CODE'].isin(df_ir['BRANCH_CODE'])]
    rating_status = pd.concat([df_ir, df_spec], ignore_index=True)
    rating_status['IR_DIFF_NUM'] = rating_status['IR'] - rating_status['INTERNI_RATING_2024']

    # ── Poměry klientů ────────────────────────────────────────────
    for c in ['POCET_KLIENTU', 'PRIMARNI_KLIENTI', 'AKTIVNI_KLIENTI']:
        if c in rating_status.columns:
            rating_status[c] = pd.to_numeric(rating_status[c], errors='coerce')
    if all(c in rating_status.columns for c in ['PRIMARNI_KLIENTI', 'AKTIVNI_KLIENTI', 'POCET_KLIENTU']):
        _tot = rating_status['POCET_KLIENTU'].replace(0, np.nan)
        rating_status['PRIM_RATIO']   = rating_status['PRIMARNI_KLIENTI'] / _tot
        rating_status['AKTIVNI_RATIO'] = rating_status['AKTIVNI_KLIENTI'] / _tot

        def _ratio_html(prim, aktiv):
            try:
                p = float(prim)  if pd.notna(prim)  else None
                a = float(aktiv) if pd.notna(aktiv) else None
                if p is None and a is None: return '—'
                bars = ''
                if p is not None:
                    pc = min(p * 100, 100)
                    bars += (f"<div style='display:flex;align-items:center;gap:4px;margin-bottom:2px;'>"
                             f"<span style='font-size:0.65rem;color:#2770f0;min-width:28px;'>Prim</span>"
                             f"<div style='flex:1;background:#eef4ff;border-radius:2px;height:7px;'>"
                             f"<div style='width:{pc:.0f}%;background:#2770f0;height:100%;border-radius:2px;'></div></div>"
                             f"<span style='font-size:0.65rem;color:#2770f0;min-width:30px;text-align:right;'>{pc:.0f}%</span></div>")
                if a is not None:
                    ac = min(a * 100, 100)
                    bars += (f"<div style='display:flex;align-items:center;gap:4px;'>"
                             f"<span style='font-size:0.65rem;color:#0bb440;min-width:28px;'>Akt.</span>"
                             f"<div style='flex:1;background:#eafaf1;border-radius:2px;height:7px;'>"
                             f"<div style='width:{ac:.0f}%;background:#0bb440;height:100%;border-radius:2px;'></div></div>"
                             f"<span style='font-size:0.65rem;color:#0bb440;min-width:30px;text-align:right;'>{ac:.0f}%</span></div>")
                return bars
            except: return '—'
        rating_status['KLIENTI_RATIO_HTML'] = rating_status.apply(
            lambda r: _ratio_html(r.get('PRIM_RATIO'), r.get('AKTIVNI_RATIO')), axis=1)

    # ── Návštěvy — poměrový HTML (4 bary, základ = celkové návštěvy) ──
    _visit_parts = [
        ('POCET_SCHUZEK_ONLINE',  'Online',   '#2770f0', '#eef4ff'),
        ('POCET_SCHUZEK_FYZICKY', 'Fyzické',  '#0bb440', '#eafaf1'),
        ('POCET_HOT_WALK_IN',     'Hot.',      '#eb4d79', '#fdecea'),
        ('POCET_BEZHOT_WALK_IN',  'Bezhot.',   '#f59e0b', '#fff8e1'),
    ]
    _total_col = 'POCET_NAVSTEV_CELKEM'

    def _navstevy_ratio_html(row):
        try:
            total = float(row.get(_total_col, 0) or 0)
            if total <= 0:
                return '—'
            bars = ''
            for col, lbl, clr, bg in _visit_parts:
                v = float(row.get(col, 0) or 0)
                pct = min(v / total * 100, 100)
                bars += (
                    f"<div style='display:flex;align-items:center;gap:4px;margin-bottom:2px;'>"
                    f"<span style='font-size:0.63rem;color:{clr};min-width:36px;white-space:nowrap;'>{lbl}</span>"
                    f"<div style='flex:1;background:{bg};border-radius:2px;height:7px;'>"
                    f"<div style='width:{pct:.0f}%;background:{clr};height:100%;border-radius:2px;'></div></div>"
                    f"<span style='font-size:0.63rem;color:{clr};min-width:30px;text-align:right;'>{pct:.0f}%</span>"
                    f"</div>"
                )
            return bars
        except:
            return '—'

    if all(c in rating_status.columns for c in [_total_col]):
        for c in [_total_col] + [col for col, *_ in _visit_parts]:
            if c in rating_status.columns:
                rating_status[c] = pd.to_numeric(rating_status[c], errors='coerce').fillna(0)
        rating_status['NAVSTEVY_RATIO_HTML'] = rating_status.apply(_navstevy_ratio_html, axis=1)

    # ── Kvintily prodejů per produkt + celkový počet prodejů ─────
    _prod_keys = [k for k, _ in PRODUCT_GROUPS]
    # Celkový počet prodejů 2025 = součet všech produktů
    _prod_sum_cols = [f'POCET_PRODEJU_{k}_2025' for k in _prod_keys
                      if f'POCET_PRODEJU_{k}_2025' in rating_status.columns]
    if _prod_sum_cols:
        for c in _prod_sum_cols:
            rating_status[c] = pd.to_numeric(rating_status[c], errors='coerce').fillna(0)
        rating_status['POCET_PRODEJU_CELKEM_2025'] = rating_status[_prod_sum_cols].sum(axis=1)

    # Kvintil pro každý produkt (1=nejméně, 5=nejvíce) + HTML tečka
    _q_dot = {
        1: ('<span style="color:#eb4d79;font-size:18px;">&#9679;</span>', 'bottom 20%'),
        2: ('<span style="color:#f0845d;font-size:18px;">&#9679;</span>', '20–40%'),
        3: ('<span style="color:#f0b55d;font-size:18px;">&#9679;</span>', '40–60%'),
        4: ('<span style="color:#42db72;font-size:18px;">&#9679;</span>', '60–80%'),
        5: ('<span style="color:#0ab33f;font-size:18px;">&#9679;</span>', 'top 20%'),
    }
    _ir_mask = (rating_status.get('IR_FLAG', pd.Series('Y', index=rating_status.index)) == 'Y')
    for _pk in _prod_keys:
        _col25  = f'POCET_PRODEJU_{_pk}_2025'
        _qcol   = f'POCET_PRODEJU_{_pk}_2025_Q'
        _qhtml  = f'POCET_PRODEJU_{_pk}_2025_Q_HTML'
        if _col25 not in rating_status.columns:
            continue
        rating_status[_col25] = pd.to_numeric(rating_status[_col25], errors='coerce').fillna(0)
        _mask2 = _ir_mask & (rating_status[_col25] > 0)
        if _mask2.sum() >= 5:
            try:
                rating_status.loc[_mask2, _qcol] = (
                    pd.qcut(rating_status.loc[_mask2, _col25], q=5, labels=False, duplicates='drop') + 1
                ).astype(float)
            except Exception:
                rating_status[_qcol] = np.nan
        else:
            rating_status[_qcol] = np.nan

        def _mk_dot(v, _qd=_q_dot):
            try:
                qi = int(float(v))
                dot, lbl = _qd.get(qi, ('', ''))
                return f'<span title="Q{qi} — {lbl}">{dot}</span>'
            except Exception:
                return ''
        rating_status[_qhtml] = rating_status[_qcol].map(_mk_dot)
    _exp_atm = globals().get('export_atm')
    _kap_atm = globals().get('kapacita_atm')

    def _atm_summary_html(branch_code):
        try:
            bc = int(branch_code)
            parts = []

            # Bankomaty — seznam
            if _exp_atm is not None and not _exp_atm.empty:
                _aid = next((c for c in ['BRANCH_ID','BRANCH_CODE','POBOCKA']
                             if c in _exp_atm.columns), None)
                if _aid:
                    _ab = _exp_atm[pd.to_numeric(_exp_atm[_aid], errors='coerce') == bc]
                    if not _ab.empty:
                        atm_list = []
                        for _, ar in _ab.iterrows():
                            def _av(col):
                                v = ar.get(col)
                                return '—' if (v is None or (isinstance(v,float) and pd.isna(v))
                                               or str(v) in ('nan','None','')) else str(v)
                            atm_id   = _av('ATM_ID')
                            atm_type = _av('ATM_TYPE')
                            loc      = _av('BRANCH_LOCATION')
                            avail    = _av('AVAILABILITY_TO_CLIENTS')
                            atm_list.append(
                                f'<span style="display:inline-block;background:#f0f4ff;color:#2770f0;'
                                f'border:1px solid #c7d9fb;border-radius:6px;padding:1px 7px;'
                                f'font-size:0.68rem;margin:1px;white-space:nowrap;">'
                                f'🏧 {atm_id} · {atm_type} · {loc} · {avail}</span>'
                            )
                        parts.append('<div style="margin-bottom:4px;">'
                                     + ''.join(atm_list) + '</div>')

            # Kapacita ATM
            if _kap_atm is not None and not _kap_atm.empty:
                _kid = next((c for c in ['BRANCH_ID','BRANCH_CODE']
                             if c in _kap_atm.columns), None)
                if _kid:
                    _kb = _kap_atm[pd.to_numeric(_kap_atm[_kid], errors='coerce') == bc]
                    if not _kb.empty:
                        kr = _kb.iloc[0]
                        def _kpct(col):
                            v = kr.get(col)
                            try: return round(float(v) * 100, 1)
                            except: return None
                        def _knum(col):
                            v = kr.get(col)
                            try: return round(float(v), 1)
                            except: return None

                        akt   = _kpct('AKTUALNI_VYTIZENI_STROJU')
                        prev  = _kpct('VYTIZENI_VSECH_ATM_PO_PREVODU')
                        volno = _kpct('ZBYVAJICI_VOLNA_KAPACITA')
                        vybery = _knum('POCET_ATM_VYBERU_PO_PREVODU_/_M')
                        vklady = _knum('POCET_ATM_VKLADU_PO_PREVODU_/_M')

                        def _pct_span(v, label, clr):
                            if v is None: return ''
                            warn = v > 100
                            c = '#eb4d79' if warn else clr
                            return (f'<span style="font-size:0.7rem;font-weight:700;color:{c};">'
                                    f'{label}: {v:.0f}%{"⚠️" if warn else ""}</span>')

                        def _num_span(v, label, icon):
                            if v is None: return ''
                            return (f'<span style="font-size:0.7rem;color:#555;">'
                                    f'{icon} {label}: {v:,.0f}/měs.</span>')

                        row_items = ' · '.join(filter(None, [
                            _pct_span(akt,   'Akt.',   '#e07020'),
                            _pct_span(prev,  'Po přev.','#2770f0'),
                            _pct_span(volno, 'Volno',   '#0bb440'),
                            _num_span(vybery, 'Výběry', '⬆️'),
                            _num_span(vklady, 'Vklady',  '⬇️'),
                        ]))
                        if row_items:
                            parts.append(f'<div style="font-size:0.72rem;line-height:1.8;">{row_items}</div>')

            return ''.join(parts) if parts else '—'
        except Exception:
            return '—'

    if _exp_atm is not None or _kap_atm is not None:
        rating_status['ATM_SUMMARY_HTML'] = rating_status['BRANCH_CODE'].apply(_atm_summary_html)
        # Samostatné numerické sloupce z kapacita_atm
        if _kap_atm is not None and not _kap_atm.empty:
            _kid = next((c for c in ['BRANCH_ID','BRANCH_CODE'] if c in _kap_atm.columns), None)
            if _kid:
                _kap = _kap_atm.copy()
                _kap[_kid] = pd.to_numeric(_kap[_kid], errors='coerce')
                _kap = _kap.drop_duplicates(subset=[_kid])
                _kap_idx = _kap.set_index(_kid)

                def _kap_val(bc, col, mult=1.0):
                    try:
                        v = _kap_idx.at[int(bc), col]
                        return round(float(v) * mult, 1) if pd.notna(v) else None
                    except Exception:
                        return None

                rating_status['ATM_VOLNA_KAPACITA']  = rating_status['BRANCH_CODE'].apply(
                    lambda bc: _kap_val(bc, 'ZBYVAJICI_VOLNA_KAPACITA', 100))
                rating_status['ATM_VYBERY_MESIC']    = rating_status['BRANCH_CODE'].apply(
                    lambda bc: _kap_val(bc, 'PRUMER._POCET_VYBERU_ATM_/_MESIC'))
                rating_status['ATM_VKLADY_MESIC']    = rating_status['BRANCH_CODE'].apply(
                    lambda bc: _kap_val(bc, 'PRUMER._POCET_VKLADU_ATM_/_MESIC'))
                rating_status['ATM_VYBERY_PO_PREV']  = rating_status['BRANCH_CODE'].apply(
                    lambda bc: _kap_val(bc, 'POCET_ATM_VYBERU_PO_PREVODU_/_M'))
                rating_status['ATM_VKLADY_PO_PREV']  = rating_status['BRANCH_CODE'].apply(
                    lambda bc: _kap_val(bc, 'POCET_ATM_VKLADU_PO_PREVODU_/_M'))
        else:
            rating_status['ATM_VOLNA_KAPACITA'] = None
            rating_status['ATM_VYBERY_MESIC']   = None
            rating_status['ATM_VKLADY_MESIC']   = None
            rating_status['ATM_VYBERY_PO_PREV'] = None
            rating_status['ATM_VKLADY_PO_PREV'] = None
    else:
        rating_status['ATM_SUMMARY_HTML']   = '—'
        rating_status['ATM_VOLNA_KAPACITA'] = None
        rating_status['ATM_VYBERY_MESIC']   = None
        rating_status['ATM_VKLADY_MESIC']   = None
        rating_status['ATM_VYBERY_PO_PREV'] = None
        rating_status['ATM_VKLADY_PO_PREV'] = None
    # Sloupce: Výnosy 21, 22, 23, 24, 25 (VYNOSY = celkové výnosy 2025)
    _rev_spark_cols  = ['REVENUES_2021', 'REVENUES_2022', 'REVENUES_2023',
                        'REVENUES_2024', 'VYNOSY']
    _rev_spark_avail = [c for c in _rev_spark_cols if c in rating_status.columns]
    _rev_spark_years = ['21', '22', '23', '24', '25'][:len(_rev_spark_avail)]

    if len(_rev_spark_avail) >= 2:
        # Převeď na numerické PŘED formátováním — sparkline musí pracovat s čísly
        for c in _rev_spark_avail:
            rating_status[c] = pd.to_numeric(rating_status[c], errors='coerce')

        def _rev_sparkline(row):
            # Načti hodnoty přímo jako čísla z row
            vals = []
            for c in _rev_spark_avail:
                v = row[c]
                try:
                    fv = float(v)
                    vals.append(fv if fv > 0 else None)
                except (TypeError, ValueError):
                    vals.append(None)

            pts = [(i, v) for i, v in enumerate(vals) if v is not None]
            if len(pts) < 2:
                return '—'

            max_v = max(v for _, v in pts) or 1
            min_v = min(v for _, v in pts)
            rng   = max_v - min_v or max_v * 0.1 or 1

            W, H, PAD = 90, 28, 3
            n = len(_rev_spark_avail)

            def _x(i): return int(i / (n - 1) * (W - PAD * 2)) + PAD
            def _y(v):  return int((1 - (v - min_v) / rng) * (H - PAD * 2 - 6)) + PAD

            first_v = pts[0][1]
            last_v  = pts[-1][1]
            line_col = '#0bb440' if last_v >= first_v else '#eb4d79'

            # Polyline — pouze dostupné body
            polypts = ' '.join(f"{_x(i)},{_y(v)}" for i, v in pts)

            # Tečky na každém dostupném bodě
            dots = ''.join(
                f'<circle cx="{_x(i)}" cy="{_y(v)}" r="1.8" fill="{line_col}" opacity="0.7"/>'
                for i, v in pts[:-1]
            )
            # Poslední bod větší a plný
            lx, ly = _x(pts[-1][0]), _y(pts[-1][1])
            last_dot = f'<circle cx="{lx}" cy="{ly}" r="2.8" fill="{line_col}"/>'

            # Popisky roků pod osou
            year_labels = ''.join(
                f'<text x="{_x(i)}" y="{H}" text-anchor="middle" '
                f'font-size="6" fill="#aaa">{_rev_spark_years[i]}</text>'
                for i in range(n)
            )

            return (
                f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" '
                f'xmlns="http://www.w3.org/2000/svg" style="display:block;">'
                f'<polyline points="{polypts}" fill="none" stroke="{line_col}" '
                f'stroke-width="1.6" stroke-linejoin="round" stroke-linecap="round"/>'
                f'{dots}{last_dot}{year_labels}'
                f'</svg>'
            )

        rating_status['REV_SPARKLINE'] = rating_status.apply(_rev_sparkline, axis=1)

    # ── Trend klientů 2021–2025 (sparkline stejný styl jako REV_SPARKLINE) ──
    _cli_spark_cols  = ['POCET_KLIENTU_EOY_2021', 'POCET_KLIENTU_EOY_2022',
                        'POCET_KLIENTU_EOY_2023', 'POCET_KLIENTU_EOY_2024', 'POCET_KLIENTU']
    _cli_spark_avail = [c for c in _cli_spark_cols if c in rating_status.columns]
    _cli_spark_years = ['21', '22', '23', '24', '25'][:len(_cli_spark_avail)]

    if len(_cli_spark_avail) >= 2:
        for c in _cli_spark_avail:
            rating_status[c] = pd.to_numeric(rating_status[c], errors='coerce')

        def _cli_sparkline(row):
            vals = []
            for c in _cli_spark_avail:
                try:
                    fv = float(row[c])
                    vals.append(fv if fv > 0 else None)
                except (TypeError, ValueError):
                    vals.append(None)
            pts = [(i, v) for i, v in enumerate(vals) if v is not None]
            if len(pts) < 2: return '—'
            max_v = max(v for _, v in pts) or 1
            min_v = min(v for _, v in pts)
            rng   = max_v - min_v or max_v * 0.1 or 1
            W, H, PAD = 90, 28, 3
            n = len(_cli_spark_avail)
            def _x(i): return int(i / (n - 1) * (W - PAD * 2)) + PAD
            def _y(v):  return int((1 - (v - min_v) / rng) * (H - PAD * 2 - 6)) + PAD
            first_v = pts[0][1]; last_v = pts[-1][1]
            line_col = '#0bb440' if last_v >= first_v else '#eb4d79'
            polypts  = ' '.join(f"{_x(i)},{_y(v)}" for i, v in pts)
            dots = ''.join(
                f'<circle cx="{_x(i)}" cy="{_y(v)}" r="1.8" fill="{line_col}" opacity="0.7"/>'
                for i, v in pts[:-1])
            lx, ly = _x(pts[-1][0]), _y(pts[-1][1])
            year_labels = ''.join(
                f'<text x="{_x(i)}" y="{H}" text-anchor="middle" '
                f'font-size="6" fill="#aaa">{_cli_spark_years[i]}</text>'
                for i in range(n))
            return (
                f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" '
                f'xmlns="http://www.w3.org/2000/svg" style="display:block;">'
                f'<polyline points="{polypts}" fill="none" stroke="{line_col}" '
                f'stroke-width="1.6" stroke-linejoin="round" stroke-linecap="round"/>'
                f'{dots}'
                f'<circle cx="{lx}" cy="{ly}" r="2.8" fill="{line_col}"/>'
                f'{year_labels}</svg>'
            )

        rating_status['KLIENTI_TREND_HTML'] = rating_status.apply(_cli_sparkline, axis=1)

    # ── Index expozice — kvintil (1=nejhorší, 5=nejlepší; větší = lepší) ──
    _ie_col = 'HODNOTA_INDEXU_EXPOZICE'
    if _ie_col in rating_status.columns:
        rating_status[_ie_col] = pd.to_numeric(rating_status[_ie_col], errors='coerce')
        _ie_mask = rating_status[_ie_col].notna() & (rating_status['IR_FLAG'] == 'Y')
        if _ie_mask.sum() >= 5:
            # větší index = lepší = Q1; qcut ascending → obrátíme: 5 - rank + 1
            _ie_rank = pd.qcut(rating_status.loc[_ie_mask, _ie_col],
                               q=5, labels=False, duplicates='drop')  # 0=nejmenší, 4=největší
            rating_status.loc[_ie_mask, 'IE_Q'] = 5 - _ie_rank  # Q1=největší, Q5=nejmenší
            # Q1=zelená (nejlepší), Q5=červená (nejhorší)
            _ie_dot_scale = {1: '#0ab33f', 2: '#42db72', 3: '#f0b55d',
                             4: '#f0845d', 5: '#eb4d79'}
            def _ie_dot(v):
                try:
                    qi = int(float(v))
                    c  = _ie_dot_scale.get(qi, '#aaa')
                    lbl = {1:'top 20% — nejlepší',2:'60–80%',3:'40–60%',
                           4:'20–40%',5:'bottom 20% — nejhorší'}.get(qi,'')
                    return (f'<span title="IE kvintil {qi}: {lbl}" '
                            f'style="color:{c};font-size:22px;line-height:1;cursor:default;">&#9679;</span>'
                            f'<span style="font-size:0.72rem;color:#888;margin-left:3px;">{qi}</span>')
                except: return ''
            rating_status['IE_Q_HTML'] = rating_status['IE_Q'].map(_ie_dot)
        else:
            rating_status['IE_Q']     = np.nan
            rating_status['IE_Q_HTML'] = ''

    existing = [c for c in COLS_TO_FIX if c in rating_status.columns]
    numeric_cols = rating_status[existing].select_dtypes(include=[np.number]).columns
    rating_status[numeric_cols] = rating_status[numeric_cols].fillna(0)
    date_cols = rating_status[existing].select_dtypes(include=['datetime']).columns
    for col in date_cols:
        rating_status[col] = rating_status[col].astype(object).fillna("")
    remaining_cols = [c for c in existing if c not in numeric_cols and c not in date_cols]
    rating_status[remaining_cols] = rating_status[remaining_cols].fillna("")

    # HTML sloupce pro změny — NaN-safe (speciální pobočky bez ratingu)
    def _safe_ci_change(r):
        try:
            a, b = float(r['PRIME_NAKLADY/VYNOSY']), float(r['C/I_RATIO_(YTD_2024)'])
            if pd.isna(a) or pd.isna(b): return DASH_HTML
            return get_change_html(a - b, mode='pct')
        except: return DASH_HTML

    def _safe_ir_change(r):
        try:
            a, b = r['IR'], r['INTERNI_RATING_2024']
            if pd.isna(a) or pd.isna(b): return DASH_HTML
            return get_change_html(int(a - b), mode='rating')
        except: return DASH_HTML

    def _safe_ir_perc_change(r):
        try:
            a, b = r['IR_PERC'], r['INTERNI_RATING_2024_-_PERCENTILY']
            if pd.isna(a) or pd.isna(b): return DASH_HTML
            return get_change_html(int(a - b), mode='rating')
        except: return DASH_HTML

    def _safe_rev_change(r):
        try:
            a, b = r['OBJEM_VYNOSU_CZK'], r['NEW_BUSINESS_2024_-_OBJEM_VYNOSU']
            if pd.isna(a) or pd.isna(b): return DASH_HTML
            return get_change_html(a - b, mode='money')
        except: return DASH_HTML

    rating_status['REV_CHANGE_HTML']      = rating_status.apply(_safe_rev_change,     axis=1)
    rating_status['CI_CHANGE_HTML']       = rating_status.apply(_safe_ci_change,      axis=1)
    rating_status['IR_CHANGE_HTML']       = rating_status.apply(_safe_ir_change,      axis=1)
    rating_status['IR_PERC_CHANGE_HTML']  = rating_status.apply(_safe_ir_perc_change, axis=1)

    # Statusy (barevné tečky + číslo kvintilu — sjednoceno s IR_Q_HTML)
    rating_status['CI_RATIO_STATUS'] = rating_status['CI_RATIO_Q'].map(
        lambda v: get_quintile_dot(v, tooltip_prefix="C/I kvintil")
    )
    rating_status['NEW_BUSINESS_VOL_STATUS'] = rating_status['NEW_BUSINESS_VOL_Q'].map(
        lambda v: get_quintile_dot(v, tooltip_prefix="Nové výnosy kvintil")
    )
    rating_status['ROCNI_SPLATKY_S_DPH_CZK_STATUS'] = rating_status['ROCNI_SPLATKY_S_DPH_CZK_Q'].map(
        lambda v: get_quintile_dot(v, tooltip_prefix="Nájemné kvintil")
    )
    if 'IR_Q' in rating_status.columns:
        rating_status['IR_Q_HTML'] = rating_status['IR_Q'].map(
            lambda v: get_quintile_dot(v, tooltip_prefix="Rating kvintil")
        )
    else:
        rating_status['IR_Q_HTML'] = ""

    # Debug: ověřit přítomnost cash sloupců v rating_status
    cash_present = [c for c in CASH_CASTKA_COLS + CASH_POCET_COLS if c in rating_status.columns]
    cash_missing  = [c for c in CASH_CASTKA_COLS + CASH_POCET_COLS if c not in rating_status.columns]
    if cash_missing:
        display(HTML(f"<b style='color:orange;'>⚠️ Chybějící cash sloupce v rating_status: {cash_missing}</b>"))
    else:
        display(HTML(f"<b style='color:green;'>✅ Cash sloupce OK: {cash_present[:4]}... ({len(cash_present)} celkem)</b>"))

    # Business rating — dílčí kvintil HTML tečky
    _cap_q_map = {
        'CAP_REV_FTE_Q':     'Výnos/FTE kvintil',
        'CAP_NEWCLI_FTE_Q':  'Noví klienti/FTE kvintil',
        'CAP_PRIMCLI_FTE_Q': 'Prim. klienti/FTE kvintil',
        'CAP_SCHUZKY_FTE_Q': 'Schůzky/FTE kvintil',
        'CAP_PEREX_Q':       'PEREX kvintil',
    }
    for _qcol, _tip in _cap_q_map.items():
        if _qcol in rating_status.columns:
            rating_status[f'{_qcol}_HTML'] = rating_status[_qcol].map(
                lambda v, t=_tip: get_quintile_dot(int(round(float(v))), tooltip_prefix=t)
                if pd.notna(v) and float(v) > 0 else ''
            )

    # Kapacitní rating HTML tečka
    if 'CAP_Q' in rating_status.columns:
        def _make_cap_q_html(v):
            try:
                fv = float(v)
                if np.isnan(fv) or fv <= 0: return ''
                return get_quintile_dot(int(round(fv)), tooltip_prefix='Business rating kvintil')
            except: return ''
        rating_status['CAP_Q_HTML'] = rating_status['CAP_Q'].map(_make_cap_q_html)

    # Přegeneruj TRN_Q_HTML ze surového TRN_Q (float 1–5, 0 = bez dat)
    if 'TRN_Q' in rating_status.columns:
        def _make_trn_q_html(v):
            try:
                fv = float(v)
                if np.isnan(fv) or fv <= 0:
                    return ''
                return get_quintile_dot(int(round(fv)), tooltip_prefix="Cash rating kvintil")
            except:
                return ''
        rating_status['TRN_Q_HTML'] = rating_status['TRN_Q'].map(_make_trn_q_html)
        trn_ok = (rating_status['TRN_Q'] > 0).sum()
        trn_rnk_ok = (rating_status['TRN_RNK'] > 0).sum() if 'TRN_RNK' in rating_status.columns else 0
        display(HTML(f"<b style='color:{'green' if trn_ok > 0 else 'red'};'>"
                     f"{'✅' if trn_ok > 0 else '❌'} Cash rating: {trn_ok} poboček s TRN_Q | "
                     f"pořadí (TRN_RNK): {trn_rnk_ok} poboček | "
                     f"distribuce Q: {dict(rating_status[rating_status['TRN_Q']>0]['TRN_Q'].value_counts().sort_index())}</b>"))
    else:
        display(HTML("<b style='color:orange;'>⚠️ TRN_Q sloupec chybí v rating_status</b>"))

    # ── Formát pobočky dle FTE ────────────────────────────────────────
    def _calc_format(fte):
        try:
            fte = float(fte)
        except (TypeError, ValueError):
            return "—"
        if fte >= 25:
            return "flagship"
        elif 10 <= fte < 25:
            return "medium"
        elif 5 <= fte < 10:
            return "medium economy"
        else:
            return "small"

    if 'FTE' in rating_status.columns:
        rating_status['BRANCH_FORMAT'] = rating_status['FTE'].apply(_calc_format)
    else:
        rating_status['BRANCH_FORMAT'] = "—"

    # ── Počet bankéřů ze specialistů ──────────────────────────────
    _BANKER_POSITIONS = [
        "osobní bankéř - junior",
        "osobní bankéř - medior",
        "osobní bankéř - senior",
        "osobní bankéř - master",
    ]
    # Obchodní pozice pro výpočet Obchodního FTE
    _OBCHODNI_POSITIONS = [
        "bankéř klientské péče - junior",
        "bankéř klientské péče - medior",
        "firemní bankéř - master",
        "firemní bankéř - medior",
        "firemní bankéř - senior",
        "hypoteční specialista - medior",
        "hypoteční specialista - senior",
        "hypoteční specialista vcb - medior",
        "hypoteční specialista vcb - senior",
        "investiční specialista - medior",
        "manaž. segm. erste premier - team leader s portfoliem",
        "osobní bankéř - junior",
        "osobní bankéř - master",
        "osobní bankéř - medior",
        "osobní bankéř - senior",
        "pobočkový specialista - hypo",
        "podpora firemních bankéřů",
        "pojišťovací specialista - medior",
        "premier bankéř - master",
        "premier bankéř - medior",
        "premier bankéř - senior",
        "privátní bankéř - medior",
        "privátní bankéř - senior",
        "privátní bankéř - wealth management",
        "remote firemní bankéř - medior",
        "remote premier bankéř - medior",
        "spec. pro firemní pojištění - senior",
    ]

    try:
        _spec_g = globals().get('df_specialiste')
        _spec_cols_g = globals().get('SPEC_POZICE_COLS', [])
        if _spec_g is not None and not _spec_g.empty:
            _bnk_cols = [c for c in _spec_cols_g
                         if c.lower().strip() in _BANKER_POSITIONS]
            # Obchodní FTE sloupce
            _obch_cols = [c for c in _spec_cols_g
                          if c.lower().strip() in _OBCHODNI_POSITIONS]
            if _bnk_cols:
                _spec_bc = _spec_g.copy()
                _spec_bc['branch_id'] = pd.to_numeric(_spec_bc['branch_id'], errors='coerce')
                for c in _bnk_cols + _obch_cols + _spec_cols_g:
                    if c in _spec_bc.columns:
                        _spec_bc[c] = pd.to_numeric(_spec_bc[c], errors='coerce').fillna(0)
                _spec_bc['BANKERS_COUNT'] = _spec_bc[_bnk_cols].sum(axis=1)
                _spec_bc['OBCHODNI_FTE']  = _spec_bc[_obch_cols].sum(axis=1) if _obch_cols else 0
                _spec_bc['_total_spec']   = _spec_bc[_spec_cols_g].apply(
                    pd.to_numeric, errors='coerce').fillna(0).sum(axis=1)
                _spec_agg = _spec_bc[['branch_id', 'BANKERS_COUNT', 'OBCHODNI_FTE', '_total_spec']].rename(
                    columns={'branch_id': 'BRANCH_CODE'})
                rating_status['BRANCH_CODE'] = pd.to_numeric(
                    rating_status['BRANCH_CODE'], errors='coerce')
                _spec_agg['BRANCH_CODE'] = pd.to_numeric(
                    _spec_agg['BRANCH_CODE'], errors='coerce')
                rating_status = rating_status.merge(_spec_agg, on='BRANCH_CODE', how='left')
                rating_status['BANKERS_COUNT'] = rating_status['BANKERS_COUNT'].fillna(0).astype(int)
                rating_status['OBCHODNI_FTE']  = rating_status['OBCHODNI_FTE'].fillna(0).astype(int)
                rating_status['_total_spec']   = rating_status['_total_spec'].fillna(0).astype(int)

                # Potřeba bankéřů — stejná metodika jako simulace 250 poboček:
                # actual = fyzické + online schůzky + 30% bezhot × (20/45) equiv
                # needed = ceil(actual / (5 sch/den × 248 dní))
                _BANKER_CAP   = 5 * 248          # 1240 schůzek/bankéř/rok
                _BEZHOT_RATE  = 0.30
                _BEZHOT_EQUIV = 20 / 45

                def _to_num(col):
                    return pd.to_numeric(
                        rating_status.get(col, pd.Series(0, index=rating_status.index)),
                        errors='coerce'
                    ).fillna(0)

                _sch_fyz  = _to_num('POCET_SCHUZEK_FYZICKY')
                _sch_onl  = _to_num('POCET_SCHUZEK_ONLINE')
                _bezhot   = _to_num('POCET_BEZHOT_WALK_IN')

                _actual_sch = (
                    _sch_fyz
                    + _sch_onl
                    + _bezhot * _BEZHOT_RATE * _BEZHOT_EQUIV
                )

                rating_status['BANKERS_NEEDED'] = np.ceil(
                    _actual_sch / _BANKER_CAP
                ).astype(int)

                _diff = rating_status['BANKERS_NEEDED'] - rating_status['BANKERS_COUNT']
                def _banker_diff_html(d):
                    if d > 0:
                        return (f"<span style='color:#eb4d79;font-weight:700;'>▲ +{int(d)}</span>")
                    elif d < 0:
                        return (f"<span style='color:#0bb440;font-weight:700;'>▼ {int(d)}</span>")
                    return "<span style='color:#aaa;'>=</span>"
                rating_status['BANKERS_DIFF_HTML'] = _diff.apply(_banker_diff_html)

                # Formát pobočky dle obchodních FTE (stejná pravidla jako celkové FTE)
                rating_status['BRANCH_FORMAT_OBCHODNI'] = rating_status['OBCHODNI_FTE'].apply(_calc_format)
    except Exception as _be:
        pass

    return rating_status


YEAR_INT_COLS = [
    'PLAN_CLOSE_(ROK)', 'INVESTICE_NF_(ROK)',
    'INVESTICE_FHC_(ROK)', 'INVESTICE_NF_OR_FHC_(ROK)',
    'PLAN_NEW_CASHIERLESS_(ROK)'
]

def _fmt_year_int(val):
    """Rok jako celé číslo, prázdné hodnoty jako prázdný řetězec."""
    if val == "" or val is None:
        return ""
    if isinstance(val, float) and np.isnan(val):
        return ""
    try:
        v = int(float(val))
        return str(v) if v > 0 else ""
    except:
        return ""


def _apply_common_formatting(d, cols_to_show):
    """Sdílená formátovací logika pro prepare_table a prepare_table_cols."""
    # Ratingy jako celá čísla
    for c in ['INTERNI_RATING_2023', 'INTERNI_RATING_2024', 'IR']:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(lambda x: int(float(x)) if x not in ('', None) and not (isinstance(x, float) and np.isnan(x)) else "")

    # Poměry klientů — procenta
    for c in ['PRIM_RATIO', 'AKTIVNI_RATIO']:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(
                lambda x: f"{float(x)*100:.1f} %" if pd.notna(x) and str(x) not in ('nan','') else '—')
        d['FTE'] = d['FTE'].apply(
            lambda x: f"{float(x):.1f}" if pd.notna(x) and x not in ('', None, '') and str(x) not in ('nan','') else '—')

    # Bankéři — celá čísla
    for c in ['BANKERS_COUNT', 'BANKERS_NEEDED', '_total_spec']:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(
                lambda x: f"{int(x):,}" if pd.notna(x) and str(x) not in ('nan','') and int(float(x if x else 0)) >= 0 else '—')

    # TRN sloupce — počet s oddělovačem tisíců, částka jako Kč, denní průměr s desetinnou
    if 'TRN_POCET_CELKEM' in cols_to_show and 'TRN_POCET_CELKEM' in d.columns:
        d['TRN_POCET_CELKEM'] = d['TRN_POCET_CELKEM'].apply(
            lambda x: f"{int(x):,}".replace(',', '\u00a0') if pd.notna(x) and x != 0 else '—')
    if 'TRN_CASTKA_CELKEM' in cols_to_show and 'TRN_CASTKA_CELKEM' in d.columns:
        d['TRN_CASTKA_CELKEM'] = d['TRN_CASTKA_CELKEM'].apply(
            lambda x: format_money(x) if pd.notna(x) and x != 0 else '—')
    if 'TRN_DENNI_POCET' in cols_to_show and 'TRN_DENNI_POCET' in d.columns:
        d['TRN_DENNI_POCET'] = d['TRN_DENNI_POCET'].apply(
            lambda x: f"{x:.1f}" if pd.notna(x) and x != 0 else '—')
    # ── Boolean / Ano–Ne formátování ────────────────────────────────
    def _fmt_bool(x):
        s = str(x).strip().upper()
        if s in ('Y','YES','ANO','1','TRUE'):
            return "<span style='color:#0ab33f;font-weight:600;'>Ano</span>"
        if s in ('N','NO','NE','0','FALSE'):
            return "<span style='color:#aaa;'>Ne</span>"
        return '—' if s in ('','NAN','NONE','NAT') else str(x)

    # Sloupce kde True/Y = pozitivní (zelená)
    _bool_cols_positive = ['ONLY_D5_FLAG', 'CASHLESS', 'BNS_FLAG', 'IR_FLAG',
                            'CALC_MIGRATION_FLAG']
    for _bc in _bool_cols_positive:
        if _bc in cols_to_show and _bc in d.columns:
            d[_bc] = d[_bc].apply(_fmt_bool)

    # Vlastnictví: L = Pronajatá (oranžová), O = Vlastní (modrá)
    if 'VLASTNICTVI' in cols_to_show and 'VLASTNICTVI' in d.columns:
        def _fmt_vlast(x):
            s = str(x).strip().upper()
            if s == 'L': return "<span style='color:#e65100;font-weight:600;'>Pronajatá</span>"
            if s == 'O': return "<span style='color:#1565c0;font-weight:600;'>Vlastní</span>"
            return '—' if s in ('','NAN','NONE') else str(x)
        d['VLASTNICTVI'] = d['VLASTNICTVI'].apply(_fmt_vlast)

    # Alokační report — formátování
    if 'ONLY_D5_FLAG' in cols_to_show and 'ONLY_D5_FLAG' in d.columns:
        pass  # již zformátováno výše přes _fmt_bool
    for _ar_col in ['WPL_MAX_2026','WPL_IN_USE_2026']:
        if _ar_col in cols_to_show and _ar_col in d.columns:
            d[_ar_col] = d[_ar_col].apply(
                lambda x: f"{int(x):,}".replace(',',' ') if pd.notna(x) and float(x) > 0 else '—')
    for _ar_col in ['CELK_PLOCHA_POBOCKY_2026','PLOCHA_POUZE_D5','PLOCHA_NAD_OPTIMALNI_POBOCKU']:
        if _ar_col in cols_to_show and _ar_col in d.columns:
            d[_ar_col] = d[_ar_col].apply(
                lambda x: f"{float(x):,.1f} m²".replace(',',' ') if pd.notna(x) and float(x) > 0 else '—')

    # Business rating — dílčí hodnoty na FTE
    if 'CAP_REV_FTE' in cols_to_show and 'CAP_REV_FTE' in d.columns:
        d['CAP_REV_FTE'] = d['CAP_REV_FTE'].apply(
            lambda x: format_money(x) if pd.notna(x) and float(x) > 0 else '—')
    if 'CAP_NEWCLI_FTE' in cols_to_show and 'CAP_NEWCLI_FTE' in d.columns:
        d['CAP_NEWCLI_FTE'] = d['CAP_NEWCLI_FTE'].apply(
            lambda x: f"{x:.2f}" if pd.notna(x) and float(x) > 0 else '—')
    if 'CAP_PRIMCLI_FTE' in cols_to_show and 'CAP_PRIMCLI_FTE' in d.columns:
        d['CAP_PRIMCLI_FTE'] = d['CAP_PRIMCLI_FTE'].apply(
            lambda x: f"{x:.1f}" if pd.notna(x) and float(x) > 0 else '—')
    if 'CAP_SCHUZKY_FTE' in cols_to_show and 'CAP_SCHUZKY_FTE' in d.columns:
        d['CAP_SCHUZKY_FTE'] = d['CAP_SCHUZKY_FTE'].apply(
            lambda x: f"{x:.2f}" if pd.notna(x) and float(x) > 0 else '—')
    if 'CAP_PEREX' in cols_to_show and 'CAP_PEREX' in d.columns:
        d['CAP_PEREX'] = d['CAP_PEREX'].apply(
            lambda x: f"{x:.3f}" if pd.notna(x) and float(x) > 0 else '—')
    if 'CAP_RNK' in cols_to_show and 'CAP_RNK' in d.columns:
        d['CAP_RNK'] = d['CAP_RNK'].apply(
            lambda x: str(int(x)) if pd.notna(x) and float(x) > 0 else '—')
    if 'CAP_Q' in cols_to_show and 'CAP_Q' in d.columns:
        d['CAP_Q'] = d['CAP_Q'].apply(
            lambda x: str(int(x)) if pd.notna(x) and float(x) > 0 else '—')
    if 'TRN_RNK' in cols_to_show and 'TRN_RNK' in d.columns:
        d['TRN_RNK'] = d['TRN_RNK'].apply(
            lambda x: str(int(x)) if pd.notna(x) and float(x) > 0 else '—')

    if 'TRN_Q' in cols_to_show and 'TRN_Q' in d.columns:
        d['TRN_Q'] = d['TRN_Q'].apply(
            lambda x: str(int(x)) if pd.notna(x) and float(x) > 0 else '—')

    # Peněžní sloupce
    for c in ['REVENUES_2021', 'REVENUES_2022', 'REVENUES_2023',
              'REVENUES_2024', 'VYNOSY', 'NEW_BUSINESS_2024_-_OBJEM_VYNOSU', 'OBJEM_VYNOSU_CZK',
              'ROCNI_SPLATKY_S_DPH_CZK', 'TOTAL_INV', 'LAST_INV', 'BIGGEST_INV']:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(format_money)

    # *** NOVÉ: hotovostní částky jako peněžní hodnoty ***
    for c in CASH_CASTKA_COLS:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(format_money)

    # *** NOVÉ: hotovostní počty jako celá čísla ***
    for c in CASH_POCET_COLS:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(format_int_sep)

    # *** NOVÉ: % převoditelnosti ***
    if 'CASH_PREVODITELNOST_PCT' in cols_to_show and 'CASH_PREVODITELNOST_PCT' in d.columns:
        def fmt_pct_prevod(v):
            try:
                if v is None or (isinstance(v, float) and np.isnan(v)):
                    return '<span style="color:#ccc;">—</span>'
                pct = float(v)
                # Barva: zelená ≥ 50 %, oranžová 20–50 %, červená < 20 %
                if pct >= 50:
                    color = '#0ab33f'
                elif pct >= 20:
                    color = '#f0b55d'
                else:
                    color = '#e64343'
                bar_width = min(int(pct), 100)
                return (
                    f'<div style="min-width:70px">'
                    f'<div style="background:#eee;border-radius:3px;height:6px;margin-bottom:2px;">'
                    f'<div style="background:{color};width:{bar_width}%;height:6px;border-radius:3px;"></div></div>'
                    f'<span style="color:{color};font-weight:600;font-size:0.85rem;">{pct:.1f} %</span>'
                    f'</div>'
                )
            except:
                return ''
        d['CASH_PREVODITELNOST_PCT'] = d['CASH_PREVODITELNOST_PCT'].apply(fmt_pct_prevod)

    # Procentuální sloupce
    for c in ['C/I_RATIO_(YTD_2021)', 'C/I_RATIO_(YTD_2022)', 'C/I_RATIO_(YTD_2023)',
              'C/I_RATIO_(YTD_2024)', 'PRIME_NAKLADY/VYNOSY']:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(format_pct)

    # Počet klientů + primární + aktivní + EOY
    for c in ['POCET_KLIENTU', 'PRIMARNI_KLIENTI', 'AKTIVNI_KLIENTI',
              'POCET_KLIENTU_EOY_2021', 'POCET_KLIENTU_EOY_2022',
              'POCET_KLIENTU_EOY_2023', 'POCET_KLIENTU_EOY_2024']:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(format_int_sep)

    # Index expozice — zaokrouhlení na 2 desetinná místa
    if 'HODNOTA_INDEXU_EXPOZICE' in cols_to_show and 'HODNOTA_INDEXU_EXPOZICE' in d.columns:
        d['HODNOTA_INDEXU_EXPOZICE'] = d['HODNOTA_INDEXU_EXPOZICE'].apply(
            lambda x: f"{float(x):.2f}".replace('.', ',') if pd.notna(x) and str(x) not in ('','nan') else '—'
        )

    # Cizinci
    for c in ['CIZINCI', 'CIZINCI_SLOVENSKO', 'CIZINCI_UKRAJINA']:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(format_int_sep)

    # Roky jako celá čísla
    for c in YEAR_INT_COLS:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(_fmt_year_int)

    # *** OPRAVA + ROZŠÍŘENÍ: všechna data v DATE_COLS_DISPLAY včetně ZACATEK_DATUM a KONEC_DATUM ***
    for c in DATE_COLS_DISPLAY:
        if c in cols_to_show and c in d.columns:
            d[c] = d[c].apply(format_date)

    # Podbarvení FOOTPRINT_STRATEGIE_(2030)
    if 'FOOTPRINT_STRATEGIE_(2030)' in cols_to_show and 'FOOTPRINT_STRATEGIE_(2030)' in d.columns:
        d['FOOTPRINT_STRATEGIE_(2030)'] = d['FOOTPRINT_STRATEGIE_(2030)'].apply(color_strategy_html)

    # Prodeje — celková čísla formát
    _prod_num_cols = (
        ['POCET_PRODEJU_CELKEM_2025'] +
        [f'POCET_PRODEJU_{k}_2025' for k, _ in PRODUCT_GROUPS]
    )
    for _pc in _prod_num_cols:
        if _pc in cols_to_show and _pc in d.columns:
            d[_pc] = d[_pc].apply(
                lambda x: f"{int(round(float(x))):,}".replace(',', '\u00a0')
                if pd.notna(x) and str(x) not in ('', 'nan', 'None') and float(x) > 0 else '—'
            )

    # Obchodní FTE — celé číslo
    if 'OBCHODNI_FTE' in cols_to_show and 'OBCHODNI_FTE' in d.columns:
        d['OBCHODNI_FTE'] = d['OBCHODNI_FTE'].apply(format_int_sep)

    # Formát pobočky dle obchodního FTE — barevný badge stejně jako BRANCH_FORMAT
    for _fmt_col in ['BRANCH_FORMAT', 'BRANCH_FORMAT_OBCHODNI']:
        if _fmt_col in cols_to_show and _fmt_col in d.columns:
            _fmt_display = NOVE_NAZVY.get(_fmt_col, _fmt_col)
            def _fmt_badge(v, col=_fmt_col):
                s = str(v or '').lower().strip()
                _colors = {'small': '#f59e0b', 'medium economy': '#fb923c',
                           'medium': '#2770f0', 'flagship': '#0bb440'}
                c = _colors.get(s, '#aaa')
                lbl = s.title() if s else '—'
                return (f'<span style="background:{c}18;color:{c};border:1px solid {c}44;'
                        f'border-radius:5px;padding:1px 7px;font-size:0.72rem;'
                        f'font-weight:600;white-space:nowrap;">{lbl}</span>')
            d[_fmt_col] = d[_fmt_col].apply(_fmt_badge)
    if 'ATM_VOLNA_KAPACITA' in cols_to_show and 'ATM_VOLNA_KAPACITA' in d.columns:
        def _fmt_atm_pct(v):
            try:
                fv = float(v)
                if pd.isna(fv): return '—'
                c = '#eb4d79' if fv < 0 else ('#f59e0b' if fv < 20 else '#0bb440')
                return f'<span style="font-weight:700;color:{c};">{fv:.1f}%</span>'
            except: return '—'
        d['ATM_VOLNA_KAPACITA'] = d['ATM_VOLNA_KAPACITA'].apply(_fmt_atm_pct)
    for _atm_num_col in ['ATM_VYBERY_MESIC','ATM_VKLADY_MESIC','ATM_VYBERY_PO_PREV','ATM_VKLADY_PO_PREV']:
        if _atm_num_col in cols_to_show and _atm_num_col in d.columns:
            d[_atm_num_col] = d[_atm_num_col].apply(
                lambda x: f"{int(round(float(x))):,}".replace(',','\u00a0')
                if pd.notna(x) and str(x) not in ('','nan','None') else '—'
            )

    # Nahradit zbývající NaN prázdným řetězcem
    for c in d.select_dtypes(include=['object']).columns:
        d[c] = d[c].fillna("")

    # ── Spádová oblast — návštěvy jako číslo se separátorem, % s pruhem ──
    for _i in [1, 2, 3]:
        _vc = f'SP_VISITS_{_i}'
        _pc = f'SP_PCT_{_i}'
        _nc = f'SP_NAZEV_{_i}'
        if _vc in cols_to_show and _vc in d.columns:
            d[_vc] = d[_vc].apply(
                lambda x: f"{int(float(x)):,}".replace(',', '\u00a0')
                if pd.notna(x) and str(x).strip() not in ('', '0', 'nan') else '—'
            )
        if _pc in cols_to_show and _pc in d.columns:
            def _fmt_sp_pct(v, _col=_pc):
                try:
                    pct = float(v)
                    if pct <= 0 or np.isnan(pct):
                        return '<span style="color:#ccc;">—</span>'
                    clr = '#2770f0' if pct >= 40 else ('#6b7fe8' if pct >= 20 else '#a5b4fc')
                    bw  = min(int(pct), 100)
                    return (
                        f'<div style="min-width:75px">'
                        f'<div style="background:#eef0f8;border-radius:3px;height:5px;margin-bottom:2px;">'
                        f'<div style="background:{clr};width:{bw}%;height:5px;border-radius:3px;"></div></div>'
                        f'<span style="color:{clr};font-weight:600;font-size:0.83rem;">{pct:.1f} %</span>'
                        f'</div>'
                    )
                except:
                    return '—'
            d[_pc] = d[_pc].apply(_fmt_sp_pct)
        if _nc in cols_to_show and _nc in d.columns:
            d[_nc] = d[_nc].apply(
                lambda x: str(x) if pd.notna(x) and str(x).strip() not in ('', 'nan') else '—'
            )

    return d


DASH_HTML = "<span style='color:#bbb;font-weight:normal;'>—</span>"

# Sloupce kde pro speciální pobočky (bez ratingu) zobrazíme pomlčku
SPECIAL_DASH_COLS = {
    'IR', 'IR_Q_HTML', 'IR_CHANGE_HTML', 'IR_PERC_CHANGE_HTML',
    'INTERNI_RATING_2023', 'INTERNI_RATING_2024',
    'CI_RATIO_STATUS', 'CI_CHANGE_HTML',
    'NEW_BUSINESS_VOL_STATUS', 'REV_CHANGE_HTML',
    'ROCNI_SPLATKY_S_DPH_CZK_STATUS',
    'C/I_RATIO_(YTD_2021)', 'C/I_RATIO_(YTD_2022)', 'C/I_RATIO_(YTD_2023)',
    'C/I_RATIO_(YTD_2024)', 'PRIME_NAKLADY/VYNOSY',
    'REVENUES_2021', 'REVENUES_2022', 'REVENUES_2023',
    'REVENUES_2024', 'VYNOSY',
    'NEW_BUSINESS_2024_-_OBJEM_VYNOSU', 'OBJEM_VYNOSU_CZK',
    'CAP_RNK', 'CAP_Q', 'CAP_Q_HTML',
    'TRN_RNK', 'TRN_Q', 'TRN_Q_HTML', 'TRN_POCET_CELKEM', 'TRN_CASTKA_CELKEM', 'TRN_DENNI_POCET',
}

def _apply_special_dashes(d, cols_to_show):
    """Pro speciální pobočky (bez IR) nahradí nulové/prázdné ratingové hodnoty pomlčkou."""
    if 'BRANCH_CODE' not in d.columns or 'IR_FLAG' not in d.columns:
        return d
    spec_mask = d['BRANCH_CODE'].isin(SPECIAL_BRANCH_CODES)
    if not spec_mask.any():
        return d
    for c in cols_to_show:
        if c not in SPECIAL_DASH_COLS or c not in d.columns:
            continue
        for idx in d[spec_mask].index:
            val = d.at[idx, c]
            v_str = str(val).strip()
            if v_str in ('', '0', '0 Kč', '0,00%', 'nan', 'None') or val == 0:
                d.at[idx, c] = DASH_HTML
    return d


def prepare_table(target_df, excluded_cols=None):
    """Naformátuje dataframe a vrátí HTML tabulku pro report."""
    d = target_df.copy()
    _excl = set(excluded_cols or [])
    cols_to_show = [c for c in NOVE_NAZVY.keys() if c in d.columns and c not in _excl]

    d = _apply_common_formatting(d, cols_to_show)
    d = _apply_special_dashes(d, cols_to_show)

    return (d[cols_to_show]
            .rename(columns=NOVE_NAZVY)
            .to_html(escape=False, index=False,
                     classes='table table-bordered table-striped table-hover table-sm'))


def prepare_table_cols(target_df, cols, sort_col=None, sort_asc=True, highlight_cols=None):
    """Jako prepare_table, ale zobrazí jen vybrané sloupce.
    sort_col:       technický název primárního sloupce řazení (zvýrazněn oranžově).
    sort_asc:       True = vzestupně (↑), False = sestupně (↓).
    highlight_cols: seznam dalších technických názvů sloupců ke zvýraznění (stejná barva).
    """
    d = target_df.copy()
    cols_to_show = [c for c in cols if c in d.columns]

    d = _apply_common_formatting(d, cols_to_show)

    rename = {k: v for k, v in NOVE_NAZVY.items() if k in cols_to_show}
    html = (d[cols_to_show]
            .rename(columns=rename)
            .to_html(escape=False, index=False,
                     classes='table table-bordered table-striped table-hover table-sm'))

    # Sestavíme seznam všech sloupců ke zvýraznění: primární + doplňkové
    hl_cols = []
    if sort_col and sort_col in rename:
        hl_cols.append((sort_col, rename[sort_col], True))   # (tech, label, přidej šipku)
    for hc in (highlight_cols or []):
        if hc in rename and hc != sort_col:
            hl_cols.append((hc, rename[hc], False))

    if not hl_cols:
        return html

    table_id = f'sort-hl-{abs(hash(str(sort_col) + str(highlight_cols) + str(sort_asc) + str(id(target_df)))) % 100000}'
    html = html.replace('<table ', f'<table id="{table_id}" ', 1)

    # Zvýrazni hlavičky — regex aby zachytil i th s atributy (rowspan apod.)
    for tech, label, add_arrow in hl_cols:
        arrow = (' ↑' if sort_asc else ' ↓') if add_arrow else ''
        html = re.sub(
            r'(<th[^>]*>)(' + re.escape(label) + r')(</th>)',
            rf'\g<1><span style="background:#f39c12;color:white;display:block;padding:2px 4px;white-space:nowrap;">{label}{arrow}</span>\3',
            html
        )

    # JavaScript: zvýrazni buňky podle indexu sloupce
    # Použijeme indexOf na textContent pro robustnost
    js_labels = [label for (_, label, _) in hl_cols]
    js_labels_json = str(js_labels).replace("'", '"')
    html += f"""
<script>
(function() {{
  var tbl = document.getElementById('{table_id}');
  if (!tbl) return;
  var ths = Array.from(tbl.querySelectorAll('thead tr:first-child th'));
  var labels = {js_labels_json};
  labels.forEach(function(lbl) {{
    var colIdx = -1;
    ths.forEach(function(th, i) {{
      var txt = th.textContent.replace(/[↑↓]/g,'').trim();
      if (txt === lbl.trim() || txt.indexOf(lbl.trim()) >= 0) colIdx = i;
    }});
    if (colIdx < 0) return;
    ths[colIdx].style.setProperty("background-color","#f39c12","important");
    ths[colIdx].style.setProperty("color","white","important");
    tbl.querySelectorAll('tbody tr').forEach(function(tr) {{
      var td = tr.cells[colIdx];
      if (td) td.style.setProperty("background-color","rgba(243,156,18,0.25)","important");
    }});
  }});
}})();
</script>"""

    return html


# Sady sloupců pro Top 10 tabulky
COLS_TOP10_REVENUE = [
    "BRANCH_CODE", "BRANCH_NAME", "REGION_NAME",
    "OBJEM_VYNOSU_CZK", "NEW_BUSINESS_2024_-_OBJEM_VYNOSU", "REV_CHANGE_HTML", "NEW_BUSINESS_VOL_STATUS",
    "VYNOSY", "REVENUES_2024",
    "POCET_PRODEJU",
]

COLS_TOP10_RATING = [
    "BRANCH_CODE", "BRANCH_NAME", "REGION_NAME",
    "IR", "INTERNI_RATING_2024", "IR_CHANGE_HTML", "IR_PERC_CHANGE_HTML",
    "OBJEM_VYNOSU_CZK", "REV_CHANGE_HTML",
    "PRIME_NAKLADY/VYNOSY", "CI_CHANGE_HTML", "CI_RATIO_STATUS",
]


def prepare_excel_df(target_df, excluded_cols=None):
    """Vrátí dataframe s čistými hodnotami pro export do Excelu."""
    d = target_df.copy()
    _excl = set(excluded_cols or [])
    cols_to_show = [c for c in NOVE_NAZVY.keys() if c in d.columns and c not in _excl]

    for c in ['INTERNI_RATING_2023', 'INTERNI_RATING_2024', 'IR']:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors='coerce')

    html_col_map = {
        'IR_CHANGE_HTML':      ('IR', 'INTERNI_RATING_2024'),
        'IR_PERC_CHANGE_HTML': ('IR_PERC', 'INTERNI_RATING_2024_-_PERCENTILY'),
        'REV_CHANGE_HTML':     ('OBJEM_VYNOSU_CZK', 'NEW_BUSINESS_2024_-_OBJEM_VYNOSU'),
        'CI_CHANGE_HTML':      ('PRIME_NAKLADY/VYNOSY', 'C/I_RATIO_(YTD_2024)'),
    }
    for html_col, (col_a, col_b) in html_col_map.items():
        if html_col in d.columns and col_a in d.columns and col_b in d.columns:
            d[html_col] = pd.to_numeric(d[col_a], errors='coerce') - pd.to_numeric(d[col_b], errors='coerce')

    status_map = {
        'CI_RATIO_STATUS':              'CI_RATIO_Q',
        'NEW_BUSINESS_VOL_STATUS':      'NEW_BUSINESS_VOL_Q',
        'ROCNI_SPLATKY_S_DPH_CZK_STATUS': 'ROCNI_SPLATKY_S_DPH_CZK_Q',
    }
    for status_col, q_col in status_map.items():
        if status_col in d.columns and q_col in d.columns:
            d[status_col] = pd.to_numeric(d[q_col], errors='coerce')

    # Produktové kvintilové HTML sloupce → nahraď číselnou hodnotou Q
    # (buňka s tečkou &#9679; neobsahuje číslo → write_excel_sheet nemůže barvit)
    for _pk, _ in PRODUCT_GROUPS:
        _qhtml = f'POCET_PRODEJU_{_pk}_2025_Q_HTML'
        _qcol  = f'POCET_PRODEJU_{_pk}_2025_Q'
        if _qhtml in d.columns and _qcol in d.columns:
            d[_qhtml] = pd.to_numeric(d[_qcol], errors='coerce')

    # Ostatní kvintilové HTML sloupce kde je jen tečka bez čísla — nahraď _Q
    _q_html_to_q = {
        'IE_Q_HTML':             'IE_Q',
        'IR_Q_HTML':             'IR_Q',
        'CAP_Q_HTML':            'CAP_Q',
        'TRN_Q_HTML':            'TRN_Q',
        'CAP_REV_FTE_Q_HTML':    'CAP_REV_FTE_Q',
        'CAP_NEWCLI_FTE_Q_HTML': 'CAP_NEWCLI_FTE_Q',
        'CAP_PRIMCLI_FTE_Q_HTML':'CAP_PRIMCLI_FTE_Q',
        'CAP_SCHUZKY_FTE_Q_HTML':'CAP_SCHUZKY_FTE_Q',
        'CAP_PEREX_Q_HTML':      'CAP_PEREX_Q',
    }
    for _qhtml, _qcol in _q_html_to_q.items():
        if _qhtml in d.columns and _qcol in d.columns:
            d[_qhtml] = pd.to_numeric(d[_qcol], errors='coerce')

    for c in DATE_COLS_DISPLAY:
        if c in d.columns:
            d[c] = d[c].apply(format_date)

    for c in YEAR_INT_COLS:
        if c in d.columns:
            d[c] = d[c].apply(_fmt_year_int)

    for c in d.select_dtypes(include=['object']).columns:
        d[c] = d[c].fillna("")

    return d[cols_to_show].rename(columns=NOVE_NAZVY)


def generate_sortable_table(html_table, table_id):
    """Obalí HTML tabulku do scrollovatelného divu."""
    return f"""
<div class="table-responsive" style="max-height:520px;overflow:auto;">
  {html_table}
</div>"""


def write_excel_sheet(ws, df_excel, freeze="D2"):
    """
    Jednotné formátování Excel listu — záhlaví, kvintily, barvy skupin, šířky.
    Sdíleno mezi statickým a regionálním exportem.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import re as _re_xl

    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    excel_cols = list(df_excel.columns)

    # Pastelová barva skupiny pro každý display název
    disp2pastel = {}
    for grp_label, grp_cols, grp_pastel in COL_GROUPS:
        ph = grp_pastel.lstrip("#")
        if len(ph) == 3: ph = "".join(c*2 for c in ph)
        ph = ph[:6].upper()
        for dc in grp_cols:
            disp2pastel[dc] = ph

    FIXED     = {"ID Pobočky", "Název pobočky", "Region"}
    FIXED_HDR = "2D5FA8"

    def darken(hx, f=0.58):
        r,g,b = int(hx[0:2],16), int(hx[2:4],16), int(hx[4:6],16)
        return f"{int(r*f):02X}{int(g*f):02X}{int(b*f):02X}"

    def lighten(hx, amt=0.60):
        r,g,b = int(hx[0:2],16), int(hx[2:4],16), int(hx[4:6],16)
        r = int(r+(255-r)*amt); g = int(g+(255-g)*amt); b = int(b+(255-b)*amt)
        return f"{min(r,255):02X}{min(g,255):02X}{min(b,255):02X}"

    # Barvy kvintilů — zelená=Q1 (nejlepší), červená=Q5 (nejhorší)
    Q_BG  = {1:"C6EFCE", 2:"CCFFCC", 3:"FFEB9C", 4:"FFCC99", 5:"FFC7CE"}
    Q_FG  = {1:"276221", 2:"276221", 3:"7A5800", 4:"8A3300", 5:"9C0006"}
    Q_LBL = {1:"Q1", 2:"Q2", 3:"Q3", 4:"Q4", 5:"Q5"}

    # Všechny sloupce kde chceme kvintilové formátování
    # — statické kvintily + dynamické produktové kvintily
    Q_COLS_STATIC = {
        "C/I kvintil", "Nové výnosy kvintil", "Nájemné kvintil",
        "Rating 25 kvintil", "Cash rating kvintil", "Business rating kvintil",
        "Výnos/FTE kvintil", "Noví klienti/FTE kvintil",
        "Prim. klienti/FTE kvintil", "Schůzky/FTE kvintil", "PEREX kvintil",
        "Index expozice kvintil",
    }
    # Produktové kvintily — všechny sloupce co končí " kvintil"
    Q_COLS = Q_COLS_STATIC | {c for c in excel_cols if c.endswith(" kvintil")}

    def strip_html(v):
        return _re_xl.sub(r"<[^>]+>", "", str(v)).replace("&#9679;","●").strip() if v else ""

    def extract_q(v):
        if v is None or v == "":
            return None
        # Přímá numerická hodnota (z prepare_excel_df)
        try:
            qi = int(float(v))
            if 1 <= qi <= 5:
                return qi
        except (ValueError, TypeError):
            pass
        # Fallback: hledej číslo v textu (pro starší formáty)
        s = strip_html(str(v))
        m = _re_xl.search(r"\b([1-5])\b", s)
        return int(m.group(1)) if m else None

    # ── Řádek 1: záhlaví ──────────────────────────────────────────
    for ci, col in enumerate(excel_cols, 1):
        hx   = FIXED_HDR if col in FIXED else darken(disp2pastel.get(col, "888888"))
        cell = ws.cell(row=1, column=ci)
        cell.fill      = PatternFill("solid", fgColor=hx)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = brd

    # ── Řádky dat ─────────────────────────────────────────────────
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=0):
        is_alt = (ri % 2 == 1)
        for ci, cell in enumerate(row, 1):
            col = excel_cols[ci-1] if ci <= len(excel_cols) else ""
            raw = cell.value
            # Kvintilové sloupce — sjednocené formátování
            if col in Q_COLS:
                q = extract_q(raw)
                if q and q in Q_BG:
                    cell.value     = Q_LBL[q]
                    cell.fill      = PatternFill("solid", fgColor=Q_BG[q])
                    cell.font      = Font(bold=True, color=Q_FG[q], size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = brd
                    continue
            # HTML → čistý text
            if isinstance(raw, str) and "<" in raw:
                cell.value = strip_html(raw)
            ph = disp2pastel.get(col)
            bg = lighten(ph, 0.75 if is_alt else 0.60) if ph else ("F4F7FF" if is_alt else "FFFFFF")
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = brd

    # ── Šířky sloupců ─────────────────────────────────────────────
    for i, col in enumerate(excel_cols, 1):
        letter = get_column_letter(i)
        best = max(
            (len(str(ws.cell(row=r, column=i).value or "")) for r in range(1, ws.max_row + 1)),
            default=8
        )
        ws.column_dimensions[letter].width = min(best + 2, 36)

    ws.freeze_panes = freeze
    ws.row_dimensions[1].height = 32


# Skupiny sloupců pro přepínání v hlavní tabulce regionu
COL_GROUPS = [
    ("📍 Adresa",              ["Oblast", "Město", "Obvod / část", "Ulice", "Č. popisné", "Č. orientační", "RUIAN ID", "ORP", "ORP kód"], "#f0f4f8"),
    ("👥 Klienti",             ["Klienti EOY 2021", "Klienti EOY 2022", "Klienti EOY 2023",
                                "Klienti EOY 2024", "Klienti EOY 2025", "Trend klientů 21–25",
                                "Primární klienti", "Aktivní klienti",
                                "Poměr klientů", "Primární / celkem", "Aktivní / celkem",
                                "Cizinci celkem", "Cizinci SK", "Cizinci UA"], "#e0f7fa"),
    ("⭐ Interní rating",      ["Rating 23", "Rating 24", "Rating 25", "Rating 25 kvintil", "Změna ratingu 25/24", "Změna ratingu perc 25/24"], "#fff8e1"),
    ("📊 C/I",                 ["C/I ratio 21", "C/I ratio 22", "C/I ratio 23", "C/I ratio 24", "C/I ratio 25", "C/I kvintil", "Změna C/I 25/24"], "#dceeff"),
    ("💰 Výnosy",              ["Výnosy 21", "Výnosy 22", "Výnosy 23", "Výnosy 24", "Výnosy 25",
                                "Trend výnosů 21–25",
                                "Nové výnosy 24", "Nové výnosy 25", "Nové výnosy kvintil", "Změna nových výnosů 25/24"], "#d4f5e0"),
    ("📈 Business rating",     [
        "Výnos/FTE", "Výnos/FTE kvintil",
        "Noví klienti/FTE", "Noví klienti/FTE kvintil",
        "Primární klienti/FTE", "Prim. klienti/FTE kvintil",
        "Schůzky/FTE", "Schůzky/FTE kvintil",
        "PEREX", "PEREX kvintil",
        "Business pořadí", "Business rating kvintil",
    ], "#f3e5f5"),
    ("🏦 Hotovostní rating",   ["Cash rating pořadí", "Vypočítaná strategie", "Cash rating kvintil", "Trn. počet celkem", "Trn. objem celkem", "Denní trn. průměr"], "#e3f2fd"),
    ("🏠 Nájemné",             ["Vlastnictví", "Roční nájemné", "Nájemné kvintil", "Začátek smlouvy", "Konec smlouvy", "Výpovědní lhůta"], "#eeddf7"),
    ("🔨 Investice",           ["Poslední investice", "Datum posl. inv.", "Výše posl. inv.", "Nejvyšší investice", "Datum nejv. inv.", "Výše nejv. inv.", "Celkem zainvestováno"], "#fce8d8"),
    ("🚶 Návštěvy",            ["Poměr návštěv", "Online schůzky", "Fyzické schůzky", "Bezhot. walkin", "Hot. walkin", "Celkové návštěvy"], "#fad8e0"),
    ("🗺️ Strategie BNS",      ["Strategie BNS", "Rok uzavření", "Status uzavření", "Rok investice", "Rok FHC investice", "Rok plánované investice"], "#faf0cc"),
    ("💵 Strategie hotovosti", ["Rok cashless přechodu", "Status cashless", "Bezhotovostní", "Datum bezhotovostní"], "#e8f5e9"),
    ("📐 Alokační report",     [
        "Pouze D5", "WPL max 2026", "WPL využito 2026",
        "Celk. plocha pobočky", "Plocha pouze D5", "Plocha nad optimál",
    ], "#fce4ec"),
    ("🏢 Budova",              ["Typologie", "Formát pobočky (celk. FTE)", "Formát pobočky (obch. FTE)", "Realizovaný formát", "Formát NF/SF", "Datum NF", "Bezhotovostní", "Datum bezhotovostní"], "#d8edd8"),
    ("💵 Transakční data",     [
        "% převoditelnosti",
        "Cash IN převod. Kč", "Cash IN nepřevod. Kč",
        "Cash IN převod. počet", "Cash IN nepřevod. počet",
        "Cash OUT převod. Kč", "Cash OUT nepřevod. Kč",
        "Cash OUT převod. počet", "Cash OUT nepřevod. počet",
    ], "#fff3cd"),
    ("🔀 Spádová oblast",     [
        "Spádová pobočka 1", "Spád. návštěvy 1", "Spád. podíl 1",
        "Spádová pobočka 2", "Spád. návštěvy 2", "Spád. podíl 2",
        "Spádová pobočka 3", "Spád. návštěvy 3", "Spád. podíl 3",
    ], "#e8f4fd"),
    ("👷 FTE",               [
        "FTE", "Obchodní FTE", "Počet bankéřů", "Potřeba bankéřů", "Bankéři — rozdíl", "Všechny pozice celkem",
    ], "#e8f0fe"),
    ("🏅 Sdružené ratingy",   [
        "Rating 25 kvintil",
        "C/I kvintil",
        "Nové výnosy kvintil",
        "Business rating kvintil",
        "Výnos/FTE kvintil",
        "Noví klienti/FTE kvintil",
        "Prim. klienti/FTE kvintil",
        "Schůzky/FTE kvintil",
        "PEREX kvintil",
        "Cash rating kvintil",
        "Nájemné kvintil",
    ], "#fef9e7"),
    ("🛍️ Prodeje",             [
        "Prodeje celkem 2025",
        "Prodeje Účty 2025",        "Účty kvintil",
        "Prodeje Hypotéky 2025",    "Hypotéky kvintil",
        "Prodeje Poj. život 2025",  "Poj. život kvintil",
        "Prodeje Poj. neživot 2025","Poj. neživot kvintil",
        "Prodeje Inv. pravid. 2025","Inv. pravid. kvintil",
        "Prodeje Inv. jednor. 2025","Inv. jednor. kvintil",
        "Prodeje Revol. úvěry 2025","Revol. úvěry kvintil",
        "Prodeje Úvěry 2025",       "Úvěry kvintil",
        "Prodeje Penze 2025",       "Penze kvintil",
    ], "#fff0e8"),
    ("📡 Index expozice",      ["Index expozice", "Index expozice kvintil"], "#e8f5e9"),
    ("🏧 ATM",                ["ATM přehled", "ATM volná kap. %",
                                "ATM výběry průměr/měs.", "ATM vklady průměr/měs.",
                                "ATM výběry po převodu/měs.", "ATM vklady po převodu/měs."], "#e8f4fd"),
]

# Odpovídající akcentové barvy pro okraje přepínačů (pořadí odpovídá COL_GROUPS)
COL_GROUP_ACCENT = [
    '#607d8b',   # Adresa
    '#0097a7',   # Klienti
    '#f9a825',   # Interní rating
    '#1565c0',   # C/I
    '#2e7d32',   # Výnosy
    '#7b1fa2',   # Business rating
    '#1565c0',   # Hotovostní rating
    '#6a1b9a',   # Nájemné
    '#e65100',   # Investice
    '#ad1457',   # Návštěvy
    '#f57f17',   # Strategie BNS
    '#33691e',   # Strategie hotovosti
    '#c2185b',   # Alokační report
    '#388e3c',   # Budova
    '#b8860b',   # Transakční data
    '#0277bd',   # Spádová oblast
    '#3f51b5',   # FTE
    '#d4a017',   # Sdružené ratingy
    '#d4682a',   # Prodeje
    '#2e7d32',   # Index expozice
    '#1565c0',   # ATM
]


def generate_column_map_html():
    """
    Kontrolní mapa sloupců: zdroj → technický název → zobrazený název → použití v tabulkách.
    """
    # ── Zdroj každého technického sloupce ────────────────────────────────────
    # (tech_col, zdroj_df, zdroj_sloupec, poznamka)
    COMPUTED_NOTE = "📐 vypočítáno"
    COL_SOURCES = []

    # Z MERGE_COLS — přímé joiny
    src_map = {
        "profitabilita": ("profitabilita", "Pobockova_profitabilita_4Q2025.xlsx"),
        "revenues":      ("revenues (LTV CSV)", "YEARLY_VYNOSY_NOVYCH_OBCHODU...csv"),
        "old_ir":        ("old_ir", "IR_2024_v06.xlsx"),
        "dbs":           ("dbs", "dbs_branch_network_epb.csv"),
        "footprint_bns": ("footprint_bns", "footprint_bns.xlsx"),
        "rent_and_ownership": ("rent_and_ownership", "export_najmy_2026.pkl"),
    }
    for src_key, (src_label, src_file) in src_map.items():
        mc = MERGE_COLS.get(src_key, {})
        for col in mc.get("cols", []):
            if col in ("BRANCH_CODE", "BRANCH_ID", "HJ", "ID_POBOCKY"):
                continue  # join klíče
            COL_SOURCES.append((col, src_label, col, ""))

    # Speciální zdroje mimo MERGE_COLS
    extra = [
        # Identita (dbs)
        ("BRANCH_CODE",             "dbs",              "BRANCH_CODE",              "primární klíč"),
        ("BRANCH_NAME",             "dbs",              "BRANCH_NAME",              ""),
        ("REGION_NAME",             "dbs",              "REGION_NAME",              ""),
        ("OBLAST",                  "dbs",              "OBLAST",                   ""),
        ("CONCATENATED_HJ",         "dbs",              "CONCATENATED_HJ",          ""),
        # Adresa (dbs)
        ("CITY",                    "dbs",              "CITY",                     ""),
        ("CITY_DISTRICT",           "dbs",              "CITY_DISTRICT",            ""),
        ("STREET",                  "dbs",              "STREET",                   ""),
        ("CISLO_POPISNE",           "dbs",              "CISLO_POPISNE",            ""),
        ("CISLO_ORIENTACNI",        "dbs",              "CISLO_ORIENTACNI",         ""),
        ("RUIAN_ID",                "dbs",              "RUIAN_ID",                 ""),
        # Cizinci (parties)
        ("CIZINCI",                 "parties",          "CIZINCI",                  ""),
        ("CIZINCI_SLOVENSKO",       "parties",          "CIZINCI_SLOVENSKO",        ""),
        ("CIZINCI_UKRAJINA",        "parties",          "CIZINCI_UKRAJINA",         ""),
        # KPI návštěvy
        ("POCET_SCHUZEK_ONLINE",    "kpis",             "POCET_SCHUZEK_ONLINE",     ""),
        ("POCET_SCHUZEK_FYZICKY",   "kpis",             "POCET_SCHUZEK_FYZICKY",    ""),
        ("POCET_BEZHOT_WALK_IN",    "kpis",             "POCET_BEZHOT_WALK_IN",     ""),
        ("POCET_HOT_WALK_IN",       "kpis",             "POCET_HOT_WALK_IN",        ""),
        ("POCET_NAVSTEV_CELKEM",    "kpis",             "POCET_NAVSTEV_CELKEM",     ""),
        ("NR_NEW_ARRIVALS",         "kpis",             "NR_NEW_ARRIVALS",          "noví klienti"),
        # Alokační report
        ("ONLY_D5_FLAG",            "ar",               "ONLY_D5_FLAG",             ""),
        ("WPL_MAX_2026",            "ar",               "WPL_MAX_2026",             ""),
        ("WPL_IN_USE_2026",         "ar",               "WPL_IN_USE_2026",          ""),
        ("CELK_PLOCHA_POBOCKY_2026","ar",               "CELK_PLOCHA_POBOCKY_2026", ""),
        ("PLOCHA_POUZE_D5",         "ar",               "PLOCHA_POUZE_D5",          ""),
        ("PLOCHA_NAD_OPTIMALNI_POBOCKU","ar",           "PLOCHA_NAD_OPTIMALNI_POBOCKU",""),
        # old_ir — historické revenues a C/I
        ("C/I_RATIO_(YTD_2021)",    "old_ir",           "C/I_RATIO_(YTD_2021)",     "C/I ratio 2021"),
        ("C/I_RATIO_(YTD_2022)",    "old_ir",           "C/I_RATIO_(YTD_2022)",     "C/I ratio 2022"),
        ("C/I_RATIO_(YTD_2023)",    "old_ir",           "C/I_RATIO_(YTD_2023)",     "C/I ratio 2023"),
        ("REVENUES_2021",           "old_ir",           "REVENUES_2021",            "celkové výnosy 2021"),
        ("REVENUES_2022",           "old_ir",           "REVENUES_2022",            "celkové výnosy 2022"),
        ("REVENUES_2023",           "old_ir",           "REVENUES_2023",            "celkové výnosy 2023"),
        ("REVENUES_2024",           "old_ir",           "REVENUES_2024",            "celkové výnosy 2024"),
        ("FORMAT_2024_(FIX_FS)",    "old_ir",           "FORMAT_2024_(FIX_FS)",     "realizovaný formát"),
        ("FORMAT_2023",             "old_ir",           "FORMAT_2023",              ""),
        ("HODNOTA_INDEXU_EXPOZICE", "old_ir",           "HODNOTA_INDEXU_EXPOZICE",  "index demografické expozice"),
        ("POCET_KLIENTU_EOY_2021",  "old_ir",           "POCET_KLIENTU_EOY_2021",   "počet klientů EOY 2021"),
        ("POCET_KLIENTU_EOY_2022",  "old_ir",           "POCET_KLIENTU_EOY_2022",   "počet klientů EOY 2022"),
        ("POCET_KLIENTU_EOY_2023",  "old_ir",           "POCET_KLIENTU_EOY_2023",   "počet klientů EOY 2023"),
        ("POCET_KLIENTU_EOY_2024",  "old_ir",           "POCET_KLIENTU_EOY_2024",   "počet klientů EOY 2024"),
        # profitabilita — klienti + náklady
        ("PRIMARNI_KLIENTI",        "profitabilita",    "PRIMARNI_KLIENTI",         ""),
        ("AKTIVNI_KLIENTI",         "profitabilita",    "AKTIVNI_KLIENTI",          ""),
        ("POCET_KLIENTU",           "profitabilita",    "POCET_KLIENTU",            "aktuální počet klientů 2025"),
        ("FTE",                     "profitabilita",    "FTE",                      "celkový přepočtený úvazek"),
        ("OSOBNI_NAKLADY",          "profitabilita",    "OSOBNI_NAKLADY",           ""),
        ("PRIME_NAKLADY",           "profitabilita",    "PRIME_NAKLADY",            ""),
        ("PRIME_NAKLADY/VYNOSY",    "profitabilita",    "PRIME_NAKLADY/VYNOSY",     "C/I ratio 2025"),
        ("VYNOSY",                  "profitabilita",    "VYNOSY",                   "celkové výnosy 2025"),
        # specialiste — FTE pozice
        ("BANKERS_COUNT",           "specialiste",      "součet bankéřských pozic", "osobní bankéř junior/medior/senior/master"),
        ("OBCHODNI_FTE",            "specialiste",      "součet obchodních pozic",  "27 definovaných obchodních rolí"),
        ("_total_spec",             "specialiste",      "součet všech pozic",       "celkový počet obsazených míst"),
        # ORP (GeoJSON + pyproj)
        ("ORP_NAZEV",               "GeoJSON ORP + GPS", "point-in-polygon WGS84→S-JTSK", "název ORP dle GPS souřadnic"),
        ("ORP_KOD",                 "GeoJSON ORP + GPS", "point-in-polygon WGS84→S-JTSK", "kód ORP"),
    ]
    COL_SOURCES.extend(extra)

    # Investice z invest pkl
    for col in ["DESC_LAST_INV","DATUM_LAST_INV","LAST_INV",
                "DESC_BIGGEST_INV","DATUM_BIGGEST_INV","BIGGEST_INV","TOTAL_INV"]:
        COL_SOURCES.append((col, "invest", col, ""))

    # Produktové obchody
    for key, label in PRODUCT_GROUPS:
        for suffix in ["POCET_PRODEJU","OBJEM_VYNOSU_CZK"]:
            for yr in ["2024","2025"]:
                col = f"{suffix}_{key}_{yr}"
                COL_SOURCES.append((col, "obchody_detail", col, ""))

    # Hotovostní transakce
    for col in ["CASH_IN_PREVODITELNA_CASTKA","CASH_IN_NEPREVODITELNA_CASTKA",
                "CASH_IN_PREVODITELNA_POCET","CASH_IN_NEPREVODITELNA_POCET",
                "CASH_OUT_PREVODITELNA_CASTKA","CASH_OUT_NEPREVODITELNA_CASTKA",
                "CASH_OUT_PREVODITELNA_POCET","CASH_OUT_NEPREVODITELNA_POCET"]:
        COL_SOURCES.append((col, "hotovostni_trn", col, ""))

    # Vypočítané sloupce
    computed = [
        # Interní rating
        ("IR_SCORE",              COMPUTED_NOTE, "CI_RATIO_PERC×0.4 + NEW_BUSINESS_VOL_RNK×0.6",        "interní rating"),
        ("IR",                    COMPUTED_NOTE, "rank(IR_SCORE) — 1=nejlepší",                           "interní rating"),
        ("IR_Q",                  COMPUTED_NOTE, "qcut(IR, q=5) — 1=nejlepší",                            "interní rating"),
        ("IR_PERC",               COMPUTED_NOTE, "percentil IR v síti",                                   "interní rating"),
        ("IR_FLAG",               COMPUTED_NOTE, "'Y' pokud má pobočka vypočítaný IR, jinak 'N'",         "interní rating"),
        ("IR_DIFF_NUM",           COMPUTED_NOTE, "IR − INTERNI_RATING_2024",                              "interní rating"),
        ("NEW_BUSINESS_VOL_Q",    COMPUTED_NOTE, "qcut(OBJEM_VYNOSU_CZK, q=5)",                           "interní rating"),
        ("CI_RATIO_Q",            COMPUTED_NOTE, "qcut(PRIME_NAKLADY/VYNOSY, q=5)",                       "interní rating"),
        # Formát pobočky
        ("BRANCH_FORMAT",         COMPUTED_NOTE, "FTE: ≥25→flagship, ≥10→medium, ≥5→medium economy, jinak small", "formát"),
        ("BRANCH_FORMAT_OBCHODNI",COMPUTED_NOTE, "OBCHODNI_FTE: stejná pravidla jako BRANCH_FORMAT",      "formát"),
        # Klienti
        ("PRIM_RATIO",            COMPUTED_NOTE, "PRIMARNI_KLIENTI / POCET_KLIENTU",                      "klienti"),
        ("AKTIVNI_RATIO",         COMPUTED_NOTE, "AKTIVNI_KLIENTI / POCET_KLIENTU",                       "klienti"),
        ("KLIENTI_RATIO_HTML",    COMPUTED_NOTE, "SVG progress bary Prim% + Akt%",                        "klienti"),
        ("KLIENTI_TREND_HTML",    COMPUTED_NOTE, "SVG sparkline z POCET_KLIENTU_EOY_2021–2024 + 2025",    "klienti"),
        # Index expozice
        ("IE_Q",                  COMPUTED_NOTE, "qcut(HODNOTA_INDEXU_EXPOZICE, q=5) — 1=nejlepší (největší)", "index expozice"),
        ("IE_Q_HTML",             COMPUTED_NOTE, "barevná tečka dle IE_Q",                                "index expozice"),
        # Výnosy
        ("REV_SPARKLINE",         COMPUTED_NOTE, "SVG sparkline z REVENUES_2021–2024 + VYNOSY",           "výnosy trend"),
        ("OBJEM_VYNOSU_CZK",      COMPUTED_NOTE, "nové výnosy 2025 dle LTV (revenues CSV)",               "výnosy"),
        # Business rating
        ("CAP_PEREX",             COMPUTED_NOTE, "(OSOBNI_NAKLADY/FTE) / (OBJEM_VYNOSU_CZK/FTE)",         "business rating"),
        ("CAP_REV_FTE",           COMPUTED_NOTE, "OBJEM_VYNOSU_CZK / FTE",                                "business rating"),
        ("CAP_NEWCLI_FTE",        COMPUTED_NOTE, "NR_NEW_ARRIVALS / FTE",                                 "business rating"),
        ("CAP_PRIMCLI_FTE",       COMPUTED_NOTE, "PRIMARNI_KLIENTI / FTE",                                "business rating"),
        ("CAP_SCHUZKY_FTE",       COMPUTED_NOTE, "(POCET_SCHUZEK_FYZICKY + POCET_SCHUZEK_ONLINE) / FTE",  "business rating"),
        ("CAP_RNK",               COMPUTED_NOTE, "rank(weighted score z 5 metrik)",                       "business rating"),
        ("CAP_Q",                 COMPUTED_NOTE, "qcut(CAP_RNK, q=5) — 1=nejlepší",                       "business rating"),
        # Cash rating
        ("TRN_SCORE",             COMPUTED_NOTE, "rank(POCET)×0.45 + rank(CASTKA)×0.55",                  "cash rating"),
        ("TRN_Q",                 COMPUTED_NOTE, "qcut(TRN_SCORE, q=5)",                                   "cash rating"),
        ("TRN_RNK",               COMPUTED_NOTE, "rank(TRN_SCORE)",                                        "cash rating"),
        ("TRN_STRATEGIE",         COMPUTED_NOTE, "top20→Var.20, top35→Var.35, top50→Var.50, jinak cashless","cash rating"),
        ("CASH_PREVODITELNOST_PCT",COMPUTED_NOTE, "převoditelné/(převod+nepřevod)×100",                    "cash rating"),
        # FTE
        ("BANKERS_NEEDED",        COMPUTED_NOTE, "ceil((fyz+onl+bezhot×0.30×0.44) / 1240) — stejná metodika jako simulace", "FTE"),
        ("BANKERS_DIFF_HTML",     COMPUTED_NOTE, "BANKERS_NEEDED − BANKERS_COUNT → HTML badge",            "FTE"),
        # Návštěvy HTML
        ("NAVSTEVY_RATIO_HTML",   COMPUTED_NOTE, "4 progress bary (Online/Fyz/Hot/Bezhot) jako % z celkových", "návštěvy"),
        # ATM
        ("ATM_SUMMARY_HTML",      COMPUTED_NOTE, "seznam ATM + kapacita z export_atm + kapacita_atm",     "ATM"),
        ("ATM_VOLNA_KAPACITA",    COMPUTED_NOTE, "ZBYVAJICI_VOLNA_KAPACITA × 100 (%)",                     "ATM"),
        ("ATM_VYBERY_MESIC",      COMPUTED_NOTE, "PRUMER._POCET_VYBERU_ATM_/_MESIC",                       "ATM"),
        ("ATM_VKLADY_MESIC",      COMPUTED_NOTE, "PRUMER._POCET_VKLADU_ATM_/_MESIC",                       "ATM"),
        ("ATM_VYBERY_PO_PREV",    COMPUTED_NOTE, "POCET_ATM_VYBERU_PO_PREVODU_/_M",                        "ATM"),
        ("ATM_VKLADY_PO_PREV",    COMPUTED_NOTE, "POCET_ATM_VKLADU_PO_PREVODU_/_M",                        "ATM"),
        # Prodeje — kvintily
        ("POCET_PRODEJU_CELKEM_2025", COMPUTED_NOTE, "součet POCET_PRODEJU_{key}_2025 přes všechny produkty", "prodeje"),
        (f"POCET_PRODEJU_{{key}}_2025_Q", COMPUTED_NOTE, "qcut(POCET_PRODEJU_{key}_2025, q=5) pro každou produktovou skupinu", "prodeje"),
        # Metriky overview kvintily
        ("POCET_NAVSTEV_CELKEM_Q",  COMPUTED_NOTE, "qcut(POCET_NAVSTEV_CELKEM, q=5)",   "metriky overview"),
        ("PRIMARNI_KLIENTI_Q",      COMPUTED_NOTE, "qcut(PRIMARNI_KLIENTI, q=5)",        "metriky overview"),
        ("POCET_KLIENTU_Q",         COMPUTED_NOTE, "qcut(POCET_KLIENTU, q=5)",           "metriky overview"),
        ("VYNOSY_Q",                COMPUTED_NOTE, "qcut(VYNOSY, q=5)",                  "metriky overview"),
        ("POCET_SCHUZEK_ONLINE_Q",  COMPUTED_NOTE, "qcut(POCET_SCHUZEK_ONLINE, q=5)",    "metriky overview"),
        ("POCET_SCHUZEK_FYZICKY_Q", COMPUTED_NOTE, "qcut(POCET_SCHUZEK_FYZICKY, q=5)",   "metriky overview"),
        ("TRN_CASTKA_CELKEM_Q",     COMPUTED_NOTE, "qcut(TRN_CASTKA_CELKEM, q=5)",       "metriky overview"),
        ("TRN_POCET_CELKEM_Q",      COMPUTED_NOTE, "qcut(TRN_POCET_CELKEM, q=5)",        "metriky overview"),
        ("PLOCHA_POUZE_D5_Q",       COMPUTED_NOTE, "qcut(PLOCHA_POUZE_D5, q=5) — nižší=lepší", "metriky overview"),
        # Spádové oblasti
        ("SP_NAZEV_1",              COMPUTED_NOTE, "název 1. spádové pobočky dle analýzy návštěvnosti",   "spádové oblasti"),
        ("SP_VISITS_1",             COMPUTED_NOTE, "počet návštěv ze spádové oblasti 1",                   "spádové oblasti"),
        ("SP_PCT_1",                COMPUTED_NOTE, "podíl návštěv ze spádové oblasti 1 (%)",               "spádové oblasti"),
        ("SP_NAZEV_2",              COMPUTED_NOTE, "název 2. spádové pobočky",                             "spádové oblasti"),
        ("SP_NAZEV_3",              COMPUTED_NOTE, "název 3. spádové pobočky",                             "spádové oblasti"),
    ]
    for item in computed:
        COL_SOURCES.append(item)

    # ── Které tabulky/sekce zobrazují který sloupec ───────────────────────────
    # display_name → list sekcí
    COL_GROUPS_FLAT = {}  # display_name → skupina
    for grp_label, cols, _ in COL_GROUPS:
        for c in cols:
            COL_GROUPS_FLAT[c] = grp_label

    def _usage(tech_col):
        display = NOVE_NAZVY.get(tech_col, "")
        usages = []
        if display and display in COL_GROUPS_FLAT:
            usages.append(f"📋 Celkový přehled → {COL_GROUPS_FLAT[display]}")
        # Specifické tabulky
        specific = {
            "IR": ["⭐ Výpočet interního ratingu", "📊 Metriky overview"],
            "IR_Q": ["📊 Metriky overview"],
            "OBJEM_VYNOSU_CZK": ["💰 Top 10 výnosy", "📊 Metriky overview"],
            "BRANCH_FORMAT": ["🏢 Přehled dle regionu — Formát"],
            "BRANCH_TYPE_NAME": ["🏢 Přehled dle regionu — Typologie"],
            "CASHLESS": ["🏢 Přehled dle regionu — Cash"],
            "POCET_NAVSTEV_CELKEM": ["📊 Metriky overview", "📊 Síťový benchmark"],
        }
        for k, v in specific.items():
            if tech_col == k:
                usages.extend(v)
        # Produktové skupiny
        for key, label in PRODUCT_GROUPS:
            if tech_col.startswith(f"POCET_PRODEJU_{key}") or tech_col.startswith(f"OBJEM_VYNOSU_CZK_{key}"):
                usages.append(f"🛒 Přehled obchodů — {label}")
                if "2025" in tech_col:
                    usages.append("📊 Metriky overview")
        return ", ".join(dict.fromkeys(usages)) if usages else "—"

    # ── Sestavení HTML ────────────────────────────────────────────────────────
    # Seskupit dle zdroje
    from collections import OrderedDict
    by_source = OrderedDict()
    for item in COL_SOURCES:
        tech, src, src_col, note = item[0], item[1], item[2], item[3] if len(item) > 3 else ""
        by_source.setdefault(src, []).append((tech, src_col, note))

    src_colors = {
        "profitabilita":      "#e3f2fd",
        "revenues (LTV CSV)": "#e8f5e9",
        "old_ir":             "#fff8e1",
        "dbs":                "#f3e5f5",
        "footprint_bns":      "#fce4ec",
        "rent_and_ownership": "#ede7f6",
        "invest":             "#fbe9e7",
        "kpis":               "#e0f7fa",
        "parties":            "#f1f8e9",
        "obchody_detail":     "#fff3e0",
        "hotovostni_trn":     "#fafafa",
        "ar":                 "#e8eaf6",
        "specialiste":        "#e8f5e9",
        "GeoJSON ORP + GPS":  "#e8f4e8",
        COMPUTED_NOTE:        "#f5f5f5",
    }

    rows_html = ""
    for src, cols in by_source.items():
        bg = src_colors.get(src, "#fafafa")
        src_label = src.replace("📐 ", "")
        is_computed = src == COMPUTED_NOTE
        icon = "📐" if is_computed else "📂"
        for j, (tech, src_col, note) in enumerate(cols):
            display = NOVE_NAZVY.get(tech, "<span style='color:#bbb;font-style:italic;'>jen interně</span>")
            usage   = _usage(tech)
            first   = j == 0
            src_cell = (f"<td rowspan='{len(cols)}' style='padding:5px 10px;border:1px solid #e0e0e0;"
                        f"background:{bg};font-weight:600;font-size:0.80rem;vertical-align:top;"
                        f"white-space:nowrap;'>{icon} {src_label}</td>") if first else ""
            diff_col = "" if src_col == tech else f" <span style='color:#999;font-size:0.75rem;'>← {src_col}</span>"
            formula  = f" <span style='color:#888;font-size:0.75rem;font-style:italic;'>{note}</span>" if note else ""
            rows_html += f"""<tr>
              {src_cell}
              <td style='padding:4px 10px;border:1px solid #e0e0e0;font-family:monospace;
                          font-size:0.79rem;color:#333;white-space:nowrap;'>{tech}{diff_col}</td>
              <td style='padding:4px 10px;border:1px solid #e0e0e0;font-size:0.81rem;
                          font-weight:500;'>{display}{formula}</td>
              <td style='padding:4px 10px;border:1px solid #e0e0e0;font-size:0.78rem;
                          color:#555;'>{usage}</td>
            </tr>"""

    tbl = f"""
<div style="overflow-x:auto;max-height:700px;overflow-y:auto;margin-top:10px;">
  <table style="width:100%;border-collapse:collapse;font-size:0.82rem;">
    <thead style="position:sticky;top:0;z-index:5;">
      <tr style="background:#2770f0;color:#fff;">
        <th style="padding:7px 12px;border:1px solid #1e5bc9;text-align:left;min-width:160px;">Zdroj</th>
        <th style="padding:7px 12px;border:1px solid #1e5bc9;text-align:left;min-width:200px;">Tech. název → zdrojový sloupec</th>
        <th style="padding:7px 12px;border:1px solid #1e5bc9;text-align:left;min-width:160px;">Zobrazený název / vzorec</th>
        <th style="padding:7px 12px;border:1px solid #1e5bc9;text-align:left;">Použití v reportu</th>
      </tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
</div>"""

    return make_collapsible(
        "col-map",
        "🗺️ Mapa sloupců — zdroje, přejmenování a použití",
        tbl,
        default_open=False
    )


def generate_rating_methodology_html():
    """Popis metodiky výpočtu všech tří ratingů s vzorci a příklady."""
    return """
<div class="age-section-card mb-4" style="font-size:0.85rem;line-height:1.7;">

  <!-- ══ INTERNÍ RATING ══════════════════════════════════════════════ -->
  <div style="margin-bottom:28px;">
    <h5 style="color:#2770f0;border-bottom:2px solid #2770f0;padding-bottom:6px;margin-bottom:12px;">
      ⭐ Interní rating (IR)
    </h5>
    <p>
      Pořadí pobočky dle kombinace <b>efektivity nákladů (C/I ratio)</b> a
      <b>objemu nových výnosů</b>. Pořadí 1 = nejlepší pobočka v síti.
    </p>
    <div style="background:#f0f4ff;border-radius:8px;padding:14px 18px;margin:10px 0;font-family:monospace;font-size:0.88rem;">
      <b>Skóre</b> = (percentil C/I ratio × 0,40) + (pořadí nových výnosů × 0,60)<br>
      <b>IR</b> = pořadí pobočky dle Skóre (rank ascending — nižší skóre = lepší)
    </div>
    <table style="width:100%;border-collapse:collapse;font-size:0.83rem;margin-top:10px;">
      <thead>
        <tr style="background:#e8f0fe;">
          <th style="padding:6px 10px;border:1px solid #c5cae9;text-align:left;">Složka</th>
          <th style="padding:6px 10px;border:1px solid #c5cae9;text-align:center;">Váha</th>
          <th style="padding:6px 10px;border:1px solid #c5cae9;text-align:left;">Popis</th>
        </tr>
      </thead>
      <tbody>
        <tr><td style="padding:5px 10px;border:1px solid #ddd;">C/I ratio percentil</td>
            <td style="padding:5px 10px;border:1px solid #ddd;text-align:center;"><b>40 %</b></td>
            <td style="padding:5px 10px;border:1px solid #ddd;">Náklady / Výnosy — percentil v síti</td></tr>
        <tr style="background:#fafafa;"><td style="padding:5px 10px;border:1px solid #ddd;">Nové výnosy pořadí</td>
            <td style="padding:5px 10px;border:1px solid #ddd;text-align:center;"><b>60 %</b></td>
            <td style="padding:5px 10px;border:1px solid #ddd;">Výnosy z nových obchodů 2025 (LTV)</td></tr>
      </tbody>
    </table>
    <div style="background:#fff8e1;border-left:4px solid #f0b55d;padding:10px 14px;margin-top:12px;border-radius:0 6px 6px 0;font-size:0.83rem;">
      <b>📌 Příklad:</b> Pobočka A má C/I ratio percentil 15 (=top 15 %, nízké náklady) a pořadí nových výnosů 8 (=8. nejvyšší v síti z 120).<br>
      Skóre = 15 × 0,40 + 8 × 0,60 = <b>6,0 + 4,8 = 10,8</b><br>
      Pokud je 10,8 nejnižší skóre v síti, IR pobočky A = <b>1</b> (nejlepší).
    </div>
  </div>

  <!-- ══ BUSINESS RATING ═════════════════════════════════════════════ -->
  <div style="margin-bottom:28px;">
    <h5 style="color:#7b3f9e;border-bottom:2px solid #7b3f9e;padding-bottom:6px;margin-bottom:12px;">
      📈 Business rating (kapacitní rating)
    </h5>
    <p>
      Měří, jak efektivně pobočka využívá svůj personál (FTE). Hodnotí pět metrik,
      všechny přepočtené na jednoho pracovníka FTE.
    </p>
    <div style="background:#f5f0ff;border-radius:8px;padding:14px 18px;margin:10px 0;font-family:monospace;font-size:0.88rem;">
      <b>Skóre</b> = rank(Výnos/FTE)×0,15 + rank(Noví klienti/FTE)×0,20
               + rank(Primární klienti/FTE)×0,15 + rank(Schůzky/FTE)×0,10
               + rank(PEREX)×0,40<br>
      <b>Business pořadí</b> = rank ascending (nižší skóre = lepší)
    </div>
    <table style="width:100%;border-collapse:collapse;font-size:0.83rem;margin-top:10px;">
      <thead>
        <tr style="background:#f3e5f5;">
          <th style="padding:6px 10px;border:1px solid #ce93d8;text-align:left;">Složka</th>
          <th style="padding:6px 10px;border:1px solid #ce93d8;text-align:center;">Váha</th>
          <th style="padding:6px 10px;border:1px solid #ce93d8;text-align:left;">Výpočet</th>
        </tr>
      </thead>
      <tbody>
        <tr><td style="padding:5px 10px;border:1px solid #e1bee7;">PEREX</td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;text-align:center;"><b>40 %</b></td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;">Osobní náklady/FTE ÷ Výnosy/FTE</td></tr>
        <tr style="background:#fafafa;"><td style="padding:5px 10px;border:1px solid #e1bee7;">Noví klienti / FTE</td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;text-align:center;"><b>20 %</b></td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;">NR_NEW_ARRIVALS ÷ FTE</td></tr>
        <tr><td style="padding:5px 10px;border:1px solid #e1bee7;">Výnos / FTE</td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;text-align:center;"><b>15 %</b></td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;">OBJEM_VYNOSU_CZK ÷ FTE</td></tr>
        <tr style="background:#fafafa;"><td style="padding:5px 10px;border:1px solid #e1bee7;">Primární klienti / FTE</td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;text-align:center;"><b>15 %</b></td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;">PRIMARNI_KLIENTI ÷ FTE</td></tr>
        <tr><td style="padding:5px 10px;border:1px solid #e1bee7;">Schůzky / FTE</td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;text-align:center;"><b>10 %</b></td>
            <td style="padding:5px 10px;border:1px solid #e1bee7;">(Fyzické + online schůzky) ÷ FTE</td></tr>
      </tbody>
    </table>
    <div style="background:#fff8e1;border-left:4px solid #f0b55d;padding:10px 14px;margin-top:12px;border-radius:0 6px 6px 0;font-size:0.83rem;">
      <b>📌 Příklad:</b> Pobočka B (3 FTE): Výnosy 6 mil. Kč, 12 nových klientů, 90 primárních klientů, 150 schůzek, osobní náklady 2,4 mil. Kč.<br>
      Výnos/FTE = 2 000 000 → rank 5 | Noví klienti/FTE = 4,0 → rank 12 | PEREX = 2 400 000/3 ÷ 2 000 000 = 0,40 → rank 3<br>
      Skóre = 5×0,15 + 12×0,20 + …+ 3×0,40 = <b>0,75 + 2,40 + … + 1,20 = ≈ 5,7</b> — čím nižší, tím lepší pořadí.
    </div>
  </div>

  <!-- ══ HOTOVOSTNÍ RATING ════════════════════════════════════════════ -->
  <div style="margin-bottom:12px;">
    <h5 style="color:#1565c0;border-bottom:2px solid #1565c0;padding-bottom:6px;margin-bottom:12px;">
      🏦 Hotovostní rating (Cash rating)
    </h5>
    <p>
      Hodnotí objem a počet hotovostních transakcí (Cash IN + Cash OUT).
      Rating se počítá zvlášť pro <b>celou populaci</b> i zvlášť pro
      <b>cashless pobočky</b> navzájem.
    </p>
    <div style="background:#e3f2fd;border-radius:8px;padding:14px 18px;margin:10px 0;font-family:monospace;font-size:0.88rem;">
      <b>Skóre</b> = rank(Počet transakcí) × 0,45 + rank(Objem transakcí) × 0,55<br>
      <b>Cash pořadí</b> = rank ascending skóre (nižší skóre = více transakcí = lepší pořadí)
    </div>
    <table style="width:100%;border-collapse:collapse;font-size:0.83rem;margin-top:10px;">
      <thead>
        <tr style="background:#e3f2fd;">
          <th style="padding:6px 10px;border:1px solid #90caf9;text-align:left;">Složka</th>
          <th style="padding:6px 10px;border:1px solid #90caf9;text-align:center;">Váha</th>
          <th style="padding:6px 10px;border:1px solid #90caf9;text-align:left;">Výpočet</th>
        </tr>
      </thead>
      <tbody>
        <tr><td style="padding:5px 10px;border:1px solid #bbdefb;">Objem transakcí</td>
            <td style="padding:5px 10px;border:1px solid #bbdefb;text-align:center;"><b>55 %</b></td>
            <td style="padding:5px 10px;border:1px solid #bbdefb;">Celkový objem Cash IN + Cash OUT (Kč)</td></tr>
        <tr style="background:#fafafa;"><td style="padding:5px 10px;border:1px solid #bbdefb;">Počet transakcí</td>
            <td style="padding:5px 10px;border:1px solid #bbdefb;text-align:center;"><b>45 %</b></td>
            <td style="padding:5px 10px;border:1px solid #bbdefb;">Celkový počet hotovostních transakcí</td></tr>
      </tbody>
    </table>
    <p style="margin-top:10px;font-size:0.82rem;color:#555;">
      <b>Strategie:</b> Top 20 poboček s nejvyšším cash ratingem → <i>Var. 20 poboček s hotovostí</i>,
      top 21–35 → <i>Var. 35</i>, top 36–50 → <i>Var. 50</i>, ostatní → přechod na cashless.
    </p>
    <div style="background:#fff8e1;border-left:4px solid #f0b55d;padding:10px 14px;margin-top:12px;border-radius:0 6px 6px 0;font-size:0.83rem;">
      <b>📌 Příklad:</b> Pobočka C zpracuje 4 200 transakcí (rank 8 ze 120) v objemu 180 mil. Kč (rank 11).<br>
      Skóre = 8×0,45 + 11×0,55 = <b>3,60 + 6,05 = 9,65</b><br>
      Finální rank = pořadí 9,65 mezi všemi pobočkami → Cash pořadí = <b>4</b> → Strategie: <i>Var. 20 poboček s hotovostí</i>.
    </div>

    <div style="margin-top:16px;padding:10px 14px;background:#f5f5f5;border-radius:6px;font-size:0.82rem;color:#666;">
      <b>ℹ️ Poznámky ke kvintilům:</b> Kvintil Q1 = nejlepší 20 % poboček, Q5 = nejhorších 20 %.
      Kvintily se počítají vždy zvlášť pro každý rating z finálního pořadí (CAP_RNK / TRN_RNK / IR).
      Cashless pobočky mají vlastní kvintily v rámci cashless subpopulace.
    </div>
  </div>

</div>
"""

def make_collapsible(section_id, header_html, content_html, default_open=False):
    """Obalí obsah do rozbalovací sekce — používá <details>/<summary> pro plnou kompatibilitu
    včetně Safari/Mac. JS pouze přidává vizuální šipku a styl, nativní toggle vždy funguje."""
    open_attr = "open" if default_open else ""
    hint = "klikni pro sbalení" if default_open else "klikni pro rozbalení"
    return f"""
<details class="collapsible-section" id="wrap-{section_id}" {open_attr}
  style="margin-bottom:6px;border:1px solid #dde4f5;border-radius:8px;background:white;">
  <summary id="hdr-{section_id}"
    style="display:flex;align-items:center;gap:10px;cursor:pointer;list-style:none;
           padding:10px 14px;border-radius:8px;
           background:linear-gradient(90deg,#f0f4ff,#f8faff);
           border-bottom:1px solid #dde4f5;user-select:none;outline:none;"
    onclick="(function(el){{
      var d=el.closest('details');
      var a=el.querySelector('.coll-arrow');
      setTimeout(function(){{ if(a) a.textContent=d.open?'▼':'▶'; }},0);
    }})(this)">
    <span class="coll-arrow" style="font-size:0.85rem;color:#2770f0;font-weight:700;min-width:14px;">{'▼' if default_open else '▶'}</span>
    <span style="flex:1;font-size:0.95rem;font-weight:700;color:#2d3748;">{header_html}</span>
    <span style="font-size:0.78rem;color:#888;">{hint}</span>
  </summary>
  <div id="cnt-{section_id}" style="padding:16px;border-radius:0 0 8px 8px;">
    {content_html}
  </div>
</details>"""


def generate_filterable_table(target_df, table_id, excluded_cols=None):
    """
    Vrátí HTML tabulku s filtrem, porovnávačem řádků a skupinovými přepínači sloupců.
    Pobočky bez IR (IR_FLAG != 'Y' a nejsou SPECIAL) jsou defaultně skryty.
    """
    _excl = set(excluded_cols or [])
    # Vyloučené sloupce jako display names pro COL_GROUPS filtr
    _excl_display = {NOVE_NAZVY.get(c, c) for c in _excl}
    html_table = prepare_table(target_df, excluded_cols=excluded_cols)
    html_table = html_table.replace('<table ', f'<table id="{table_id}" ', 1)

    import json
    col_groups_js = json.dumps([
        (label, [c for c in cols if c not in _excl_display], color, COL_GROUP_ACCENT[i % len(COL_GROUP_ACCENT)])
        for i, (label, cols, color) in enumerate(COL_GROUPS)
    ])

    # Kódy poboček bez ratingu — detekujeme podle prázdného sloupce IR (Rating 25)
    if 'BRANCH_CODE' in target_df.columns and 'IR' in target_df.columns:
        no_rating_mask = target_df['IR'].isna() | (target_df['IR'].astype(str).str.strip() == '') | (target_df['IR'] == 0)
        no_rating_codes = json.dumps([int(c) for c in target_df.loc[no_rating_mask, 'BRANCH_CODE'].tolist()])
    else:
        no_rating_codes = '[]'
    no_rating_count = len(json.loads(no_rating_codes))

    return f"""
<style>
  #wrapper-{table_id} .ft-bar {{
    display:flex; flex-wrap:wrap; gap:8px; align-items:center; margin-bottom:8px;
  }}
  #wrapper-{table_id} .ft-input {{
    padding:7px 12px; border:1px solid #ccc; border-radius:6px;
    font-size:0.88rem; width:280px; outline:none;
    box-shadow:inset 0 1px 3px rgba(0,0,0,.08);
  }}
  #wrapper-{table_id} .ft-input:focus {{ border-color:#2770f0; }}
  #wrapper-{table_id} .ft-btn {{
    padding:6px 14px; border:1px solid #ccc; border-radius:6px;
    background:#f5f5f5; font-size:0.85rem; cursor:pointer;
  }}
  #wrapper-{table_id} .ft-btn-blue {{ border:none; background:#2770f0; color:white; }}
  #wrapper-{table_id} .ft-btn-red  {{ border:none; background:#e64343; color:white; }}
  #wrapper-{table_id} .ft-count {{ font-size:0.82rem; color:#888; }}
  #wrapper-{table_id} .ft-label {{ font-size:0.85rem; color:#555; font-weight:600; }}
  #wrapper-{table_id} .col-toggle-bar {{
    display:flex; flex-wrap:wrap; gap:6px; margin-bottom:10px;
    padding:10px 12px; background:#f8faff; border:1px solid #dde4f5;
    border-radius:8px;
  }}
  #wrapper-{table_id} .col-toggle-btn {{
    display:flex; align-items:center; gap:5px;
    padding:4px 10px; border-radius:20px; border:2px solid transparent;
    font-size:0.82rem; cursor:pointer; user-select:none;
    transition: opacity 0.15s;
    color: #333;
  }}
  #wrapper-{table_id} .col-toggle-btn.inactive {{
    opacity: 0.35;
  }}
  #wrapper-{table_id} .col-toggle-btn input {{
    width:13px; height:13px; cursor:pointer; margin:0;
  }}
  #wrapper-{table_id} .col-toggle-all {{
    padding:4px 12px; border-radius:20px; border:1px solid #aaa;
    background:#f0f0f0; font-size:0.82rem; cursor:pointer; font-weight:600; color:#333;
  }}
  /* ── Expanded (fullscreen) mode ── */
  #wrapper-{table_id}.tbl-expanded {{
    position: fixed;
    top: 0; left: 0; right: 0; bottom: 0;
    z-index: 9999;
    background: #f7f9fc;
    padding: 16px 20px;
    overflow-y: auto;
    box-shadow: 0 0 0 9999px rgba(0,0,0,0.45);
  }}
  #wrapper-{table_id}.tbl-expanded .tbl-scroll-wrap {{
    max-height: calc(100vh - 220px);
  }}
</style>

<div id="wrapper-{table_id}">

  <!-- Skupinové přepínače sloupců -->
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:5px;gap:8px;flex-wrap:wrap;">
    <div style="font-size:0.82rem;color:#555;font-weight:600;">📂 Zobrazit skupiny sloupců:</div>
    <button id="ft-expand-{table_id}" title="Rozšířit tabulku na celou šířku obrazovky"
            style="padding:5px 12px;border:1px solid #2770f0;border-radius:6px;
                   background:#eef4ff;color:#2770f0;font-size:0.82rem;font-weight:600;
                   cursor:pointer;display:flex;align-items:center;gap:5px;white-space:nowrap;">
      ⤢ Rozšířit
    </button>
  </div>
  <div class="col-toggle-bar" id="col-toggles-{table_id}">
    <button class="col-toggle-all" id="col-all-{table_id}">✔ Vše</button>
    <button class="col-toggle-all" id="col-none-{table_id}">✕ Nic</button>
  </div>

  <!-- Filtr a porovnání -->
  <div class="ft-bar">
    <input id="ft-input-{table_id}" class="ft-input"
           placeholder="🔍 Filtruj název nebo kód pobočky...">
    <button id="ft-clear-{table_id}" class="ft-btn">✖ Zrušit filtr</button>
    <button id="ft-no-rating-{table_id}" class="ft-btn"
            style="background:#fff3e0;border-color:#fb8c00;color:#e65100;font-size:0.82rem;"
            title="Pobočky bez vypočítaného interního ratingu">
      👁 Zobrazit bez ratingu
    </button>
    <span   id="ft-count-{table_id}" class="ft-count"></span>
  </div>

  <!-- Multi-filtr (Excel-like) -->
  <div style="margin-bottom:8px;">
    <button id="ft-multifilter-toggle-{table_id}" class="ft-btn"
            style="background:#f0f4ff;border-color:#2770f0;color:#2770f0;font-weight:600;">
      🔽 Rozšířené filtry sloupců
    </button>
    <button id="ft-multifilter-clear-{table_id}" class="ft-btn ft-btn-red"
            style="display:none;font-size:0.82rem;">✖ Zrušit všechny filtry</button>
    <span id="ft-multifilter-active-{table_id}"
          style="font-size:0.78rem;color:#2770f0;font-weight:600;margin-left:6px;display:none;"></span>
  </div>
  <div id="ft-multifilter-panel-{table_id}"
       style="display:none;background:#f8faff;border:1px solid #dde4f5;border-radius:8px;
              padding:12px 14px;margin-bottom:10px;">
    <div style="font-size:0.78rem;color:#666;margin-bottom:8px;font-weight:600;">
      Přidej filtr pro libovolný sloupec — hodnoty se automaticky doplní:
    </div>
    <div id="ft-mf-rows-{table_id}"></div>
    <button id="ft-mf-add-{table_id}" class="ft-btn ft-btn-blue"
            style="font-size:0.82rem;margin-top:6px;">＋ Přidat filtr</button>
  </div>

  <div class="ft-bar">
    <span class="ft-label">📌 Porovnání:</span>
    <span id="ft-hint-{table_id}" style="font-size:0.82rem;color:#888;">Zaškrtni řádky v tabulce</span>
    <button id="ft-show-{table_id}"  class="ft-btn ft-btn-blue" style="display:none;">👁 Zobrazit jen vybrané</button>
    <button id="ft-desel-{table_id}" class="ft-btn ft-btn-red"  style="display:none;">✖ Zrušit výběr</button>
  </div>

  <div class="table-responsive tbl-scroll-wrap" style="max-height:900px; min-height:120px; overflow-y:auto; overflow-x:auto; position:relative;">
    {html_table}
  </div>

</div>

<script>
setTimeout(function() {{
  var wrap    = document.getElementById("wrapper-{table_id}");
  if (!wrap) return;
  var tbl     = document.getElementById("{table_id}");
  var input   = document.getElementById("ft-input-{table_id}");
  var btnClr  = document.getElementById("ft-clear-{table_id}");
  var countEl = document.getElementById("ft-count-{table_id}");
  var hintEl  = document.getElementById("ft-hint-{table_id}");
  var btnShow = document.getElementById("ft-show-{table_id}");
  var btnDes  = document.getElementById("ft-desel-{table_id}");
  var toggleBar = document.getElementById("col-toggles-{table_id}");
  var btnAll      = document.getElementById("col-all-{table_id}");
  var btnNone     = document.getElementById("col-none-{table_id}");
  var btnNoRating = document.getElementById("ft-no-rating-{table_id}");
  var noRatingCodes = new Set({no_rating_codes});
  var showNoRating  = false;

  if (!tbl) return;

  var theadRow = tbl.querySelector("thead tr");
  var thCb = document.createElement("th");
  thCb.style.cssText = "width:34px; background:#2770f0; border:1px solid #1e5bc9; position:sticky; top:0; z-index:13;";
  theadRow.insertBefore(thCb, theadRow.firstChild);

  // Ukotvit záhlaví (pouze vertikálně — top:0)
  var allThs = Array.from(theadRow.querySelectorAll("th"));
  allThs.forEach(function(th) {{
    th.style.position   = "sticky";
    th.style.top        = "0";
    th.style.background = "#2770f0";
  }});

  var stickyOffsets = [0, 34, 110, 310];  // zachováno kvůli doFilter referenci
  var rows = Array.from(tbl.querySelectorAll("tbody tr"));
  var STICKY_BG = "";
  rows.forEach(function(row, ri) {{
    var td = document.createElement("td");
    td.style.cssText = "text-align:center; vertical-align:middle; background:transparent; width:34px; min-width:34px; border-right:1px solid #dee2e6;";
    var cb = document.createElement("input");
    cb.type = "checkbox";
    cb.style.cssText = "cursor:pointer; width:15px; height:15px; accent-color:#2770f0;";
    td.appendChild(cb);
    row.insertBefore(td, row.firstChild);
  }});

  var headerCells = Array.from(theadRow.querySelectorAll("th")).slice(1);
  var headerMap = {{}};
  headerCells.forEach(function(th, i) {{
    headerMap[th.textContent.trim().toLowerCase()] = i + 1;
  }});

  var colGroups = {col_groups_js};
  var alwaysVisible = ["id pobočky", "název pobočky", "region"];

  var groupColIdxs = colGroups.map(function(g) {{
    var label = g[0], names = g[1], pastel = g[2], accent = g[3];
    var idxs = [];
    names.forEach(function(n) {{
      var idx = headerMap[n.toLowerCase()];
      if (idx !== undefined) idxs.push(idx);
    }});
    return {{label: label, idxs: idxs, pastel: pastel, accent: accent}};
  }}).filter(function(g) {{ return g.idxs.length > 0; }});

  var groupActive = {{}};
  groupColIdxs.forEach(function(g) {{ groupActive[g.label] = true; }});

  // Označ každý TH atributem data-col-idx pro spolehlivé párování s TD
  headerCells.forEach(function(th, i) {{
    th.setAttribute("data-col-idx", i + 1);
  }});
  rows.forEach(function(row) {{
    var tds = Array.from(row.querySelectorAll("td")).slice(1); // skip checkbox td
    tds.forEach(function(td, i) {{
      td.setAttribute("data-col-idx", i + 1);
    }});
  }});

  // Barva skupiny — uložena jako data-group-color na TH i TD
  function applyGroupColors() {{
    var colColorMap = {{}};
    var colSortActiveMap = {{}};  // pro sort — neuplatňovat group color na sort sloupec
    groupColIdxs.forEach(function(g) {{
      g.idxs.forEach(function(idx) {{ colColorMap[idx] = g.pastel; }});
    }});
    // Ulož na TH
    headerCells.forEach(function(th) {{
      var idx = parseInt(th.getAttribute("data-col-idx"));
      if (colColorMap[idx]) th.setAttribute("data-group-color", colColorMap[idx]);
      else th.removeAttribute("data-group-color");
    }});
    // Aplikuj na TD
    rows.forEach(function(row) {{
      Array.from(row.querySelectorAll("td[data-col-idx]")).forEach(function(td) {{
        var idx = parseInt(td.getAttribute("data-col-idx"));
        if (colColorMap[idx]) {{
          td.setAttribute("data-group-color", colColorMap[idx]);
          td.style.setProperty("background-color", colColorMap[idx], "important");
        }} else {{
          td.removeAttribute("data-group-color");
        }}
      }});
    }});
  }}

  function applyColVisibility() {{
    var visible = new Set();
    headerCells.forEach(function(th) {{
      if (alwaysVisible.indexOf(th.textContent.trim().replace(/[↕↑↓]/g,"").trim().toLowerCase()) >= 0) {{
        visible.add(parseInt(th.getAttribute("data-col-idx")));
      }}
    }});
    groupColIdxs.forEach(function(g) {{
      if (groupActive[g.label]) g.idxs.forEach(function(idx) {{ visible.add(idx); }});
    }});
    headerCells.forEach(function(th) {{
      var idx = parseInt(th.getAttribute("data-col-idx"));
      th.style.display = visible.has(idx) ? "" : "none";
    }});
    thCb.style.display = "";
    rows.forEach(function(row) {{
      Array.from(row.querySelectorAll("td[data-col-idx]")).forEach(function(td) {{
        var idx = parseInt(td.getAttribute("data-col-idx"));
        td.style.display = visible.has(idx) ? "" : "none";
      }});
    }});
  }}

  groupColIdxs.forEach(function(g) {{
    var btn = document.createElement("label");
    btn.className = "col-toggle-btn";
    btn.title = "Zobrazit / skrýt: " + g.label;
    btn.style.backgroundColor = g.pastel;
    btn.style.borderColor = g.accent;
    btn.style.color = "#333";
    var cb = document.createElement("input");
    cb.type = "checkbox";
    cb.checked = true;
    cb.style.accentColor = g.accent;
    cb.addEventListener("change", function() {{
      groupActive[g.label] = cb.checked;
      btn.className = "col-toggle-btn" + (cb.checked ? "" : " inactive");
      applyColVisibility();
    }});
    btn.appendChild(cb);
    btn.appendChild(document.createTextNode(" " + g.label));
    toggleBar.appendChild(btn);
  }});

  btnAll.addEventListener("click", function() {{
    groupColIdxs.forEach(function(g) {{ groupActive[g.label] = true; }});
    toggleBar.querySelectorAll("input[type=checkbox]").forEach(function(cb) {{ cb.checked = true; }});
    toggleBar.querySelectorAll(".col-toggle-btn").forEach(function(b) {{
      b.className = "col-toggle-btn";
    }});
    applyColVisibility();
  }});

  btnNone.addEventListener("click", function() {{
    groupColIdxs.forEach(function(g) {{ groupActive[g.label] = false; }});
    toggleBar.querySelectorAll("input[type=checkbox]").forEach(function(cb) {{ cb.checked = false; }});
    toggleBar.querySelectorAll(".col-toggle-btn").forEach(function(b) {{
      b.className = "col-toggle-btn inactive";
    }});
    applyColVisibility();
  }});

  applyGroupColors();
  applyColVisibility();

  // Defaultně skryj pobočky bez ratingu
  doFilter();

  function updateCount() {{
    var vis = rows.filter(function(r) {{ return r.style.display !== "none"; }}).length;
    countEl.textContent = vis + "\u00a0/\u00a0" + rows.length + " poboček";
  }}

  function doFilter() {{
    var q = input.value.toLowerCase().trim();
    rows.forEach(function(r) {{
      var cells   = r.querySelectorAll("td");
      var codeRaw = cells[1] ? cells[1].textContent.trim() : "";
      var codeNum = parseInt(codeRaw, 10);
      var name    = cells[2] ? cells[2].textContent.toLowerCase() : "";
      var isNoRat = noRatingCodes.has(codeNum);
      if (isNoRat && !showNoRating) {{ r.style.display = "none"; return; }}
      var basicOk = (!q || codeRaw.toLowerCase().includes(q) || name.includes(q));
      if (!basicOk) {{ r.style.display = "none"; return; }}
      // Multi-filtry (AND mezi sloupci, OR uvnitř sloupce)
      var active  = (typeof mfFilters !== "undefined") ? mfFilters.filter(function(f) {{ return f.values && f.values.size > 0; }}) : [];
      var multiOk = active.every(function(f) {{
        var td = cells[f.idx];
        if (!td) return false;
        var cellVal = td.textContent.trim();
        return f.values.has(cellVal);
      }});
      r.style.display = multiOk ? "" : "none";
    }});
    updateCount();
  }}

  input.addEventListener("input", doFilter);
  btnClr.addEventListener("click", function() {{
    input.value = "";
    doFilter();
  }});

  // ─── Multi-filtr (Excel-like) ─────────────────────────────────────────────
  var mfPanel      = document.getElementById("ft-multifilter-panel-{table_id}");
  var mfToggle     = document.getElementById("ft-multifilter-toggle-{table_id}");
  var mfClearAll   = document.getElementById("ft-multifilter-clear-{table_id}");
  var mfActiveInfo = document.getElementById("ft-multifilter-active-{table_id}");
  var mfRowsDiv    = document.getElementById("ft-mf-rows-{table_id}");
  var mfAddBtn     = document.getElementById("ft-mf-add-{table_id}");

  // Sesbírej unikátní hodnoty pro každý sloupec (z textContent buněk)
  var colValues = {{}};  // idx -> Set of strings
  var colNames  = [];    // [{{idx, label}}] — jen viditelné sloupce (dle headerCells)
  headerCells.forEach(function(th, i) {{
    var label = th.textContent.trim().replace(/[↕↑↓]/g,"").trim();
    var idx   = i + 1;
    if (!label) return;
    colNames.push({{idx: idx, label: label}});
    colValues[idx] = new Set();
    rows.forEach(function(r) {{
      var td = r.querySelectorAll("td")[idx];
      if (td) {{
        var v = td.textContent.trim();
        if (v && v !== "—") colValues[idx].add(v);
      }}
    }});
  }});

  var mfFilters = [];   // pole {{idx, values: Set}} aktivních filtrů

  function applyMultiFilter() {{
    var active = mfFilters.filter(function(f) {{ return f.values && f.values.size > 0; }});
    rows.forEach(function(r) {{
      if (r.style.display === "none" && active.length === 0) return;
      var cells     = r.querySelectorAll("td");
      var codeRaw   = cells[1] ? cells[1].textContent.trim() : "";
      var codeNum   = parseInt(codeRaw, 10);
      var name      = cells[2] ? cells[2].textContent.toLowerCase() : "";
      var q         = input.value.toLowerCase().trim();
      var isNoRat   = noRatingCodes.has(codeNum);
      if (isNoRat && !showNoRating) {{ r.style.display = "none"; return; }}
      var basicOk   = (!q || codeRaw.toLowerCase().includes(q) || name.includes(q));
      if (!basicOk) {{ r.style.display = "none"; return; }}
      // Multi-filtr: AND mezi sloupci, OR uvnitř sloupce (hodnoty v Set)
      var multiOk = active.every(function(f) {{
        var td = r.querySelectorAll("td")[f.idx];
        if (!td) return false;
        return f.values.has(td.textContent.trim());
      }});
      r.style.display = multiOk ? "" : "none";
    }});
    updateCount();
    var n = active.length;
    if (n > 0) {{
      mfActiveInfo.textContent = n + (n === 1 ? " aktivní filtr" : n < 5 ? " aktivní filtry" : " aktivních filtrů");
      mfActiveInfo.style.display = "inline";
      mfClearAll.style.display   = "inline-block";
      mfToggle.style.background  = "#dbeafe";
    }} else {{
      mfActiveInfo.style.display = "none";
      mfClearAll.style.display   = "none";
      mfToggle.style.background  = "#f0f4ff";
    }}
  }}

  function buildMfRow(filterObj) {{
    var rowDiv = document.createElement("div");
    rowDiv.style.cssText = "margin-bottom:8px;";

    var controlRow = document.createElement("div");
    controlRow.style.cssText = "display:flex;align-items:center;gap:8px;flex-wrap:wrap;";

    // Select — výběr sloupce
    var sel = document.createElement("select");
    sel.style.cssText = "padding:5px 8px;border:1px solid #ccc;border-radius:6px;font-size:0.84rem;" +
                        "max-width:220px;background:#fff;cursor:pointer;";
    sel.innerHTML = '<option value="">— Vyberte sloupec —</option>';
    colNames.forEach(function(cn) {{
      var opt = document.createElement("option");
      opt.value = cn.idx;
      opt.textContent = cn.label;
      if (filterObj.idx === cn.idx) opt.selected = true;
      sel.appendChild(opt);
    }});

    // ── Checkbox-dropdown pro výběr více hodnot ──────────────────
    var dropWrap = document.createElement("div");
    dropWrap.style.cssText = "position:relative;";

    var dropBtn = document.createElement("button");
    dropBtn.style.cssText = "padding:5px 12px;border:1px solid #ccc;border-radius:6px;" +
                            "background:#f5f5f5;font-size:0.84rem;cursor:pointer;white-space:nowrap;" +
                            "min-width:180px;text-align:left;";
    dropBtn.textContent = "— Vyberte hodnoty —";

    var dropPanel = document.createElement("div");
    dropPanel.style.cssText = "display:none;position:absolute;top:100%;left:0;z-index:999;" +
                              "background:#fff;border:1px solid #ccd;border-radius:8px;" +
                              "padding:8px 10px;max-height:240px;overflow-y:auto;" +
                              "min-width:220px;box-shadow:0 4px 14px rgba(0,0,0,.12);";

    dropWrap.appendChild(dropBtn);
    dropWrap.appendChild(dropPanel);

    function updateDropBtn() {{
      var n = filterObj.values.size;
      if (n === 0) {{
        dropBtn.textContent  = "— Vyberte hodnoty (vše) —";
        dropBtn.style.background   = "#f5f5f5";
        dropBtn.style.borderColor  = "#ccc";
        dropBtn.style.color        = "#333";
      }} else {{
        dropBtn.textContent  = n + " " + (n === 1 ? "hodnota vybrána" : n < 5 ? "hodnoty vybrány" : "hodnot vybráno") + " ▾";
        dropBtn.style.background   = "#dbeafe";
        dropBtn.style.borderColor  = "#2770f0";
        dropBtn.style.color        = "#1d4ed8";
        dropBtn.style.fontWeight   = "600";
      }}
    }}

    function rebuildPanel() {{
      dropPanel.innerHTML = "";
      filterObj.values = new Set();
      var idx = parseInt(sel.value);
      if (isNaN(idx) || !colValues[idx] || colValues[idx].size === 0) {{
        dropPanel.innerHTML = "<div style='color:#aaa;font-size:0.8rem;padding:4px 0;'>Žádné hodnoty</div>";
        updateDropBtn();
        return;
      }}
      var sorted = Array.from(colValues[idx]).sort(function(a, b) {{
        var na = parseFloat(a), nb = parseFloat(b);
        if (!isNaN(na) && !isNaN(nb)) return na - nb;
        return a.localeCompare(b, "cs");
      }});

      // Řádek "Vybrat vše / Zrušit vše"
      var ctrlRow = document.createElement("div");
      ctrlRow.style.cssText = "display:flex;gap:6px;margin-bottom:6px;padding-bottom:6px;" +
                               "border-bottom:1px solid #eee;";
      var btnAll  = document.createElement("button");
      btnAll.textContent = "✔ Vše";
      btnAll.style.cssText = "font-size:0.75rem;padding:2px 8px;border:1px solid #aaa;" +
                             "border-radius:4px;background:#f5f5f5;cursor:pointer;";
      var btnNone = document.createElement("button");
      btnNone.textContent = "✕ Nic";
      btnNone.style.cssText = btnAll.style.cssText;
      ctrlRow.appendChild(btnAll);
      ctrlRow.appendChild(btnNone);
      dropPanel.appendChild(ctrlRow);

      var cbs = [];
      sorted.forEach(function(v) {{
        var lbl = document.createElement("label");
        lbl.style.cssText = "display:flex;align-items:center;gap:6px;padding:2px 0;" +
                            "font-size:0.82rem;cursor:pointer;white-space:nowrap;";
        var cb = document.createElement("input");
        cb.type = "checkbox";
        cb.style.cursor = "pointer";
        cb.addEventListener("change", function() {{
          if (cb.checked) filterObj.values.add(v);
          else            filterObj.values.delete(v);
          updateDropBtn();
          applyMultiFilter();
        }});
        cbs.push({{cb: cb, v: v}});
        lbl.appendChild(cb);
        lbl.appendChild(document.createTextNode(v));
        dropPanel.appendChild(lbl);
      }});

      btnAll.addEventListener("click", function(e) {{
        e.preventDefault();
        cbs.forEach(function(o) {{ o.cb.checked = true; filterObj.values.add(o.v); }});
        updateDropBtn(); applyMultiFilter();
      }});
      btnNone.addEventListener("click", function(e) {{
        e.preventDefault();
        cbs.forEach(function(o) {{ o.cb.checked = false; }});
        filterObj.values = new Set();
        updateDropBtn(); applyMultiFilter();
      }});

      updateDropBtn();
    }}

    // Otevření/zavření dropdownu
    dropBtn.addEventListener("click", function(e) {{
      e.stopPropagation();
      dropPanel.style.display = dropPanel.style.display === "none" ? "block" : "none";
    }});
    dropPanel.addEventListener("click", function(e) {{ e.stopPropagation(); }});
    document.addEventListener("click", function() {{ dropPanel.style.display = "none"; }});

    sel.addEventListener("change", function() {{
      filterObj.idx = parseInt(sel.value) || 0;
      rebuildPanel();
      applyMultiFilter();
    }});

    // Inicializuj panel pokud je sloupec předvybraný
    if (filterObj.idx) rebuildPanel();

    // Tlačítko smazat tento řádek
    var del = document.createElement("button");
    del.textContent = "✖";
    del.title = "Odebrat filtr";
    del.style.cssText = "padding:4px 8px;border:1px solid #e64343;background:#fff0f0;color:#e64343;" +
                        "border-radius:6px;cursor:pointer;font-size:0.82rem;flex-shrink:0;";
    del.addEventListener("click", function() {{
      var fi = mfFilters.indexOf(filterObj);
      if (fi >= 0) mfFilters.splice(fi, 1);
      mfRowsDiv.removeChild(rowDiv);
      applyMultiFilter();
    }});

    controlRow.appendChild(sel);
    controlRow.appendChild(dropWrap);
    controlRow.appendChild(del);
    rowDiv.appendChild(controlRow);
    mfRowsDiv.appendChild(rowDiv);
  }}

  mfAddBtn.addEventListener("click", function() {{
    var filterObj = {{idx: 0, values: new Set()}};
    mfFilters.push(filterObj);
    buildMfRow(filterObj);
  }});

  mfToggle.addEventListener("click", function() {{
    var open = mfPanel.style.display !== "none";
    mfPanel.style.display = open ? "none" : "block";
    mfToggle.textContent  = (open ? "🔽" : "🔼") + " Rozšířené filtry sloupců";
  }});

  mfClearAll.addEventListener("click", function() {{
    mfFilters.length = 0;
    mfRowsDiv.innerHTML = "";
    applyMultiFilter();
  }});

  if (btnNoRating) {{
    // Aktualizuj label tlačítka
    function _updateNoRatingBtn() {{
      btnNoRating.textContent = showNoRating
        ? "🙈 Skrýt bez ratingu (" + noRatingCodes.size + ")"
        : "👁 Zobrazit bez ratingu (" + noRatingCodes.size + ")";
      btnNoRating.style.background  = showNoRating ? "#e8f5e9" : "#fff3e0";
      btnNoRating.style.borderColor = showNoRating ? "#43a047" : "#fb8c00";
      btnNoRating.style.color       = showNoRating ? "#2e7d32" : "#e65100";
    }}
    _updateNoRatingBtn();
    btnNoRating.addEventListener("click", function() {{
      showNoRating = !showNoRating;
      _updateNoRatingBtn();
      doFilter();
    }});
  }}

  function onCheck() {{
    var sel = rows.filter(function(r) {{ return r.querySelector("input[type=checkbox]").checked; }});
    var any = sel.length > 0;
    btnShow.style.display = any ? "inline-block" : "none";
    btnDes.style.display  = any ? "inline-block" : "none";
    if (any) {{
      var n = sel.length;
      hintEl.textContent = n + "\u00a0" +
        (n === 1 ? "pobočka vybrána" : n < 5 ? "pobočky vybrány" : "poboček vybráno");
    }} else {{
      hintEl.textContent = "Zaškrtni řádky v tabulce";
    }}
    rows.forEach(function(r) {{
      r.style.backgroundColor = r.querySelector("input[type=checkbox]").checked ? "#fff8dc" : "";
    }});
  }}

  tbl.querySelector("tbody").addEventListener("change", function(e) {{
    if (e.target.type === "checkbox") onCheck();
  }});

  btnShow.addEventListener("click", function() {{
    rows.forEach(function(r) {{
      r.style.display = r.querySelector("input[type=checkbox]").checked ? "" : "none";
    }});
    updateCount();
  }});

  btnDes.addEventListener("click", function() {{
    rows.forEach(function(r) {{
      r.querySelector("input[type=checkbox]").checked = false;
      r.style.display = "";
      r.style.backgroundColor = "";
    }});
    btnShow.style.display = "none";
    btnDes.style.display  = "none";
    hintEl.textContent = "Zaškrtni řádky v tabulce";
    updateCount();
  }});

  updateCount();

  // ─── Klikatelné řazení sloupců ────────────────────────────────────
  var _sortState = {{col: -1, asc: true}};

  // Přidej data-sort každé buňce podle viditelného textu nebo čísla
  function _addSortData() {{
    var allRows = Array.from(tbl.querySelectorAll("tbody tr"));
    allRows.forEach(function(row) {{
      Array.from(row.cells).forEach(function(td, ci) {{
        if (ci === 0) return; // checkbox cell
        var txt = td.textContent.trim();
        var num = parseFloat(txt.replace(/[^0-9.]/g, "").replace(",", "."));
        td.setAttribute("data-sort", isNaN(num) ? txt.toLowerCase() : num);
        td.setAttribute("data-sort-num", isNaN(num) ? "0" : "1");
      }});
    }});
  }}
  _addSortData();

  function _reapplyColors(allRows, sortColIdx) {{
    // Po sortu: obnov group colors + zebra + oranžový sort sloupec
    allRows.forEach(function(row, i) {{
      var cbInput = row.querySelector("input[type=checkbox]");
      var isChecked = cbInput && cbInput.checked;
      Array.from(row.querySelectorAll("td[data-col-idx]")).forEach(function(td) {{
        var colidx = parseInt(td.getAttribute("data-col-idx"));
        if (isChecked) {{
          // checked rows: žlutá přes vše
          td.style.setProperty("background-color", "#fff8dc", "important");
        }} else if (colidx === sortColIdx) {{
          // sort sloupec: oranžová
          td.style.setProperty("background-color", "rgba(243,156,18,0.13)", "important");
        }} else {{
          // group color nebo zebra
          var gc = td.getAttribute("data-group-color");
          if (gc) {{
            td.style.setProperty("background-color", gc, "important");
          }} else {{
            var zebra = (i % 2 === 0) ? "#ffffff" : "#eef4ff";
            td.style.setProperty("background-color", zebra, "important");
          }}
        }}
      }});
      // Checkbox td (index 0) - zebra
      if (!isChecked && row.cells[0]) {{
        row.cells[0].style.setProperty("background-color", (i%2===0)?"#ffffff":"#eef4ff","important");
      }}
    }});
  }}

  function _sortBy(colIdx) {{
    var asc = (_sortState.col === colIdx) ? !_sortState.asc : true;
    _sortState = {{col: colIdx, asc: asc}};
    var tbody = tbl.querySelector("tbody");
    var allRows = Array.from(tbody.querySelectorAll("tr"));

    // Najdi TD na colIdx pomocí data-col-idx
    function getSortVal(row, idx) {{
      var td = row.querySelector("td[data-col-idx='" + idx + "']");
      if (!td) return "";
      return td.getAttribute("data-sort") || td.textContent.trim().toLowerCase();
    }}
    var isNum = (function() {{
      for (var ri = 0; ri < allRows.length; ri++) {{
        var td = allRows[ri].querySelector("td[data-col-idx='" + colIdx + "']");
        if (td) return td.getAttribute("data-sort-num") === "1";
      }}
      return false;
    }})();

    allRows.sort(function(a, b) {{
      var va = getSortVal(a, colIdx), vb = getSortVal(b, colIdx);
      if (isNum) return asc ? parseFloat(va) - parseFloat(vb) : parseFloat(vb) - parseFloat(va);
      return asc ? va.localeCompare(vb, "cs") : vb.localeCompare(va, "cs");
    }});

    allRows.forEach(function(row) {{ tbody.appendChild(row); }});
    _reapplyColors(allRows, colIdx);

    // Šipky + oranžová hlavička
    Array.from(tbl.querySelectorAll("thead th")).forEach(function(th) {{
      var arrow = th.querySelector(".ft-sort-arrow");
      if (!arrow) return;
      var thIdx = parseInt(th.getAttribute("data-col-idx") || "-1");
      if (thIdx === colIdx) {{
        th.style.setProperty("background-color", "#f39c12", "important");
        arrow.textContent = asc ? " ↑" : " ↓";
        arrow.style.opacity = "1";
      }} else {{
        th.style.removeProperty("background-color");
        arrow.textContent = " ↕";
        arrow.style.opacity = "0.4";
      }}
    }});

    updateCount();
  }}

  // Přidej šipku a data-col-idx na každou TH hlavičky
  Array.from(tbl.querySelectorAll("thead th")).forEach(function(th, i) {{
    if (i === 0) return; // přeskoč checkbox th
    var arrow = document.createElement("span");
    arrow.className = "ft-sort-arrow";
    arrow.textContent = " ↕";
    arrow.style.cssText = "font-size:0.7rem;opacity:0.4;pointer-events:none;";
    th.appendChild(arrow);
    th.style.cursor = "pointer";
    th.title = "Klikni pro řazení";
    th.setAttribute("data-col-idx", i);   // i=1 for first real col (checkbox is 0)
    th.addEventListener("click", function() {{
      _sortBy(parseInt(th.getAttribute("data-col-idx")));
    }});
  }});
  // ─────────────────────────────────────────────────────────────────

  // ─── Expand / Collapse tlačítko ──────────────────────────────────
  var btnExpand  = document.getElementById("ft-expand-{table_id}");
  var scrollWrap = wrap.querySelector(".tbl-scroll-wrap");
  var _expanded  = false;

  function _setExpanded(on) {{
    _expanded = on;
    if (on) {{
      wrap.classList.add("tbl-expanded");
      scrollWrap.style.maxHeight = "calc(100vh - 220px)";
      btnExpand.textContent = "⤡ Zmenšit";
      btnExpand.style.background = "#2770f0";
      btnExpand.style.color      = "white";
      btnExpand.style.borderColor= "#1a4db5";
      document.body.style.overflow = "hidden";
    }} else {{
      wrap.classList.remove("tbl-expanded");
      scrollWrap.style.maxHeight = "900px";
      btnExpand.textContent = "⤢ Rozšířit";
      btnExpand.style.background  = "#eef4ff";
      btnExpand.style.color       = "#2770f0";
      btnExpand.style.borderColor = "#2770f0";
      document.body.style.overflow = "";
    }}
  }}

  if (btnExpand) {{
    btnExpand.addEventListener("click", function() {{ _setExpanded(!_expanded); }});
  }}

  // Zavři expand přes Escape
  document.addEventListener("keydown", function(e) {{
    if (e.key === "Escape" && _expanded) _setExpanded(false);
  }});

  updateCount();
}}, 0);
</script>
"""




# =============================================================================
# MAPA POBOČEK — scatter_mapbox
# =============================================================================

MAPBOX_TOKEN = (
    'pk.eyJ1IjoiY2VydmVueWppIiwiYSI6ImNqM2tsbTl6ajA'
    'wazMyd3FzeGZxa2VxZzcifQ.z-Ruzxhj76p-f84Ti4r3Gw'
)

def _mapbox_layout(center_lat=49.8, center_lon=15.5, zoom=6.5):
    """Vrátí dict pro mapbox layout — bez tokenu použije open-street-map."""
    if MAPBOX_TOKEN:
        return dict(style='light', accesstoken=MAPBOX_TOKEN,
                    center=dict(lat=center_lat, lon=center_lon), zoom=zoom)
    return dict(style='open-street-map',
                center=dict(lat=center_lat, lon=center_lon), zoom=zoom)

def generate_map_html(df, title="🗺️ Mapa poboček"):
    """
    Scatter_mapbox s pobočkami. Každá kombinace (hotovostní/bezhotovostní × strategie) = vlastní trace.
    Strategie: Close rok, Cashless rok, Keep, Ostatní.
    Symbol: circle = hotovostní, circle-stroke = bezhotovostní (outline).
    """
    import plotly.graph_objects as go

    _df = df.copy()
    for _gc in ['GPS_X', 'GPS_Y']:
        if _gc in _df.columns:
            _df[_gc] = pd.to_numeric(_df[_gc], errors='coerce')

    if 'GPS_X' not in _df.columns or 'GPS_Y' not in _df.columns:
        return '<p style="color:#aaa;font-style:italic;">GPS souřadnice nejsou k dispozici.</p>'

    _df = _df[_df['GPS_X'].notna() & _df['GPS_Y'].notna()].copy()
    if _df.empty:
        return '<p style="color:#aaa;font-style:italic;">Žádné pobočky s GPS souřadnicemi.</p>'

    def _sv(row, col, default='—'):
        v = row.get(col, default)
        return default if (v is None or (isinstance(v, float) and pd.isna(v)) or str(v) in ('nan','None','')) else str(v)

    def _year(v):
        try:
            y = int(float(str(v)))
            return str(y) if 1900 < y < 2100 else None
        except: return None

    # ── Kategorizace ─────────────────────────────────────────────────────────
    # Strategie: pevné kategorie s rokem
    # Cash: hotovostní (CASHLESS != Y) vs bezhotovostní (CASHLESS = Y)
    def _get_strat(row):
        strat = str(row.get('FOOTPRINT_STRATEGIE_(2030)', '')).lower().strip()
        cl_r  = _year(row.get('PLAN_CLOSE_(ROK)', ''))
        cs_r  = _year(row.get('PLAN_NEW_CASHIERLESS_(ROK)', ''))
        if 'close' in strat:
            return ('close', cl_r or '?', f'🔴 Close {cl_r or "?"}')
        if cs_r:
            return ('cashless_plan', cs_r, f'🔵 → Cashless {cs_r}')
        if 'keep' in strat:
            return ('keep', None, '🟢 Keep')
        return ('other', None, '⚪ Ostatní / bez strategie')

    def _is_cashless(row):
        return str(row.get('CASHLESS', 'N')).upper() in ['Y', 'YES', 'ANO', '1', 'TRUE']

    _df['_strat_key']   = [_get_strat(r)[0] for _, r in _df.iterrows()]
    _df['_strat_year']  = [_get_strat(r)[1] for _, r in _df.iterrows()]
    _df['_strat_label'] = [_get_strat(r)[2] for _, r in _df.iterrows()]
    _df['_is_cashless'] = [_is_cashless(r) for _, r in _df.iterrows()]
    _df['_cash_label']  = _df['_is_cashless'].map({True: 'Bezhotovostní', False: 'Hotovostní'})

    # Barva dle strategie
    def _strat_color(key):
        return {'close': '#d62728', 'cashless_plan': '#2770f0', 'keep': '#2ca02c'}.get(key, '#999999')

    # scatter_mapbox podporované symboly: circle, square, marker (=pin)
    # Hotovostní = circle (plný), Bezhotovostní = circle s outline = větší prázdný kruh
    # Realizujeme přes 2 parametry: size + opacity

    # ── Hover text ───────────────────────────────────────────────────────────
    hovers = []
    for _, row in _df.iterrows():
        nm    = _sv(row, 'BRANCH_NAME')
        bc    = _sv(row, 'BRANCH_CODE')
        fmt   = _sv(row, 'BRANCH_FORMAT')
        nfsf  = _sv(row, 'BRANCH_BUILDING_NF_SF')
        reg   = _sv(row, 'REGION_NAME')
        ir    = _sv(row, 'IR')
        strat = _sv(row, 'FOOTPRINT_STRATEGIE_(2030)')
        cl_r  = _year(row.get('PLAN_CLOSE_(ROK)', ''))
        cs_r  = _year(row.get('PLAN_NEW_CASHIERLESS_(ROK)', ''))
        cs_st = _sv(row, 'NEW_CASHIERLESS_STATUS')
        cash_lbl = '🏦 Bezhotovostní' if _is_cashless(row) else '💵 Hotovostní'
        lines = [
            f'<b>{nm}</b> <span style="color:#aaa;">#{bc}</span>',
            f'📍 {reg} &nbsp;·&nbsp; 📐 {fmt} &nbsp;·&nbsp; {nfsf}',
            f'{cash_lbl} &nbsp;·&nbsp; ⭐ IR 2025: <b>{ir}</b>',
            f'🏗️ {strat}',
        ]
        if cl_r:  lines.append(f'🔴 Plán uzavření: <b>{cl_r}</b>')
        if cs_r:  lines.append(f'🔵 Cashless přechod: <b>{cs_r}</b> ({cs_st})')
        hovers.append('<br>'.join(lines))
    _df['_hover'] = hovers

    # ── Pevné pořadí kategorií pro legendu ───────────────────────────────────
    # Seřadíme: close roky vzestupně, cashless roky vzestupně, keep, ostatní
    def _sort_key(label):
        if label.startswith('🔴'): return (0, label)
        if label.startswith('🔵'): return (1, label)
        if label.startswith('🟢'): return (2, label)
        return (3, label)

    strat_labels_ordered = sorted(_df['_strat_label'].unique(), key=_sort_key)

    # ── Traces ───────────────────────────────────────────────────────────────
    traces = []
    for strat_lbl in strat_labels_ordered:
        strat_key = _df[_df['_strat_label'] == strat_lbl]['_strat_key'].iloc[0]
        color = _strat_color(strat_key)

        for is_cash_val, cash_lbl, size, opacity, border_w in [
            (False, 'Hotovostní',     10, 0.90, 0),    # plný kruh
            (True,  'Bezhotovostní',  11, 0.35, 2),    # průhledný s okrajem
        ]:
            mask = (_df['_strat_label'] == strat_lbl) & (_df['_is_cashless'] == is_cash_val)
            sub  = _df[mask]
            if sub.empty: continue

            trace_name = f'{strat_lbl} — {cash_lbl}'
            traces.append(go.Scattermapbox(
                lat=sub['GPS_X'].tolist(),
                lon=sub['GPS_Y'].tolist(),
                mode='markers',
                marker=dict(
                    size=size,
                    color=color,
                    opacity=opacity,
                ),
                text=sub['_hover'].tolist(),
                hovertemplate='%{text}<extra></extra>',
                name=trace_name,
                legendgroup=strat_lbl,
                legendgrouptitle_text=strat_lbl if cash_lbl == 'Hotovostní' else None,
            ))
            # Přidáme outline vrstvu pro bezhotovostní (druhý trace přes stejné body, jen outline)
            if is_cash_val:
                traces.append(go.Scattermapbox(
                    lat=sub['GPS_X'].tolist(),
                    lon=sub['GPS_Y'].tolist(),
                    mode='markers',
                    marker=dict(size=size, color=color, opacity=0.95),
                    text=sub['_hover'].tolist(),
                    hovertemplate='%{text}<extra></extra>',
                    name=trace_name + ' (outline)',
                    legendgroup=strat_lbl,
                    showlegend=False,
                ))

    fig = go.Figure(traces)
    fig.update_layout(
        mapbox=_mapbox_layout(center_lat=49.8, center_lon=15.5, zoom=6.5),
        margin=dict(l=0, r=0, t=0, b=0),
        height=560,
        legend=dict(
            bgcolor='rgba(255,255,255,0.93)',
            bordercolor='#ccc',
            borderwidth=1,
            font=dict(size=11),
            tracegroupgap=4,
            x=0.01, y=0.99,
            xanchor='left', yanchor='top',
        ),
        paper_bgcolor='white',
    )

    _map_div = fig.to_html(
        full_html=False,
        include_plotlyjs='cdn',
        config={'displayModeBar': True, 'scrollZoom': True,
                'modeBarButtonsToRemove': ['toImage'], 'displaylogo': False},
    )

    n_close    = (_df['_strat_key'] == 'close').sum()
    n_cashless = (_df['_strat_key'] == 'cashless_plan').sum()
    n_keep     = (_df['_strat_key'] == 'keep').sum()
    n_cash_hot = (~_df['_is_cashless']).sum()
    n_cash_bez = _df['_is_cashless'].sum()

    return f"""
<div style="margin-top:24px;padding-top:20px;border-top:2px solid #e0e6f5;">
  <div class="section-header">{title}</div>
  <div style="display:flex;gap:10px;flex-wrap:wrap;margin:6px 0 8px 0;font-size:0.78rem;">
    <span style="background:#fdecea;color:#d62728;border:1px solid #f5b8b8;border-radius:12px;padding:2px 10px;font-weight:600;">🔴 Close: {n_close}</span>
    <span style="background:#e8f0ff;color:#2770f0;border:1px solid #b8ccf7;border-radius:12px;padding:2px 10px;font-weight:600;">🔵 → Cashless: {n_cashless}</span>
    <span style="background:#eaf7ec;color:#2ca02c;border:1px solid #a8dfc8;border-radius:12px;padding:2px 10px;font-weight:600;">🟢 Keep: {n_keep}</span>
    <span style="background:#f5f5f5;color:#555;border:1px solid #ddd;border-radius:12px;padding:2px 10px;">💵 Hotovostní: {n_cash_hot} &nbsp;|&nbsp; 🏦 Bezhotovostní: {n_cash_bez}</span>
  </div>
  <div style="border-radius:10px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.1);">
    {_map_div}
  </div>
  <div style="font-size:0.71rem;color:#aaa;margin-top:4px;">
    Plný kruh = hotovostní &nbsp;|&nbsp; Průhledný kruh = bezhotovostní &nbsp;|&nbsp;
    Kliknutím na legendu zobrazíte/skryjete skupinu. Scroll = zoom.
  </div>
</div>"""


def _render_top_revenues_table(df, region_label="", n=10):
    """
    Přehled oblastí s nejvyššími výnosy.
    Pokud je k dispozici sloupec OBLAST, zobrazí agregát dle oblasti.
    Jinak zobrazí top pobočky.
    Dvě tabulky vedle sebe: dle celkových výnosů (VYNOSY) a nových výnosů (OBJEM_VYNOSU_CZK).
    """
    d = df.copy()
    for c in ['VYNOSY', 'OBJEM_VYNOSU_CZK', 'IR', 'PRIME_NAKLADY/VYNOSY']:
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors='coerce')

    use_oblast = 'OBLAST' in d.columns and d['OBLAST'].notna().any() and (d['OBLAST'].astype(str).str.strip() != '').any()

    if use_oblast:
        # ── Oblast agregace ───────────────────────────────────────────
        def _tbl(src_df, val_col, val_label, accent):
            if val_col not in src_df.columns:
                return "<p style='color:#aaa;font-style:italic;font-size:0.82rem;'>Žádná data.</p>"
            g = (src_df.dropna(subset=[val_col])
                 .groupby('OBLAST', observed=True)
                 .agg(
                     vynosy_sum = (val_col, 'sum'),
                     pocet      = ('BRANCH_CODE', 'count'),
                     ci_mean    = ('PRIME_NAKLADY/VYNOSY', 'mean') if 'PRIME_NAKLADY/VYNOSY' in src_df.columns else ('BRANCH_CODE', 'count'),
                 )
                 .reset_index()
                 .query('vynosy_sum > 0')
                 .sort_values('vynosy_sum', ascending=False)
                 .head(n))
            if g.empty:
                return "<p style='color:#aaa;font-style:italic;font-size:0.82rem;'>Žádná data.</p>"
            max_val = float(g['vynosy_sum'].max()) or 1
            rows = ''
            for i, (_, r) in enumerate(g.iterrows()):
                oblast = str(r['OBLAST'])
                val    = float(r['vynosy_sum'])
                pocet  = int(r['pocet'])
                pct    = val / max_val * 100
                bg     = '#fff' if i % 2 == 0 else '#fafbff'
                medal  = ['🥇','🥈','🥉'][i] if i < 3 else f'<span style="color:#aaa;font-size:0.72rem;">#{i+1}</span>'
                try: ci_s = f"⌀ C/I {float(r['ci_mean'])*100:.1f} %" if 'PRIME_NAKLADY/VYNOSY' in src_df.columns else ''
                except: ci_s = ''
                rows += (
                    f'<tr style="background:{bg};">'
                    f'<td style="padding:5px 8px;text-align:center;width:28px;">{medal}</td>'
                    f'<td style="padding:5px 8px;font-weight:600;color:#0f172a;">{oblast}'
                    f'<span style="color:#aaa;font-size:0.7rem;margin-left:6px;">{pocet} pb.</span></td>'
                    f'<td style="padding:5px 8px;">'
                    f'<div style="display:flex;align-items:center;gap:6px;">'
                    f'<div style="flex:1;background:#eee;border-radius:3px;height:6px;">'
                    f'<div style="width:{pct:.0f}%;background:{accent};height:100%;border-radius:3px;"></div></div>'
                    f'<span style="font-weight:700;color:{accent};white-space:nowrap;font-size:0.82rem;">'
                    f'{format_money(val)}</span></div></td>'
                    f'<td style="padding:5px 8px;text-align:right;color:#888;font-size:0.75rem;">{ci_s}</td>'
                    f'</tr>'
                )
            return (
                f'<table style="width:100%;border-collapse:collapse;font-size:0.8rem;">'
                f'<thead><tr style="background:#f4f8ff;">'
                f'<th style="padding:5px 8px;"></th>'
                f'<th style="padding:5px 8px;text-align:left;color:#555;">Oblast</th>'
                f'<th style="padding:5px 8px;text-align:left;color:{accent};">{val_label}</th>'
                f'<th style="padding:5px 8px;text-align:right;color:#888;">C/I průměr</th>'
                f'</tr></thead><tbody>{rows}</tbody></table>'
            )
        entity_label = "Oblast"
    else:
        # ── Fallback: top pobočky ─────────────────────────────────────
        def _tbl(src_df, val_col, val_label, accent):
            if val_col not in src_df.columns:
                return "<p style='color:#aaa;font-style:italic;font-size:0.82rem;'>Žádná data.</p>"
            top = src_df.dropna(subset=[val_col]).query(f'{val_col} > 0').sort_values(val_col, ascending=False).head(n)
            if top.empty:
                return "<p style='color:#aaa;font-style:italic;font-size:0.82rem;'>Žádná data.</p>"
            max_val = float(top[val_col].max()) or 1
            rows = ''
            for i, (_, r) in enumerate(top.iterrows()):
                bc  = int(r.get('BRANCH_CODE', 0))
                bn  = str(r.get('BRANCH_NAME', bc))
                val = float(r[val_col])
                ir  = r.get('IR'); ci = r.get('PRIME_NAKLADY/VYNOSY')
                try: ir_s = str(int(float(ir)))
                except: ir_s = '—'
                try: ci_s = f"{float(ci)*100:.1f} %"
                except: ci_s = '—'
                pct  = val / max_val * 100
                bg   = '#fff' if i % 2 == 0 else '#fafbff'
                medal = ['🥇','🥈','🥉'][i] if i < 3 else f'<span style="color:#aaa;font-size:0.72rem;">#{i+1}</span>'
                rows += (
                    f'<tr style="background:{bg};">'
                    f'<td style="padding:5px 8px;text-align:center;width:28px;">{medal}</td>'
                    f'<td style="padding:5px 8px;font-weight:600;color:#0f172a;">'
                    f'{bn}<span style="color:#aaa;font-size:0.7rem;margin-left:5px;">({bc})</span></td>'
                    f'<td style="padding:5px 8px;">'
                    f'<div style="display:flex;align-items:center;gap:6px;">'
                    f'<div style="flex:1;background:#eee;border-radius:3px;height:6px;">'
                    f'<div style="width:{pct:.0f}%;background:{accent};height:100%;border-radius:3px;"></div></div>'
                    f'<span style="font-weight:700;color:{accent};white-space:nowrap;font-size:0.82rem;">'
                    f'{format_money(val)}</span></div></td>'
                    f'<td style="padding:5px 8px;text-align:center;color:#888;font-size:0.75rem;">{ir_s}</td>'
                    f'<td style="padding:5px 8px;text-align:center;color:#888;font-size:0.75rem;">{ci_s}</td>'
                    f'</tr>'
                )
            return (
                f'<table style="width:100%;border-collapse:collapse;font-size:0.8rem;">'
                f'<thead><tr style="background:#f4f8ff;">'
                f'<th style="padding:5px 8px;"></th>'
                f'<th style="padding:5px 8px;text-align:left;color:#555;">Pobočka</th>'
                f'<th style="padding:5px 8px;text-align:left;color:{accent};">{val_label}</th>'
                f'<th style="padding:5px 8px;text-align:center;color:#888;">Rating</th>'
                f'<th style="padding:5px 8px;text-align:center;color:#888;">C/I</th>'
                f'</tr></thead><tbody>{rows}</tbody></table>'
            )
        entity_label = "Pobočka"

    left  = (f'<div style="flex:1;min-width:300px;">'
             f'<div style="font-size:0.75rem;font-weight:700;color:#0bb440;text-transform:uppercase;'
             f'letter-spacing:.4px;margin-bottom:8px;">💰 Top {n} — celkové výnosy (VYNOSY)</div>'
             f'{_tbl(d,"VYNOSY","Výnosy celkem","#0bb440")}</div>') if 'VYNOSY' in d.columns else ''
    right = (f'<div style="flex:1;min-width:300px;">'
             f'<div style="font-size:0.75rem;font-weight:700;color:#2770f0;text-transform:uppercase;'
             f'letter-spacing:.4px;margin-bottom:8px;">🆕 Top {n} — nové výnosy (OBJEM_VYNOSU_CZK)</div>'
             f'{_tbl(d,"OBJEM_VYNOSU_CZK","Nové výnosy","#2770f0")}</div>') if 'OBJEM_VYNOSU_CZK' in d.columns else ''

    if not left and not right:
        return "<p style='color:#aaa;font-style:italic;'>Žádná výnosová data.</p>"
    src_note = f'<div style="font-size:0.72rem;color:#aaa;margin-bottom:10px;">{"Seskupeno dle oblasti — součet výnosů všech poboček v oblasti" if use_oblast else "Top pobočky dle výnosů — OBLAST není k dispozici"}</div>'
    return f'{src_note}<div style="display:flex;gap:24px;flex-wrap:wrap;align-items:flex-start;">{left}{right}</div>'


def generate_report(rating_status, mode='static', output_prefix="report"):
    """
    Generuje HTML report(y).
      mode='static'   -> jeden souhrnný soubor
      mode='regional' -> soubory per region
    """
    # ── Připoj spádovky do rating_status ─────────────────────────────────────
    if 'spadovky' in dir() or 'spadovky' in globals():
        _sp_src = globals().get('spadovky')
        if _sp_src is not None and not _sp_src.empty:
            _sp = _sp_src.copy()
            _sp.columns = [c.strip().upper() for c in _sp.columns]
            _sp['BRANCH_ID'] = pd.to_numeric(_sp['BRANCH_ID'], errors='coerce')
            # Vypočítej celkové návštěvy a procentuální podíl pro každou pobočku
            for _i in [1, 2, 3]:
                _sp[f'TOP_POBOCKA_VISITS_{_i}'] = pd.to_numeric(_sp.get(f'TOP_POBOCKA_VISITS_{_i}', 0), errors='coerce').fillna(0)
            _sp['_SP_VIS_TOTAL'] = _sp['TOP_POBOCKA_VISITS_1'] + _sp['TOP_POBOCKA_VISITS_2'] + _sp['TOP_POBOCKA_VISITS_3']
            for _i in [1, 2, 3]:
                _sp[f'SP_NAZEV_{_i}']  = _sp.get(f'TOP_POBOCKA_NAZEV_{_i}', '')
                _sp[f'SP_VISITS_{_i}'] = _sp[f'TOP_POBOCKA_VISITS_{_i}'].astype(int)
                _sp[f'SP_PCT_{_i}']    = (_sp[f'TOP_POBOCKA_VISITS_{_i}'] / _sp['_SP_VIS_TOTAL'].replace(0, float('nan')) * 100).round(1)
            _sp_cols = ['BRANCH_ID'] + [f'SP_NAZEV_{i}' for i in [1,2,3]] + [f'SP_VISITS_{i}' for i in [1,2,3]] + [f'SP_PCT_{i}' for i in [1,2,3]]
            _sp_merge = _sp[[c for c in _sp_cols if c in _sp.columns]].copy()
            # Odstraň staré spádové sloupce pokud existují, pak merguj
            _drop = [c for c in _sp_merge.columns if c != 'BRANCH_ID' and c in rating_status.columns]
            if _drop:
                rating_status = rating_status.drop(columns=_drop)
            rating_status = rating_status.merge(
                _sp_merge.rename(columns={'BRANCH_ID': 'BRANCH_CODE'}),
                on='BRANCH_CODE', how='left'
            )
    bootstrap_css = '''
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
    :root {
        --brand:    #2770f0;
        --brand-dk: #1a4db5;
        --brand-bg: #eef4ff;
        --pos:      #0bb440;
        --pos-bg:   #eafaf1;
        --accent:   #eb4d79;
        --accent-bg:#fdf0f4;
        --warn:     #f59e0b;
        --warn-bg:  #fffbeb;
        --muted:    #6b7280;
        --border:   #dde4f5;
        --surface:  #ffffff;
        --page-bg:  #f7f9fc;
    }

    /* ── Základní layout ───────────────────────────────────────────── */
    .report-wrapper {
        font-family: 'Segoe UI', Arial, sans-serif;
        padding: 32px 36px;
        background: var(--page-bg);
        max-width: 1600px;
        margin: 0 auto;
    }
    .region-title {
        font-size: 1.45rem;
        font-weight: 700;
        color: #0f172a;
        border-left: 6px solid var(--brand);
        padding: 6px 0 6px 16px;
        margin-bottom: 28px;
        background: var(--brand-bg);
        border-radius: 0 6px 6px 0;
    }
    .section-header {
        margin-top: 36px;
        margin-bottom: 12px;
        font-size: 1.1rem;
        font-weight: 700;
        color: #0f172a;
        padding-bottom: 6px;
        border-bottom: 2px solid var(--border);
    }

    /* ── Tabulky ───────────────────────────────────────────────────── */
    .table {
        font-size: 0.82rem;
        background: var(--surface);
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
        border-radius: 6px;
        overflow: hidden;
    }
    .table thead th {
        background-color: var(--brand) !important;
        color: white !important;
        text-align: center;
        vertical-align: middle;
        white-space: nowrap;
        border: 1px solid var(--brand-dk);
        position: sticky;
        top: 0;
        z-index: 10;
        font-weight: 600;
        letter-spacing: 0.2px;
    }
    .table tbody td {
        vertical-align: middle;
        text-align: center;
        white-space: nowrap;
        border: 1px solid #e8edf5;
    }
    .table-striped tbody tr:nth-of-type(even) { background-color: var(--brand-bg) !important; }
    .table-hover tbody tr:hover td,
    .table tbody tr:hover td { background-color: #d1fae5 !important; }

    /* ── Karty a panely ────────────────────────────────────────────── */
    .age-section-card {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: 10px;
        padding: 20px 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    .benchmark-card { border: none; border-radius: 10px; overflow: hidden; }
    .benchmark-val { font-weight: 700; color: var(--brand); }

    /* ── Metodika / legenda ────────────────────────────────────────── */
    .legend-box {
        margin-top: 48px;
        padding: 24px 28px;
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    .legend-title {
        color: var(--brand);
        font-weight: 700;
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 0.4px;
        margin-bottom: 12px;
        border-bottom: 2px solid var(--border);
        padding-bottom: 6px;
    }

    /* ── Collapsible ───────────────────────────────────────────────── */
    details.collapsible-section > summary { list-style: none; }
    details.collapsible-section > summary::-webkit-details-marker { display: none; }
    details.collapsible-section > summary::marker { display: none; }
    details.collapsible-section > summary {
        font-weight: 600;
        font-size: 0.95rem;
        background: var(--brand-bg);
        color: var(--brand);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 10px 16px;
        cursor: pointer;
        user-select: none;
        display: flex;
        align-items: center;
        gap: 8px;
        transition: background 0.15s;
    }
    details.collapsible-section > summary:hover { background: #dbeafe; }
    details.collapsible-section[open] > summary { border-radius: 8px 8px 0 0; border-bottom: 1px solid var(--border); }
    details.collapsible-section:not([open]) > summary { border-radius: 8px; border-bottom: none; }
    details.collapsible-section > .collapsible-body {
        border: 1px solid var(--border);
        border-top: none;
        border-radius: 0 0 8px 8px;
        padding: 16px 20px;
        background: var(--surface);
    }

    /* ── Oddělovač ─────────────────────────────────────────────────── */
    .report-divider {
        margin: 32px 0;
        border: none;
        border-top: 2px solid var(--border);
    }
</style>'''

    def _fmt_validity(v):
        """EOY2025 → 1.1.2025–31.12.2025, EOY2024 → 31.12.2024, jinak ponech."""
        if v == "EOY2025": return "1.1.2025–31.12.2025"
        if v == "EOY2024": return "31.12.2024"
        return v

    datasets_rows = "".join(
        f"<tr>"
        f"<td style='padding:7px 12px;border:1px solid #e8edf5;font-weight:700;"
        f"color:#1a2a4a;white-space:nowrap;'>{jmeno}</td>"
        f"<td style='padding:7px 12px;border:1px solid #e8edf5;color:#555;"
        f"font-size:0.8rem;font-family:monospace;'>{info.get('path','—')}</td>"
        f"<td style='padding:7px 12px;border:1px solid #e8edf5;text-align:center;white-space:nowrap;'>"
        f"<span style='background:#eef4ff;color:#2770f0;border:1px solid #c7d9fb;"
        f"border-radius:5px;padding:2px 9px;font-size:0.78rem;font-weight:600;'>"
        f"{_fmt_validity(info.get('validity','—'))}</span></td>"
        f"<td style='padding:7px 12px;border:1px solid #e8edf5;color:#444;"
        f"font-size:0.82rem;'>{info.get('source_desc','—')}</td>"
        f"</tr>"
        for jmeno, info in soubory.items()
    )

    html_legenda = f'''
<div class="legend-box">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:20px;">
    <span style="font-size:1.3rem;">📖</span>
    <h3 style="margin:0;font-size:1.05rem;font-weight:700;color:#0f172a;">Metodika a vysvětlivky</h3>
  </div>

  {make_collapsible("legenda-metodika", "🧮 Metodika výpočtu ratingů", generate_rating_methodology_html(), default_open=False)}

  <div style="margin-top:20px;">
    <div class="legend-title">🗂️ Použité datasety a jejich platnost</div>
    <div style="overflow-x:auto;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,0.06);">
      <table style="width:100%;border-collapse:collapse;font-size:0.83rem;">
        <thead>
          <tr style="background:#2770f0;">
            <th style="padding:8px 12px;text-align:left;color:white;font-weight:600;border:1px solid #1a4db5;">Dataset</th>
            <th style="padding:8px 12px;text-align:left;color:white;font-weight:600;border:1px solid #1a4db5;">Zdroj (cesta)</th>
            <th style="padding:8px 12px;text-align:center;color:white;font-weight:600;border:1px solid #1a4db5;white-space:nowrap;">Platnost dat</th>
            <th style="padding:8px 12px;text-align:left;color:white;font-weight:600;border:1px solid #1a4db5;">Popis</th>
          </tr>
        </thead>
        <tbody>{datasets_rows}</tbody>
      </table>
    </div>
  </div>

  <div style="margin-top:16px;">
    {generate_column_map_html()}
  </div>
</div>'''

    if mode == 'static':
        df_sorted    = rating_status.sort_values(by="IR")
        df_by_branch = rating_status.sort_values(by="BRANCH_CODE")

        # --- věková tabulka a graf ---
        age_chart_id   = "apex-age-static"
        age_table_html = generate_age_segment_table(df_by_branch, table_id="age-tbl-static")
        age_chart_html = generate_age_apex_chart(rating_status, age_chart_id)

        # --- obchodní tabulka ---
        obchody_cols  = ['BRANCH_CODE'] + [c for c in obchody_detail.columns if c != 'BRANCH_CODE']
        obchody_avail = [c for c in obchody_cols if c in obchody_detail.columns]
        _fmt_col_avail = [c for c in ['BRANCH_FORMAT', 'BRANCH_TYPE_NAME'] if c in df_by_branch.columns]
        _obch_base_cols = ['BRANCH_CODE', 'BRANCH_NAME'] + _fmt_col_avail
        df_obchody = df_by_branch[_obch_base_cols].merge(
            obchody_detail[obchody_avail], on='BRANCH_CODE', how='left'
        ).sort_values(by="BRANCH_CODE")
        obchody_raw  = generate_obchody_table(df_obchody, table_id="obchody-static")
        obchody_html = generate_sortable_table(obchody_raw, "obchody-sort-static")
        obchody_bench_html = generate_obchody_benchmark_chart(df_obchody, chart_id="obch-bench-static")

        # --- přehledové tabulky ---
        branch_summary_html  = generate_branch_summary_tables(rating_status)
        close_branches_html       = generate_close_branches_table(rating_status)
        invest_branches_html      = generate_invest_branches_table(rating_status)
        cashierless_branches_html = generate_cashierless_table(rating_status)

        # --- hlavní tabulka ---
        html_table = prepare_table(df_sorted)
        html_table = html_table.replace(
            '<th>Rating 25</th>',
            '<th style="background:#f39c12;color:white;white-space:nowrap;">Rating 25 ↑</th>'
        )
        html_table = html_table.replace('<table ', '<table id="main-static-tbl" ', 1)
        html_table += """
<script>
(function() {
  var tbl = document.getElementById('main-static-tbl');
  if (!tbl) return;
  var ths = Array.from(tbl.querySelectorAll('thead th'));
  var colIdx = ths.findIndex(function(th) { return th.textContent.trim().startsWith('Rating 25 ↑'); });
  if (colIdx < 0) return;
  tbl.querySelectorAll('tbody tr').forEach(function(tr) {
    var td = tr.cells[colIdx];
    if (td) td.style.background = 'rgba(243,156,18,0.12)';
  });
})();
</script>"""

        # --- síťový benchmark (celá síť — bez srovnání s "ostatními") ---
        def net_bench_row(label, value_html, note=""):
            return (f"<tr><td style='text-align:left;'>{label}</td>"
                    f"<td class='benchmark-val' style='text-align:right;'>{value_html}</td>"
                    f"<td style='text-align:right;color:#888;font-size:0.8rem;'>{note}</td></tr>")

        html_benchmark_static = f'''
<style>
  .benchmark-hover-tbl tbody tr:hover td {{ background-color: #d1fae5 !important; }}
</style>
<div class="age-section-card mb-4">
  <div class="row align-items-start">
    <div class="col-md-8">
      <table class="table table-sm mb-0 benchmark-hover-tbl" style="font-size:0.83rem;">
        <thead class="table-dark">
          <tr>
            <th style="text-align:left;">Ukazatel</th>
            <th style="color:#7ec8f7;text-align:right;">Průměr — celá síť</th>
            <th style="text-align:right;color:#aaa;">Pozn.</th>
          </tr>
        </thead>
        <tbody>
          <tr><td colspan="3" style="background:#f0f4ff;font-weight:600;font-size:0.78rem;color:#2770f0;padding:4px 8px;text-align:left;">💰 Finanční ukazatele</td></tr>
          {net_bench_row("Výnosy 25",      format_money(rating_status["VYNOSY"].mean()),              f"∑ {format_money(rating_status['VYNOSY'].sum())}")}
          {net_bench_row("Nové výnosy 25", format_money(rating_status["OBJEM_VYNOSU_CZK"].mean()),    f"∑ {format_money(rating_status['OBJEM_VYNOSU_CZK'].sum())}")}
          {net_bench_row("C/I Ratio 25",   format_pct(rating_status["PRIME_NAKLADY/VYNOSY"].mean()), "")}

          <tr><td colspan="3" style="background:#e8f5e9;font-weight:600;font-size:0.78rem;color:#1a7a4a;padding:4px 8px;text-align:left;">🛒 Obchody dle produktů — průměrné výnosy a počty prodejů 2025</td></tr>
          {"".join(
            f"<tr style='background:#f9fffe;'>"
            f"<td style='text-align:left;padding-left:16px;'>{label} — výnosy 25</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_money(rating_status[f'OBJEM_VYNOSU_CZK_{key}_2025'].mean()) if f'OBJEM_VYNOSU_CZK_{key}_2025' in rating_status.columns else '—'}</td>"
            f"<td style='text-align:right;color:#888;font-size:0.8rem;'>∑ {format_money(rating_status[f'OBJEM_VYNOSU_CZK_{key}_2025'].sum()) if f'OBJEM_VYNOSU_CZK_{key}_2025' in rating_status.columns else '—'}</td>"
            f"</tr>"
            f"<tr style='background:#f5fffc;'>"
            f"<td style='text-align:left;padding-left:16px;'>{label} — počet prodejů 25</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_num_round(rating_status[f'POCET_PRODEJU_{key}_2025'].mean()) if f'POCET_PRODEJU_{key}_2025' in rating_status.columns else '—'}</td>"
            f"<td style='text-align:right;color:#888;font-size:0.8rem;'>∑ {format_num_round(rating_status[f'POCET_PRODEJU_{key}_2025'].sum()) if f'POCET_PRODEJU_{key}_2025' in rating_status.columns else '—'}</td>"
            f"</tr>"
            for key, label in PRODUCT_GROUPS
            if f'OBJEM_VYNOSU_CZK_{key}_2025' in rating_status.columns or f'POCET_PRODEJU_{key}_2025' in rating_status.columns
          )}

          <tr><td colspan="3" style="background:#f0f4ff;font-weight:600;font-size:0.78rem;color:#2770f0;padding:4px 8px;text-align:left;">🚶 Návštěvnost</td></tr>
          {net_bench_row("Návštěvy celkem",       format_num_round(rating_status["POCET_NAVSTEV_CELKEM"].mean()),  f"∑ {format_num_round(rating_status['POCET_NAVSTEV_CELKEM'].sum())}")}
          {net_bench_row("Hotovostní walk-in",    format_num_round(rating_status["POCET_HOT_WALK_IN"].mean()),     f"∑ {format_num_round(rating_status['POCET_HOT_WALK_IN'].sum())}")}
          {net_bench_row("Bezhotovostní walk-in", format_num_round(rating_status["POCET_BEZHOT_WALK_IN"].mean()),  f"∑ {format_num_round(rating_status['POCET_BEZHOT_WALK_IN'].sum())}")}

          <tr><td colspan="3" style="background:#f0e6ff;font-weight:600;font-size:0.78rem;color:#6c3483;padding:4px 8px;text-align:left;">🌍 Cizinci</td></tr>
          {"".join(
            f"<tr><td style='text-align:left;'>{lbl}</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_num_round(rating_status[col].mean()) if col in rating_status.columns else '—'}</td>"
            f"<td style='text-align:right;color:#888;font-size:0.8rem;'>∑ {format_num_round(rating_status[col].sum()) if col in rating_status.columns else '—'}</td>"
            f"</tr>"
            for col, lbl in [('CIZINCI','Cizinci celkem'),('CIZINCI_SLOVENSKO','Cizinci SK'),('CIZINCI_UKRAJINA','Cizinci UA')]
          )}

          <tr><td colspan="3" style="background:#f0f4ff;font-weight:600;font-size:0.78rem;color:#2770f0;padding:4px 8px;text-align:left;">👥 Věkové skupiny klientů — průměrné počty a výnosy</td></tr>
          {"".join(
            f"<tr>"
            f"<td style='text-align:left;'>{g} — počet</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_num_round(rating_status[f'{g}_POCET_KLIENTU'].mean()) if f'{g}_POCET_KLIENTU' in rating_status.columns else '—'}</td>"
            f"<td style='text-align:right;color:#888;font-size:0.8rem;'>∑ {format_num_round(rating_status[f'{g}_POCET_KLIENTU'].sum()) if f'{g}_POCET_KLIENTU' in rating_status.columns else '—'}</td>"
            f"</tr>"
            f"<tr>"
            f"<td style='text-align:left;'>{g} — prům. výnos</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_money(rating_status[f'{g}_PRUMERNY_VYNOS'].mean()) if f'{g}_PRUMERNY_VYNOS' in rating_status.columns else '—'}</td>"
            f"<td></td>"
            f"</tr>"
            for g in AGE_GROUPS
          )}

          <tr><td colspan="3" style="background:#fff3cd;font-weight:600;font-size:0.78rem;color:#b8860b;padding:4px 8px;text-align:left;">💵 Transakční data — hotovostní operace (průměr na pobočku)</td></tr>
          {"".join(
            f"<tr style='background:#fffdf0;'>"
            f"<td style='text-align:left;padding-left:16px;'>{lbl}</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_money(rating_status[col].mean()) if col in rating_status.columns and col.endswith('_CASTKA') else format_num_round(rating_status[col].mean()) if col in rating_status.columns else '—'}</td>"
            f"<td style='text-align:right;color:#888;font-size:0.8rem;'>∑ {format_money(rating_status[col].sum()) if col in rating_status.columns and col.endswith('_CASTKA') else format_num_round(rating_status[col].sum()) if col in rating_status.columns else '—'}</td>"
            f"</tr>"
            for col, lbl in [
              ('CASH_IN_PREVODITELNA_CASTKA','Cash IN — převoditelné Kč'),
              ('CASH_IN_NEPREVODITELNA_CASTKA','Cash IN — nepřevoditelné Kč'),
              ('CASH_IN_PREVODITELNA_POCET','Cash IN — převoditelné počet'),
              ('CASH_IN_NEPREVODITELNA_POCET','Cash IN — nepřevoditelné počet'),
              ('CASH_OUT_PREVODITELNA_CASTKA','Cash OUT — převoditelné Kč'),
              ('CASH_OUT_NEPREVODITELNA_CASTKA','Cash OUT — nepřevoditelné Kč'),
              ('CASH_OUT_PREVODITELNA_POCET','Cash OUT — převoditelné počet'),
              ('CASH_OUT_NEPREVODITELNA_POCET','Cash OUT — nepřevoditelné počet'),
            ]
            if col in rating_status.columns
          )}

          {"" if "PLOCHA_POUZE_D5" not in rating_status.columns else f"""
          <tr><td colspan='3' style='background:#fce4ec;font-weight:600;font-size:0.78rem;color:#c2185b;padding:4px 8px;text-align:left;'>📐 Alokační report</td></tr>
          <tr><td style='text-align:left;'>Plocha pouze D5 (průměr)</td>
              <td class='benchmark-val' style='text-align:right;'>{rating_status['PLOCHA_POUZE_D5'].mean():.1f} m²</td>
              <td style='text-align:right;color:#888;font-size:0.8rem;'>∑ {rating_status['PLOCHA_POUZE_D5'].sum():.0f} m²</td></tr>
          <tr><td style='text-align:left;'>Celková plocha pobočky (průměr)</td>
              <td class='benchmark-val' style='text-align:right;'>{(rating_status['CELK_PLOCHA_POBOCKY_2026'].mean() if 'CELK_PLOCHA_POBOCKY_2026' in rating_status.columns else 0):.1f} m²</td>
              <td style='text-align:right;color:#888;font-size:0.8rem;'>∑ {(rating_status['CELK_PLOCHA_POBOCKY_2026'].sum() if 'CELK_PLOCHA_POBOCKY_2026' in rating_status.columns else 0):.0f} m²</td></tr>
          <tr><td style='text-align:left;'>Plocha nad optimální pobočku (průměr)</td>
              <td class='benchmark-val' style='text-align:right;'>{(rating_status['PLOCHA_NAD_OPTIMALNI_POBOCKU'].mean() if 'PLOCHA_NAD_OPTIMALNI_POBOCKU' in rating_status.columns else 0):.1f} m²</td>
              <td style='text-align:right;color:#888;font-size:0.8rem;'>∑ {(rating_status['PLOCHA_NAD_OPTIMALNI_POBOCKU'].sum() if 'PLOCHA_NAD_OPTIMALNI_POBOCKU' in rating_status.columns else 0):.0f} m²</td></tr>
          """}

        </tbody>
      </table>
    </div>
  </div>
</div>'''

        # ── Pre-výpočet pomalých bloků PŘED sestavením f-stringu ────────────
        print("  🚨 Počítám nebezpečnou zónu...")
        _danger_zone_html = generate_danger_zone_table(df_sorted)

        print("  🗺️  Generuji analýzu pokrytí ORP...")
        _orp_coverage_html = generate_orp_coverage_html(df_sorted)

        print("  👥 Generuji věkovou pyramidu...")
        _age_pyramid_html = generate_age_pyramid(df_sorted)

        # Věkový blok: celá síť + per region záložky
        _regions_list = sorted(df_sorted['REGION_NAME'].dropna().unique()) if 'REGION_NAME' in df_sorted.columns else []
        _age_tabs_html = ''
        for _rn in _regions_list:
            _df_r = df_sorted[df_sorted['REGION_NAME'] == _rn]
            _slug  = _rn.replace(' ', '_').replace('/', '_')
            _pyr   = generate_age_pyramid(_df_r)
            _age_tabs_html += (
                f'<div style="margin-bottom:20px;">'
                f'<div style="font-size:0.8rem;font-weight:700;color:#2770f0;'
                f'text-transform:uppercase;letter-spacing:.3px;margin-bottom:8px;">{_rn}</div>'
                f'{_pyr}</div>'
            )
        _age_combined_html = (
            f'<div style="margin-bottom:16px;">'
            f'<div style="font-size:0.8rem;font-weight:700;color:#0f172a;'
            f'text-transform:uppercase;letter-spacing:.3px;margin-bottom:8px;">Celá síť</div>'
            f'{_age_pyramid_html}</div>'
            + _age_tabs_html
            + f'<div class="age-section-card mb-4">{age_table_html}</div>'
        )

        print("  📊 Generuji kohortovou analýzu...")
        _cohort_html = generate_cohort_analysis(df_sorted)

        print("  📡 Generuji přehled indexu expozice...")
        _ie_top5_static_html = generate_ie_top5_html(df_sorted, "Top 5 — Index expozice (celá síť)", n=5)
        _ie_section_static = (
            '<div class="section-header">&#x1F4E1; Index expozice &mdash; top 5 cel&#225; s&#237;&#357;</div>'
            f'<div style="display:flex;gap:20px;flex-wrap:wrap;">{_ie_top5_static_html}</div>'
        ) if _ie_top5_static_html else ''

        print("  🏧 Generuji ATM přehled...")
        _atm_static_html = generate_atm_region_html(
            df_sorted,
            export_atm=globals().get('export_atm'),
            kapacita_atm=globals().get('kapacita_atm'),
        )
        _atm_static_section = (
            '<div class="section-header">&#x1F3E7; Bankomaty &mdash; celková s&#237;&#357;</div>'
            + _atm_static_html
        ) if _atm_static_html else ''

        print("  🔨 Generuji analýzu korelace investice → výkon...")
        _sp_arg = globals().get('spadovky')
        _inv_impact_html = generate_network_simulation_200(df_sorted, _sp_arg)

        print("  💎 Generuji heatmapu výnosových klientů...")
        _pf_arg = globals().get('parties_full')
        _rev_heatmap_html = generate_revenue_client_heatmap(df_sorted, _pf_arg)
        print("  ✅ Bloky připraveny, sestavuji HTML...")

        full_html = f'''
<div class="report-wrapper">
    {bootstrap_css}
    <div class="region-title">📊 Celkový přehled ratingů 2026 — celá síť</div>
    <div style="font-size:0.85rem;color:#444;margin:-8px 0 20px 0;padding:8px 16px;
                background:#f0f6ff;border-left:4px solid #2770f0;border-radius:0 6px 6px 0;
                display:inline-block;">
      📊 <strong>{len(rating_status)}</strong> poboček &nbsp;·&nbsp;
      DBS ke dni <strong>{DBS_DATE}</strong>
    </div>

    <!-- ═══ ČÁST 1: Vždy viditelné ═══ -->

    <!-- 1. Celkový přehled -->
    <div class="section-header">📋 Celkový přehled pobočkových ratingů</div>
    <p style="font-size:0.82rem;color:#666;margin:-6px 0 10px 0;">
      Kompletní tabulka všech poboček s ratingy, výnosy, C/I, klienty a dalšími metrikami.
      Seřazeno vzestupně dle <strong>Ratingu 25</strong> — nejlepší rating nahoře.
      Pomocí tlačítek nad tabulkou přepínáte viditelné skupiny sloupců.
    </p>
    {generate_filterable_table(df_sorted, "main-static-tbl")}

    <div style="margin-bottom:28px;"></div>

    <!-- 2. Obchody -->
    <div class="section-header">🛒 Přehled obchodů dle produktové skupiny — srovnání 2024 / 2025</div>
    <p style="font-size:0.82rem;color:#666;margin:-6px 0 10px 0;">
      Meziroční srovnání počtu obchodů a objemu nových výnosů dle produktové skupiny.
      Sloupce <b>Δ</b> ukazují změnu: <b style="color:#0ab33f;">zelená = růst</b>,
      <b style="color:#e64343;">červená = pokles</b>.
    </p>
    <div class="age-section-card mb-4">{obchody_html}</div>
    {obchody_bench_html}

    <div style="margin-bottom:20px;"></div>

    <!-- 3. Přehled poboček dle regionu -->
    <div class="section-header">🏢 Přehled poboček dle regionu — Typologie, Bezhotovostní provoz, Formát NF/SF</div>
    <p style="font-size:0.82rem;color:#666;margin:-6px 0 10px 0;">
      Počty poboček v jednotlivých regionech rozdělené dle typu budovy, cashless statusu a formátu NF/SF.
    </p>
    <div style="margin:0 0 20px 0;width:100%;">{branch_summary_html}</div>

    <!-- 3b. ORP pokrytí -->
    {make_collapsible("orp-static", "&#x1F5FA; Pokrytí ORP — analýza",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Přiřazení poboček k ORP (obce s rozšířenou působností) dle GPS souřadnic. Přehled pokrytí: ve kterých ORP pobočku máme a kde chybí.</p>'
        + _orp_coverage_html, default_open=False)}

    <!-- 4. Mapa -->
    {generate_map_html(df_sorted, title="🗺️ Mapa poboček — celá síť")}

    <hr class="report-divider">

    <!-- ═══ ČÁST 2: Rozbalovací sekce ═══ -->

    <!-- 5. Benchmark -->
    {make_collapsible("bench-static", "📊 Síťový benchmark",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Porovnání průměrných hodnot klíčových metrik napříč celou sítí — výnosy, C/I, FTE, návštěvy a další.</p>'
        + html_benchmark_static, default_open=False)}

    <!-- 6. Strategie -->
    {make_collapsible("strategy-static", "📋 Strategie poboček — Close · Investice · Cashless",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Přehled plánovaných uzavření, investic do nových formátů a přechodů na bezhotovostní provoz dle schváleného footprintu BNS.</p>'
        + render_strategy_columns(rating_status), default_open=False)}

    <!-- 7. Nejvyšší nové výnosy -->
    {make_collapsible("top-revenue-static", "💰 Nejvyšší nové výnosy",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Top 5 poboček s nejvyššími novými výnosy v roce 2025 — výnosy z nově uzavřených obchodů.</p>'
        + render_top5_cards(df_sorted, "revenue", "💰 Nejvyšší nové výnosy 2025", "seřazeno sestupně dle objemu nových výnosů", n=5),
        default_open=False)}

    <!-- 8. Top 5 zlepšení a zhoršení -->
    {make_collapsible("top-rating-static", "🚀⚠️ Top 5: Zlepšení a zhoršení ratingu", f"""
<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Pobočky s největším pohybem v ratingu oproti předchozímu roku — jak nahoru (zlepšení), tak dolů (zhoršení).</p>
<div style="display:flex;gap:20px;flex-wrap:wrap;align-items:flex-start;">
  <div style="flex:1;min-width:280px;">{render_top5_cards(df_sorted, 'improve', '🚀 Největší zlepšení ratingu', 'sestupně dle posunu pořadí nahoru')}</div>
  <div style="flex:1;min-width:280px;">{render_top5_cards(df_sorted, 'worsen', '⚠️ Největší zhoršení ratingu', 'sestupně dle posunu pořadí dolů')}</div>
</div>""", default_open=False)}

    <!-- 9. Věkové segmenty + pyramida (sloučené) -->
    {make_collapsible("age-static", "👥 Věková struktura klientů",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Rozložení klientů dle věkových skupin a průměrné výnosy — celá síť i po regionech. Věková pyramida ukazuje strukturu dle pohlaví.</p>"
        + _age_combined_html,
        default_open=False)}

    <!-- 10. Top investice -->
    {make_collapsible("top-inv-static", "🔨 Top investice — celá síť", f"""
<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Přehled největších a nejnovějších investic do pobočkové sítě — rekonstrukce, nové formáty, technologie.</p>
<div style="display:flex;gap:20px;flex-wrap:wrap;align-items:flex-start;">
  <div style="flex:1;min-width:280px;">{render_top5_cards(rating_status, 'inv_new', '🕐 Nejnovější investice', 'seřazeno dle data realizace')}</div>
  <div style="flex:1;min-width:280px;">{render_top5_cards(rating_status, 'inv_big', '💰 Největší investice', 'seřazeno dle výše investice')}</div>
</div>""", default_open=False)}

    <!-- 11. Nebezpečná zóna -->
    {make_collapsible("danger-zone-static", "🚨 Pobočky v nebezpečné zóně",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Pobočky s kombinací špatného ratingu, vysokého C/I a nízkých výnosů — vyžadují zvýšenou pozornost.</p>'
        + _danger_zone_html, default_open=False)}

    <!-- 13. Kohortová analýza -->
    {make_collapsible("cohort-static", "📊 Kohortová analýza výnosů",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Průměrné nové výnosy a výnos/FTE seskupené dle formátu, velikosti, ratingového kvintilu, nájemného a plochy — odpovídá na otázku: co nejvíce ovlivňuje výkonnost pobočky?</p>'
        + _cohort_html, default_open=False)}

    <!-- 14. ATM přehled -->
    {make_collapsible("atm-static", "🏧 Bankomaty — celková síť",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Přehled bankomatů v síti, kapacitní vytížení, počty výběrů a vkladů — aktuálně i po plánovaném převodu hotovostních operací.</p>'
        + (_atm_static_html if _atm_static_html else '<p style="color:#aaa;font-style:italic;">Data ATM nejsou k dispozici.</p>'),
        default_open=False)}

    <!-- 15. Index expozice -->
    {make_collapsible("ie-static", "📡 Index expozice — top 5 celá síť",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Index expozice vyjadřuje míru „viditelnosti" pobočky vůči potenciálním klientům — kombinace demografické dostupnosti, konkurence a spádu. Q1 = nejlepší pozice.</p>'
        + (f'<div style="display:flex;gap:20px;flex-wrap:wrap;">{_ie_top5_static_html}</div>' if _ie_top5_static_html else ''),
        default_open=False)}

    <!-- 16. Kde bydlí klienti -->
    {make_collapsible("revenue-heatmap-static", "💎 Kde bydlí klienti s nejvyššími výnosy — top 20 % poboček",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Geografická heatmapa bydliště klientů generujících nejvyšší výnosy — identifikuje spádové oblasti a potenciál pro rozvoj sítě.</p>'
        + _rev_heatmap_html, default_open=False)}

    <!-- 17. Simulace sítě -->
    {make_collapsible("inv-impact-static", "🏦 Simulace sítě 250 poboček",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Simulace optimalizované sítě 250 poboček pro rok 2030 — výnosy, C/I, bankéřská kapacita a přesun klientů ze zavřených poboček.</p>'
        + _inv_impact_html, default_open=False)}

    <!-- 18. Obsazení pozic -->
    {make_collapsible("spec-static", "👷 Přehled obsazení pozic",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Aktuální obsazení bankéřských a specialistických pozic v pobočkové síti dle regionu.</p>'
        + generate_specialiste_summary_table(df_sorted["BRANCH_CODE"].dropna().astype(int).tolist(), title=""),
        default_open=False)}

    <!-- 19. Výkonnostní metriky -->
    {make_collapsible("metrics-static", "📊 Přehled výkonnostních metrik",
        '<p style="font-size:0.82rem;color:#666;margin:0 0 12px 0;">Detailní přehled výkonnostních metrik poboček — business rating, schůzky/FTE, PEREX a další ukazatele obchodní aktivity.</p>'
        + generate_metrics_overview(df_sorted, title=""),
        default_open=False)}

    <hr class="report-divider">

    <!-- ═══ ČÁST 3: Metodika ═══ -->
    {html_legenda}
</div>'''

        filename = f"{output_prefix}_staticky.html"
        with open(filename, "w", encoding="utf-8") as f:
            f.write(full_html)
        print(f"✅ Uložen: {filename}")

        xlsx_filename = f"{output_prefix}_staticky.xlsx"
        df_excel = prepare_excel_df(df_sorted)

        with pd.ExcelWriter(xlsx_filename, engine="openpyxl") as writer:
            df_excel.to_excel(writer, sheet_name="Celkový přehled", index=False)
            ws = writer.sheets["Celkový přehled"]
            write_excel_sheet(ws, df_excel, freeze="D2")

        print(f"✅ Uložen: {xlsx_filename}")

        # Export PKL pro další použití
        pkl_filename = f"{output_prefix}_staticky.pkl"
        df_sorted.to_pickle(pkl_filename)
        print(f"✅ Uložen: {pkl_filename}")

    elif mode == 'regional':
        for region in sorted(rating_status['REGION_NAME'].unique()):
            df_reg    = rating_status[rating_status['REGION_NAME'] == region].copy()
            df_others = rating_status[rating_status['REGION_NAME'] != region].copy()

            region_slug  = region.replace(' ', '-').replace('/', '-')
            chart_id     = f"apex-{region_slug}"
            age_chart_id = f"apex-age-{region_slug}"

            chart_html        = generate_apex_chart_html(df_reg, chart_id)
            age_table_html_raw = generate_age_segment_table(df_reg.sort_values(by="BRANCH_CODE"), table_id=f"age-tbl-{region_slug}")
            age_chart_html    = generate_age_apex_chart(df_reg, age_chart_id)

            obchody_avail_cols = (
                ['BRANCH_CODE'] +
                [c for c in obchody_detail.columns if c != 'BRANCH_CODE']
            )
            obchody_avail_cols = [c for c in obchody_avail_cols if c in obchody_detail.columns]
            _fmt_reg_avail = [c for c in ['BRANCH_FORMAT', 'BRANCH_TYPE_NAME'] if c in df_reg.columns]
            _obch_reg_base = ['BRANCH_CODE', 'BRANCH_NAME'] + _fmt_reg_avail
            df_obchody_reg = df_reg[_obch_reg_base].merge(
                obchody_detail[obchody_avail_cols], on='BRANCH_CODE', how='left'
            ).sort_values(by="BRANCH_CODE")
            obchody_html_reg_raw = generate_obchody_table(
                df_obchody_reg, table_id=f"obchody-{region_slug}"
            )
            age_table_sortable    = age_table_html_raw
            obchody_html_sortable = generate_sortable_table(obchody_html_reg_raw, f"obchody-sort-{region_slug}")
            obchody_bench_html_reg = generate_obchody_benchmark_chart(
                df_obchody_reg, chart_id=f"obch-bench-{region_slug}"
            )

            # Pre-výpočet věkové pyramidy a kohortové analýzy pro region
            _reg_age_pyramid_html = generate_age_pyramid(df_reg)
            _reg_cohort_html      = generate_cohort_analysis(df_reg)
            _reg_atm_html         = generate_atm_region_html(
                df_reg,
                export_atm=globals().get('export_atm'),
                kapacita_atm=globals().get('kapacita_atm'),
            )

            # Index expozice top 5 pro region
            _ie_top5_reg = generate_ie_top5_html(df_reg, f"Top 5 — Index expozice ({region})")
            _ie_section_reg = (
                f'<div class="section-header">&#x1F4E1; Index expozice &mdash; top 5</div>'
                f'<div style="display:flex;gap:20px;flex-wrap:wrap;">{_ie_top5_reg}</div>'
                f'<div style="margin-bottom:20px;"></div>'
            ) if _ie_top5_reg else ''
            _atm_hdr = f'<div class="section-header">&#x1F3E7; Bankomaty &mdash; {region}</div>'
            _reg_atm_section_html = (
                _atm_hdr + _reg_atm_html + '<div style="margin-bottom:28px;"></div>'
                if _reg_atm_html else ''
            )

            # ORP pokrytí pro region
            _reg_orp_html = generate_orp_coverage_html(df_reg)

            branch_summary_html_reg  = generate_branch_summary_tables(df_reg)
            close_branches_html_reg       = generate_close_branches_table(df_reg)
            invest_branches_html_reg      = generate_invest_branches_table(df_reg)
            cashierless_branches_html_reg = generate_cashierless_table(df_reg)

            def bench_diff(val_reg, val_others, mode='num'):
                diff = val_reg - val_others
                return get_change_html(diff, mode=mode)

            # *** NOVÉ: řádky hotovostních operací pro benchmark ***
            def cash_bench_rows(df_r, df_o):
                """Generuje HTML řádky tabulky benchmarku pro hotovostní transakce."""
                rows = []
                label_map = {
                    'CASH_IN_PREVODITELNA_CASTKA':    'Cash IN — převoditelné Kč',
                    'CASH_IN_NEPREVODITELNA_CASTKA':  'Cash IN — nepřevoditelné Kč',
                    'CASH_IN_PREVODITELNA_POCET':     'Cash IN — převoditelné počet',
                    'CASH_IN_NEPREVODITELNA_POCET':   'Cash IN — nepřevoditelné počet',
                    'CASH_OUT_PREVODITELNA_CASTKA':   'Cash OUT — převoditelné Kč',
                    'CASH_OUT_NEPREVODITELNA_CASTKA': 'Cash OUT — nepřevoditelné Kč',
                    'CASH_OUT_PREVODITELNA_POCET':    'Cash OUT — převoditelné počet',
                    'CASH_OUT_NEPREVODITELNA_POCET':  'Cash OUT — nepřevoditelné počet',
                }
                is_castka = {k: k.endswith('_CASTKA') for k in label_map}

                rows.append(
                    "<tr><td colspan='4' style='background:#fff3cd;font-weight:600;font-size:0.78rem;"
                    "color:#b8860b;padding:4px 8px;text-align:left;'>💵 Transakční data — hotovostní operace (průměr na pobočku)</td></tr>"
                )
                for col, lbl in label_map.items():
                    if col not in df_r.columns:
                        continue
                    val_r = df_r[col].mean() if col in df_r.columns else 0
                    val_o = df_o[col].mean() if col in df_o.columns else 0
                    if is_castka[col]:
                        fmt_r = format_money(val_r)
                        fmt_o = format_money(val_o)
                        diff_html = bench_diff(val_r, val_o, mode='money')
                    else:
                        fmt_r = format_num_round(val_r)
                        fmt_o = format_num_round(val_o)
                        diff_html = bench_diff(val_r, val_o, mode='num')
                    rows.append(
                        f"<tr style='background:#fffdf0;'>"
                        f"<td style='text-align:left;padding-left:16px;'>{lbl}</td>"
                        f"<td class='benchmark-val' style='text-align:right;'>{fmt_r}</td>"
                        f"<td style='text-align:right;'>{fmt_o}</td>"
                        f"<td style='text-align:right;'>{diff_html}</td>"
                        f"</tr>"
                    )
                return "\n".join(rows)

            html_benchmark = f'''
<style>
  .benchmark-hover-tbl tbody tr:hover td {{ background-color: #d1fae5 !important; }}
</style>
<div class="age-section-card mb-4">
  <div class="row align-items-start">
    <div class="col-md-8">
      <table class="table table-sm mb-0 benchmark-hover-tbl" style="font-size:0.83rem;">
        <thead class="table-dark">
          <tr>
            <th style="text-align:left;">Ukazatel</th>
            <th style="color:#7ec8f7;text-align:right;">Region&nbsp;{region}</th>
            <th style="text-align:right;">Ostatní regiony</th>
            <th style="text-align:right;">Rozdíl</th>
          </tr>
        </thead>
        <tbody>
          <tr><td colspan="4" style="background:#f0f4ff;font-weight:600;font-size:0.78rem;color:#2770f0;padding:4px 8px;text-align:left;">💰 Finanční ukazatele</td></tr>
          <tr>
            <td style="text-align:left;">Výnosy 25</td>
            <td class="benchmark-val" style="text-align:right;">{format_money(df_reg["VYNOSY"].mean())}</td>
            <td style="text-align:right;">{format_money(df_others["VYNOSY"].mean())}</td>
            <td style="text-align:right;">{bench_diff(df_reg["VYNOSY"].mean(), df_others["VYNOSY"].mean(), mode="money")}</td>
          </tr>
          <tr>
            <td style="text-align:left;">Nové výnosy 25</td>
            <td class="benchmark-val" style="text-align:right;">{format_money(df_reg["OBJEM_VYNOSU_CZK"].mean())}</td>
            <td style="text-align:right;">{format_money(df_others["OBJEM_VYNOSU_CZK"].mean())}</td>
            <td style="text-align:right;">{bench_diff(df_reg["OBJEM_VYNOSU_CZK"].mean(), df_others["OBJEM_VYNOSU_CZK"].mean(), mode="money")}</td>
          </tr>
          <tr>
            <td style="text-align:left;">C/I Ratio 25</td>
            <td class="benchmark-val" style="text-align:right;">{format_pct(df_reg["PRIME_NAKLADY/VYNOSY"].mean())}</td>
            <td style="text-align:right;">{format_pct(df_others["PRIME_NAKLADY/VYNOSY"].mean())}</td>
            <td style="text-align:right;">{bench_diff(df_reg["PRIME_NAKLADY/VYNOSY"].mean(), df_others["PRIME_NAKLADY/VYNOSY"].mean(), mode="pct")}</td>
          </tr>

          <tr><td colspan="4" style="background:#e8f5e9;font-weight:600;font-size:0.78rem;color:#1a7a4a;padding:4px 8px;text-align:left;">🛒 Obchody dle produktů — průměrné výnosy a počty prodejů 2025</td></tr>
          {"".join(
            f"<tr style='background:#f9fffe;'>"
            f"<td style='text-align:left;padding-left:16px;'>{label} — výnosy 25</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_money(df_reg[f'OBJEM_VYNOSU_CZK_{key}_2025'].mean()) if f'OBJEM_VYNOSU_CZK_{key}_2025' in df_reg.columns else '—'}</td>"
            f"<td style='text-align:right;'>{format_money(df_others[f'OBJEM_VYNOSU_CZK_{key}_2025'].mean()) if f'OBJEM_VYNOSU_CZK_{key}_2025' in df_others.columns else '—'}</td>"
            f"<td style='text-align:right;'>{bench_diff(df_reg[f'OBJEM_VYNOSU_CZK_{key}_2025'].mean(), df_others[f'OBJEM_VYNOSU_CZK_{key}_2025'].mean(), mode='money') if f'OBJEM_VYNOSU_CZK_{key}_2025' in df_reg.columns and f'OBJEM_VYNOSU_CZK_{key}_2025' in df_others.columns else '—'}</td>"
            f"</tr>"
            f"<tr style='background:#f5fffc;'>"
            f"<td style='text-align:left;padding-left:16px;'>{label} — počet prodejů 25</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_num_round(df_reg[f'POCET_PRODEJU_{key}_2025'].mean()) if f'POCET_PRODEJU_{key}_2025' in df_reg.columns else '—'}</td>"
            f"<td style='text-align:right;'>{format_num_round(df_others[f'POCET_PRODEJU_{key}_2025'].mean()) if f'POCET_PRODEJU_{key}_2025' in df_others.columns else '—'}</td>"
            f"<td style='text-align:right;'>{bench_diff(df_reg[f'POCET_PRODEJU_{key}_2025'].mean(), df_others[f'POCET_PRODEJU_{key}_2025'].mean(), mode='num') if f'POCET_PRODEJU_{key}_2025' in df_reg.columns and f'POCET_PRODEJU_{key}_2025' in df_others.columns else '—'}</td>"
            f"</tr>"
            for key, label in PRODUCT_GROUPS
            if f'OBJEM_VYNOSU_CZK_{key}_2025' in df_reg.columns or f'POCET_PRODEJU_{key}_2025' in df_reg.columns
          )}

          <tr><td colspan="4" style="background:#f0f4ff;font-weight:600;font-size:0.78rem;color:#2770f0;padding:4px 8px;text-align:left;">🚶 Návštěvnost</td></tr>
          <tr>
            <td style="text-align:left;">Návštěvy celkem</td>
            <td class="benchmark-val" style="text-align:right;">{format_num_round(df_reg["POCET_NAVSTEV_CELKEM"].mean())}</td>
            <td style="text-align:right;">{format_num_round(df_others["POCET_NAVSTEV_CELKEM"].mean())}</td>
            <td style="text-align:right;">{bench_diff(df_reg["POCET_NAVSTEV_CELKEM"].mean(), df_others["POCET_NAVSTEV_CELKEM"].mean(), mode="num")}</td>
          </tr>
          <tr>
            <td style="text-align:left;">Hotovostní walk-in</td>
            <td class="benchmark-val" style="text-align:right;">{format_num_round(df_reg["POCET_HOT_WALK_IN"].mean())}</td>
            <td style="text-align:right;">{format_num_round(df_others["POCET_HOT_WALK_IN"].mean())}</td>
            <td style="text-align:right;">{bench_diff(df_reg["POCET_HOT_WALK_IN"].mean(), df_others["POCET_HOT_WALK_IN"].mean(), mode="num")}</td>
          </tr>
          <tr>
            <td style="text-align:left;">Bezhotovostní walk-in</td>
            <td class="benchmark-val" style="text-align:right;">{format_num_round(df_reg["POCET_BEZHOT_WALK_IN"].mean())}</td>
            <td style="text-align:right;">{format_num_round(df_others["POCET_BEZHOT_WALK_IN"].mean())}</td>
            <td style="text-align:right;">{bench_diff(df_reg["POCET_BEZHOT_WALK_IN"].mean(), df_others["POCET_BEZHOT_WALK_IN"].mean(), mode="num")}</td>
          </tr>

          <tr><td colspan="4" style="background:#f0e6ff;font-weight:600;font-size:0.78rem;color:#6c3483;padding:4px 8px;text-align:left;">🌍 Cizinci</td></tr>
          {"".join(
            f"<tr><td style='text-align:left;'>{lbl}</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_num_round(df_reg[col].mean()) if col in df_reg.columns else '—'}</td>"
            f"<td style='text-align:right;'>{format_num_round(df_others[col].mean()) if col in df_others.columns else '—'}</td>"
            f"<td style='text-align:right;'>{bench_diff(df_reg[col].mean(), df_others[col].mean(), mode='num') if col in df_reg.columns and col in df_others.columns else '—'}</td>"
            f"</tr>"
            for col, lbl in [('CIZINCI','Cizinci celkem'),('CIZINCI_SLOVENSKO','Cizinci SK'),('CIZINCI_UKRAJINA','Cizinci UA')]
          )}

          <tr><td colspan="4" style="background:#f0f4ff;font-weight:600;font-size:0.78rem;color:#2770f0;padding:4px 8px;text-align:left;">👥 Věkové skupiny klientů — průměrné počty a výnosy</td></tr>
          {"".join(
            f"<tr>"
            f"<td style='text-align:left;'>{g} — počet</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_num_round(df_reg[f'{g}_POCET_KLIENTU'].mean()) if f'{g}_POCET_KLIENTU' in df_reg.columns else '—'}</td>"
            f"<td style='text-align:right;'>{format_num_round(df_others[f'{g}_POCET_KLIENTU'].mean()) if f'{g}_POCET_KLIENTU' in df_others.columns else '—'}</td>"
            f"<td style='text-align:right;'>{bench_diff(df_reg[f'{g}_POCET_KLIENTU'].mean(), df_others[f'{g}_POCET_KLIENTU'].mean(), mode='num') if f'{g}_POCET_KLIENTU' in df_reg.columns and f'{g}_POCET_KLIENTU' in df_others.columns else '—'}</td>"
            f"</tr>"
            f"<tr>"
            f"<td style='text-align:left;'>{g} — prům. výnos</td>"
            f"<td class='benchmark-val' style='text-align:right;'>{format_money(df_reg[f'{g}_PRUMERNY_VYNOS'].mean()) if f'{g}_PRUMERNY_VYNOS' in df_reg.columns else '—'}</td>"
            f"<td style='text-align:right;'>{format_money(df_others[f'{g}_PRUMERNY_VYNOS'].mean()) if f'{g}_PRUMERNY_VYNOS' in df_others.columns else '—'}</td>"
            f"<td style='text-align:right;'>{bench_diff(df_reg[f'{g}_PRUMERNY_VYNOS'].mean(), df_others[f'{g}_PRUMERNY_VYNOS'].mean(), mode='money') if f'{g}_PRUMERNY_VYNOS' in df_reg.columns and f'{g}_PRUMERNY_VYNOS' in df_others.columns else '—'}</td>"
            f"</tr>"
            for g in AGE_GROUPS
          )}

          {cash_bench_rows(df_reg, df_others)}

          {"" if "PLOCHA_POUZE_D5" in df_reg.columns else f"""
          <tr><td colspan='4' style='background:#fce4ec;font-weight:600;font-size:0.78rem;color:#c2185b;padding:4px 8px;text-align:left;'>📐 Alokační report</td></tr>
          <tr style='background:#fff5f8;'>
            <td style='text-align:left;'>Plocha pouze D5 (průměr)</td>
            <td class='benchmark-val' style='text-align:right;'>{float(df_reg['PLOCHA_POUZE_D5'].mean()):.1f} m²</td>
            <td style='text-align:right;'>{float(df_others['PLOCHA_POUZE_D5'].mean()) if 'PLOCHA_POUZE_D5' in df_others.columns else 0:.1f} m²</td>
            <td style='text-align:right;'>{bench_diff(df_reg['PLOCHA_POUZE_D5'].mean(), df_others['PLOCHA_POUZE_D5'].mean() if 'PLOCHA_POUZE_D5' in df_others.columns else 0, mode='num')}</td>
          </tr>
          <tr style='background:#fff5f8;'>
            <td style='text-align:left;'>Celková plocha pobočky (průměr)</td>
            <td class='benchmark-val' style='text-align:right;'>{float(df_reg['CELK_PLOCHA_POBOCKY_2026'].mean()):.1f} m²</td>
            <td style='text-align:right;'>{float(df_others['CELK_PLOCHA_POBOCKY_2026'].mean()) if 'CELK_PLOCHA_POBOCKY_2026' in df_others.columns else 0:.1f} m²</td>
            <td style='text-align:right;'>{bench_diff(df_reg['CELK_PLOCHA_POBOCKY_2026'].mean(), df_others['CELK_PLOCHA_POBOCKY_2026'].mean() if 'CELK_PLOCHA_POBOCKY_2026' in df_others.columns else 0, mode='num')}</td>
          </tr>
          <tr style='background:#fff5f8;'>
            <td style='text-align:left;'>Plocha nad optimální pobočku (průměr)</td>
            <td class='benchmark-val' style='text-align:right;'>{float(df_reg['PLOCHA_NAD_OPTIMALNI_POBOCKU'].mean()):.1f} m²</td>
            <td style='text-align:right;'>{float(df_others['PLOCHA_NAD_OPTIMALNI_POBOCKU'].mean()) if 'PLOCHA_NAD_OPTIMALNI_POBOCKU' in df_others.columns else 0:.1f} m²</td>
            <td style='text-align:right;'>{bench_diff(df_reg['PLOCHA_NAD_OPTIMALNI_POBOCKU'].mean(), df_others['PLOCHA_NAD_OPTIMALNI_POBOCKU'].mean() if 'PLOCHA_NAD_OPTIMALNI_POBOCKU' in df_others.columns else 0, mode='num')}</td>
          </tr>"""}

        </tbody>
      </table>
    </div>
  </div>
</div>'''

            full_html = f'''
<div class="report-wrapper">
    {bootstrap_css}
    <div class="region-title">🗺️ Regionální report 2026 — {region}</div>
    <div style="font-size:0.85rem;color:#444;margin:-8px 0 20px 0;padding:8px 16px;
                background:#f0f6ff;border-left:4px solid #2770f0;border-radius:0 6px 6px 0;
                display:inline-block;">
      📊 <strong>{len(df_reg)}</strong> poboček v regionu &nbsp;·&nbsp;
      DBS ke dni <strong>{DBS_DATE}</strong>
    </div>

    <!-- ═══ ČÁST 1: Vždy viditelné ═══ -->

    <!-- 1. Celkový přehled -->
    <div class="section-header">📋 Celkový přehled pobočkových ratingů</div>
    <p style="font-size:0.82rem;color:#666;margin:-6px 0 10px 0;">
      Kompletní tabulka poboček regionu <strong>{region}</strong> s ratingy, výnosy, C/I a dalšími metrikami.
      Seřazeno vzestupně dle <strong>Ratingu 25</strong> — nejlepší rating nahoře.
    </p>
    {generate_filterable_table(df_reg.sort_values(by="IR"), f"main-table-{region_slug}",
        excluded_cols=['ROCNI_SPLATKY_S_DPH_CZK'])}

    <div style="margin-bottom:20px;"></div>

    <!-- 2. Obchody -->
    <div class="section-header">🛒 Přehled obchodů dle produktové skupiny — srovnání 2024 / 2025</div>
    <p style="font-size:0.82rem;color:#666;margin:-6px 0 10px 0;">
      Meziroční srovnání počtu obchodů a objemu nových výnosů dle produktové skupiny v regionu {region}.
      Sloupce <b>Δ</b>: <b style="color:#0ab33f;">zelená = růst</b>, <b style="color:#e64343;">červená = pokles</b>.
    </p>
    <div class="age-section-card mb-4">{obchody_html_sortable}</div>
    {obchody_bench_html_reg}

    <div style="margin-bottom:20px;"></div>

    <!-- 3. Přehled poboček dle typologie -->
    <div class="section-header">🏢 Přehled poboček regionu — Typologie, Bezhotovostní provoz, Formát NF/SF</div>
    <p style="font-size:0.82rem;color:#666;margin:-6px 0 10px 0;">
      Počty poboček v regionu {region} rozdělené dle typu budovy, cashless statusu a formátu NF/SF.
    </p>
    <div style="margin:0 0 20px 0;width:100%;">{branch_summary_html_reg}</div>

    <!-- 4. Mapa -->
    {generate_map_html(df_reg.sort_values(by="IR"), title=f"🗺️ Mapa poboček — {region}")}

    <hr class="report-divider">

    <!-- ═══ ČÁST 2: Rozbalovací sekce ═══ -->

    <!-- 5. Benchmark -->
    {make_collapsible(f"bench-{region_slug}", "📊 Regionální benchmark",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Porovnání průměrných hodnot klíčových metrik regionu {region} vůči zbytku sítě — výnosy, C/I, FTE, návštěvy.</p>"
        + html_benchmark, default_open=False)}

    <!-- 6. Strategie -->
    {make_collapsible(f"strategy-{region_slug}", "📋 Strategie poboček — Close · Investice · Cashless",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Přehled plánovaných uzavření, investic a přechodů na bezhotovostní provoz v regionu {region}.</p>"
        + render_strategy_columns(df_reg), default_open=False)}

    <!-- 7. Nejvyšší výnosy a nové výnosy -->
    {make_collapsible(f"top-revenue-{region_slug}", "💰 Přehled oblastí s nejvyššími výnosy",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Top pobočky regionu {region} dle celkových výnosů (VYNOSY) a nových výnosů (OBJEM_VYNOSU_CZK) v roce 2025.</p>"
        + _render_top_revenues_table(df_reg, region),
        default_open=False)}

    <!-- 8. Top 5 zlepšení a zhoršení -->
    {make_collapsible(f"top-rating-{region_slug}", "🚀⚠️ Top 5: Zlepšení a zhoršení ratingu", f"""
<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Pobočky regionu {region} s největším pohybem v ratingu oproti předchozímu roku.</p>
<div style="display:flex;gap:20px;flex-wrap:wrap;align-items:flex-start;">
  <div style="flex:1;min-width:280px;">{render_top5_cards(df_reg, 'improve', '🚀 Největší zlepšení ratingu', 'sestupně dle posunu pořadí nahoru')}</div>
  <div style="flex:1;min-width:280px;">{render_top5_cards(df_reg, 'worsen', '⚠️ Největší zhoršení ratingu', 'sestupně dle posunu pořadí dolů')}</div>
</div>""", default_open=False)}

    <!-- 9. Věkové segmenty -->
    {make_collapsible(f"age-{region_slug}", "👥 Věkové segmenty klientů",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Rozložení klientů regionu {region} dle věkových skupin a průměrné výnosy na klienta v každém segmentu.</p><div class='age-section-card mb-4'>{age_table_sortable}</div>",
        default_open=False)}

    <!-- 10. Top investice -->
    {make_collapsible(f"top-inv-{region_slug}", f"🔨 Top investice — {region}", f"""
<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Největší a nejnovější investice do poboček v regionu {region}.</p>
<div style="display:flex;gap:20px;flex-wrap:wrap;align-items:flex-start;">
  <div style="flex:1;min-width:280px;">{render_top5_cards(df_reg, 'inv_new', '🕐 Nejnovější investice', 'seřazeno dle data realizace')}</div>
  <div style="flex:1;min-width:280px;">{render_top5_cards(df_reg, 'inv_big', '💰 Největší investice', 'seřazeno dle výše investice')}</div>
</div>""", default_open=False)}

    <!-- 11. Věková pyramida -->
    {make_collapsible(f"age-pyramid-{region_slug}", f"👥 Věková pyramida klientů — {region}",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Věková struktura klientů regionu {region} — rozložení dle pohlaví a věkových skupin.</p>"
        + _reg_age_pyramid_html, default_open=False)}

    <!-- 12. Kohortová analýza -->
    {make_collapsible(f"cohort-{region_slug}", f"📊 Kohortová analýza výnosů — {region}",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Průměrné nové výnosy a výnos/FTE seskupené dle formátu, velikosti a ratingového kvintilu pro region {region}.</p>"
        + _reg_cohort_html, default_open=False)}

    <!-- 13. ATM přehled -->
    {make_collapsible(f"atm-{region_slug}", f"🏧 Bankomaty — {region}",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Přehled bankomatů v regionu {region}, kapacitní vytížení a počty výběrů a vkladů.</p>"
        + (_reg_atm_html if _reg_atm_html else '<p style="color:#aaa;font-style:italic;">Data ATM nejsou k dispozici.</p>'),
        default_open=False)}

    <!-- 14. Index expozice -->
    {make_collapsible(f"ie-{region_slug}", f"📡 Index expozice — top 5 {region}",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Top 5 poboček regionu {region} s nejlepším indexem expozice — demografická dostupnost, konkurence a spád. Q1 = nejlepší.</p>"
        + (f'<div style="display:flex;gap:20px;flex-wrap:wrap;">{_ie_top5_reg}</div>' if _ie_top5_reg else ''),
        default_open=False)}

    <!-- 15. Obsazení pozic -->
    {make_collapsible(f"spec-{region_slug}", f"👷 Obsazení pozic — {region}",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Aktuální obsazení bankéřských a specialistických pozic v pobočkách regionu {region}.</p>"
        + generate_specialiste_summary_table(df_reg["BRANCH_CODE"].dropna().astype(int).tolist(), title=""),
        default_open=False)}

    <!-- 16. Výkonnostní metriky -->
    {make_collapsible(f"metrics-{region_slug}", f"📊 Výkonnostní metriky — {region}",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Detailní přehled výkonnostních metrik poboček regionu {region} — business rating, schůzky/FTE, PEREX a další ukazatele obchodní aktivity.</p>"
        + generate_metrics_overview(df_reg, title=""),
        default_open=False)}

    <!-- 17. ORP pokrytí regionu -->
    {make_collapsible(f"orp-{region_slug}", f"&#x1F5FA; Pokrytí ORP — {region}",
        f"<p style='font-size:0.82rem;color:#666;margin:0 0 12px 0;'>Přiřazení poboček regionu {region} k ORP dle GPS souřadnic. Přepínač: všechny pobočky / pouze Keep (bez plánovaných uzavření).</p>"
        + _reg_orp_html,
        default_open=False)}

    <hr class="report-divider">

    <!-- ═══ ČÁST 3: Metodika ═══ -->
    {html_legenda}
</div>'''

            filename = f"{output_prefix}_{region.replace(' ', '_')}.html"
            with open(filename, "w", encoding="utf-8") as f:
                f.write(full_html)
            print(f"✅ Uložen: {filename}")

            # Excel export pro region (bez výše nájemného)
            _reg_xlsx = f"{output_prefix}_{region.replace(' ', '_')}.xlsx"
            _df_reg_excel = prepare_excel_df(
                df_reg, excluded_cols=['ROCNI_SPLATKY_S_DPH_CZK']
            )
            try:
                with pd.ExcelWriter(_reg_xlsx, engine="openpyxl") as _writer:
                    _sheet_name = region[:31]
                    _df_reg_excel.to_excel(_writer, sheet_name=_sheet_name, index=False)
                    _ws = _writer.sheets[_sheet_name]
                    write_excel_sheet(_ws, _df_reg_excel, freeze="D2")
                print(f"✅ Uložen: {_reg_xlsx}")
            except Exception as _xlsx_err:
                print(f"  ⚠️  Excel pro region {region} selhal: {_xlsx_err}")


# =============================================================================
# 1. NAČTENÍ DAT
# =============================================================================

pd.set_option('display.max_columns', 500)
validities = {}

_load_log = []   # záznamy pro souhrnnou tabulku

for jmeno, info in soubory.items():
    try:
        cesta = info["path"]
        validity = info.get("validity", "")
        source_desc = info.get("source_desc", "")
        all_params = info.get("extra_params", {}).copy()
        target_date = all_params.pop("filter_date", None)

        if cesta.endswith(".csv"):
            if 'usecols' in all_params:
                try:
                    df = pd.read_csv(cesta, **all_params)
                except ValueError as _uce:
                    _params_fallback = {k: v for k, v in all_params.items() if k != 'usecols'}
                    df = pd.read_csv(cesta, **_params_fallback)
                    _load_log.append({"jmeno": jmeno, "status": "⚠️", "pozn": f"usecols selhal, načteno bez omezení sloupců"})
            else:
                df = pd.read_csv(cesta, **all_params)
        elif cesta.endswith(".pkl"):
            df = pd.read_pickle(cesta)
        else:
            df = pd.read_excel(cesta, **all_params)

        if not df.empty:
            df = normalize_columns(df)

            if jmeno == "revenues":
                # Pokud CSV obsahuje sloupec ROK nebo YEAR, filtruj jen data za rok 2025
                _rok_col = next((c for c in df.columns if c in ('ROK', 'YEAR', 'YEAR_')), None)
                if _rok_col:
                    df[_rok_col] = pd.to_numeric(df[_rok_col], errors='coerce')
                    _avail_years = sorted(df[_rok_col].dropna().unique())
                    _target_year = 2025 if 2025 in _avail_years else (max(_avail_years) if _avail_years else None)
                    if _target_year is not None:
                        df = df[df[_rok_col] == _target_year].copy()
                df = df.groupby('BRANCH_CODE').agg(
                    POCET_PRODEJU=('POCET_PRODEJU', 'sum'),
                    OBJEM_VYNOSU_CZK=('OBJEM_VYNOSU_CZK', 'sum'),
                    BRANCH_NAME=('BRANCH_NAME', 'first')
                ).reset_index()
                branch_fix = {451: 'Praha 1 - Národní', 515: 'Praha 9 - OC Galerie Harfa'}
                df['BRANCH_NAME'] = df.apply(lambda x: branch_fix.get(x['BRANCH_CODE'], x['BRANCH_NAME']), axis=1)

            filter_note = ""
            if target_date and "EFFECTIVE_DATE" in df.columns:
                original_len = len(df)
                df["EFFECTIVE_DATE"] = pd.to_datetime(df["EFFECTIVE_DATE"], errors='coerce')
                df = df[df["EFFECTIVE_DATE"] == pd.to_datetime(target_date)]
                filter_note = f"filtr na {target_date} ({original_len} → {len(df)} řádků)"

            globals()[jmeno] = df
            validities[jmeno] = validity

            if jmeno == "parties" and "NR_NEW_ARRIVALS" not in df.columns:
                _cols_needed = {'CALC_CLIENT_STATE_SHORT', 'CALC_MIGRATION_FLAG',
                                'CALC_CLIENT_CLOSED_BRANCH_FLAG', 'DBS_HOME_BRANCH_CODE',
                                'DBS_HOME_BRANCH_NAME'}
                if _cols_needed.issubset(set(df.columns)):
                    _new_arr = (
                        df.query('CALC_CLIENT_STATE_SHORT == "NOVÝ KLIENT"'
                                 ' and CALC_MIGRATION_FLAG != "Y"'
                                 ' and CALC_CLIENT_CLOSED_BRANCH_FLAG != "Y"')
                        .groupby(['DBS_HOME_BRANCH_CODE', 'DBS_HOME_BRANCH_NAME'], observed=True)
                        .size()
                        .reset_index(name='NR_NEW_ARRIVALS')
                    )
                    df = df.merge(_new_arr, on=['DBS_HOME_BRANCH_CODE', 'DBS_HOME_BRANCH_NAME'], how='left')
                    df['NR_NEW_ARRIVALS'] = df['NR_NEW_ARRIVALS'].fillna(0).astype(int)
                    globals()[jmeno] = df
                    filter_note = (filter_note + " | " if filter_note else "") + f"NR_NEW_ARRIVALS dopočítán ({int(df['NR_NEW_ARRIVALS'].sum()):,} nových klientů)"
                else:
                    _missing_cols = _cols_needed - set(df.columns)
                    filter_note = (filter_note + " | " if filter_note else "") + f"⚠️ NR_NEW_ARRIVALS nelze dopočítat — chybí: {_missing_cols}"

            _ext = cesta.rsplit(".", 1)[-1].upper() if "." in cesta else "?"
            _load_log.append({
                "jmeno":   jmeno,
                "status":  "✅",
                "typ":     _ext,
                "řádků":   f"{len(df):,}",
                "sloupců": str(len(df.columns)),
                "platnost": validity,
                "pozn":    filter_note or "—",
            })
        else:
            _load_log.append({"jmeno": jmeno, "status": "⚠️", "typ": "?", "řádků": "0", "sloupců": "0", "platnost": validity, "pozn": "prázdný soubor"})
    except Exception as e:
        _load_log.append({"jmeno": jmeno, "status": "❌", "typ": "?", "řádků": "—", "sloupců": "—", "platnost": "", "pozn": str(e)})

# ── Souhrnná tabulka načtených datasetů ──────────────────────────────────────
_load_rows = ""
for _r in _load_log:
    _bg = "#eafaf1" if _r["status"] == "✅" else ("#fff8e1" if "⚠️" in _r["status"] else "#fff0f0")
    _sc = "#0bb440" if _r["status"] == "✅" else ("#f59e0b" if "⚠️" in _r["status"] else "#e64343")
    _load_rows += (
        f"<tr style='background:{_bg};'>"
        f"<td style='padding:6px 11px;border:1px solid #e8edf5;font-weight:700;color:#1a2a4a;'>"
        f"<span style='color:{_sc};margin-right:5px;'>{_r['status']}</span>{_r['jmeno']}</td>"
        f"<td style='padding:6px 11px;border:1px solid #e8edf5;text-align:center;color:#555;'>{_r.get('typ','?')}</td>"
        f"<td style='padding:6px 11px;border:1px solid #e8edf5;text-align:right;font-variant-numeric:tabular-nums;'>{_r.get('řádků','—')}</td>"
        f"<td style='padding:6px 11px;border:1px solid #e8edf5;text-align:center;'>{_r.get('sloupců','—')}</td>"
        f"<td style='padding:6px 11px;border:1px solid #e8edf5;text-align:center;white-space:nowrap;"
        f"font-size:0.8rem;color:#2770f0;font-weight:600;'>{_r.get('platnost','')}</td>"
        f"<td style='padding:6px 11px;border:1px solid #e8edf5;font-size:0.8rem;color:#666;'>{_r.get('pozn','—')}</td>"
        f"</tr>"
    )

_ok  = sum(1 for r in _load_log if r["status"] == "✅")
_warn = sum(1 for r in _load_log if "⚠️" in r["status"])
_err = sum(1 for r in _load_log if r["status"] == "❌")

def _show_load_summary():
    _load_rows = ""
    for _r in _load_log:
        _bg = "#eafaf1" if _r["status"] == "✅" else ("#fff8e1" if "⚠️" in _r["status"] else "#fff0f0")
        _sc = "#0bb440" if _r["status"] == "✅" else ("#f59e0b" if "⚠️" in _r["status"] else "#e64343")
        _load_rows += (
            f"<tr style='background:{_bg};'>"
            f"<td style='padding:6px 11px;border:1px solid #e8edf5;font-weight:700;color:#1a2a4a;'>"
            f"<span style='color:{_sc};margin-right:5px;'>{_r['status']}</span>{_r['jmeno']}</td>"
            f"<td style='padding:6px 11px;border:1px solid #e8edf5;text-align:center;color:#555;'>{_r.get('typ','?')}</td>"
            f"<td style='padding:6px 11px;border:1px solid #e8edf5;text-align:right;font-variant-numeric:tabular-nums;'>{_r.get('řádků','—')}</td>"
            f"<td style='padding:6px 11px;border:1px solid #e8edf5;text-align:center;'>{_r.get('sloupců','—')}</td>"
            f"<td style='padding:6px 11px;border:1px solid #e8edf5;text-align:center;white-space:nowrap;"
            f"font-size:0.8rem;color:#2770f0;font-weight:600;'>{_r.get('platnost','')}</td>"
            f"<td style='padding:6px 11px;border:1px solid #e8edf5;font-size:0.8rem;color:#666;'>{_r.get('pozn','—')}</td>"
            f"</tr>"
        )
    _ok2  = sum(1 for r in _load_log if r["status"] == "✅")
    _warn2= sum(1 for r in _load_log if "⚠️" in r["status"])
    _err2 = sum(1 for r in _load_log if r["status"] == "❌")
    display(HTML(f"""
<div style="border-left:5px solid #2770f0;padding:14px 18px;margin:16px 0 4px 0;
            background:#f8fbff;border-radius:0 8px 8px 0;">
  <div style="display:flex;align-items:center;gap:16px;margin-bottom:12px;flex-wrap:wrap;">
    <span style="font-size:1rem;font-weight:700;color:#0f172a;">📦 Načítání datasetů</span>
    <span style="background:#eafaf1;color:#0bb440;border:1px solid #a7f3d0;border-radius:5px;
                 padding:2px 10px;font-size:0.82rem;font-weight:700;">✅ {_ok2} načteno</span>
    {"" if not _warn2 else f"<span style='background:#fff8e1;color:#f59e0b;border:1px solid #fde68a;border-radius:5px;padding:2px 10px;font-size:0.82rem;font-weight:700;'>⚠️ {_warn2} varování</span>"}
    {"" if not _err2  else f"<span style='background:#fff0f0;color:#e64343;border:1px solid #fecaca;border-radius:5px;padding:2px 10px;font-size:0.82rem;font-weight:700;'>❌ {_err2} chyb</span>"}
  </div>
  <table style="width:100%;border-collapse:collapse;font-size:0.83rem;">
    <thead>
      <tr style="background:#2770f0;color:white;">
        <th style="padding:7px 11px;text-align:left;border:1px solid #1a4db5;">Dataset</th>
        <th style="padding:7px 11px;text-align:center;border:1px solid #1a4db5;">Typ</th>
        <th style="padding:7px 11px;text-align:right;border:1px solid #1a4db5;">Řádků</th>
        <th style="padding:7px 11px;text-align:center;border:1px solid #1a4db5;">Sloupců</th>
        <th style="padding:7px 11px;text-align:center;border:1px solid #1a4db5;">Platnost</th>
        <th style="padding:7px 11px;text-align:left;border:1px solid #1a4db5;">Poznámka</th>
      </tr>
    </thead>
    <tbody>{_load_rows}</tbody>
  </table>
</div>
"""))


# =============================================================================
# 1b. PIVOT HOTOVOSTNÍCH TRANSAKCÍ
# =============================================================================

try:
    hotovostni_trn_wide = pivot_hotovostni_trn(hotovostni_trn)
    _load_log.append({"jmeno": "hotovostni_trn_wide", "status": "✅", "typ": "pivot",
        "řádků": f"{len(hotovostni_trn_wide):,}", "sloupců": str(len(hotovostni_trn_wide.columns)),
        "platnost": "", "pozn": "wide formát z pivot_hotovostni_trn"})
except Exception as e:
    import traceback
    _load_log.append({"jmeno": "hotovostni_trn_wide", "status": "❌", "typ": "pivot",
        "řádků": "—", "sloupců": "—", "platnost": "", "pozn": str(e)})
    display(HTML(f"<b style='color:red;'>❌ Pivot hotovostni_trn selhal:</b> {e}<br><pre>{traceback.format_exc()}</pre>"))
    hotovostni_trn_wide = pd.DataFrame(columns=['BRANCH_CODE'] + CASH_CASTKA_COLS + CASH_POCET_COLS)


# =============================================================================
# 1b2. NAČTENÍ DETAILNÍHO SHEETU S PREVODITELNOST_LOG
# =============================================================================
# Sheet "prevoditelnost_cash_trn" má jeden řádek per POBOCKA+POHYB+PREVODITELNOST_LOG
# — používáme ho přímo v pobočkových reportech pro detail operací.

hotovostni_trn_detail = None
try:
    _path_trn = soubory["hotovostni_trn"]["path"]
    _df_log = pd.read_excel(_path_trn, sheet_name="prevoditelnost_cash_trn")
    _df_log = normalize_columns(_df_log)

    _log_id   = next((c for c in ['POBOCKA','BRANCH_CODE','POBOCKA_ID'] if c in _df_log.columns), None)
    _log_poh  = next((c for c in ['POHYB','TYP_POHYBU'] if c in _df_log.columns), None)
    _log_prv  = next((c for c in ['PREVODITELNOST','PREV'] if c in _df_log.columns), None)
    _log_col  = next((c for c in ['PREVODITELNOST_LOG','POPIS_OPERACE'] if c in _df_log.columns), None)
    _log_cas  = next((c for c in ['SUM_CASTKA_TRANSAKCE','CASTKA','SUM_CASTKA'] if c in _df_log.columns), None)
    _log_poc  = next((c for c in ['POCET_OPERACI','POCET'] if c in _df_log.columns), None)

    if all(x for x in [_log_id, _log_poh, _log_prv, _log_cas, _log_poc]):
        _df_log[_log_id]  = pd.to_numeric(_df_log[_log_id],  errors='coerce')
        _df_log[_log_poh] = _df_log[_log_poh].astype(str).str.strip().str.upper()
        _df_log[_log_prv] = _df_log[_log_prv].astype(str).str.strip().str.upper()
        _df_log[_log_cas] = pd.to_numeric(_df_log[_log_cas], errors='coerce').fillna(0)
        _df_log[_log_poc] = pd.to_numeric(_df_log[_log_poc], errors='coerce').fillna(0)
        hotovostni_trn_detail = _df_log.rename(columns={
            _log_id:  'POBOCKA', _log_poh: 'POHYB', _log_prv: 'PREVODITELNOST',
            _log_cas: 'SUM_CASTKA_TRANSAKCE', _log_poc: 'POCET_OPERACI',
            **({_log_col: 'PREVODITELNOST_LOG'} if _log_col else {}),
        })
        _load_log.append({"jmeno": "hotovostni_trn_detail", "status": "✅", "typ": "XLSX",
            "řádků": f"{len(hotovostni_trn_detail):,}",
            "sloupců": str(len(hotovostni_trn_detail.columns)),
            "platnost": "", "pozn": f"{hotovostni_trn_detail['POBOCKA'].nunique()} poboček, sheet prevoditelnost_cash_trn"})
    else:
        _load_log.append({"jmeno": "hotovostni_trn_detail", "status": "⚠️", "typ": "XLSX",
            "řádků": "—", "sloupců": "—", "platnost": "", "pozn": f"chybí povinné sloupce: {list(_df_log.columns)}"})
except Exception as e:
    _load_log.append({"jmeno": "hotovostni_trn_detail", "status": "⚠️", "typ": "XLSX",
        "řádků": "—", "sloupců": "—", "platnost": "", "pozn": str(e)})


# =============================================================================
# 1c. AGREGACE HOTOVOSTNÍCH TRANSAKCÍ PRO CASH RATING (CASH_IN + CASH_OUT + OTHER)
# =============================================================================

try:
    _trn = hotovostni_trn.copy()
    id_col     = next((c for c in ['POBOCKA', 'BRANCH_CODE', 'POBOCKA_ID'] if c in _trn.columns), None)
    castka_col = next((c for c in ['SUM_CASTKA_TRANSAKCE', 'CASTKA', 'SUM_CASTKA'] if c in _trn.columns), None)
    pocet_col  = next((c for c in ['POCET_OPERACI', 'POCET', 'POCET_TRANSAKCI'] if c in _trn.columns), None)

    if not all([id_col, castka_col, pocet_col]):
        raise ValueError(f"Chybí sloupce. Dostupné: {list(_trn.columns)}")

    _trn[id_col]     = pd.to_numeric(_trn[id_col],     errors='coerce')
    _trn[castka_col] = pd.to_numeric(_trn[castka_col], errors='coerce').fillna(0)
    _trn[pocet_col]  = pd.to_numeric(_trn[pocet_col],  errors='coerce').fillna(0)
    _trn = _trn.dropna(subset=[id_col])

    transakce_trn_agg = (
        _trn.groupby(id_col)
        .agg(TRN_POCET_CELKEM=(pocet_col, 'sum'),
             TRN_CASTKA_CELKEM=(castka_col, 'sum'))
        .reset_index()
        .rename(columns={id_col: 'BRANCH_CODE'})
    )
    transakce_trn_agg['BRANCH_CODE']     = transakce_trn_agg['BRANCH_CODE'].astype(int)
    transakce_trn_agg['TRN_DENNI_POCET'] = (transakce_trn_agg['TRN_POCET_CELKEM'] / WORKING_DAYS_PER_YEAR).round(1)

    _load_log.append({"jmeno": "transakce_trn_agg", "status": "✅", "typ": "agg",
        "řádků": f"{len(transakce_trn_agg):,}", "sloupců": "3",
        "platnost": "", "pozn": f"cash rating agregát, {int(transakce_trn_agg['TRN_POCET_CELKEM'].sum()):,} transakcí celkem"})

except Exception as e:
    import traceback
    _load_log.append({"jmeno": "transakce_trn_agg", "status": "❌", "typ": "agg",
        "řádků": "—", "sloupců": "—", "platnost": "", "pozn": str(e)})
    display(HTML(f"<b style='color:red;'>❌ Agregace cash ratingu selhala:</b> {e}<br><pre>{traceback.format_exc()}</pre>"))
    transakce_trn_agg = pd.DataFrame(columns=['BRANCH_CODE', 'TRN_POCET_CELKEM', 'TRN_CASTKA_CELKEM', 'TRN_DENNI_POCET'])

# ── Zobraz souhrnnou tabulku načtených datasetů ───────────────────────────────
_show_load_summary()


# =============================================================================
# 2. FILTRACE A PŘÍPRAVA DBS
# =============================================================================

# =============================================================================
# SNAPSHOT POPULACE POBOČEK
# Soubor branch_snapshot.csv uzamkne populaci poboček k datu první kalkulace.
# Při dalších spuštěních (např. po aktualizaci DBS se sloupcem CASHLESS)
# se populace načte ze snapshotu — nové uzavření pobočky v DBS ji neodstraní.
# Chceš-li snapshot záměrně obnovit, smaž soubor branch_snapshot.csv.
# =============================================================================
BRANCH_SNAPSHOT_PATH = "branch_snapshot.csv"

try:
    dbs = dbs.query('BRANCH_CLOSED == False').copy()

    branch_types_y = ['BP', 'BZ', 'PP', 'UP', 'ZP']
    special_codes = [902, 904, 905, 906, 909, 919]

    condition = (
        ((dbs['BRANCH_CODE'] >= 1) & (dbs['BRANCH_CODE'] <= 675) & (dbs['BRANCH_TYPE'].isin(branch_types_y)))
        | ((dbs['BRANCH_CODE'].isin(special_codes)) & (dbs['BRANCH_TYPE'] == 'PR'))
    )
    dbs['BNS_FLAG'] = np.where(condition, 'Y', 'N')
    dbs_filtered = dbs.query('BNS_FLAG == "Y"').copy()

    # ── Snapshot logika ───────────────────────────────────────────────────────
    print(f"📂 Pracovní adresář: {os.getcwd()}")
    print(f"🔍 Hledám snapshot: {os.path.abspath(BRANCH_SNAPSHOT_PATH)}")

    if os.path.exists(BRANCH_SNAPSHOT_PATH):
        # Načti uzamčenou populaci
        _snap = pd.read_csv(BRANCH_SNAPSHOT_PATH)
        _snap['BRANCH_CODE'] = pd.to_numeric(_snap['BRANCH_CODE'], errors='coerce').astype(int)
        _locked_codes = set(_snap['BRANCH_CODE'].tolist())

        # Z aktuálního DBS (již filtrovaného na DBS_DATE) vezmi řádky dle snapshotu
        # — dostaneme aktualizované hodnoty sloupců (CASHLESS atd.) ke dni DBS_DATE
        dbs_snap_rows = dbs[dbs['BRANCH_CODE'].isin(_locked_codes)].copy()

        # Pobočky ze snapshotu které v novém DBS chybí — doplň z dbs_filtered jako fallback
        _missing = _locked_codes - set(dbs_snap_rows['BRANCH_CODE'].tolist())
        if _missing:
            print(f"⚠️  {len(_missing)} poboček ze snapshotu nenalezeno v DBS ke dni {DBS_DATE}: {sorted(_missing)}")
            _fallback = dbs_filtered[dbs_filtered['BRANCH_CODE'].isin(_missing)].copy()
            dbs_snap_rows = pd.concat([dbs_snap_rows, _fallback], ignore_index=True)

        dbs_snap_rows['BNS_FLAG'] = 'Y'
        dbs = dbs_snap_rows.copy()
        _snap_date = _snap['SNAPSHOT_DATE'].iloc[0] if 'SNAPSHOT_DATE' in _snap.columns else '?'
        print(f"🔒 Populace načtena ze snapshotu ({_snap_date}) — {len(dbs)} poboček, DBS aktualizován ke dni {DBS_DATE}")
    else:
        # První spuštění — ulož snapshot
        dbs = dbs_filtered.copy()
        _snap_df = pd.DataFrame({
            'BRANCH_CODE':   dbs['BRANCH_CODE'].astype(int).tolist(),
            'BRANCH_NAME':   dbs['BRANCH_NAME'].tolist() if 'BRANCH_NAME' in dbs.columns else [''] * len(dbs),
            'SNAPSHOT_DATE': pd.Timestamp.today().strftime('%Y-%m-%d'),
        })
        _snap_df.to_csv(BRANCH_SNAPSHOT_PATH, index=False)
        print(f"✅ Snapshot vytvořen: {os.path.abspath(BRANCH_SNAPSHOT_PATH)} — {len(_snap_df)} poboček uzamčeno")

    print(f"🎯 DBS po filtraci (BNS_FLAG == Y): {len(dbs)} poboček")


# =============================================================================
# 3. MERGE
# =============================================================================

    _dbs_base_cols = ["BRANCH_CODE", "CONCATENATED_HJ", "BRANCH_NAME", "REGION_NAME", "BNS_FLAG"]
    if "OBLAST" in dbs.columns:
        _dbs_base_cols.insert(_dbs_base_cols.index("REGION_NAME") + 1, "OBLAST")
    dbs_final = dbs[_dbs_base_cols].copy()
    _base_codes = set(dbs_final['BRANCH_CODE'].dropna().astype(int))
    _total = len(dbs_final)

    _merge_log = []   # seznam dict pro výslednou tabulku

    for source_name, cfg in MERGE_COLS.items():
        source_df = globals()[source_name]
        _before   = len(dbs_final)

        # Zjisti merge klíč pro source_df
        _cfg_peek = cfg.copy()
        _cfg_peek.pop("cols", None); _cfg_peek.pop("drop_key", None)
        _right_key = _cfg_peek.get("right_on") or _cfg_peek.get("on")

        # Počet unikátních kódů v source_df
        _src_codes = set()
        if _right_key and _right_key in source_df.columns:
            _src_codes = set(pd.to_numeric(source_df[_right_key], errors='coerce').dropna().astype(int))

        try:
            dbs_final = merge_with_source(dbs_final, source_df, cfg)
            _after = len(dbs_final)

            # Kolik poboček z DBS má alespoň jeden nenulový sloupec z tohoto zdroje
            _new_cols = [c for c in cfg.get("cols", [])
                         if c != _right_key and c in dbs_final.columns]
            _filled = 0
            if _new_cols:
                _filled = dbs_final[_new_cols[0]].notna().sum()

            _missing_codes = _base_codes - _src_codes if _src_codes else set()
            _status = "✅" if _after == _before and len(_missing_codes) == 0 else (
                      "⚠️" if _after == _before else "❌")

            _merge_log.append({
                "zdroj":       source_name,
                "status":      _status,
                "dbs_poboček": _total,
                "zdroj_kódů":  len(_src_codes) if _src_codes else "—",
                "spárováno":   _filled,
                "chybí_kódů":  len(_missing_codes),
                "chybí_list":  sorted(_missing_codes)[:10] if _missing_codes else [],
                "řádky_před":  _before,
                "řádky_po":    _after,
            })
        except Exception as _me:
            _merge_log.append({
                "zdroj": source_name, "status": "❌ CHYBA",
                "dbs_poboček": _total, "zdroj_kódů": "—",
                "spárováno": 0, "chybí_kódů": "—", "chybí_list": [],
                "řádky_před": _before, "řádky_po": _before,
            })
            display(HTML(f"<b style='color:red;'>❌ Merge selhal pro <code>{source_name}</code>: {_me}</b>"))
            import traceback; traceback.print_exc()

    dbs_final = dbs_final.sort_values(by="BRANCH_CODE")
    total_rows = len(dbs_final)

    # ── Detailní merge tabulka ─────────────────────────────────────────────
    _tbl_rows = ""
    for _m in _merge_log:
        _chybi_html = ""
        if _m["chybí_list"]:
            _chybi_html = f"<br><small style='color:#c0392b;'>ID: {_m['chybí_list']}" + \
                          ("…" if _m["chybí_kódů"] > 10 else "") + "</small>"
        _pct = f"{_m['spárováno']/_total*100:.1f} %" if isinstance(_m['spárováno'], int) and _total else "—"
        _row_bg = "#fff" if _m["status"] == "✅" else ("#fff8e1" if "⚠️" in _m["status"] else "#fff0f0")
        _tbl_rows += f"""
<tr style="background:{_row_bg};">
  <td style="padding:6px 10px;border:1px solid #ddd;font-weight:600;">{_m['status']} {_m['zdroj']}</td>
  <td style="padding:6px 10px;border:1px solid #ddd;text-align:center;">{_m['dbs_poboček']}</td>
  <td style="padding:6px 10px;border:1px solid #ddd;text-align:center;">{_m['zdroj_kódů']}</td>
  <td style="padding:6px 10px;border:1px solid #ddd;text-align:center;font-weight:700;
             color:{'#27ae60' if isinstance(_m['spárováno'],int) and _m['spárováno']==_total else '#e67e22'};">
    {_m['spárováno']} / {_total} &nbsp;({_pct})</td>
  <td style="padding:6px 10px;border:1px solid #ddd;text-align:center;
             color:{'#c0392b' if isinstance(_m['chybí_kódů'],int) and _m['chybí_kódů']>0 else '#27ae60'};">
    {_m['chybí_kódů']}{_chybi_html}</td>
  <td style="padding:6px 10px;border:1px solid #ddd;text-align:center;font-size:0.82rem;color:#888;">
    {_m['řádky_před']} → {_m['řádky_po']}</td>
</tr>"""

    display(HTML(f"""
<div style="border-left:5px solid #3498db;padding:15px;margin-top:20px;background:#f8fbff;">
  <h3 style="margin-top:0;color:#2980b9;">🔗 Merge diagnostika — krok za krokem</h3>
  <p style="margin:4px 0 10px 0;font-size:0.85rem;color:#666;">
    Základna: <b>{_total} poboček</b> z DBS (snapshot) &nbsp;|&nbsp; DBS datum: <b>{DBS_DATE}</b>
  </p>
  <table style="width:100%;border-collapse:collapse;font-size:0.84rem;">
    <thead>
      <tr style="background:#2770f0;color:white;">
        <th style="padding:7px 10px;text-align:left;border:1px solid #1e5bc9;">Zdroj</th>
        <th style="padding:7px 10px;text-align:center;border:1px solid #1e5bc9;">DBS pobočky</th>
        <th style="padding:7px 10px;text-align:center;border:1px solid #1e5bc9;">Kódů ve zdroji</th>
        <th style="padding:7px 10px;text-align:center;border:1px solid #1e5bc9;">Spárováno</th>
        <th style="padding:7px 10px;text-align:center;border:1px solid #1e5bc9;">Chybí kódů</th>
        <th style="padding:7px 10px;text-align:center;border:1px solid #1e5bc9;">Řádky před→po</th>
      </tr>
    </thead>
    <tbody>{_tbl_rows}</tbody>
  </table>
</div>"""))

    # ── Pobočky s chybějícími klíčovými daty ──────────────────────────────
    _key_cols = [c for c in ["FTE", "POCET_PRODEJU", "INTERNI_RATING_2024",
                              "PRIME_NAKLADY/VYNOSY", "OBJEM_VYNOSU_CZK"] if c in dbs_final.columns]
    _missing_key = dbs_final[dbs_final[_key_cols].isna().any(axis=1)][
        ["BRANCH_CODE", "BRANCH_NAME"] + _key_cols
    ]
    if not _missing_key.empty:
        display(HTML(f"""
<div style="border-left:5px solid #e67e22;padding:10px 15px;margin-top:14px;background:#fff8e1;">
  <b>⚠️ {len(_missing_key)} poboček s chybějícími klíčovými daty (FTE / výnosy / rating / C/I):</b>
</div>"""))
        display(_missing_key.reset_index(drop=True))
    else:
        display(HTML("""
<div style="border-left:5px solid #27ae60;padding:8px 15px;margin-top:14px;background:#eafaf1;">
  <b>✅ Všechny pobočky mají kompletní klíčová data.</b>
</div>"""))

    display(HTML("<h4 style='margin-top:20px;color:#2c3e50;'>📋 Ukázka dbs_final (prvních 20 řádků):</h4>"))
    display(dbs_final.head(20))

except Exception as e:
    import traceback
    display(HTML(f"<b style='color: red;'>❌ Chyba při zpracování/merge:</b> {e}"))
    traceback.print_exc()


# =============================================================================
# 4. ANALÝZA ROZDÍLŮ V ID
# =============================================================================

ids_final    = set(dbs_final['BRANCH_CODE'])
ids_profit   = set(profitabilita['ID_POBOCKY'])
ids_revenues = set(revenues['BRANCH_CODE'])


def analyze_diff(target_set, source_set, source_name):
    missing = target_set - source_set
    extra   = source_set - target_set
    display(HTML(f"""
        <div style="margin-top:15px;border-bottom:1px solid #ddd;padding-bottom:10px;">
            <h4 style="margin-bottom:5px;">🔍 Srovnání: dbs_final vs {source_name}</h4>
            <ul style="margin-top:5px;">
                <li><b>Chybí v {source_name}:</b> <span style="color:{'red' if missing else 'green'};">{len(missing)} ID</span>
                    {f'→ <small>{sorted(list(missing))}</small>' if missing else ''}</li>
                <li><b>Přebývá v {source_name}:</b> <span style="color:#3498db;">{len(extra)} ID</span>
                    <small>(pravděpodobně pobočky vyfiltrované dotazem)</small></li>
            </ul>
        </div>
    """))
    return missing, extra


missing_prof, _ = analyze_diff(ids_final, ids_profit,   "PROFITABILITA")
missing_rev,  _ = analyze_diff(ids_final, ids_revenues, "REVENUES")

if not missing_prof and not missing_rev:
    display(HTML("<b style='color:green;'>✅ Datová integrita: Finální tabulka je kompletně pokryta oběma zdroji.</b>"))
else:
    display(HTML("<b style='color:#e67e22;'>⚠️ Pozor: Některé pobočky mají prázdné hodnoty (NaN) kvůli chybějícím ID ve zdrojích.</b>"))


# =============================================================================
# 5. VÝPOČET IR_FLAG A RANKINGŮ
# =============================================================================

df = dbs_final

# Explicitní přiřazení výnosových sloupců ze správných zdrojů (přes map — bez merge, bez duplikátů)
# Nové výnosy 2024: old_ir → NEW_BUSINESS_2024_-_OBJEM_VYNOSU
_nb24_key = next((c for c in old_ir.columns if c in ('BRANCH_ID', 'BRANCH_CODE')), None)
if _nb24_key and 'NEW_BUSINESS_2024_-_OBJEM_VYNOSU' in old_ir.columns:
    _nb24_map = (old_ir
                 .dropna(subset=[_nb24_key])
                 .assign(**{_nb24_key: lambda x: pd.to_numeric(x[_nb24_key], errors='coerce')})
                 .dropna(subset=[_nb24_key])
                 .set_index(_nb24_key)['NEW_BUSINESS_2024_-_OBJEM_VYNOSU'])
    df['NEW_BUSINESS_2024_-_OBJEM_VYNOSU'] = (
        pd.to_numeric(df['BRANCH_CODE'], errors='coerce').map(_nb24_map))

# Nové výnosy 2025: revenues → OBJEM_VYNOSU_CZK
if 'OBJEM_VYNOSU_CZK' in revenues.columns and 'BRANCH_CODE' in revenues.columns:
    _rev25_map = (revenues
                  .dropna(subset=['BRANCH_CODE'])
                  .assign(BRANCH_CODE=lambda x: pd.to_numeric(x['BRANCH_CODE'], errors='coerce'))
                  .dropna(subset=['BRANCH_CODE'])
                  .set_index('BRANCH_CODE')['OBJEM_VYNOSU_CZK'])
    df['OBJEM_VYNOSU_CZK'] = (
        pd.to_numeric(df['BRANCH_CODE'], errors='coerce').map(_rev25_map))

condition_flag = (
    df['OBJEM_VYNOSU_CZK'].isna() | (df['OBJEM_VYNOSU_CZK'] == 0) |
    df['PRIME_NAKLADY/VYNOSY'].isna() | (df['PRIME_NAKLADY/VYNOSY'] == 0)
)
df['IR_FLAG'] = np.where(condition_flag, 'N', 'Y')

flag_counts = df['IR_FLAG'].value_counts()
display(HTML(f"""
    <div style="background-color:#fcf3cf;padding:10px;border-left:5px solid #f1c40f;margin-top:20px;">
        <h4 style="margin:0;">🚩 Distribuce IR_FLAG</h4>
        <p style="margin:5px 0;"><b>Y (Připraveno k výpočtu):</b> {flag_counts.get('Y', 0)} | <b>N (Chybějící/Nulová data):</b> {flag_counts.get('N', 0)}</p>
    </div>
"""))

if flag_counts.get('N', 0) > 0:
    print("Ukázka poboček vyřazených z IR (IR_FLAG == 'N'):")
    display(df[df['IR_FLAG'] == 'N'][['BRANCH_CODE', 'BRANCH_NAME', 'OBJEM_VYNOSU_CZK', 'PRIME_NAKLADY/VYNOSY']].head(10))

mask = df["IR_FLAG"] == 'Y'
df.loc[mask, "NEW_BUSINESS_VOL_RNK"] = df.loc[mask, "OBJEM_VYNOSU_CZK"].rank(ascending=False, method='first').astype(int)
df.loc[mask, "CI_RATIO_RNK"]         = df.loc[mask, "PRIME_NAKLADY/VYNOSY"].rank(ascending=True,  method='first').astype(int)
df.loc[mask, "NEW_BUSINESS_VOL_Q"]   = 6 - (pd.qcut(df.loc[mask, "OBJEM_VYNOSU_CZK"], q=5, labels=False) + 1)
df.loc[mask, "CI_RATIO_Q"]           = pd.qcut(df.loc[mask, "PRIME_NAKLADY/VYNOSY"], q=5, labels=False) + 1
df.loc[mask, "NEW_BUSINESS_VOL_PERC"] = (df.loc[mask, "OBJEM_VYNOSU_CZK"].rank(method='first', pct=True, ascending=False) * 100).apply(lambda x: max(1, int(round(x))))
df.loc[mask, "CI_RATIO_PERC"]         = (df.loc[mask, "PRIME_NAKLADY/VYNOSY"].rank(method='first', pct=True, ascending=True) * 100).apply(lambda x: max(1, int(round(x))))
df.loc[mask, "IR_SCORE"]              = (df.loc[mask, "CI_RATIO_RNK"] * 0.4) + (df.loc[mask, "NEW_BUSINESS_VOL_RNK"] * 0.6)
df.loc[mask, "IR_SCORE_CI_RATIO_PERC"] = (df.loc[mask, "CI_RATIO_PERC"] * 0.4) + (df.loc[mask, "NEW_BUSINESS_VOL_RNK"] * 0.6)
df.loc[mask, "IR"]                    = df.loc[mask, "IR_SCORE_CI_RATIO_PERC"].rank(ascending=True, method='first').astype(int)
df.loc[mask, "IR_Q"]                  = pd.qcut(df.loc[mask, "IR"], q=5, labels=False) + 1
df.loc[mask, "IR_PERC"]               = (df.loc[mask, "IR_SCORE_CI_RATIO_PERC"].rank(method='first', pct=True, ascending=True) * 100).apply(lambda x: max(1, int(round(x))))

# ── Doplňkové kvintily pro metrics overview ───────────────────────────────────
def _safe_qcut(series, ascending=False):
    """Kvintil 1=nejlepší.
    ascending=False → vyšší hodnota = lepší → nejvyšší hodnota dostane Q1.
    ascending=True  → nižší hodnota = lepší → nejnižší hodnota dostane Q1.
    """
    try:
        ranked = series.rank(ascending=ascending, method='first', pct=True) * 100
        return pd.cut(ranked, bins=[0, 20, 40, 60, 80, 100],
                      labels=[1, 2, 3, 4, 5],
                      include_lowest=True).astype(float)
    except Exception:
        return pd.Series(np.nan, index=series.index)

for _col, _asc in [
    ("POCET_NAVSTEV_CELKEM", False),   # více návštěv = lepší
    ("PRIMARNI_KLIENTI",     False),   # více = lepší
    ("POCET_KLIENTU",        False),   # více = lepší
    ("VYNOSY",               False),   # vyšší = lepší
    ("PRIME_NAKLADY/VYNOSY", True),    # nižší C/I = lepší → Q1
]:
    _qcol = f"{_col}_Q"
    if _col in df.columns:
        df.loc[mask, _qcol] = _safe_qcut(df.loc[mask, _col], ascending=_asc)

# Alias: CI_RATIO_Q (z pd.qcut výše) → PRIME_NAKLADY/VYNOSY_Q (pro _get_q)
# Oba sloupce drž synchronizované — preferuj _safe_qcut verzi
if "CI_RATIO_Q" in df.columns and "PRIME_NAKLADY/VYNOSY_Q" in df.columns:
    pass  # obě existují ze safe_qcut
elif "CI_RATIO_Q" in df.columns:
    df["PRIME_NAKLADY/VYNOSY_Q"] = df["CI_RATIO_Q"]

# (sync _Q sloupců do rating_status se provede níže, po prepare_rating_status)


# =============================================================================
# 6. NÁJEMNÉ, INVESTICE A KPI (join pkl zdrojů)
# =============================================================================

def explode_hj_merge(base_df, right_df, right_on='HJ'):
    """Rozbalí CONCATENATED_HJ, merguje s right_df a deduplikuje zpět na BRANCH_CODE."""
    right_df = right_df.copy()
    right_df[right_on] = right_df[right_on].astype(str).str.strip()

    expanded = base_df.copy()
    expanded['HJ_temp'] = expanded['CONCATENATED_HJ'].astype(str).str.split(';')
    expanded = expanded.explode('HJ_temp')
    expanded['HJ_temp'] = expanded['HJ_temp'].str.strip()

    merged = pd.merge(expanded, right_df, left_on='HJ_temp', right_on=right_on, how='left')

    # Prioritizuj řádky kde se podařilo spárovat (HJ není NaN)
    right_col = right_on if right_on in merged.columns else right_on + '_y'
    if right_col in merged.columns:
        merged['_sort_priority'] = merged[right_col].isna().astype(int)
    else:
        merged['_sort_priority'] = 0
    merged = merged.sort_values(by=['BRANCH_CODE', '_sort_priority'])
    result = merged.drop_duplicates(subset=['BRANCH_CODE'], keep='first').copy()

    # Ukliď pomocné a duplicitní sloupce
    drop_cols = ['_sort_priority', 'HJ_temp']
    # Odstraň _x / _y duplikáty — preferuj _x (z base_df) nebo přejmenuj
    for col in list(result.columns):
        if col.endswith('_x'):
            base_col = col[:-2]
            result = result.rename(columns={col: base_col})
            if base_col + '_y' in result.columns:
                drop_cols.append(base_col + '_y')
        elif col.endswith('_y') and col[:-2] in result.columns:
            drop_cols.append(col)
    # Odstraň pravý klíč pokud zůstal jako samostatný sloupec
    if right_on in result.columns and right_on != 'BRANCH_CODE':
        drop_cols.append(right_on)
    drop_cols = [c for c in drop_cols if c in result.columns]
    result = result.drop(columns=drop_cols)
    return result


for jmeno, info in soubory.items():
    if not info["path"].endswith(".pkl"):
        continue
    source_df = globals()[jmeno]

    if info.get("hj_merge"):
        df = explode_hj_merge(df, source_df)
    elif "merge_on" in info:
        merge_params = info["merge_on"].copy()
        drop_key = info.get("drop_key")
        df = pd.merge(df, source_df, how="left", **merge_params)
        if drop_key and drop_key in df.columns:
            df = df.drop(columns=[drop_key])

mask = df["IR_FLAG"] == 'Y'
if "ROCNI_SPLATKY_S_DPH_CZK" in df.columns:
    _rent_valid = df.loc[mask, "ROCNI_SPLATKY_S_DPH_CZK"].notna() & (df.loc[mask, "ROCNI_SPLATKY_S_DPH_CZK"] > 0)
    if _rent_valid.sum() >= 5:
        df.loc[mask & _rent_valid, "ROCNI_SPLATKY_S_DPH_CZK_Q"] = pd.qcut(
            df.loc[mask & _rent_valid, "ROCNI_SPLATKY_S_DPH_CZK"], q=5, labels=False) + 1
    else:
        display(HTML("<b style='color:orange;'>⚠️ ROCNI_SPLATKY_S_DPH_CZK: nedostatek hodnot pro qcut (méně než 5) — kvintil přeskočen.</b>"))
else:
    # Diagnostika — ukaž co z nájemného vůbec přišlo
    _rent_cols = [c for c in df.columns if "SPLATK" in c or "NAJEMN" in c or "VLASTNICTVI" in c or "HJ" in c]
    display(HTML(f"""
    <div style="border-left:4px solid #e67e22;padding:10px;background:#fff8e1;margin:8px 0;">
      <b>⚠️ Sloupec ROCNI_SPLATKY_S_DPH_CZK chybí v df — merge nájemného zřejmě selhal nebo sloupec má jiný název.</b><br>
      <span style="font-size:0.85rem;">Sloupce z nájemného které dorazily: <code>{_rent_cols}</code></span><br>
      <span style="font-size:0.85rem;">Kvintil nájemného (ROCNI_SPLATKY_S_DPH_CZK_Q) nebude vypočítán.</span>
    </div>"""))
    df["ROCNI_SPLATKY_S_DPH_CZK_Q"] = None

# =============================================================================
# 6b. MERGE HOTOVOSTNÍCH TRANSAKCÍ — až po všech pkl mergech
# =============================================================================

hotovostni_trn_wide['BRANCH_CODE'] = hotovostni_trn_wide['BRANCH_CODE'].astype(int)
df['BRANCH_CODE'] = df['BRANCH_CODE'].astype(int)

# Odstraň případné cash sloupce které mohly vzniknout jako nuly dříve
for col in CASH_CASTKA_COLS + CASH_POCET_COLS:
    if col in df.columns:
        df = df.drop(columns=[col])

df = df.merge(hotovostni_trn_wide, on='BRANCH_CODE', how='left')

for col in CASH_CASTKA_COLS + CASH_POCET_COLS:
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

# Merge transakce_trn_agg (CASH_IN + CASH_OUT + OTHER)
# ── Merge transakce_trn_agg ─────────────────────────────────────────
for col in ['TRN_POCET_CELKEM', 'TRN_CASTKA_CELKEM', 'TRN_DENNI_POCET']:
    if col in df.columns:
        df = df.drop(columns=[col])

if len(transakce_trn_agg) > 0 and 'BRANCH_CODE' in transakce_trn_agg.columns:
    # Explicitní přetypování obou klíčů na int64
    _agg = transakce_trn_agg.copy()
    _agg['BRANCH_CODE'] = _agg['BRANCH_CODE'].astype('int64')
    df['BRANCH_CODE']   = df['BRANCH_CODE'].astype('int64')

    df = df.merge(_agg[['BRANCH_CODE','TRN_POCET_CELKEM','TRN_CASTKA_CELKEM','TRN_DENNI_POCET']],
                  on='BRANCH_CODE', how='left')
    for col in ['TRN_POCET_CELKEM', 'TRN_CASTKA_CELKEM', 'TRN_DENNI_POCET']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    matched_trn  = (df['TRN_POCET_CELKEM'] > 0).sum()
    total_trn    = int(df['TRN_POCET_CELKEM'].sum())
    avg_daily    = df.loc[df['TRN_POCET_CELKEM']>0, 'TRN_DENNI_POCET'].mean()
    # Kontrola překryvu klíčů
    overlap = set(_agg['BRANCH_CODE']) & set(df['BRANCH_CODE'])
    display(HTML(f"""
        <div style="background:#e8f8f0;padding:8px 12px;border-left:4px solid #27ae60;font-size:0.85rem;margin-top:10px;">
            🏦 <b>Merge transakce_trn:</b> {matched_trn} poboček s transakcemi |
            Celk. transakcí: <b>{total_trn:,}</b> |
            Průměr denních: <b>{avg_daily:.1f}</b><br>
            🔍 Klíče: agg má <b>{len(_agg)}</b> poboček |
            df má <b>{len(df)}</b> poboček |
            Překryv: <b>{len(overlap)}</b> shodných BRANCH_CODE
        </div>
    """))
    if matched_trn == 0:
        display(HTML(f"""
            <div style="background:#fdecea;padding:8px 12px;border-left:4px solid #e74c3c;font-size:0.85rem;">
                ❌ <b>Merge selhal — žádná shoda!</b><br>
                Ukázka agg BRANCH_CODE: {_agg['BRANCH_CODE'].head(5).tolist()}<br>
                Ukázka df BRANCH_CODE: {df['BRANCH_CODE'].head(5).tolist()}
            </div>
        """))
else:
    for col in ['TRN_POCET_CELKEM', 'TRN_CASTKA_CELKEM', 'TRN_DENNI_POCET']:
        df[col] = 0.0
    display(HTML("<b style='color:orange;'>⚠️ transakce_trn_agg prázdný nebo chybí BRANCH_CODE — sloupce nastaveny na 0.</b>"))

# Výpočet hotovostního ratingu (45 % počet + 55 % objem)
df = compute_cash_rating(df)

# Výpočet business ratingu — debug vstupních dat
_cap_cols = ['FTE', 'OBJEM_VYNOSU_CZK', 'OSOBNI_NAKLADY', 'PRIMARNI_KLIENTI',
             'POCET_SCHUZEK_FYZICKY', 'POCET_SCHUZEK_ONLINE', 'NR_NEW_ARRIVALS']
_cap_present  = [c for c in _cap_cols if c in df.columns]
_cap_missing  = [c for c in _cap_cols if c not in df.columns]
_cap_fte_ok   = (df['FTE'] > 0).sum() if 'FTE' in df.columns else 0
_cap_rev_ok   = (df['OBJEM_VYNOSU_CZK'] > 0).sum() if 'OBJEM_VYNOSU_CZK' in df.columns else 0
_cap_onakl_ok = (df['OSOBNI_NAKLADY'] > 0).sum() if 'OSOBNI_NAKLADY' in df.columns else 0
_cap_arr_ok   = (df['NR_NEW_ARRIVALS'] > 0).sum() if 'NR_NEW_ARRIVALS' in df.columns else 0
display(HTML(f"""
    <div style="background:#f3e5f5;padding:8px 12px;border-left:4px solid #7b1fa2;font-size:0.85rem;margin-top:10px;">
        📈 <b>Business rating — vstupní data:</b><br>
        Přítomné sloupce: <b>{_cap_present}</b><br>
        {'<b style="color:red;">❌ Chybějící: ' + str(_cap_missing) + '</b><br>' if _cap_missing else '✅ Všechny vstupní sloupce nalezeny<br>'}
        FTE &gt; 0: <b>{_cap_fte_ok}</b> | OBJEM_VYNOSU_CZK &gt; 0: <b>{_cap_rev_ok}</b> |
        OSOBNI_NAKLADY &gt; 0: <b>{_cap_onakl_ok}</b> | NR_NEW_ARRIVALS &gt; 0: <b>{_cap_arr_ok}</b>
    </div>
"""))
df = compute_capacity_rating(df)
cap_ok = (df['CAP_RNK'] > 0).sum() if 'CAP_RNK' in df.columns else 0
display(HTML(f"""
    <div style='background:#f3e5f5;padding:8px 12px;border-left:4px solid #7b1fa2;font-size:0.85rem;margin-top:10px;'>
        📈 <b>Business rating (kapacitní):</b> {cap_ok} poboček s výsledkem |
        distribuce Q: {dict(df[df['CAP_Q']>0]['CAP_Q'].value_counts().sort_index()) if cap_ok > 0 else 'N/A'}
    </div>
"""))
trn_q_counts = df['TRN_Q'].value_counts().sort_index() if 'TRN_Q' in df.columns else {}
display(HTML(f"""
    <div style="background:#e3f2fd;padding:8px 12px;border-left:4px solid #1565c0;font-size:0.85rem;margin-top:10px;">
        🏦 <b>Cash rating kvintily:</b>
        Q1 (top): <b>{int(trn_q_counts.get(1.0, 0))}</b> |
        Q2: <b>{int(trn_q_counts.get(2.0, 0))}</b> |
        Q3: <b>{int(trn_q_counts.get(3.0, 0))}</b> |
        Q4: <b>{int(trn_q_counts.get(4.0, 0))}</b> |
        Q5 (bottom): <b>{int(trn_q_counts.get(5.0, 0))}</b>
    </div>
"""))

# Výpočet % převoditelnosti = převoditelné počty / (převoditelné + nepřevoditelné) × 100
# Zahrnuje CASH_IN i CASH_OUT dohromady
df['_prev_pocet'] = (
    df['CASH_IN_PREVODITELNA_POCET'] + df['CASH_OUT_PREVODITELNA_POCET']
)
df['_total_pocet'] = (
    df['CASH_IN_PREVODITELNA_POCET']   + df['CASH_IN_NEPREVODITELNA_POCET'] +
    df['CASH_OUT_PREVODITELNA_POCET']  + df['CASH_OUT_NEPREVODITELNA_POCET']
)
df['CASH_PREVODITELNOST_PCT'] = np.where(
    df['_total_pocet'] > 0,
    (df['_prev_pocet'] / df['_total_pocet'] * 100).round(1),
    np.nan
)
df = df.drop(columns=['_prev_pocet', '_total_pocet'])

matched = (df['CASH_IN_PREVODITELNA_CASTKA'] > 0).sum()
display(HTML(f"""
    <div style="background:#e8f8f0;padding:8px 12px;border-left:4px solid #27ae60;font-size:0.85rem;margin-top:10px;">
        💵 <b>Merge hotovostní transakce (sekce 6b):</b> {len(df)} řádků |
        Pobočky s nenulovým Cash IN: <b>{matched}</b> |
        Ukázka pobočka 1 — Cash IN převod.:
        <b>{int(df.loc[df['BRANCH_CODE']==1, 'CASH_IN_PREVODITELNA_CASTKA'].values[0]):,} Kč</b>
        (očekáváno 313 610 200 Kč)
    </div>
"""))


# =============================================================================
# 6b. NAČTENÍ SPECIALISTŮ
# =============================================================================

try:
    df_specialiste = pd.read_pickle("export_specialiste.pkl")
    # Pozicové sloupce = vše kromě identifikátorů
    _spec_id_cols = {"branch_id", "branch_name", "gps_x", "gps_y", "Evidenční stav"}
    SPEC_POZICE_COLS = [c for c in df_specialiste.columns if c not in _spec_id_cols]
    # Normalizujeme branch_id na int pro merge
    df_specialiste["branch_id"] = pd.to_numeric(df_specialiste["branch_id"], errors="coerce")
    print(f"✅ Specialisté načteni: {len(df_specialiste)} poboček, {len(SPEC_POZICE_COLS)} pozic")
except Exception as _e:
    df_specialiste = None
    SPEC_POZICE_COLS = []
    print(f"⚠️ Specialisté nenačteni: {_e}")

# -----------------------------------------------------------------------------
# Detailní investice po položkách (export_investice_all_2026.pkl)
# Sloupce: HJ, JEDNOTNE_ZNACENI, SCHVALENO_IKCS_DATUM, INVESTICE
# Jedna pobočka (HJ) může mít více řádků — zobrazujeme všechny seřazené dle data.
# -----------------------------------------------------------------------------
try:
    df_investice_all = pd.read_pickle("export_investice_all_2026.pkl")
    df_investice_all["SCHVALENO_IKCS_DATUM"] = pd.to_datetime(
        df_investice_all["SCHVALENO_IKCS_DATUM"], errors="coerce"
    )
    df_investice_all["INVESTICE"] = pd.to_numeric(
        df_investice_all["INVESTICE"], errors="coerce"
    )
    df_investice_all["HJ"] = df_investice_all["HJ"].astype(str).str.strip()
    df_investice_all = df_investice_all.sort_values("SCHVALENO_IKCS_DATUM", ascending=False)
    print(f"✅ Investice (all) načteny: {len(df_investice_all)} záznamů, "
          f"{df_investice_all['HJ'].nunique()} unikátních HJ")
except Exception as _e:
    df_investice_all = None
    print(f"⚠️ Investice (all) nenačteny: {_e}")


def generate_specialiste_table(branch_codes, title="👥 Obsazení pozic"):
    """
    Vygeneruje HTML tabulku specialistů pro zadané branch_codes (list nebo single int).
    Zobrazí jen pozice, kde je alespoň 1 pracovník u některé z poboček.
    """
    if df_specialiste is None or df_specialiste.empty:
        return ""

    if not isinstance(branch_codes, (list, tuple)):
        branch_codes = [branch_codes]
    branch_codes = [int(b) for b in branch_codes if pd.notna(b)]

    subset = df_specialiste[df_specialiste["branch_id"].isin(branch_codes)].copy()
    if subset.empty:
        return ""

    # Jen pozice s alespoň 1 pracovníkem
    active_cols = [c for c in SPEC_POZICE_COLS if pd.to_numeric(subset[c], errors="coerce").fillna(0).sum() > 0]
    if not active_cols:
        return ""

    rows_html = ""
    for _, row in subset.iterrows():
        for pozice in active_cols:
            val = pd.to_numeric(row.get(pozice, 0), errors="coerce")
            if pd.isna(val) or val == 0:
                continue
            val = int(val)
            dots = "●" * min(val, 8) + ("…" if val > 8 else "")
            rows_html += f"""
            <tr>
              <td style="padding:4px 10px;border:1px solid #e0e0e0;font-size:0.82rem;color:#333;">{pozice}</td>
              <td style="padding:4px 10px;border:1px solid #e0e0e0;text-align:center;font-weight:bold;font-size:0.85rem;">{val}</td>
              <td style="padding:4px 10px;border:1px solid #e0e0e0;color:#1976d2;letter-spacing:2px;font-size:0.9rem;">{dots}</td>
            </tr>"""

    if not rows_html:
        return ""

    return f"""
    <div style="margin:18px 0 10px 0;">
      <div class="section-header">{title}</div>
      <table style="width:100%;border-collapse:collapse;font-size:0.83rem;margin-top:6px;">
        <thead>
          <tr style="background:#1565C0;">
            <th style="padding:6px 10px;border:1px solid #1565C0;color:#fff;text-align:left;">Pozice</th>
            <th style="padding:6px 10px;border:1px solid #1565C0;color:#fff;text-align:center;width:60px;">Počet</th>
            <th style="padding:6px 10px;border:1px solid #1565C0;color:#fff;text-align:left;">Vizualizace</th>
          </tr>
        </thead>
        <tbody>{rows_html}
        </tbody>
      </table>
    </div>"""


def generate_metrics_overview(df_input, title="📊 Přehled výkonnostních metrik"):
    """
    Select pobočky (stejný styl jako specialiste) → jednořádková tabulka s kvintilními kuličkami.
    """
    import uuid as _uuid, json as _json

    wid = "mtr_" + _uuid.uuid4().hex[:8]
    d = df_input.copy()

    # ── Výpočet kvintilů ──────────────────────────────────────────────────────
    def _q_hi(col):
        if col not in d.columns: return
        s = pd.to_numeric(d[col], errors='coerce')
        if s.dropna().empty: return
        try: d[col + '_Q'] = 6 - (pd.qcut(s, q=5, labels=False, duplicates='drop') + 1)
        except: d[col + '_Q'] = np.nan

    def _q_lo(col):
        if col not in d.columns: return
        s = pd.to_numeric(d[col], errors='coerce')
        if s.dropna().empty: return
        try: d[col + '_Q'] = pd.qcut(s, q=5, labels=False, duplicates='drop') + 1
        except: d[col + '_Q'] = np.nan

    _q_hi('POCET_NAVSTEV_CELKEM'); _q_hi('PRIMARNI_KLIENTI')
    _q_hi('POCET_KLIENTU');        _q_hi('VYNOSY')
    _q_lo("PRIME_NAKLADY/VYNOSY"); _q_lo('CAP_PEREX')
    _q_hi('POCET_SCHUZEK_ONLINE'); _q_hi('POCET_SCHUZEK_FYZICKY')
    _q_lo('PLOCHA_POUZE_D5')

    # TRN kvintily jen pro pobočky s transakcemi (ne bezhotovostní)
    for _trn_col in ['TRN_CASTKA_CELKEM', 'TRN_POCET_CELKEM', 'TRN_DENNI_POCET']:
        if _trn_col not in d.columns: continue
        s = pd.to_numeric(d[_trn_col], errors='coerce')
        _cash_mask = s > 0
        if _cash_mask.sum() < 2: continue
        try:
            d[_trn_col + '_Q'] = np.nan
            d.loc[_cash_mask, _trn_col + '_Q'] = 6 - (
                pd.qcut(s[_cash_mask], q=5, labels=False, duplicates='drop') + 1
            )
        except Exception:
            d[_trn_col + '_Q'] = np.nan

    # ── Definice metrik ───────────────────────────────────────────────────────
    import re as _re

    def _q_num(row, qsrc):
        """Vrátí int 1-5 nebo None — prohledá _Q, _STATUS, _Q_HTML."""
        for suffix in ['_Q', '_STATUS', '_Q_HTML', '_HTML']:
            v = row.get(qsrc + suffix)
            if v is None or (isinstance(v, float) and np.isnan(v)): continue
            s = str(v)
            try:
                vi = int(float(s))
                if 1 <= vi <= 5: return vi
            except: pass
            m = _re.search(r'title="[^"]*\s([1-5]):', s)
            if m: return int(m.group(1))
            m = _re.search(r'>([1-5])<', s)
            if m: return int(m.group(1))
        # přímý _Q sloupec
        v = row.get(qsrc + '_Q')
        if v is not None:
            try:
                vi = int(float(v))
                if 1 <= vi <= 5: return vi
            except: pass
        return None

    ALL_METRICS = [
        ("Rating 25",        "IR"),
        ("Nové výnosy",      "NEW_BUSINESS_VOL"),
        ("C/I ratio",        "PRIME_NAKLADY/VYNOSY"),
        ("PEREX",            "CAP_PEREX"),
        ("Výnosy 25",        "VYNOSY"),
        ("Celk. návštěvy",   "POCET_NAVSTEV_CELKEM"),
        ("Primární klienti", "PRIMARNI_KLIENTI"),
        ("Počet klientů",    "POCET_KLIENTU"),
        ("Schůzky online",   "POCET_SCHUZEK_ONLINE"),
        ("Schůzky fyzicky",  "POCET_SCHUZEK_FYZICKY"),
        ("Objem transakcí",  "TRN_CASTKA_CELKEM"),
        ("Počet operací",    "TRN_POCET_CELKEM"),
        ("Denní trn. prům.", "TRN_DENNI_POCET"),
        ("Plocha D5",        "PLOCHA_POUZE_D5"),
    ]

    # Barvy kuličky — shodné s get_quintile_dot
    DOT_COLOR = {1:"#2ca02c", 2:"#6abf69", 3:"#f0a500", 4:"#e07020", 5:"#d62728"}
    TOOLTIP   = {1:"top 20 %", 2:"20–40 %", 3:"40–60 %", 4:"60–80 %", 5:"bottom 20 %"}

    # ── Sestavit data ─────────────────────────────────────────────────────────
    branches_data = {}
    for _, row in d.iterrows():
        bid   = int(row.get("BRANCH_CODE", 0))
        bname = str(row.get("BRANCH_NAME", bid))
        qs = [_q_num(row, qsrc) for _, qsrc in ALL_METRICS]
        branches_data[bid] = {"name": bname, "qs": qs}

    ordered = sorted(branches_data.items(), key=lambda x: x[1]["name"])
    options_html = "\n".join(
        f'<option value="{bid}">{info["name"]}</option>'
        for bid, info in ordered
    )
    data_js    = _json.dumps(branches_data, ensure_ascii=False)
    dot_col_js = _json.dumps(DOT_COLOR)
    tooltip_js = _json.dumps(TOOLTIP)
    labels_js  = _json.dumps([m[0] for m in ALL_METRICS], ensure_ascii=False)

    HDR_COLOR = "#2770f0"
    HDR_DARK  = "#1a4db5"

    sub_headers = "".join(
        f'<th style="background:{HDR_DARK};filter:brightness(1.15);color:#fff;'
        f'border:1px solid rgba(255,255,255,0.15);padding:4px 7px;text-align:center;'
        f'font-size:0.74rem;white-space:nowrap;">{label}</th>'
        for label, _ in ALL_METRICS
    )

    return f"""
<div style="margin-top:36px;margin-bottom:14px;padding-top:24px;
            border-top:2px solid #e0e6f5;">
  <div class="section-header">{title}</div>

  <!-- Výběr pobočky — stejný styl jako přehled pozic -->
  <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;
              margin:10px 0 12px 0;padding:10px 14px;
              background:#f0f4ff;border:1px solid #d0d8f0;border-radius:8px;">
    <span style="font-size:0.85rem;font-weight:600;color:#333;white-space:nowrap;">🔍 Vyberte pobočku:</span>
    <select id="{wid}_sel"
            style="padding:6px 12px;border:1px solid #b0bcdf;border-radius:6px;
                   font-size:0.87rem;min-width:260px;outline:none;cursor:pointer;
                   background:#fff;box-shadow:inset 0 1px 2px rgba(0,0,0,.06);">
      <option value="">— vyberte pobočku —</option>
      {options_html}
    </select>
    <span id="{wid}_badge"
          style="display:none;background:#dce8ff;border:1px solid #a0b8f0;
                 border-radius:20px;padding:3px 12px;font-size:0.81rem;
                 color:{HDR_COLOR};font-weight:600;"></span>
  </div>

  <!-- Tabulka s kuličkami -->
  <div id="{wid}_card" style="display:none;overflow-x:auto;">
    <table style="border-collapse:collapse;font-size:0.82rem;
                  border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.10);">
      <thead>
        <tr style="background:{HDR_COLOR};">
          <th style="padding:9px 14px;color:#fff;font-size:0.88rem;font-weight:bold;
                     text-align:left;min-width:170px;border:1px solid rgba(255,255,255,0.15);"
              id="{wid}_hdr"></th>
          {sub_headers}
        </tr>
      </thead>
      <tbody id="{wid}_rows"></tbody>
    </table>
  </div>
</div>

<script>
(function() {{
  var data    = {data_js};
  var DOT_COL = {dot_col_js};
  var TT      = {tooltip_js};
  var LABELS  = {labels_js};
  var sel     = document.getElementById('{wid}_sel');
  var card    = document.getElementById('{wid}_card');
  var hdr     = document.getElementById('{wid}_hdr');
  var badge   = document.getElementById('{wid}_badge');
  var tbody   = document.getElementById('{wid}_rows');
  if (!sel) return;

  function dot(q, label) {{
    if (!q || !DOT_COL[q])
      return '<span style="color:#ccc;font-size:20px;line-height:1;" title="' + label + ': bez dat">●</span>';
    return '<span title="' + label + ' — Q' + q + ': ' + (TT[q]||'') + '" '
      + 'style="color:' + DOT_COL[q] + ';font-size:22px;line-height:1;cursor:default;">●</span>'
      + '<span style="font-size:0.70rem;color:#888;margin-left:2px;">' + q + '</span>';
  }}

  sel.addEventListener('change', function() {{
    var bid = this.value;
    card.style.display = 'none';
    badge.style.display = 'none';
    if (!bid || !data[bid]) return;
    var b = data[bid];
    hdr.textContent = '📊 ' + b.name;
    badge.textContent = b.name;
    badge.style.display = 'inline-block';

    var cells = b.qs.map(function(q, i) {{
      return '<td style="padding:5px 8px;border:1px solid #dee2e6;text-align:center;'
           + 'vertical-align:middle;background:#fff;">' + dot(q, LABELS[i]) + '</td>';
    }}).join('');

    tbody.innerHTML = '<tr>'
      + '<td style="padding:6px 14px;font-size:0.82rem;font-weight:600;color:#333;'
      + 'background:#f0f4ff;position:sticky;left:0;z-index:5;white-space:nowrap;'
      + 'box-shadow:2px 0 4px rgba(0,0,0,.07);">' + b.name + '</td>'
      + cells + '</tr>';

    card.style.display = 'block';
  }});
}})();
</script>"""


def generate_specialiste_summary_table(branch_codes_list, title="👥 Přehled obsazení pozic — region", single_branch_code=None):
    """
    Interaktivní výběr pobočky + tabulka pozic ve stylu obchody tabulky.
    Pokud je zadán single_branch_code, vykreslí tabulku rovnou bez selectu.
    """
    if df_specialiste is None or df_specialiste.empty:
        return ""

    import uuid as _uuid, json as _json
    wid = "spec_" + _uuid.uuid4().hex[:8]

    branch_codes_list = [int(b) for b in branch_codes_list if pd.notna(b)]
    subset = df_specialiste[df_specialiste["branch_id"].isin(branch_codes_list)].copy()
    if subset.empty:
        return ""

    subset["branch_id"] = subset["branch_id"].astype(int)

    branches_data = {}
    for _, row in subset.iterrows():
        bid   = int(row["branch_id"])
        bname = str(row.get("branch_name", bid)) if pd.notna(row.get("branch_name")) else str(bid)
        pozice, total = [], 0
        for c in SPEC_POZICE_COLS:
            val = int(pd.to_numeric(row.get(c, 0), errors="coerce") or 0)
            if val > 0:
                pozice.append({"label": c, "val": val})
                total += val
        branches_data[bid] = {"name": bname, "pozice": pozice, "total": total}

    HDR_COLOR  = "#2770f0"
    HDR_DARK   = "#1a4db5"
    SUB_COLOR  = "#e8edf8"

    # ── Režim: přímé vykreslení pro jednu pobočku ─────────────────────────────
    if single_branch_code is not None:
        bid = int(single_branch_code)
        b   = branches_data.get(bid)
        if not b or not b["pozice"]:
            return f"""
<div style="margin:18px 0 14px 0;">
  <div class="section-header">{title}</div>
  <p style="color:#999;font-size:0.84rem;font-style:italic;padding:8px 0;">
    Pro tuto pobočku nejsou evidováni žádní specialisté.
  </p>
</div>"""

        rows_html = ""
        for i, p in enumerate(b["pozice"]):
            bg = "#ffffff" if i % 2 == 0 else "#f4f7ff"
            rows_html += (
                f"<tr style='background:{bg};'>"
                f"<td style='padding:6px 14px;border-bottom:1px solid #e8ecf5;'>{p['label']}</td>"
                f"<td style='padding:6px 14px;border-bottom:1px solid #e8ecf5;text-align:center;"
                f"font-weight:bold;color:{HDR_COLOR};'>{p['val']}</td>"
                f"</tr>"
            )

        return f"""
<div style="margin:18px 0 14px 0;">
  <div class="section-header">{title}</div>
  <table style="width:100%;max-width:500px;border-collapse:collapse;font-size:0.84rem;
                border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.10);">
    <thead>
      <tr style="background:{HDR_COLOR};">
        <th colspan="2"
            style="padding:9px 16px;color:#fff;text-align:left;
                   font-size:0.90rem;font-weight:bold;letter-spacing:0.2px;">👥 {b['name']}</th>
      </tr>
      <tr style="background:{SUB_COLOR};">
        <th style="padding:6px 14px;border-bottom:2px solid #c5cae9;text-align:left;
                   color:#444;font-size:0.80rem;font-weight:600;letter-spacing:0.3px;
                   text-transform:uppercase;">Pozice</th>
        <th style="padding:6px 14px;border-bottom:2px solid #c5cae9;text-align:center;
                   color:#444;font-size:0.80rem;font-weight:600;letter-spacing:0.3px;
                   text-transform:uppercase;width:80px;">Počet</th>
      </tr>
    </thead>
    <tbody>{rows_html}</tbody>
    <tfoot>
      <tr style="background:{HDR_DARK};">
        <td style="padding:8px 14px;font-weight:bold;color:#fff;font-size:0.86rem;">∑ Celkem pracovníků</td>
        <td style="padding:8px 14px;text-align:center;font-weight:bold;color:#fff;font-size:0.92rem;">{b['total']}</td>
      </tr>
    </tfoot>
  </table>
</div>"""

    # ── Režim: interaktivní select (původní chování) ───────────────────────────
    ordered = sorted(branches_data.items(), key=lambda x: x[1]["name"])
    options_html = "\n".join(
        f'<option value="{bid}">{info["name"]}</option>'
        for bid, info in ordered
    )
    data_js = _json.dumps(branches_data, ensure_ascii=False)

    HDR_COLOR  = "#2770f0"   # stejná modrá jako hlavní tabulka
    HDR_DARK   = "#1a4db5"
    SUB_COLOR  = "#e8edf8"

    return f"""
<div style="margin:18px 0 14px 0;">
  <div class="section-header">{title}</div>

  <!-- Výběr pobočky -->
  <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;
              margin:10px 0 12px 0;padding:10px 14px;
              background:#f0f4ff;border:1px solid #d0d8f0;border-radius:8px;">
    <span style="font-size:0.85rem;font-weight:600;color:#333;white-space:nowrap;">🔍 Vyberte pobočku:</span>
    <select id="{wid}_sel"
            style="padding:6px 12px;border:1px solid #b0bcdf;border-radius:6px;
                   font-size:0.87rem;min-width:260px;outline:none;cursor:pointer;
                   background:#fff;box-shadow:inset 0 1px 2px rgba(0,0,0,.06);">
      <option value="">— vyberte pobočku —</option>
      {options_html}
    </select>
    <span id="{wid}_badge"
          style="display:none;background:#dce8ff;border:1px solid #a0b8f0;
                 border-radius:20px;padding:3px 12px;font-size:0.81rem;
                 color:{HDR_COLOR};font-weight:600;"></span>
  </div>

  <!-- Karta s výsledky -->
  <div id="{wid}_card" style="display:none;">
    <table style="width:100%;max-width:500px;border-collapse:collapse;font-size:0.84rem;
                  border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.10);">
      <thead>
        <tr style="background:{HDR_COLOR};">
          <th colspan="2" id="{wid}_card_hdr"
              style="padding:9px 16px;color:#fff;text-align:left;
                     font-size:0.90rem;font-weight:bold;letter-spacing:0.2px;"></th>
        </tr>
        <tr style="background:{SUB_COLOR};">
          <th style="padding:6px 14px;border-bottom:2px solid #c5cae9;
                     text-align:left;color:#444;font-size:0.80rem;font-weight:600;
                     letter-spacing:0.3px;text-transform:uppercase;">Pozice</th>
          <th style="padding:6px 14px;border-bottom:2px solid #c5cae9;
                     text-align:center;color:#444;font-size:0.80rem;font-weight:600;
                     letter-spacing:0.3px;text-transform:uppercase;width:80px;">Počet</th>
        </tr>
      </thead>
      <tbody id="{wid}_rows"></tbody>
      <tfoot>
        <tr style="background:{HDR_DARK};">
          <td style="padding:8px 14px;font-weight:bold;color:#fff;font-size:0.86rem;">
            ∑ Celkem pracovníků
          </td>
          <td id="{wid}_total"
              style="padding:8px 14px;text-align:center;font-weight:bold;
                     color:#fff;font-size:0.92rem;"></td>
        </tr>
      </tfoot>
    </table>
  </div>

  <div id="{wid}_empty"
       style="display:none;margin-top:8px;color:#999;font-size:0.84rem;font-style:italic;
              padding:8px 14px;background:#fafafa;border:1px solid #eee;border-radius:6px;">
    Pro tuto pobočku nejsou evidováni žádní specialisté.
  </div>
</div>

<script>
(function() {{
  var data  = {data_js};
  var sel   = document.getElementById('{wid}_sel');
  var card  = document.getElementById('{wid}_card');
  var hdr   = document.getElementById('{wid}_card_hdr');
  var tbody = document.getElementById('{wid}_rows');
  var tfoot = document.getElementById('{wid}_total');
  var badge = document.getElementById('{wid}_badge');
  var empty = document.getElementById('{wid}_empty');
  if (!sel) return;

  sel.addEventListener('change', function() {{
    var bid = this.value;
    card.style.display  = 'none';
    empty.style.display = 'none';
    badge.style.display = 'none';
    if (!bid || !data[bid]) return;
    var b = data[bid];
    if (!b.pozice || b.pozice.length === 0) {{
      empty.style.display = 'block';
      return;
    }}
    hdr.textContent = '👥 ' + b.name;
    tbody.innerHTML = b.pozice.map(function(p, i) {{
      var bg = i % 2 === 0 ? '#ffffff' : '#f4f7ff';
      return '<tr style="background:' + bg + ';">'
        + '<td style="padding:6px 14px;border-bottom:1px solid #e8ecf5;">' + p.label + '</td>'
        + '<td style="padding:6px 14px;border-bottom:1px solid #e8ecf5;text-align:center;'
        + 'font-weight:bold;color:{HDR_COLOR};">' + p.val + '</td>'
        + '</tr>';
    }}).join('');
    tfoot.textContent  = b.total;
    badge.textContent  = b.pozice.length + ' pozic · ' + b.total + ' pracovníků';
    badge.style.display = 'inline-block';
    card.style.display  = 'block';
  }});
}})();
</script>"""

def generate_branch_reports(rating_status, output_dir="report_pobocky", hotovostni_trn=None, hotovostni_trn_detail=None, export_atm=None, kapacita_atm=None, visits=None, parties_full=None, spadovky=None):
    """
    Generuje jeden HTML report pro každou pobočku v rating_status.
    Každý report obsahuje:
      - Hlavičku s klíčovými metrikami pobočky (scorecards)
      - Benchmark: pobočka vs. region vs. celá síť
      - Obchody dle produktové skupiny (jen tato pobočka)
      - Věkové segmenty klientů
      - Metriky overview (kuličky kvintilů)
      - Informace o pozicích specialistů
      - Strategické informace (BNS, investice, nájemné)
    Soubory jsou uloženy do output_dir/BRANCH_CODE_BRANCH_NAME.html
    """
    import os
    os.makedirs(output_dir, exist_ok=True)

    bootstrap_css = '''
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
    body { font-family: 'Segoe UI', Arial, sans-serif; background:#f8f9fb; }
    .report-wrapper { padding: 28px; max-width: 1400px; margin: 0 auto; }
    .section-header { margin-top: 32px; margin-bottom: 12px; font-size: 1.15rem;
                      font-weight: 600; color: #2c3e50; }
    .table { font-size: 0.82rem; background: white; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }
    .table thead th { background-color: #2770f0 !important; color: white !important;
                      text-align: center; vertical-align: middle; white-space: nowrap;
                      border: 1px solid #1e5bc9; position: sticky; top: 0; z-index: 10; }
    .table tbody td { vertical-align: middle; text-align: center; white-space: nowrap;
                      border: 1px solid #dee2e6; }
    .benchmark-val { font-weight: bold; color: #2770f0; }
    .legend-box { margin-top: 40px; padding: 20px; background: #fff;
                  border: 1px solid #dee2e6; border-radius: 8px; }
    .age-section-card { background: #fff; border: 1px solid #dee2e6; border-radius: 8px;
                        padding: 18px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }
    .scorecard { background: #fff; border-radius: 10px; padding: 14px 18px;
                 box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 5px solid #2770f0;
                 min-width: 140px; }
    .scorecard .sc-label { font-size: 0.75rem; color: #888; font-weight: 600;
                           text-transform: uppercase; letter-spacing: 0.4px; }
    .scorecard .sc-value { font-size: 1.35rem; font-weight: 700; color: #1a2a4a;
                           line-height: 1.2; margin-top: 2px; }
    .scorecard .sc-sub { font-size: 0.75rem; color: #aaa; margin-top: 2px; }
    .bm-tbl td, .bm-tbl th { padding: 5px 12px !important; }
    /* details/summary collapsible — cross-browser vč. Safari */
    details.collapsible-section > summary { list-style: none; }
    details.collapsible-section > summary::-webkit-details-marker { display: none; }
    details.collapsible-section > summary::marker { display: none; }
    details.collapsible-section[open] > summary { border-radius: 8px 8px 0 0; border-bottom: 1px solid #dde4f5; }
    details.collapsible-section:not([open]) > summary { border-radius: 8px; border-bottom: none; }
</style>'''

    def _v(row, col, default=None):
        v = row.get(col, default)
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        return v

    def _sc(label, value, sub="", color="#2770f0"):
        return (f'<div class="scorecard" style="border-left-color:{color};">'
                f'<div class="sc-label">{label}</div>'
                f'<div class="sc-value">{value}</div>'
                f'<div class="sc-sub">{sub}</div></div>')

    def _bm_row(label, val_b, val_r, val_s, fmt="num"):
        fmts = {"money": format_money, "pct": format_pct, "num": format_num_round}
        f = fmts.get(fmt, format_num_round)
        diff = get_change_html(val_b - val_r, mode=fmt) if val_b is not None and val_r is not None else "—"
        return (f"<tr><td style='text-align:left;'>{label}</td>"
                f"<td class='benchmark-val' style='text-align:right;'>{f(val_b) if val_b is not None else '—'}</td>"
                f"<td style='text-align:right;'>{f(val_r) if val_r is not None else '—'}</td>"
                f"<td style='text-align:right;'>{f(val_s) if val_s is not None else '—'}</td>"
                f"<td style='text-align:right;'>{diff}</td></tr>")

    def _col(df, col):
        return df[col].mean() if col in df.columns and len(df) > 0 else None

    # ── Definice metrik pro radar / overview — musí být před for loop ────────
    ALL_METRICS_B = [
        ("Rating 25",        "IR"),
        ("Nové výnosy",      "NEW_BUSINESS_VOL"),
        ("C/I ratio",        "PRIME_NAKLADY/VYNOSY"),
        ("PEREX",            "CAP_PEREX"),
        ("Výnosy 25",        "VYNOSY"),
        ("Celk. návštěvy",   "POCET_NAVSTEV_CELKEM"),
        ("Primární klienti", "PRIMARNI_KLIENTI"),
        ("Počet klientů",    "POCET_KLIENTU"),
        ("Schůzky online",   "POCET_SCHUZEK_ONLINE"),
        ("Schůzky fyzicky",  "POCET_SCHUZEK_FYZICKY"),
        ("Objem transakcí",  "TRN_CASTKA_CELKEM"),
        ("Počet operací",    "TRN_POCET_CELKEM"),
        ("Denní trn. prům.", "TRN_DENNI_POCET"),
        ("Plocha D5",        "PLOCHA_POUZE_D5"),
    ]
    DOT_COLOR_M = {1:"#2ca02c", 2:"#6abf69", 3:"#f0a500", 4:"#e07020", 5:"#d62728"}
    TOOLTIP_M   = {1:"top 20 %", 2:"20–40 %", 3:"40–60 %", 4:"60–80 %", 5:"bottom 20 %"}

    total_branches = len(rating_status)
    print(f"Generuji {total_branches} pobočkových reportů do '{output_dir}/'...")

    for idx, (_, row) in enumerate(rating_status.iterrows(), 1):
        branch_code = int(_v(row, "BRANCH_CODE", 0))
        branch_name = str(_v(row, "BRANCH_NAME", str(branch_code)))
        region      = str(_v(row, "REGION_NAME", "—"))

        df_region  = rating_status[rating_status["REGION_NAME"] == region]
        df_b       = rating_status[rating_status["BRANCH_CODE"] == branch_code]

        slug = f"{branch_code}_{branch_name.replace(' ', '_').replace('/', '-')}"

        ir_val      = _v(row, "IR")
        ir_24       = _v(row, "INTERNI_RATING_2024")
        ir_diff     = _v(row, "IR_DIFF_NUM")
        ir_change   = _v(row, "IR_CHANGE_HTML", "")
        fte         = _v(row, "FTE")
        fte_float   = float(pd.to_numeric(fte, errors="coerce")) if fte is not None else None
        vynosy      = _v(row, "VYNOSY")
        novevyn     = _v(row, "OBJEM_VYNOSU_CZK")
        ci_ratio    = _v(row, "PRIME_NAKLADY/VYNOSY")
        navstevy    = _v(row, "POCET_NAVSTEV_CELKEM")
        poc_klientu  = _v(row, "POCET_KLIENTU")
        prim_klienti = _v(row, "PRIMARNI_KLIENTI")
        aktiv_klienti= _v(row, "AKTIVNI_KLIENTI")
        strategie   = _v(row, "FOOTPRINT_STRATEGIE_(2030)", "—")
        format_p    = _v(row, "BRANCH_FORMAT", "—")
        cashless    = _v(row, "CASHLESS", "N")
        vlastnictvi = _v(row, "VLASTNICTVI", "—")
        najemne     = _v(row, "ROCNI_SPLATKY_S_DPH_CZK")

        is_cashless_branch = str(cashless).upper() in ['Y', 'YES', 'ANO', '1', 'TRUE']

        # Přepočítáme kvintily z celé sítě (stejná logika jako generate_metrics_overview)
        _d_net = rating_status.copy()
        def _qhi_net(col):
            if col not in _d_net.columns: return
            s = pd.to_numeric(_d_net[col], errors='coerce')
            if s.dropna().empty: return
            try: _d_net[col + '_Q'] = 6 - (pd.qcut(s, q=5, labels=False, duplicates='drop') + 1)
            except: _d_net[col + '_Q'] = np.nan
        def _qlo_net(col):
            if col not in _d_net.columns: return
            s = pd.to_numeric(_d_net[col], errors='coerce')
            if s.dropna().empty: return
            try: _d_net[col + '_Q'] = pd.qcut(s, q=5, labels=False, duplicates='drop') + 1
            except: _d_net[col + '_Q'] = np.nan

        _qhi_net('POCET_NAVSTEV_CELKEM'); _qhi_net('PRIMARNI_KLIENTI')
        _qhi_net('POCET_KLIENTU');        _qhi_net('VYNOSY')
        _qlo_net("PRIME_NAKLADY/VYNOSY"); _qlo_net('CAP_PEREX')
        _qhi_net('POCET_SCHUZEK_ONLINE'); _qhi_net('POCET_SCHUZEK_FYZICKY')
        _qlo_net('PLOCHA_POUZE_D5')
        for _tc in ['TRN_CASTKA_CELKEM', 'TRN_POCET_CELKEM', 'TRN_DENNI_POCET']:
            if _tc not in _d_net.columns: continue
            _s2 = pd.to_numeric(_d_net[_tc], errors='coerce')
            _cm = _s2 > 0
            if _cm.sum() < 2: continue
            try:
                _d_net[_tc + '_Q'] = np.nan
                _d_net.loc[_cm, _tc + '_Q'] = 6 - (pd.qcut(_s2[_cm], q=5, labels=False, duplicates='drop') + 1)
            except: _d_net[_tc + '_Q'] = np.nan

        # Řádek pro tuto pobočku
        _brow = _d_net[_d_net["BRANCH_CODE"] == branch_code]

        # ── Kvintily — přímo z globálních _Q sloupců ─────────────────────────
        DOT_COLOR_SC = {1:"#2ca02c", 2:"#6abf69", 3:"#f0a500", 4:"#e07020", 5:"#d62728"}
        DOT_BG_SC    = {1:"#eaf7ec", 2:"#f2faf0", 3:"#fdf8ec", 4:"#fdf1e8", 5:"#fdecea"}
        TOOLTIP_SC   = {1:"top 20 %", 2:"20–40 %", 3:"40–60 %", 4:"60–80 %", 5:"bottom 20 %"}

        # Mapování alias metriky → skutečný název _Q sloupce v rating_status
        _Q_ALIAS = {
            "PRIME_NAKLADY/VYNOSY": "CI_RATIO_Q",
            "OBJEM_VYNOSU_CZK":     "NEW_BUSINESS_VOL_Q",
            "IR":                   "IR_Q",
        }

        def _get_q(col):
            """Int 1-5 — čte z _brow (přepočítané Q), fallback na row z rating_status."""
            q_col = _Q_ALIAS.get(col, col + "_Q")
            # primárně z _brow (obsahuje čerstvě přepočítané kvintily stejnou logikou jako metrics overview)
            if len(_brow) > 0 and q_col in _brow.columns:
                v = _brow.iloc[0][q_col]
                if v is not None and not (isinstance(v, float) and np.isnan(v)):
                    try:
                        vi = int(float(v))
                        if 1 <= vi <= 5: return vi
                    except: pass
            # fallback na row
            v = row.get(q_col)
            if v is None or (isinstance(v, float) and np.isnan(v)): return None
            try:
                vi = int(float(v))
                return vi if 1 <= vi <= 5 else None
            except: return None

        def _q_badge(q):
            if not q: return '<span style="color:#ddd;font-size:13px;">● bez dat</span>'
            c = DOT_COLOR_SC[q]
            return (f'<span style="color:{c};font-size:15px;vertical-align:middle;">●</span>'
                    f'<span style="font-size:0.78rem;color:{c};font-weight:700;margin-left:4px;">Q{q}</span>'
                    f'<span style="font-size:0.72rem;color:#999;margin-left:3px;">{TOOLTIP_SC[q]}</span>')

        # 6 metrikových karet s kvintily ze 📊 Výkonnostní metriky — pozice v síti
        _metrics = [
            ("Rating 25",       "IR",                   ir_val,      format_num_round, "pořadí v síti"),
            ("Nové výnosy",     "NEW_BUSINESS_VOL",     novevyn,     format_money,     "LTV new business"),
            ("Výnosy 25",       "VYNOSY",               vynosy,      format_money,     "celkové výnosy"),
            ("C/I Ratio",       "PRIME_NAKLADY/VYNOSY", ci_ratio,    format_pct,       "náklady / výnosy"),
            ("Počet klientů",   "POCET_KLIENTU",        poc_klientu, format_num_round, "celkem klientů"),
            ("Celk. návštěvy",  "POCET_NAVSTEV_CELKEM", navstevy,    format_num_round, "návštěvy za rok"),
        ]

        def _metric_card(label, q_col, value, fmt_fn, sub):
            q  = _get_q(q_col)
            bc = DOT_COLOR_SC.get(q, "#2770f0")
            bg = DOT_BG_SC.get(q, "#fff")
            vs = fmt_fn(value) if value is not None else "—"
            return (f'<div style="background:{bg};border-radius:10px;padding:12px 16px;'
                    f'border-left:4px solid {bc};box-shadow:0 2px 8px rgba(0,0,0,.07);'
                    f'min-width:140px;flex:1;">'
                    f'<div style="font-size:0.72rem;color:#888;font-weight:600;'
                    f'text-transform:uppercase;letter-spacing:0.4px;">{label}</div>'
                    f'<div style="font-size:1.2rem;font-weight:700;color:#1a2a4a;margin:3px 0 4px;">{vs}</div>'
                    f'<div>{_q_badge(q)}</div>'
                    f'<div style="font-size:0.70rem;color:#bbb;margin-top:2px;">{sub}</div>'
                    f'</div>')

        metric_cards_html = "".join(
            _metric_card(lbl, qcol, val, fn, sub)
            for lbl, qcol, val, fn, sub in _metrics
        )

        # 3 Hero kartičky (změna ratingu, počet klientů, FTE)
        def _ir_diff_html():
            if ir_diff is None: return "—"
            try: d = int(float(ir_diff))
            except: return "—"
            if d < 0:   return f'<span style="color:#2ca02c;font-weight:800;">▲ {abs(d)}</span>'
            elif d > 0: return f'<span style="color:#d62728;font-weight:800;">▼ {d}</span>'
            else:       return '<span style="color:#888;">= 0</span>'

        ir_safe   = f"{int(float(ir_val)):,}" if ir_val is not None else "—"
        ir24_safe = f"{int(float(ir_24)):,}"  if ir_24  is not None else "—"
        q_ir      = _get_q("IR")
        _ir_q_color = DOT_COLOR_SC.get(q_ir, "#2770f0")
        _ir_q_bg    = DOT_BG_SC.get(q_ir, "#f0f4ff")
        _ir_q_label = f"Q{q_ir} — {TOOLTIP_SC[q_ir]}" if q_ir else "—"

        hero_cards_html = f"""
<div style="display:flex;gap:14px;flex-wrap:wrap;margin-bottom:10px;">
  <!-- IR hero — prominentní rating + kvintil -->
  <div style="background:{_ir_q_bg};border-radius:10px;padding:14px 20px;
              border-left:5px solid {_ir_q_color};box-shadow:0 2px 8px rgba(0,0,0,.07);
              min-width:180px;flex:1.5;">
    <div style="font-size:0.72rem;color:{_ir_q_color};font-weight:600;text-transform:uppercase;letter-spacing:.4px;">Interní Rating 2025</div>
    <div style="display:flex;align-items:baseline;gap:12px;margin:4px 0 2px;">
      <span style="font-size:2.6rem;font-weight:900;color:#1a2a4a;line-height:1;">{ir_safe}</span>
      <span style="background:{_ir_q_color};color:#fff;border-radius:20px;padding:3px 12px;
                   font-size:0.9rem;font-weight:800;white-space:nowrap;">{_ir_q_label}</span>
    </div>
    <div style="font-size:0.78rem;color:#888;margin-top:2px;">
      Změna: {_ir_diff_html()} &nbsp;·&nbsp; Rating 2024: <b>{ir24_safe}</b>
    </div>
  </div>
  <!-- Počet klientů -->
  <div style="background:#f0faf3;border-radius:10px;padding:14px 20px;
              border-left:5px solid #2ca02c;box-shadow:0 2px 8px rgba(0,0,0,.07);min-width:150px;flex:1;">
    <div style="font-size:0.72rem;color:#2ca02c;font-weight:600;text-transform:uppercase;letter-spacing:.4px;">Počet klientů</div>
    <div style="font-size:2rem;font-weight:800;color:#1a2a4a;margin:4px 0 2px;line-height:1.1;">{format_num_round(poc_klientu) if poc_klientu else "—"}</div>
    <div style="font-size:0.78rem;color:#888;">celkem aktivních klientů</div>
  </div>
  <!-- Poměr klientů -->
  <div style="background:#f0f4ff;border-radius:10px;padding:14px 20px;
              border-left:5px solid #2770f0;box-shadow:0 2px 8px rgba(0,0,0,.07);min-width:160px;flex:1;">
    <div style="font-size:0.72rem;color:#2770f0;font-weight:600;text-transform:uppercase;letter-spacing:.4px;">Struktura klientů</div>
    <div style="margin:8px 0;">
      {"" if not prim_klienti or not poc_klientu else
        (lambda p, a, t: (
          f"<div style='display:flex;align-items:center;gap:5px;margin-bottom:4px;'>"
          f"<span style='font-size:0.72rem;color:#2770f0;min-width:36px;'>Prim.</span>"
          f"<div style='flex:1;background:#dde8ff;border-radius:3px;height:9px;'>"
          f"<div style='width:{min(p/t*100,100):.0f}%;background:#2770f0;height:100%;border-radius:3px;'></div></div>"
          f"<span style='font-size:0.72rem;font-weight:700;color:#2770f0;min-width:36px;text-align:right;'>{p/t*100:.0f}%</span></div>"
          f"<div style='display:flex;align-items:center;gap:5px;'>"
          f"<span style='font-size:0.72rem;color:#0bb440;min-width:36px;'>Akt.</span>"
          f"<div style='flex:1;background:#d4f5e0;border-radius:3px;height:9px;'>"
          f"<div style='width:{min(a/t*100,100) if a else 0:.0f}%;background:#0bb440;height:100%;border-radius:3px;'></div></div>"
          f"<span style='font-size:0.72rem;font-weight:700;color:#0bb440;min-width:36px;text-align:right;'>{a/t*100:.0f}%</span></div>"
        ) if t and t > 0 else '—')(
          float(pd.to_numeric(prim_klienti, errors='coerce') or 0),
          float(pd.to_numeric(aktiv_klienti, errors='coerce') or 0),
          float(pd.to_numeric(poc_klientu,   errors='coerce') or 0)
        )
      }
    </div>
    <div style="font-size:0.72rem;color:#aaa;">ze {format_num_round(poc_klientu) if poc_klientu else "?"} celkem</div>
  </div>
  <!-- FTE -->
  <div style="background:#f5f0ff;border-radius:10px;padding:14px 20px;
              border-left:5px solid #7b5ea7;box-shadow:0 2px 8px rgba(0,0,0,.07);min-width:150px;flex:1;">
    <div style="font-size:0.72rem;color:#7b5ea7;font-weight:600;text-transform:uppercase;letter-spacing:.4px;">FTE</div>
    <div style="font-size:2rem;font-weight:800;color:#1a2a4a;margin:4px 0 2px;line-height:1.1;">{f"{fte_float:.1f}" if fte_float else "—"}</div>
    <div style="font-size:0.78rem;color:#888;">formát: {format_p}</div>
  </div>
</div>"""

        scorecards_html = hero_cards_html + f"""
<div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:24px;">
  {metric_cards_html}
</div>"""

        # ── Graf historických výnosů (2022–2025) — line/area + benchmark regionu a sítě ──
        import json as _json_rev
        _rev_years = ["2022", "2023", "2024", "2025"]
        _rev_cols  = ["REVENUES_2022", "REVENUES_2023", "REVENUES_2024", "VYNOSY"]
        _rev_vals  = []
        for _rc in _rev_cols:
            _rv = _v(row, _rc)
            try:    _rev_vals.append(round(float(pd.to_numeric(_rv, errors="coerce") or 0), 0))
            except: _rev_vals.append(0)
        _rev_chart_id = f"rev_hist_{branch_code}"
        _has_rev = any(v > 0 for v in _rev_vals)

        # Benchmark průměry — region a celá síť
        def _rev_bench_avgs(df_subset, cols):
            avgs = []
            for _rc in cols:
                if _rc in df_subset.columns:
                    _s = pd.to_numeric(df_subset[_rc], errors='coerce').dropna()
                    avgs.append(round(float(_s.mean()), 0) if len(_s) > 0 else 0)
                else:
                    avgs.append(0)
            return avgs

        _df_region_all = rating_status[rating_status["REGION_NAME"] == region]
        _rev_reg  = _rev_bench_avgs(_df_region_all, _rev_cols)
        _rev_net  = _rev_bench_avgs(rating_status, _rev_cols)

        if _has_rev:
            rev_history_html = f"""
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3"></script>
<div style="background:#fff;border-radius:10px;padding:16px;
            box-shadow:0 2px 8px rgba(0,0,0,.07);margin-bottom:18px;">
  <div style="font-size:0.85rem;font-weight:600;color:#444;margin-bottom:8px;">
    📈 Vývoj výnosů 2022–2025
    <span style="font-size:0.72rem;font-weight:400;color:#aaa;margin-left:8px;">
      — — prům. region &nbsp;·&nbsp; ··· prům. síť
    </span>
  </div>
  <div id="{_rev_chart_id}" style="min-height:240px;"></div>
</div>
<script>
(function() {{
  new ApexCharts(document.getElementById("{_rev_chart_id}"), {{
    chart: {{ type:"line", height:240, toolbar:{{show:false}},
              background:"transparent", animations:{{enabled:true,speed:700}} }},
    series: [
      {{
        name:"Výnosy pobočky",
        data:{_json_rev.dumps(_rev_vals)},
        type:"area"
      }},
      {{
        name:"Prům. region ({region})",
        data:{_json_rev.dumps(_rev_reg)},
        type:"line"
      }},
      {{
        name:"Prům. celá síť",
        data:{_json_rev.dumps(_rev_net)},
        type:"line"
      }}
    ],
    xaxis: {{ categories:{_json_rev.dumps(_rev_years)},
              labels:{{style:{{fontSize:"13px",fontWeight:"600",colors:"#555"}}}} }},
    yaxis: {{ labels:{{ formatter:function(v){{
      return (v/1e6).toFixed(1)+"\\u00a0M";
    }}, style:{{fontSize:"11px",colors:"#888"}} }} }},
    colors: ["#2770f0", "#e07020", "#9b59b6"],
    stroke: {{ curve:"smooth", width:[3, 2, 2], dashArray:[0, 6, 3] }},
    markers: {{ size:[7, 4, 4], strokeColors:"#fff", strokeWidth:2, hover:{{size:9}} }},
    fill: {{
      type:["gradient","solid","solid"],
      gradient:{{ shadeIntensity:1, opacityFrom:0.25, opacityTo:0.02, stops:[0,100] }}
    }},
    dataLabels: {{ enabled:true,
      enabledOnSeries:[0],
      formatter:function(v){{
        if(v===0) return "";
        return (v/1e6).toFixed(1)+"\\u00a0M";
      }},
      offsetY:-10,
      style:{{fontSize:"11px",colors:["#2770f0"],fontWeight:"600"}},
      background:{{enabled:false}}
    }},
    legend:{{ position:"top", fontSize:"12px", labels:{{colors:"#555"}} }},
    grid:{{borderColor:"#e8e8e8",strokeDashArray:4}},
    tooltip:{{ y:{{ formatter:function(v){{
      return new Intl.NumberFormat("cs-CZ").format(Math.round(v))+"\\u00a0Kč";
    }} }} }}
  }}).render();
}})();
</script>"""
        else:
            rev_history_html = ""

        # ── Detail hotovostních operací (jen pro hotovostní pobočky) ─────────
        cash_detail_html = ""
        _htrn = hotovostni_trn_detail if (hotovostni_trn_detail is not None and not hotovostni_trn_detail.empty) else hotovostni_trn
        _is_first_cash = not is_cashless_branch and not getattr(generate_branch_reports, '_cash_diag_done', False)
        if not is_cashless_branch:
            if _htrn is None or (hasattr(_htrn, 'empty') and _htrn.empty):
                if _is_first_cash:
                    generate_branch_reports._cash_diag_done = True
                    print(f"[cash_detail] ⚠️  hotovostni_trn je None nebo prázdný!\n"
                          f"             Zkontroluj volání: generate_branch_reports(..., hotovostni_trn=hotovostni_trn)\n"
                          f"             Typ: {type(_htrn)}")
            else:
                import json as _jc

                if _is_first_cash:
                    generate_branch_reports._cash_diag_done = True
                    print(f"[cash_detail] ✅  hotovostni_trn OK — {len(_htrn)} řádků")
                    print(f"              sloupce: {list(_htrn.columns)}")
                    print(f"              unikátní POHYB: {_htrn.iloc[:,1].unique().tolist() if len(_htrn.columns) > 1 else '?'}")

                # Detekce sloupců
                _hid  = next((c for c in ['POBOCKA','BRANCH_CODE','POBOCKA_ID'] if c in _htrn.columns), None)
                _hpoh = next((c for c in ['POHYB','TYP_POHYBU'] if c in _htrn.columns), None)
                _hprv = next((c for c in ['PREVODITELNOST','PREV'] if c in _htrn.columns), None)
                _hlog = next((c for c in ['PREVODITELNOST_LOG','POPIS_OPERACE'] if c in _htrn.columns), None)
                _hcas = next((c for c in ['SUM_CASTKA_TRANSAKCE','CASTKA','SUM_CASTKA'] if c in _htrn.columns), None)
                _hpoc = next((c for c in ['POCET_OPERACI','POCET'] if c in _htrn.columns), None)
                # PREVODITELNOST_LOG je nepovinný — pokud chybí, vytvoříme ho z PREVODITELNOST
                _hlog_synthetic = _hlog is None
                if _hlog_synthetic:
                    _hlog = '__LOG__'  # syntetický název sloupce

                if not all(x is not None for x in [_hid, _hpoh, _hprv, _hcas, _hpoc]):
                    if _is_first_cash:
                        print(f"[cash_detail] ⚠️  Chybí povinné sloupce: "
                              f"id={_hid}, pohyb={_hpoh}, prev={_hprv}, castka={_hcas}, pocet={_hpoc}")
                else:
                    _df_h = _htrn.copy()
                    _df_h[_hid] = pd.to_numeric(_df_h[_hid], errors='coerce')
                    _df_h_filt = _df_h[_df_h[_hid] == branch_code].copy()
                    if len(_df_h_filt) == 0:
                        _df_h_filt = _df_h[_df_h[_hid].astype(str).str.strip() == str(branch_code)].copy()
                    _df_h = _df_h_filt
                    if _is_first_cash:
                        _sample_ids = _htrn[_hid].dropna().unique()[:5].tolist()
                        print(f"[cash_detail] branch_code={branch_code} (type={type(branch_code).__name__})")
                        print(f"              řádků po filtru: {len(_df_h)}")
                        print(f"              vzorové hodnoty {_hid}: {_sample_ids}")
                    _df_h[_hpoh] = _df_h[_hpoh].astype(str).str.strip().str.upper()
                    _df_h[_hprv] = _df_h[_hprv].astype(str).str.strip().str.upper()
                    _df_h[_hcas] = pd.to_numeric(_df_h[_hcas], errors='coerce').fillna(0)
                    _df_h[_hpoc] = pd.to_numeric(_df_h[_hpoc], errors='coerce').fillna(0)
                    _df_h = _df_h[_df_h[_hpoh].isin(['CASH_IN','CASH_OUT'])]
                    # Syntetický popis operace z PREVODITELNOST pokud LOG chybí
                    if _hlog_synthetic:
                        _df_h[_hlog] = _df_h[_hprv].map({'Y': 'Převoditelná na ATM', 'N': 'Nepřevoditelná na ATM'}).fillna('Ostatní')

                    def _cash_section(pohyb_val, title, icon, accent):
                        _df_p = _df_h[_df_h[_hpoh] == pohyb_val].copy()
                        if _df_p.empty:
                            return f'<p style="color:#aaa;font-style:italic;font-size:0.83rem;">Žádná data pro {pohyb_val}.</p>'

                        tot_cas  = _df_p[_hcas].sum()
                        tot_poc  = _df_p[_hpoc].sum()
                        prev_cas = _df_p[_df_p[_hprv]=='Y'][_hcas].sum()
                        prev_poc = _df_p[_df_p[_hprv]=='Y'][_hpoc].sum()
                        pct_cas  = (prev_cas / tot_cas * 100) if tot_cas > 0 else 0
                        pct_poc  = (prev_poc / tot_poc * 100) if tot_poc > 0 else 0
                        # Průměr přes objem i počet — jedno souhrnné číslo
                        pct_avg  = (pct_cas + pct_poc) / 2

                        def _pct_color(pct):
                            if pct >= 70: return '#1a8a6a'
                            if pct >= 40: return '#f0a500'
                            return '#d62728'

                        def _bar_wide(pct, color, height='14px'):
                            return (
                                f'<div style="background:#eee;border-radius:20px;height:{height};'
                                f'width:100%;overflow:hidden;position:relative;">'
                                f'<div style="width:{pct:.1f}%;background:{color};height:100%;'
                                f'border-radius:20px;transition:width .8s;"></div>'
                                f'</div>'
                            )

                        # ── Souhrnný KPI banner ──
                        _kpi = f"""
<div style="background:linear-gradient(135deg,{accent}18,{accent}08);border:1px solid {accent}44;
            border-radius:10px;padding:14px 16px;margin-bottom:12px;display:flex;align-items:center;gap:16px;">
  <div style="text-align:center;min-width:80px;">
    <div style="font-size:2rem;font-weight:800;color:{_pct_color(pct_avg)};line-height:1;">{pct_avg:.0f} %</div>
    <div style="font-size:0.68rem;color:#888;font-weight:600;text-transform:uppercase;margin-top:2px;">převoditelné</div>
  </div>
  <div style="flex:1;">
    {_bar_wide(pct_avg, _pct_color(pct_avg), '16px')}
    <div style="display:flex;justify-content:space-between;margin-top:6px;font-size:0.72rem;color:#888;">
      <span>dle objemu: <b style="color:{_pct_color(pct_cas)};">{pct_cas:.1f} %</b></span>
      <span>dle počtu: <b style="color:{_pct_color(pct_poc)};">{pct_poc:.1f} %</b></span>
      <span>celkem: <b>{tot_cas/1e6:.1f} M Kč</b> / <b>{int(tot_poc):,} op.</b></span>
    </div>
  </div>
</div>"""

                        # ── Tabulka dle PREVODITELNOST_LOG ──
                        _df_grp = _df_p.groupby(_hlog, dropna=False).agg(
                            _castka=(_hcas, 'sum'),
                            _pocet=(_hpoc, 'sum'),
                            _prev_flag=(_hprv, lambda x: 'Y' if (x == 'Y').any() else 'N')
                        ).reset_index().sort_values('_castka', ascending=False)

                        _rows = ""
                        for _ri, (_, _r) in enumerate(_df_grp.iterrows()):
                            _plog   = str(_r[_hlog]) if pd.notna(_r[_hlog]) else "—"
                            _cas_r  = float(_r['_castka'])
                            _poc_r  = int(_r['_pocet'])
                            _is_prv = _r['_prev_flag'] == 'Y'
                            _pct_r  = (_cas_r / tot_cas * 100) if tot_cas > 0 else 0
                            _pill   = (
                                '<span style="background:#edf9f5;color:#1a8a6a;border:1px solid #a8dfc8;'
                                'border-radius:10px;padding:1px 8px;font-size:0.68rem;font-weight:700;white-space:nowrap;">✓ ATM</span>'
                                if _is_prv else
                                '<span style="background:#fff5ef;color:#c05a10;border:1px solid #f5c9a8;'
                                'border-radius:10px;padding:1px 8px;font-size:0.68rem;font-weight:700;white-space:nowrap;">✗ přepážka</span>'
                            )
                            _bar_r = (
                                f'<div style="background:#eee;border-radius:6px;height:6px;width:100%;overflow:hidden;">'
                                f'<div style="width:{_pct_r:.1f}%;background:{"#27ae60" if _is_prv else "#e07020"};'
                                f'height:100%;border-radius:6px;"></div></div>'
                                f'<span style="font-size:0.68rem;color:#999;">{_pct_r:.1f} %</span>'
                            )
                            _bg = "#f9f9f9" if _ri % 2 == 0 else "#fff"
                            _rows += (
                                f'<tr style="background:{_bg};">'
                                f'<td style="padding:5px 8px;font-size:0.76rem;color:#333;">{_plog}</td>'
                                f'<td style="padding:5px 8px;text-align:center;">{_pill}</td>'
                                f'<td style="padding:5px 8px;text-align:right;font-size:0.78rem;font-weight:600;white-space:nowrap;">{_cas_r/1e6:.2f} M</td>'
                                f'<td style="padding:5px 8px;min-width:90px;">{_bar_r}</td>'
                                f'<td style="padding:5px 8px;text-align:right;font-size:0.76rem;color:#666;">{_poc_r:,}</td>'
                                f'</tr>'
                            )

                        _tbl = f"""
<table style="width:100%;border-collapse:collapse;font-size:0.78rem;">
  <thead>
    <tr style="background:{accent};color:#fff;">
      <th style="padding:6px 8px;text-align:left;">Typ operace</th>
      <th style="padding:6px 8px;text-align:center;">Kanál</th>
      <th style="padding:6px 8px;text-align:right;white-space:nowrap;">Objem</th>
      <th style="padding:6px 8px;">Podíl</th>
      <th style="padding:6px 8px;text-align:right;">Počet op.</th>
    </tr>
  </thead>
  <tbody>{_rows}</tbody>
</table>"""
                        return _kpi + _tbl

                    _html_in  = _cash_section('CASH_IN',  'Vklady hotovosti',  '⬇️', '#2770f0')
                    _html_out = _cash_section('CASH_OUT', 'Výběry hotovosti',  '⬆️', '#e07020')

                    cash_detail_html = f"""
<div style="margin-top:28px;padding-top:20px;border-top:2px solid #e0e6f5;">
  <div class="section-header">💵 Detail hotovostních operací</div>
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-top:4px;">
    <div style="background:#fff;border-radius:10px;padding:16px;box-shadow:0 2px 8px rgba(0,0,0,.07);
                border-top:3px solid #2770f0;">
      <div style="font-weight:700;color:#2770f0;font-size:0.9rem;margin-bottom:12px;">⬇️ CASH IN — Vklady</div>
      {_html_in}
    </div>
    <div style="background:#fff;border-radius:10px;padding:16px;box-shadow:0 2px 8px rgba(0,0,0,.07);
                border-top:3px solid #e07020;">
      <div style="font-weight:700;color:#e07020;font-size:0.9rem;margin-bottom:12px;">⬆️ CASH OUT — Výběry</div>
      {_html_out}
    </div>
  </div>
</div>"""

        # ── FTE peer group — podobně velké pobočky (±1 FTE) ──────────────────
        if fte_float is not None:
            fte_lo, fte_hi = max(0, fte_float - 1.0), fte_float + 1.0
            df_fte = rating_status[
                (pd.to_numeric(rating_status["FTE"], errors="coerce") >= fte_lo) &
                (pd.to_numeric(rating_status["FTE"], errors="coerce") <= fte_hi) &
                (rating_status["BRANCH_CODE"] != branch_code)
            ]
        else:
            df_fte = rating_status.iloc[0:0]

        fte_count = len(df_fte)
        fte_label = (f"Prům. FTE±1 ({fte_count} poboček)"
                     if fte_count > 0 else "FTE skupina (N/A)")

        # ── Formátová peer group ──────────────────────────────────────────────
        df_fmt = rating_status[
            (rating_status["BRANCH_FORMAT"] == format_p) &
            (rating_status["BRANCH_CODE"] != branch_code)
        ] if format_p and format_p != "—" else rating_status.iloc[0:0]
        fmt_count  = len(df_fmt)
        fmt_label  = (f"Prům. {format_p} ({fmt_count} poboček)"
                      if fmt_count > 0 else f"Formát {format_p} (N/A)")

        def _bmrow6_(label, val_b, val_r, val_s, val_f, val_fmt, fmt="num"):
            """Benchmark řádek: pobočka | region | síť | FTE±1 | formát peer | Δ region."""
            _fmts = {"money": format_money, "pct": format_pct, "num": format_num_round}
            _ff = _fmts.get(fmt, format_num_round)
            _diff = get_change_html(val_b - val_r, mode=fmt) if val_b is not None and val_r is not None else "—"
            return (f"<tr><td style='text-align:left;'>{label}</td>"
                    f"<td class='benchmark-val' style='text-align:right;'>{_ff(val_b) if val_b is not None else '—'}</td>"
                    f"<td style='text-align:right;'>{_ff(val_r) if val_r is not None else '—'}</td>"
                    f"<td style='text-align:right;'>{_ff(val_s) if val_s is not None else '—'}</td>"
                    f"<td style='text-align:right;color:#7b5ea7;'>{_ff(val_f) if val_f is not None else '—'}</td>"
                    f"<td style='text-align:right;color:#c0692a;'>{_ff(val_fmt) if val_fmt is not None else '—'}</td>"
                    f"<td style='text-align:right;'>{_diff}</td></tr>")

        def _bmsec6_(label, color, rows_html):
            return (f"<tr><td colspan='7' style='background:{color};font-weight:600;"
                    f"font-size:0.78rem;padding:5px 10px;text-align:left;'>{label}</td></tr>"
                    + rows_html)

        def _bmprods6_():
            rows = ""
            for key, lbl in PRODUCT_GROUPS:
                for suffix, fmt, lbl_s in [
                    (f"OBJEM_VYNOSU_CZK_{key}_2025", "money", "výnosy 25"),
                    (f"POCET_PRODEJU_{key}_2025",    "num",   "prodeje 25"),
                ]:
                    rows += _bmrow6_(
                        f"{lbl} — {lbl_s}",
                        _v(row, suffix),
                        _col(df_region, suffix),
                        _col(rating_status, suffix),
                        _col(df_fte, suffix),
                        _col(df_fmt, suffix),
                        fmt=fmt
                    )
            return rows

        trn_section = ""
        if _v(row, "TRN_POCET_CELKEM"):
            trn_section = _bmsec6_("💵 Hotovostní transakce", "#fff3cd; color:#b8860b",
                _bmrow6_("Počet transakcí", _v(row,"TRN_POCET_CELKEM"),  _col(df_region,"TRN_POCET_CELKEM"),  _col(rating_status,"TRN_POCET_CELKEM"),  _col(df_fte,"TRN_POCET_CELKEM"),  _col(df_fmt,"TRN_POCET_CELKEM"),  "num") +
                _bmrow6_("Objem transakcí", _v(row,"TRN_CASTKA_CELKEM"), _col(df_region,"TRN_CASTKA_CELKEM"), _col(rating_status,"TRN_CASTKA_CELKEM"), _col(df_fte,"TRN_CASTKA_CELKEM"), _col(df_fmt,"TRN_CASTKA_CELKEM"), "money") +
                _bmrow6_("Denní průměr",    _v(row,"TRN_DENNI_POCET"),   _col(df_region,"TRN_DENNI_POCET"),   _col(rating_status,"TRN_DENNI_POCET"),   _col(df_fte,"TRN_DENNI_POCET"),   _col(df_fmt,"TRN_DENNI_POCET"),   "num")
            )

        # DIAGNOSTIKA — typ kontrola těsně před voláním
        import sys as _sd
        def _chk(name, fn):
            if not callable(fn):
                print(f"[DIAG] {branch_code} {name} is {type(fn)}: {str(fn)[:80]!r}", file=_sd.stderr)
                raise TypeError(f"{name} is not callable: {type(fn)}")
            return fn
        _bm_financni = _bmsec6_("💰 Finanční ukazatele", "#f0f4ff; color:#2770f0",
            _bmrow6_("Výnosy 25",      _v(row,"VYNOSY"),               _col(df_region,"VYNOSY"),               _col(rating_status,"VYNOSY"),               _col(df_fte,"VYNOSY"),               _col(df_fmt,"VYNOSY"),               "money") +
            _bmrow6_("Nové výnosy 25", _v(row,"OBJEM_VYNOSU_CZK"),     _col(df_region,"OBJEM_VYNOSU_CZK"),     _col(rating_status,"OBJEM_VYNOSU_CZK"),     _col(df_fte,"OBJEM_VYNOSU_CZK"),     _col(df_fmt,"OBJEM_VYNOSU_CZK"),     "money") +
            _bmrow6_("C/I Ratio 25",   _v(row,"PRIME_NAKLADY/VYNOSY"), _col(df_region,"PRIME_NAKLADY/VYNOSY"), _col(rating_status,"PRIME_NAKLADY/VYNOSY"), _col(df_fte,"PRIME_NAKLADY/VYNOSY"), _col(df_fmt,"PRIME_NAKLADY/VYNOSY"), "pct"))
        _bm_obchody   = _bmsec6_("🛒 Obchody dle produktů", "#e8f5e9; color:#1a7a4a", _bmprods6_())
        _bm_navstev   = _bmsec6_("🚶 Návštěvnost", "#f0f4ff; color:#2770f0",
            _bmrow6_("Návštěvy celkem",   _v(row,"POCET_NAVSTEV_CELKEM"),  _col(df_region,"POCET_NAVSTEV_CELKEM"),  _col(rating_status,"POCET_NAVSTEV_CELKEM"),  _col(df_fte,"POCET_NAVSTEV_CELKEM"),  _col(df_fmt,"POCET_NAVSTEV_CELKEM"),  "num") +
            _bmrow6_("Schůzky fyzicky",   _v(row,"POCET_SCHUZEK_FYZICKY"), _col(df_region,"POCET_SCHUZEK_FYZICKY"), _col(rating_status,"POCET_SCHUZEK_FYZICKY"), _col(df_fte,"POCET_SCHUZEK_FYZICKY"), _col(df_fmt,"POCET_SCHUZEK_FYZICKY"), "num") +
            _bmrow6_("Schůzky online",    _v(row,"POCET_SCHUZEK_ONLINE"),  _col(df_region,"POCET_SCHUZEK_ONLINE"),  _col(rating_status,"POCET_SCHUZEK_ONLINE"),  _col(df_fte,"POCET_SCHUZEK_ONLINE"),  _col(df_fmt,"POCET_SCHUZEK_ONLINE"),  "num") +
            _bmrow6_("Hot. walk-in",      _v(row,"POCET_HOT_WALK_IN"),     _col(df_region,"POCET_HOT_WALK_IN"),     _col(rating_status,"POCET_HOT_WALK_IN"),     _col(df_fte,"POCET_HOT_WALK_IN"),     _col(df_fmt,"POCET_HOT_WALK_IN"),     "num") +
            _bmrow6_("Bezhot. walk-in",   _v(row,"POCET_BEZHOT_WALK_IN"),  _col(df_region,"POCET_BEZHOT_WALK_IN"),  _col(rating_status,"POCET_BEZHOT_WALK_IN"),  _col(df_fte,"POCET_BEZHOT_WALK_IN"),  _col(df_fmt,"POCET_BEZHOT_WALK_IN"),  "num"))
        _bm_klienti   = _bmsec6_("👥 Klienti", "#f0e6ff; color:#6c3483",
            _bmrow6_("Počet klientů",    _v(row,"POCET_KLIENTU"),    _col(df_region,"POCET_KLIENTU"),    _col(rating_status,"POCET_KLIENTU"),    _col(df_fte,"POCET_KLIENTU"),    _col(df_fmt,"POCET_KLIENTU"),    "num") +
            _bmrow6_("Primární klienti", _v(row,"PRIMARNI_KLIENTI"), _col(df_region,"PRIMARNI_KLIENTI"), _col(rating_status,"PRIMARNI_KLIENTI"), _col(df_fte,"PRIMARNI_KLIENTI"), _col(df_fmt,"PRIMARNI_KLIENTI"), "num") +
            _bmrow6_("Cizinci",          _v(row,"CIZINCI"),          _col(df_region,"CIZINCI"),          _col(rating_status,"CIZINCI"),          _col(df_fte,"CIZINCI"),          _col(df_fmt,"CIZINCI"),          "num"))

        benchmark_html = f"""
<div class="age-section-card mb-4">
  <table class="table table-sm bm-tbl mb-0" style="font-size:0.82rem;">
    <thead class="table-dark">
      <tr>
        <th style="text-align:left;">Ukazatel</th>
        <th style="color:#7ec8f7;text-align:right;">Tato pobočka</th>
        <th style="text-align:right;">Prům. region</th>
        <th style="text-align:right;">Prům. síť</th>
        <th style="text-align:right;color:#c9b3f5;">{fte_label}</th>
        <th style="text-align:right;color:#f5c18a;">{fmt_label}</th>
        <th style="text-align:right;">Δ vs. region</th>
      </tr>
    </thead>
    <tbody>
      {_bm_financni}
      {_bm_obchody}
      {_bm_navstev}
      {_bm_klienti}
      {trn_section}
    </tbody>
  </table>
</div>"""

        # ── Podobné pobočky (nejbližší v regionu a v síti) ───────────────────
        _sim_cols = ["VYNOSY", "POCET_KLIENTU", "FTE", "POCET_NAVSTEV_CELKEM",
                     "OBJEM_VYNOSU_CZK", "PRIME_NAKLADY/VYNOSY"]
        _sim_avail = [c for c in _sim_cols if c in rating_status.columns]

        def _similarity_score(ref_row, cmp_df, cols):
            """Normalizovaná Euklidovská vzdálenost (0=identická, vyšší=vzdálenější)."""
            ref = pd.to_numeric(pd.Series({c: ref_row.get(c) for c in cols}), errors="coerce").fillna(0)
            cmp = cmp_df[cols].apply(pd.to_numeric, errors="coerce").fillna(0)
            stds = cmp.std().replace(0, 1)
            diff = (cmp - ref) / stds
            return (diff ** 2).sum(axis=1) ** 0.5

        def _q_profile_for_branch(bc):
            """Vrátí seznam score [0-5] pro každou metriku z ALL_METRICS_B dle _d_net."""
            brow = _d_net[_d_net["BRANCH_CODE"] == bc]
            if len(brow) == 0:
                return [0] * len(ALL_METRICS_B)
            brow_r = brow.iloc[0]
            profile = []
            for _, qsrc in ALL_METRICS_B:
                q_col = _Q_ALIAS.get(qsrc, qsrc + "_Q")
                v = brow_r.get(q_col)
                try:
                    q = int(float(v))
                    profile.append((6 - q) if 1 <= q <= 5 else 0)
                except:
                    profile.append(0)
            return profile

        def _get_top_similar(candidates_df, n=1):
            """Vrátí seznam (branch_code, branch_name) top-n nejpodobnějších."""
            if candidates_df.empty: return []
            scores = _similarity_score(row, candidates_df, _sim_avail)
            top = candidates_df.assign(_score=scores).nsmallest(n, "_score")
            return [
                (int(r.get("BRANCH_CODE", 0)), str(r.get("BRANCH_NAME", "—")))
                for _, r in top.iterrows()
            ]

        def _make_similar_table(candidates_df, title, n=5):
            if candidates_df.empty: return ""
            scores = _similarity_score(row, candidates_df, _sim_avail)
            top = candidates_df.assign(_score=scores).nsmallest(n, "_score")
            rows_s = ""
            for rank_i, (_, sr) in enumerate(top.iterrows(), 1):
                sc = sr.get("_score", 0)
                sim_pct = max(0, round(100 - sc * 10))
                sim_bar_w = max(4, min(100, sim_pct))
                sim_color = "#2ca02c" if sim_pct >= 80 else "#f0a500" if sim_pct >= 60 else "#e07020"
                b_name = str(sr.get("BRANCH_NAME", "—"))
                b_code = str(sr.get("BRANCH_CODE", "—"))
                b_reg  = str(sr.get("REGION_NAME", "—"))
                b_fmt  = str(sr.get("BRANCH_FORMAT", "—"))
                b_fte  = sr.get("FTE")
                b_vyn  = sr.get("VYNOSY")
                b_ir   = sr.get("IR")
                rows_s += (
                    f"<tr>"
                    f"<td style='padding:6px 10px;font-weight:600;'>{rank_i}. {b_name}"
                    f"<br><span style='font-size:0.75rem;color:#999;'>ID {b_code} · {b_reg} · {b_fmt}</span></td>"
                    f"<td style='padding:6px 10px;text-align:center;'>"
                    f"<div style='background:#eee;border-radius:8px;height:10px;width:80px;display:inline-block;vertical-align:middle;'>"
                    f"<div style='background:{sim_color};border-radius:8px;height:10px;width:{sim_bar_w*0.8:.0f}px;'></div></div>"
                    f"&nbsp;<span style='font-size:0.8rem;color:{sim_color};font-weight:700;'>{sim_pct} %</span></td>"
                    f"<td style='padding:6px 10px;text-align:right;'>{f'{float(b_fte):.1f}' if b_fte else '—'}</td>"
                    f"<td style='padding:6px 10px;text-align:right;'>{format_money(b_vyn) if b_vyn else '—'}</td>"
                    f"<td style='padding:6px 10px;text-align:right;'>{format_num_round(b_ir) if b_ir else '—'}</td>"
                    f"</tr>"
                )
            return f"""
<div style="margin-bottom:10px;">
  <div style="font-size:0.82rem;font-weight:600;color:#555;margin-bottom:6px;">{title}</div>
  <table style="width:100%;font-size:0.82rem;border-collapse:collapse;">
    <thead><tr style="background:#f5f7fa;border-bottom:2px solid #dee2e6;">
      <th style="padding:5px 10px;text-align:left;">Pobočka</th>
      <th style="padding:5px 10px;text-align:center;">Podobnost</th>
      <th style="padding:5px 10px;text-align:right;">FTE</th>
      <th style="padding:5px 10px;text-align:right;">Výnosy 25</th>
      <th style="padding:5px 10px;text-align:right;">Rating 25</th>
    </tr></thead>
    <tbody>{rows_s}</tbody>
  </table>
</div>"""

        # Region (bez aktuální pobočky)
        _df_reg_other = df_region[df_region["BRANCH_CODE"] != branch_code]
        _sim_region_html = _make_similar_table(_df_reg_other, f"📍 Nejpodobnější v regionu {region}", n=5)

        # Celá síť (bez regionu a bez aktuální pobočky)
        _df_net_other = rating_status[
            (rating_status["BRANCH_CODE"] != branch_code) &
            (rating_status["REGION_NAME"] != region)
        ]
        _sim_net_html = _make_similar_table(_df_net_other, "🌐 Nejpodobnější v celé síti (mimo region)", n=5)

        similar_html = f"""
<div class="age-section-card mb-4" style="display:grid;grid-template-columns:1fr 1fr;gap:18px;">
  <div>{_sim_region_html}</div>
  <div>{_sim_net_html}</div>
</div>"""

        # Q profily nejpodobnější pobočky v regionu a v síti (pro radar)
        _top_reg = _get_top_similar(_df_reg_other, n=1)
        _top_net = _get_top_similar(_df_net_other, n=1)
        _sim_reg_bc,  _sim_reg_name  = (_top_reg[0] if _top_reg else (None, "—"))
        _sim_net_bc,  _sim_net_name  = (_top_net[0] if _top_net else (None, "—"))
        _sim_reg_scores = _q_profile_for_branch(_sim_reg_bc) if _sim_reg_bc else [0]*len(ALL_METRICS_B)
        _sim_net_scores = _q_profile_for_branch(_sim_net_bc) if _sim_net_bc else [0]*len(ALL_METRICS_B)

        import json as _json_radar

        # ── Nejlepší pobočka v regionu stejného formátu (dle IR) ───────────────
        _df_best_fmt = rating_status[
            (rating_status["BRANCH_CODE"] != branch_code) &
            (rating_status["REGION_NAME"] == region) &
            (rating_status["BRANCH_FORMAT"] == format_p)
        ].copy()
        _best_ir_s = pd.to_numeric(_df_best_fmt["IR"], errors="coerce")
        _df_best_fmt = _df_best_fmt[_best_ir_s.notna()]
        if not _df_best_fmt.empty:
            _best_row      = _df_best_fmt.loc[pd.to_numeric(_df_best_fmt["IR"], errors="coerce").idxmin()]
            _best_fmt_bc   = int(_best_row["BRANCH_CODE"])
            _best_fmt_name = str(_best_row["BRANCH_NAME"])
        else:
            _best_fmt_bc   = None
            _best_fmt_name = f"Žádná shodná pobočka ({format_p})"
        _best_fmt_scores = _q_profile_for_branch(_best_fmt_bc) if _best_fmt_bc else [0]*len(ALL_METRICS_B)

        # ── Všechny pobočky pro dropdown (seřazené dle jména) ──────────────────
        _all_branches_list = []
        for _, _abr in rating_status.sort_values("BRANCH_NAME").iterrows():
            _abc = int(_abr["BRANCH_CODE"])
            if _abc == branch_code:
                continue
            _all_branches_list.append({
                "code":   _abc,
                "name":   str(_abr["BRANCH_NAME"]),
                "format": str(_abr.get("BRANCH_FORMAT", "")),
                "region": str(_abr.get("REGION_NAME", "")),
                "scores": _q_profile_for_branch(_abc),
            })
        _all_branches_json = _json_radar.dumps(_all_branches_list)
        obchody_avail = ['BRANCH_CODE'] + [c for c in obchody_detail.columns if c != 'BRANCH_CODE']
        obchody_avail = [c for c in obchody_avail if c in obchody_detail.columns]
        df_obch_b = df_b[['BRANCH_CODE','BRANCH_NAME']].merge(
            obchody_detail[obchody_avail], on='BRANCH_CODE', how='left'
        )
        obchody_html_b = generate_sortable_table(
            generate_obchody_table(df_obch_b, table_id=f"obch-{branch_code}"),
            f"obch-sort-{branch_code}"
        )

        # ── Produktové skupiny — přepínatelný pohled ─────────────────────────
        _prod_rev_html = ""
        _prod_items = []
        for _pk, _pl in PRODUCT_GROUPS:
            _pcol_v = f'OBJEM_VYNOSU_CZK_{_pk}_2025'
            _pcol_n = f'POCET_PRODEJU_{_pk}_2025'
            _pv = float(pd.to_numeric(df_obch_b[_pcol_v].iloc[0], errors='coerce') if _pcol_v in df_obch_b.columns and len(df_obch_b) > 0 else 0) or 0
            _pn = float(pd.to_numeric(df_obch_b[_pcol_n].iloc[0], errors='coerce') if _pcol_n in df_obch_b.columns and len(df_obch_b) > 0 else 0) or 0
            _prod_items.append((_pk, _pl, _pv, _pn))

        _total_vynosy = sum(v for _, _, v, _ in _prod_items)
        _total_pocty  = sum(n for _, _, _, n in _prod_items)

        if _total_vynosy > 0 or _total_pocty > 0:
            _prod_group_colors = [
                '#4a90d9','#45b065','#9b6bbf','#d4682a',
                '#2aacbd','#b89a00','#a83250','#4a8c4a','#7a5c3a'
            ]
            _prod_color_map = {k: _prod_group_colors[i % len(_prod_group_colors)]
                               for i, (k, _, __, ___) in enumerate(sorted(_prod_items, key=lambda x: x[0]))}
            _pid = f"prod-{branch_code}"

            # Seřadit dle výnosů pro výchozí zobrazení
            _prod_sorted_v = sorted(_prod_items, key=lambda x: x[2], reverse=True)
            _prod_sorted_n = sorted(_prod_items, key=lambda x: x[3], reverse=True)
            _max_avg = max((p[2]/p[3] for p in _prod_items if p[3] > 0), default=1)
            _prod_sorted_a = sorted(_prod_items, key=lambda x: (x[2]/x[3] if x[3]>0 else 0), reverse=True)

            def _prod_bars(items_sorted, mode, total):
                """Generuje HTML bary pro daný mode: 'v'=výnosy, 'n'=počty, 'a'=průměr"""
                _html = ""
                _mx = max((i[2] if mode=='v' else i[3] if mode=='n' else (i[2]/i[3] if i[3]>0 else 0)) for i in items_sorted) or 1
                for _pk2, _pl2, _pv2, _pn2 in items_sorted:
                    _color2 = _prod_color_map.get(_pk2, '#4a90d9')
                    if mode == 'v':
                        _val   = _pv2
                        _pct   = _pv2 / total * 100 if total > 0 else 0
                        _bar_w = _pv2 / _mx * 100
                        _pct_s = f"{_pct:.1f} %"
                        _val_s = f"{_pv2/1e6:.2f} M Kč" if _pv2 >= 1e5 else f"{_pv2:,.0f} Kč".replace(",","\u00a0")
                    elif mode == 'n':
                        _val   = _pn2
                        _pct   = _pn2 / total * 100 if total > 0 else 0
                        _bar_w = _pn2 / _mx * 100
                        _pct_s = f"{_pct:.1f} %"
                        _val_s = f"{int(_pn2):,} ks".replace(",","\u00a0")
                    else:  # průměr
                        _avg   = _pv2 / _pn2 if _pn2 > 0 else 0
                        _pct   = 0
                        _bar_w = _avg / _mx * 100
                        _pct_s = f"{_avg/1e3:.1f} tis." if _avg >= 1000 else f"{_avg:.0f} Kč"
                        _val_s = f"{int(_pn2):,} prodejů".replace(",","\u00a0") if _pn2 > 0 else "—"
                    _html += f"""
<div style="margin-bottom:7px;">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:2px;">
    <span style="font-size:0.78rem;font-weight:600;color:#333;flex:1;min-width:0;
                 overflow:hidden;text-overflow:ellipsis;white-space:nowrap;"
          title="{_pl2}">{_pl2}</span>
    <span style="font-size:0.82rem;font-weight:800;color:{_color2};
                 min-width:60px;text-align:right;">{_pct_s}</span>
    <span style="font-size:0.72rem;color:#999;min-width:90px;text-align:right;">{_val_s}</span>
  </div>
  <div style="background:#f0f0f0;border-radius:5px;height:9px;overflow:hidden;">
    <div style="width:{max(_bar_w,0.4):.1f}%;background:{_color2};height:100%;
                border-radius:5px;transition:width .4s;"></div>
  </div>
</div>"""
                return _html

            _bars_v = _prod_bars(_prod_sorted_v, 'v', _total_vynosy)
            _bars_n = _prod_bars(_prod_sorted_n, 'n', _total_pocty)
            _bars_a = _prod_bars(_prod_sorted_a, 'a', 0)

            _prod_rev_html = f"""
<div id="{_pid}-wrap" style="background:#fff;border-radius:10px;padding:14px 16px;
     box-shadow:0 2px 8px rgba(0,0,0,.06);margin-bottom:12px;">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;flex-wrap:wrap;">
    <span style="font-size:0.78rem;font-weight:700;color:#555;text-transform:uppercase;
                 letter-spacing:.4px;">📊 Produktové skupiny 2025</span>
    <div style="display:flex;gap:4px;margin-left:auto;">
      <button onclick="prodSwitch('{_pid}','v')" id="{_pid}-btn-v"
        style="padding:3px 10px;border-radius:5px;border:1px solid #2770f0;
               background:#2770f0;color:#fff;font-size:0.75rem;font-weight:600;cursor:pointer;">
        💰 Výnosy
      </button>
      <button onclick="prodSwitch('{_pid}','n')" id="{_pid}-btn-n"
        style="padding:3px 10px;border-radius:5px;border:1px solid #ccc;
               background:#f5f5f5;color:#555;font-size:0.75rem;font-weight:600;cursor:pointer;">
        🔢 Počty
      </button>
      <button onclick="prodSwitch('{_pid}','a')" id="{_pid}-btn-a"
        style="padding:3px 10px;border-radius:5px;border:1px solid #ccc;
               background:#f5f5f5;color:#555;font-size:0.75rem;font-weight:600;cursor:pointer;">
        ⌀ Výnos/prodej
      </button>
    </div>
  </div>
  <div id="{_pid}-sub" style="font-size:0.72rem;color:#aaa;margin-bottom:10px;">
    celkem {_total_vynosy/1e6:.2f} M Kč
  </div>
  <div id="{_pid}-v">{_bars_v}</div>
  <div id="{_pid}-n" style="display:none;">{_bars_n}</div>
  <div id="{_pid}-a" style="display:none;">{_bars_a}</div>
</div>
<script>
(function(){{
  var _subs = {{
    'v': 'celkem {_total_vynosy/1e6:.2f}\u00a0M\u00a0Kč',
    'n': 'celkem {int(_total_pocty):,} prodejů'.replace(/,/g,'\u00a0'),
    'a': 'průměrný výnos na 1 prodej'
  }};
  window.prodSwitch = window.prodSwitch || function(pid, mode){{
    ['v','n','a'].forEach(function(m){{
      var el = document.getElementById(pid+'-'+m);
      var btn = document.getElementById(pid+'-btn-'+m);
      if(el) el.style.display = (m===mode)?'block':'none';
      if(btn){{
        btn.style.background = (m===mode)?'#2770f0':'#f5f5f5';
        btn.style.color      = (m===mode)?'#fff':'#555';
        btn.style.borderColor= (m===mode)?'#2770f0':'#ccc';
      }}
    }});
    var sub = document.getElementById(pid+'-sub');
    if(sub) sub.textContent = (_subs[mode]||'');
  }};
}})();
</script>"""

        # ── Věkové segmenty — 2 ApexCharts + cizinci ─────────────────────────
        import json as _json_age
        _cid_bar = f"age_bar_{branch_code}"
        _cid_line = f"age_line_{branch_code}"

        _age_counts = [
            int(pd.to_numeric(_v(row, f"{g}_POCET_KLIENTU", 0), errors="coerce") or 0)
            for g in AGE_GROUPS
        ]
        _age_revs = [
            round(float(pd.to_numeric(_v(row, f"{g}_PRUMERNY_VYNOS", 0), errors="coerce") or 0), 0)
            for g in AGE_GROUPS
        ]
        _age_colors = ["#4361ee","#3a86ff","#48cae4","#06d6a0","#ffd166"]
        _has_age = sum(_age_counts) > 0

        # Cizinci — počet a podíl z celkového počtu klientů
        _poc_klientu = _v(row, "POCET_KLIENTU") or 0
        _poc_klientu = float(pd.to_numeric(_poc_klientu, errors="coerce") or 0)
        def _ciz_card(col, label, color, icon):
            val = float(pd.to_numeric(_v(row, col, 0) or 0, errors="coerce"))
            pct = f"{val/_poc_klientu*100:.1f} %" if _poc_klientu > 0 else "—"
            return (f'<div style="background:#fff;border-radius:8px;padding:12px 18px;'
                    f'border-left:4px solid {color};box-shadow:0 1px 6px rgba(0,0,0,.07);'
                    f'min-width:130px;">'
                    f'<div style="font-size:0.72rem;color:#888;font-weight:600;'
                    f'text-transform:uppercase;">{icon} {label}</div>'
                    f'<div style="font-size:1.25rem;font-weight:700;color:#1a2a4a;margin-top:2px;">'
                    f'{format_num_round(val)}</div>'
                    f'<div style="font-size:0.75rem;color:{color};font-weight:600;">'
                    f'{pct} z klientů</div></div>')

        _cizinci_html = f"""
<div style="display:flex;flex-wrap:wrap;gap:12px;margin-top:18px;">
  {_ciz_card("CIZINCI",           "Cizinci celkem", "#6c5ce7", "🌍")}
  {_ciz_card("CIZINCI_SLOVENSKO", "Cizinci SK",     "#00b894", "🇸🇰")}
  {_ciz_card("CIZINCI_UKRAJINA",  "Cizinci UA",     "#e17055", "🇺🇦")}
</div>"""

        if _has_age:
            age_charts_html = f"""
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3"></script>
<div style="display:flex;gap:18px;flex-wrap:wrap;align-items:flex-start;">

  <!-- Bar: počty klientů -->
  <div style="flex:1;min-width:280px;background:#fff;border-radius:10px;padding:14px;
              box-shadow:0 2px 8px rgba(0,0,0,.07);">
    <div style="font-size:0.82rem;font-weight:600;color:#444;margin-bottom:6px;">
      👥 Počet klientů dle věkové skupiny
    </div>
    <div id="{_cid_bar}" style="min-height:240px;"></div>
  </div>

  <!-- Line: průměrný výnos -->
  <div style="flex:1;min-width:280px;background:#fff;border-radius:10px;padding:14px;
              box-shadow:0 2px 8px rgba(0,0,0,.07);">
    <div style="font-size:0.82rem;font-weight:600;color:#444;margin-bottom:6px;">
      💰 Průměrný výnos / klient dle věkové skupiny
    </div>
    <div id="{_cid_line}" style="min-height:240px;"></div>
  </div>

</div>

{_cizinci_html}

<script>
(function() {{
  // Bar chart — počty
  new ApexCharts(document.getElementById("{_cid_bar}"), {{
    chart: {{ type:"bar", height:240, toolbar:{{show:false}}, background:"transparent",
              animations:{{enabled:true,speed:500}} }},
    series: [{{ name:"Počet klientů", data:{_json_age.dumps(_age_counts)} }}],
    xaxis: {{ categories:{_json_age.dumps(AGE_GROUPS)},
              labels:{{style:{{fontSize:"12px",colors:"#555"}}}} }},
    yaxis: {{ labels:{{ formatter:function(v){{return v.toLocaleString("cs-CZ");}},
                        style:{{fontSize:"11px",colors:"#888"}} }} }},
    colors: {_json_age.dumps(_age_colors)},
    plotOptions: {{ bar:{{ distributed:true, borderRadius:5, columnWidth:"55%",
                          dataLabels:{{position:"top"}} }} }},
    dataLabels: {{ enabled:true, formatter:function(v){{return v.toLocaleString("cs-CZ");}},
                  offsetY:-20, style:{{fontSize:"10px",colors:["#333"],fontWeight:"bold"}} }},
    legend: {{ show:false }},
    grid: {{ borderColor:"#e8e8e8", strokeDashArray:4 }},
    tooltip: {{ y:{{ formatter:function(v){{return v.toLocaleString("cs-CZ")+" klientů";}} }} }}
  }}).render();

  // Line chart — výnosy
  new ApexCharts(document.getElementById("{_cid_line}"), {{
    chart: {{ type:"line", height:240, toolbar:{{show:false}}, background:"transparent",
              animations:{{enabled:true,speed:500}} }},
    series: [{{ name:"Prům. výnos", data:{_json_age.dumps(_age_revs)} }}],
    xaxis: {{ categories:{_json_age.dumps(AGE_GROUPS)},
              labels:{{style:{{fontSize:"12px",colors:"#555"}}}} }},
    yaxis: {{ labels:{{ formatter:function(v){{return (v/1000).toFixed(0)+"\\u00a0Kč";}},
                        style:{{fontSize:"11px",colors:"#888"}} }} }},
    colors: ["#2770f0"],
    stroke: {{ curve:"smooth", width:3 }},
    markers: {{ size:6, colors:["#2770f0"], strokeColors:"#fff", strokeWidth:2,
                hover:{{size:8}} }},
    fill: {{ type:"gradient", gradient:{{shadeIntensity:1,opacityFrom:0.3,
             opacityTo:0.03,stops:[0,100]}} }},
    dataLabels: {{ enabled:true,
                  formatter:function(v){{return (v/1000).toFixed(0)+"\\u00a0K";}},
                  style:{{fontSize:"10px",colors:["#2770f0"],fontWeight:"600"}},
                  offsetY:-8, background:{{enabled:false}} }},
    grid: {{ borderColor:"#e8e8e8", strokeDashArray:4 }},
    tooltip: {{ y:{{ formatter:function(v){{
      return new Intl.NumberFormat("cs-CZ").format(Math.round(v))+"\\u00a0Kč";
    }} }} }}
  }}).render();
}})();
</script>"""
        else:
            age_charts_html = (
                "<p style='color:#aaa;font-size:0.85rem;font-style:italic;'>"
                "Žádná data o věkových skupinách.</p>"
                + _cizinci_html
            )

        # ── Strategické info ──────────────────────────────────────────────────
        def info_row(label, value, icon=""):
            if not value or str(value).strip() in ("", "nan", "0", "0.0", "None"):
                return ""
            return (f"<tr><td style='padding:5px 12px;color:#666;font-size:0.82rem;"
                    f"white-space:nowrap;width:220px;'>{icon} {label}</td>"
                    f"<td style='padding:5px 12px;font-weight:600;font-size:0.82rem;'>{value}</td></tr>")

        def _info_cell(icon, label, value, hl_bg=None, hl_border=None):
            """Jedna buňka v gridu strategických informací."""
            if not value or str(value).strip() in ("", "nan", "0", "0.0", "None"):
                return ""
            bg     = hl_bg     or "#fff"
            border = f"border-left:3px solid {hl_border};" if hl_border else ""
            return (f'<div style="background:{bg};border-radius:7px;padding:10px 14px;{border}'
                    f'box-shadow:0 1px 4px rgba(0,0,0,.05);">'
                    f'<div style="font-size:0.70rem;color:#999;font-weight:600;text-transform:uppercase;'
                    f'letter-spacing:.3px;">{icon} {label}</div>'
                    f'<div style="font-size:0.88rem;font-weight:700;color:#1a2a4a;margin-top:3px;">{value}</div>'
                    f'</div>')

        _nf_rok  = _v(row, "INVESTICE_NF_(ROK)")
        _fhc_rok = _v(row, "INVESTICE_FHC_(ROK)")
        _inv_rok = _v(row, "INVESTICE_NF_OR_FHC_(ROK)")
        _cl_rok  = _v(row, "PLAN_NEW_CASHIERLESS_(ROK)")
        _cl_stat = _v(row, "NEW_CASHIERLESS_STATUS", "")

        def _rok_str(v):
            if not v or str(v).strip() in ("", "nan", "0", "0.0", "None"): return None
            try: return str(int(float(v)))
            except: return str(v).strip()

        _nf_str  = _rok_str(_nf_rok)
        _fhc_str = _rok_str(_fhc_rok)
        _inv_str = _rok_str(_inv_rok)
        _cl_str  = _rok_str(_cl_rok)

        _inv_parts = []
        if _nf_str:  _inv_parts.append(f"NF v roce {_nf_str}")
        if _fhc_str: _inv_parts.append(f"FHC v roce {_fhc_str}")
        _inv_label = " + ".join(_inv_parts) if _inv_parts else (_inv_str or None)

        _cl_label = None
        if _cl_str:
            _cl_label = f"Rok {_cl_str}"
            if _cl_stat and str(_cl_stat).strip() not in ("", "nan"):
                _cl_label += f" · {_cl_stat}"

        _close_rok = _rok_str(_v(row, "PLAN_CLOSE_(ROK)"))

        _cells = [
            _info_cell("🗺️",  "Region",              region),
            _info_cell("🏢",  "Typologie",            _v(row,"BRANCH_TYPE_NAME")),
            _info_cell("📐",  "Formát pobočky",       format_p),
            _info_cell("🔨",  "Formát NF/SF",         _v(row,"BRANCH_BUILDING_NF_SF")),
            _info_cell("🎯",  "Strategie BNS 2030",   strategie),
            _info_cell("🔑",  "Vlastnictví",          vlastnictvi),
            _info_cell("💰",  "Roční nájemné",        format_money(najemne) if najemne else None),
            _info_cell("📅",  "Konec nájemní smlouvy",_v(row,"KONEC_DATUM")),
            _info_cell("🔨",  "Poslední investice",   _v(row,"DESC_LAST_INV")),
            _info_cell("💰",  "Výše poslední inv.",   format_money(_v(row,"LAST_INV")) if _v(row,"LAST_INV") else None),
            _info_cell("🔴",  "Plán uzavření",        _close_rok),
            _info_cell("🔨",  "Plánovaná investice",  _inv_label,  "#fff8ec", "#e07020"),
            _info_cell("🏦",  "Plán cashless přechod",_cl_label,   "#edf9f5", "#1a8a6a"),
        ]
        _cells_html = "".join(c for c in _cells if c)

        strat_html = f"""
<div class="age-section-card mb-4">
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;">
    {_cells_html}
  </div>
</div>"""

        # ── Detailní investice z export_investice_all_2026.pkl ────────────────
        investice_html = ""
        _inv_count = 0
        if df_investice_all is not None:
            hj_raw = _v(row, "CONCATENATED_HJ", "")
            hj_list = [h.strip() for h in str(hj_raw).split(";") if h.strip() and h.strip() != "nan"]
            if hj_list:
                df_inv_b = df_investice_all[
                    df_investice_all["HJ"].isin(hj_list)
                ].sort_values("SCHVALENO_IKCS_DATUM", ascending=False).copy()
            else:
                df_inv_b = pd.DataFrame()

            _inv_count = len(df_inv_b)
            if _inv_count > 0:
                celkem = df_inv_b["INVESTICE"].sum()
                rows_inv = ""
                for i, (_, irow) in enumerate(df_inv_b.iterrows()):
                    bg = "#ffffff" if i % 2 == 0 else "#f8f9ff"
                    datum = irow["SCHVALENO_IKCS_DATUM"]
                    datum_str = datum.strftime("%-d. %-m. %Y") if pd.notna(datum) else "—"
                    castka_str = format_money(irow["INVESTICE"]) if pd.notna(irow["INVESTICE"]) else "—"
                    jednotne = str(irow.get("JEDNOTNE_ZNACENI", "—") or "—")
                    hj_val   = str(irow.get("HJ", "—") or "—")
                    rows_inv += (
                        f"<tr style='background:{bg};'>"
                        f"<td style='padding:5px 10px;text-align:left;color:#555;white-space:nowrap;'>{datum_str}</td>"
                        f"<td style='padding:5px 10px;text-align:left;'>{jednotne}</td>"
                        f"<td style='padding:5px 10px;text-align:left;color:#888;font-size:0.8rem;'>{hj_val}</td>"
                        f"<td style='padding:5px 10px;text-align:right;font-weight:600;color:#1a3a5a;'>{castka_str}</td>"
                        f"</tr>"
                    )
                investice_html = f"""
<div class="age-section-card mb-4">
  <div style="margin-bottom:8px;font-size:0.85rem;color:#666;">
    {_inv_count} záznamů &nbsp;·&nbsp; celkem:
    <b style="color:#2770f0;">{format_money(celkem)}</b>
  </div>
  <div style="overflow-x:auto;max-height:500px;overflow-y:auto;">
    <table style="width:100%;border-collapse:collapse;font-size:0.83rem;">
      <thead>
        <tr style="background:#2770f0;color:white;position:sticky;top:0;z-index:5;">
          <th style="padding:7px 10px;text-align:left;white-space:nowrap;">Datum schválení</th>
          <th style="padding:7px 10px;text-align:left;">Jednotné značení</th>
          <th style="padding:7px 10px;text-align:left;color:#aed6ff;">HJ</th>
          <th style="padding:7px 10px;text-align:right;white-space:nowrap;">Investice (Kč)</th>
        </tr>
      </thead>
      <tbody>{rows_inv}</tbody>
      <tfoot>
        <tr style="background:#f0f4ff;font-weight:700;">
          <td colspan="3" style="padding:7px 10px;text-align:right;color:#444;">Celkem</td>
          <td style="padding:7px 10px;text-align:right;color:#2770f0;">{format_money(celkem)}</td>
        </tr>
      </tfoot>
    </table>
  </div>
</div>"""
            else:
                investice_html = "<p style='color:#aaa;font-size:0.85rem;font-style:italic;margin:8px 0;'>Žádné záznamy o investicích.</p>"

        # ── Metriky overview — přímo pro tuto pobočku (bez selectu) ──────────
        # Sestavíme kuličky kvintilů z rating_status (celá síť = základ kvintilů)
        import uuid as _uuid2, json as _json2, re as _re2
        _wid2 = "mtr_b_" + _uuid2.uuid4().hex[:8]


        def _get_q_for_metric(qsrc):
            for suffix in ['_Q', '_STATUS', '_Q_HTML']:
                col = qsrc + suffix
                if col not in _brow.columns: continue
                v = _brow.iloc[0].get(col) if len(_brow) > 0 else None
                if v is None or (isinstance(v, float) and np.isnan(v)): continue
                try:
                    vi = int(float(str(v)))
                    if 1 <= vi <= 5: return vi
                except: pass
                m = _re2.search(r'title="[^"]*\s([1-5]):', str(v))
                if m: return int(m.group(1))
            return None

        _hdr_color_m = "#2770f0"
        _hdr_dark_m  = "#1a4db5"

        _sub_hdrs_m = "".join(
            f'<th style="background:{_hdr_dark_m};filter:brightness(1.15);color:#fff;'
            f'border:1px solid rgba(255,255,255,0.15);padding:4px 7px;text-align:center;'
            f'font-size:0.74rem;white-space:nowrap;">{lbl}</th>'
            for lbl, _ in ALL_METRICS_B
        )

        # ── Radar chart — Výkonnostní metriky ────────────────────────────────

        _radar_labels = []
        _radar_scores = []      # tato pobočka
        _radar_qs     = []

        _RADAR_COLOR = {1:"#2ca02c", 2:"#6abf69", 3:"#f0a500", 4:"#e07020", 5:"#d62728", 0:"#cccccc"}

        for lbl, qsrc in ALL_METRICS_B:
            q_col = _Q_ALIAS.get(qsrc, qsrc + "_Q")
            q = _get_q_for_metric(qsrc)
            score = (6 - q) if q else 0
            _radar_labels.append(lbl)
            _radar_scores.append(score)
            _radar_qs.append(q)

        _radar_has_data = any(s > 0 for s in _radar_scores)
        _radar_chart_id = f"radar_{branch_code}"

        _tooltip_qs = [
            f"Q{q} — {TOOLTIP_M[q]}" if q else "bez dat"
            for q in _radar_qs
        ]

        _metric_pills = "".join(
            f'<div style="display:flex;align-items:center;gap:6px;padding:3px 0;">'
            f'<span style="color:{_RADAR_COLOR[q or 0]};font-size:16px;line-height:1;">●</span>'
            f'<span style="font-size:0.78rem;font-weight:600;color:#333;min-width:130px;">{lbl}</span>'
            f'<span style="font-size:0.75rem;font-weight:700;color:{_RADAR_COLOR[q or 0]};">'
            f'{"Q" + str(q) + " — " + TOOLTIP_M[q] if q else "bez dat"}</span>'
            f'</div>'
            for (lbl, _), q in zip(ALL_METRICS_B, _radar_qs)
        )

        if _radar_has_data:
            metrics_html = f"""
<div style="margin-top:36px;margin-bottom:14px;padding-top:24px;border-top:2px solid #e0e6f5;">
  <div class="section-header">📊 Výkonnostní metriky — pozice v síti</div>
  <div style="background:#fff;border-radius:10px;padding:18px;box-shadow:0 2px 8px rgba(0,0,0,.08);">

    <!-- Toggle tlačítka pro benchmark série -->
    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:14px;align-items:center;">
      <span style="font-size:0.78rem;color:#888;font-weight:600;">Porovnat s:</span>
      <button id="btn-reg-{branch_code}" onclick="toggleRadarSeries_{branch_code}('reg')"
        style="padding:4px 14px;border-radius:20px;border:2px solid #e07020;background:#fff8ec;
               color:#e07020;font-size:0.78rem;font-weight:700;cursor:pointer;transition:all .2s;">
        📍 Nejpodobnější v regionu<br>
        <span style="font-size:0.72rem;font-weight:400;color:#aaa;">{_sim_reg_name}</span>
      </button>
      <button id="btn-net-{branch_code}" onclick="toggleRadarSeries_{branch_code}('net')"
        style="padding:4px 14px;border-radius:20px;border:2px solid #7b5ea7;background:#fff;
               color:#7b5ea7;font-size:0.78rem;font-weight:700;cursor:pointer;transition:all .2s;">
        🌐 Nejpodobnější v síti<br>
        <span style="font-size:0.72rem;font-weight:400;color:#aaa;">{_sim_net_name}</span>
      </button>
      <button id="btn-best-{branch_code}" onclick="toggleRadarSeries_{branch_code}('best')"
        style="padding:4px 14px;border-radius:20px;border:2px solid #0b8a4e;background:#fff;
               color:#0b8a4e;font-size:0.78rem;font-weight:700;cursor:pointer;transition:all .2s;">
        🏆 Nejlepší v regionu ({format_p})<br>
        <span style="font-size:0.72rem;font-weight:400;color:#aaa;">{_best_fmt_name}</span>
      </button>
      <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;">
        <select id="sel-custom-{branch_code}"
          onchange="toggleRadarSeries_{branch_code}('custom-select')"
          style="padding:4px 10px;border-radius:20px;border:2px solid #bbb;background:#fff;
                 color:#444;font-size:0.78rem;cursor:pointer;max-width:230px;outline:none;">
          <option value="">➕ Přidat libovolnou pobočku…</option>
        </select>
        <button id="btn-custom-{branch_code}" onclick="toggleRadarSeries_{branch_code}('custom')"
          style="display:none;padding:4px 14px;border-radius:20px;border:2px solid #1a7abf;
                 background:#fff;color:#1a7abf;font-size:0.78rem;font-weight:700;cursor:pointer;
                 transition:all .2s;">
          ✕ Odebrat
        </button>
      </div>
    </div>

    <div style="display:flex;gap:24px;flex-wrap:wrap;align-items:flex-start;">
      <!-- Radar chart -->
      <div style="flex:1;min-width:320px;">
        <script src="https://cdn.jsdelivr.net/npm/apexcharts@3"></script>
        <div id="{_radar_chart_id}" style="min-height:400px;"></div>
      </div>
      <!-- Metric list -->
      <div style="flex:0 0 240px;padding-top:8px;">
        <div style="font-size:0.78rem;font-weight:700;color:#555;text-transform:uppercase;
                    letter-spacing:.5px;margin-bottom:10px;">Detail metrik — {branch_name}</div>
        {_metric_pills}
        <div style="margin-top:14px;padding-top:10px;border-top:1px solid #eee;
                    font-size:0.72rem;color:#aaa;line-height:1.6;">
          Osa: Q1 = nejlepší (vrchol) → Q5 = nejhorší (střed)<br>
          Šedá = data nejsou k dispozici
        </div>
      </div>
    </div>
  </div>
</div>

<script>
(function() {{
  var cid   = "{_radar_chart_id}";
  var branchName  = {_json_radar.dumps(branch_name)};
  var simRegName  = {_json_radar.dumps(_sim_reg_name)};
  var simNetName  = {_json_radar.dumps(_sim_net_name)};
  var bestFmtName = {_json_radar.dumps(_best_fmt_name)};
  var labels      = {_json_radar.dumps(_radar_labels)};
  var scBranch    = {_json_radar.dumps(_radar_scores)};
  var scReg       = {_json_radar.dumps(_sim_reg_scores)};
  var scNet       = {_json_radar.dumps(_sim_net_scores)};
  var scBest      = {_json_radar.dumps(_best_fmt_scores)};
  var tooltips    = {_json_radar.dumps(_tooltip_qs)};
  var allBranches = {_all_branches_json};

  var showReg = false, showNet = false, showBest = false;
  var customData = null;  // {{name, scores}} nebo null
  var chart = null;

  // Naplníme select
  var sel = document.getElementById("sel-custom-{branch_code}");
  allBranches.forEach(function(b) {{
    var opt = document.createElement("option");
    opt.value = b.code;
    opt.text  = b.name + " (" + b.format + ", " + b.region + ")";
    sel.appendChild(opt);
  }});

  function buildSeries() {{
    var s = [{{ name: branchName, data: scBranch }}];
    if (showReg)  s.push({{ name: "📍 " + simRegName,  data: scReg }});
    if (showNet)  s.push({{ name: "🌐 " + simNetName,  data: scNet }});
    if (showBest) s.push({{ name: "🏆 " + bestFmtName, data: scBest }});
    if (customData) s.push({{ name: "➕ " + customData.name, data: customData.scores }});
    return s;
  }}

  function buildColors() {{
    var c = ["#2770f0"];
    if (showReg)  c.push("#e07020");
    if (showNet)  c.push("#7b5ea7");
    if (showBest) c.push("#0b8a4e");
    if (customData) c.push("#1a7abf");
    return c;
  }}

  function buildDash() {{
    var d = [0];
    if (showReg)  d.push(4);
    if (showNet)  d.push(6);
    if (showBest) d.push(3);
    if (customData) d.push(8);
    return d;
  }}

  function buildFill() {{
    var f = [0.35];
    if (showReg)  f.push(0.15);
    if (showNet)  f.push(0.15);
    if (showBest) f.push(0.15);
    if (customData) f.push(0.12);
    return f;
  }}

  function renderChart() {{
    var options = {{
      chart: {{
        type: "radar", height: 400,
        toolbar: {{ show: false }},
        background: "transparent",
        animations: {{ enabled: true, speed: 500 }}
      }},
      series: buildSeries(),
      colors: buildColors(),
      xaxis: {{
        categories: labels,
        labels: {{ style: {{ fontSize: "11px", fontWeight: "600",
          colors: Array(labels.length).fill("#444") }} }}
      }},
      yaxis: {{
        min: 0, max: 5, tickAmount: 5,
        labels: {{
          formatter: function(v) {{
            return {{5:"Q1",4:"Q2",3:"Q3",2:"Q4",1:"Q5",0:""}}[Math.round(v)] || "";
          }},
          style: {{ fontSize: "10px", colors: "#aaa" }}
        }}
      }},
      fill: {{ opacity: buildFill() }},
      stroke: {{ width: Array(buildSeries().length).fill(2.5), dashArray: buildDash() }},
      markers: {{ size: 4, strokeColors: "#fff", strokeWidth: 2, hover: {{ size: 7 }} }},
      legend: {{
        show: true, position: "top",
        fontSize: "12px",
        labels: {{ colors: "#444" }}
      }},
      tooltip: {{
        custom: function(o) {{
          var di = o.dataPointIndex;
          var lbl = labels[di];
          var html = '<div style="padding:8px 12px;font-size:12px;background:#fff;'
                   + 'border:1px solid #ddd;border-radius:6px;min-width:160px;">'
                   + '<b>' + lbl + '</b><br>';
          o.series.forEach(function(sArr, si) {{
            var s = sArr[di];
            var qn = s > 0 ? "Q" + (6 - s) : "—";
            var clr = buildColors()[si];
            var sName = o.w.config.series[si].name;
            html += '<span style="color:' + clr + ';font-weight:700;">'
                  + sName.replace(/^[^A-Za-z0-9]*/, "").substring(0,22) + ': ' + qn
                  + '</span><br>';
          }});
          html += '<span style="color:#aaa;font-size:11px;">' + tooltips[di] + '</span>';
          html += '</div>';
          return html;
        }}
      }},
      plotOptions: {{
        radar: {{
          size: 155,
          polygons: {{
            strokeColors: "#e8e8e8",
            fill: {{ colors: ["#f8f9ff", "#fff"] }}
          }}
        }}
      }}
    }};
    if (chart) {{ chart.destroy(); }}
    chart = new ApexCharts(document.getElementById(cid), options);
    chart.render();
  }}

  window["toggleRadarSeries_{branch_code}"] = function(type) {{
    var btnReg    = document.getElementById("btn-reg-{branch_code}");
    var btnNet    = document.getElementById("btn-net-{branch_code}");
    var btnBest   = document.getElementById("btn-best-{branch_code}");
    var btnCustom = document.getElementById("btn-custom-{branch_code}");
    var selCustom = document.getElementById("sel-custom-{branch_code}");

    if (type === "reg") {{
      showReg = !showReg;
      btnReg.style.background    = showReg ? "#fff0e0" : "#fff";
      btnReg.style.borderWidth   = showReg ? "2.5px" : "2px";
    }} else if (type === "net") {{
      showNet = !showNet;
      btnNet.style.background    = showNet ? "#f3eeff" : "#fff";
      btnNet.style.borderWidth   = showNet ? "2.5px" : "2px";
    }} else if (type === "best") {{
      showBest = !showBest;
      btnBest.style.background   = showBest ? "#e6f7f0" : "#fff";
      btnBest.style.borderWidth  = showBest ? "2.5px" : "2px";
    }} else if (type === "custom-select") {{
      var val = selCustom.value;
      if (!val) {{ customData = null; btnCustom.style.display = "none"; renderChart(); return; }}
      var found = allBranches.find(function(b) {{ return String(b.code) === String(val); }});
      if (found) {{
        customData = {{ name: found.name, scores: found.scores }};
        btnCustom.style.display = "inline-block";
        btnCustom.textContent = "✕ " + found.name.substring(0,20);
      }}
    }} else if (type === "custom") {{
      customData = null;
      btnCustom.style.display = "none";
      selCustom.value = "";
    }}
    renderChart();
  }};

  renderChart();
}})();
</script>"""
        else:
            metrics_html = f"""
<div style="margin-top:36px;margin-bottom:14px;padding-top:24px;border-top:2px solid #e0e6f5;">
  <div class="section-header">📊 Výkonnostní metriky — pozice v síti</div>
  <p style="color:#aaa;font-style:italic;font-size:0.85rem;">Nejsou k dispozici data pro výpočet kvintilů.</p>
</div>"""

        # ── Mapa návštěvnosti (visits) ────────────────────────────────────────
        visits_html = ""
        if visits is not None and not visits.empty:
            _vcols = [c.upper() for c in visits.columns]
            _vdf = visits.copy()
            _vdf.columns = _vcols

            # Detekce ID sloupce
            _vid = next((c for c in ['BRANCH_ID','BRANCH_CODE','POBOCKA'] if c in _vcols), None)
            if _vid:
                _vb = _vdf[pd.to_numeric(_vdf[_vid], errors='coerce') == branch_code].copy()
            else:
                _vb = pd.DataFrame()

            if not _vb.empty:
                # Parsování datumu a času
                if 'VISIT_DATE' in _vb.columns:
                    _vb['_dt'] = pd.to_datetime(_vb['VISIT_DATE'], errors='coerce')
                    _vb['_month']   = _vb['_dt'].dt.month
                    _vb['_weekday'] = _vb['_dt'].dt.weekday   # 0=Po … 4=Pá
                    _vb['_day']     = _vb['_dt'].dt.day
                if 'VISIT_TIME' in _vb.columns:
                    _vb['_hour'] = pd.to_numeric(
                        _vb['VISIT_TIME'].astype(str).str.split(':').str[0], errors='coerce')
                else:
                    _vb['_hour'] = None

                _total_v = len(_vb)

                # ── Agregace ──────────────────────────────────────────────────
                # 1) Po měsících (1-12)
                _by_month = _vb.groupby('_month').size().reindex(range(1,13), fill_value=0)
                _month_names = ['Led','Úno','Bře','Dub','Kvě','Čvn','Čvc','Srp','Zář','Říj','Lis','Pro']
                _month_max = max(_by_month.max(), 1)

                # 2) Po dnech v týdnu (Po-Pá, event. So)
                _by_wd = _vb.groupby('_weekday').size().reindex(range(7), fill_value=0)
                _wd_names = ['Po','Út','St','Čt','Pá','So','Ne']
                _wd_max = max(_by_wd.max(), 1)

                # 3) Po hodinách (6-19)
                _by_hour = _vb.groupby('_hour').size() if _vb['_hour'].notna().any() else pd.Series(dtype=int)
                _hour_range = list(range(6, 20))
                _by_hour = _by_hour.reindex(_hour_range, fill_value=0)
                _hour_max = max(_by_hour.max(), 1)

                # 4) Heatmapa: weekday × hour
                _has_hour = _vb['_hour'].notna().any()
                if _has_hour:
                    _hm = _vb.groupby(['_weekday','_hour']).size().unstack(fill_value=0)
                    _hm = _hm.reindex(index=range(5), columns=_hour_range, fill_value=0)
                    _hm_max = max(_hm.values.max(), 1)

                # ── Segment mix ───────────────────────────────────────────────
                _seg_html = ""
                if 'CLIENT_SEGMENT' in _vb.columns:
                    _by_seg = _vb['CLIENT_SEGMENT'].value_counts().head(6)
                    _seg_total = _by_seg.sum()
                    _seg_colors = ['#2770f0','#45b065','#e07020','#9b6bbf','#2aacbd','#b89a00']
                    _seg_bars = ""
                    for _si, (_sn, _sc2) in enumerate(_by_seg.items()):
                        _sp = _sc2/_seg_total*100 if _seg_total>0 else 0
                        _sc3 = _seg_colors[_si % len(_seg_colors)]
                        _seg_bars += f"""
<div style="margin-bottom:5px;display:flex;align-items:center;gap:8px;">
  <span style="font-size:0.75rem;font-weight:600;color:#444;min-width:36px;">{_sn}</span>
  <div style="flex:1;background:#f0f0f0;border-radius:4px;height:10px;overflow:hidden;">
    <div style="width:{_sp:.1f}%;background:{_sc3};height:100%;border-radius:4px;"></div>
  </div>
  <span style="font-size:0.73rem;color:#888;min-width:42px;text-align:right;">{_sp:.1f}%</span>
</div>"""
                    _seg_html = f"""
<div style="background:#fafafa;border-radius:8px;padding:12px 14px;border:1px solid #eee;">
  <div style="font-size:0.75rem;font-weight:700;color:#666;text-transform:uppercase;
              letter-spacing:.4px;margin-bottom:8px;">Segmenty klientů</div>
  {_seg_bars}
</div>"""

                # ── Typ návštěvy ───────────────────────────────────────────────
                _att_html = ""
                if 'ATTENDANCE_TYPE' in _vb.columns:
                    _by_att = _vb['ATTENDANCE_TYPE'].value_counts().head(6)
                    _att_total = _by_att.sum()
                    _att_colors = ['#e07020','#2770f0','#45b065','#9b6bbf','#2aacbd','#b89a00']
                    _att_bars = ""
                    for _ai, (_an, _ac) in enumerate(_by_att.items()):
                        _ap = _ac/_att_total*100 if _att_total>0 else 0
                        _ac2 = _att_colors[_ai % len(_att_colors)]
                        _att_bars += f"""
<div style="margin-bottom:5px;display:flex;align-items:center;gap:8px;">
  <span style="font-size:0.72rem;color:#444;flex:1;min-width:0;overflow:hidden;
               text-overflow:ellipsis;white-space:nowrap;" title="{_an}">{_an}</span>
  <div style="width:90px;background:#f0f0f0;border-radius:4px;height:10px;overflow:hidden;flex-shrink:0;">
    <div style="width:{_ap:.1f}%;background:{_ac2};height:100%;border-radius:4px;"></div>
  </div>
  <span style="font-size:0.73rem;color:#888;min-width:38px;text-align:right;">{_ap:.1f}%</span>
</div>"""
                    _att_html = f"""
<div style="background:#fafafa;border-radius:8px;padding:12px 14px;border:1px solid #eee;margin-top:10px;">
  <div style="font-size:0.75rem;font-weight:700;color:#666;text-transform:uppercase;
              letter-spacing:.4px;margin-bottom:8px;">Typ návštěvy</div>
  {_att_bars}
</div>"""

                # ── Pomocná funkce: bar buňka s heat barvou ───────────────────
                def _hcell(val, max_val, height='28px', width='100%'):
                    _p = val/max_val*100 if max_val>0 else 0
                    # interpolace: nízká=#e8f4fd → střední=#2770f0 → vysoká=#c0392b
                    _r = int(40 + (_p/100)*195)
                    _g = int(112 - (_p/100)*72)
                    _b = int(240 - (_p/100)*197)
                    _col2 = f'rgb({_r},{_g},{_b})'
                    _txt = '#fff' if _p > 45 else '#333'
                    return (f'<td style="padding:3px 4px;text-align:center;">'
                            f'<div style="background:{_col2 if val>0 else "#f5f5f5"};'
                            f'border-radius:4px;height:{height};width:{width};'
                            f'display:flex;align-items:center;justify-content:center;'
                            f'font-size:0.68rem;font-weight:{"700" if _p>55 else "400"};'
                            f'color:{_txt if val>0 else "#ccc"};">'
                            f'{val:,}</div></td>')

                # ── HTML: barchart po měsících ────────────────────────────────
                _month_cells = ""
                for _mi in range(1,13):
                    _mv = int(_by_month.get(_mi, 0))
                    _mp = _mv/_month_max*100
                    _mc = f'rgb({int(40+_mp/100*195)},{int(112-_mp/100*72)},{int(240-_mp/100*197)})'
                    _month_cells += f"""
<div style="display:flex;flex-direction:column;align-items:center;gap:3px;flex:1;min-width:0;">
  <div style="font-size:0.65rem;color:#888;">{_month_names[_mi-1]}</div>
  <div style="width:100%;background:#f0f0f0;border-radius:4px;height:80px;
              display:flex;align-items:flex-end;overflow:hidden;">
    <div style="width:100%;background:{_mc if _mv>0 else '#f0f0f0'};
                height:{max(_mp,2):.1f}%;border-radius:4px 4px 0 0;
                transition:height .4s;"></div>
  </div>
  <div style="font-size:0.63rem;color:#555;font-weight:600;">{_mv:,}</div>
</div>"""

                # ── HTML: barchart po dnech v týdnu ──────────────────────────
                _wd_cells = ""
                for _wi in range(5):   # Po–Pá (So a Ne vynechat pokud 0)
                    _wv = int(_by_wd.get(_wi, 0))
                    _wp = _wv/_wd_max*100
                    _wc = f'rgb({int(40+_wp/100*195)},{int(112-_wp/100*72)},{int(240-_wp/100*197)})'
                    _wd_cells += f"""
<div style="display:flex;flex-direction:column;align-items:center;gap:3px;flex:1;">
  <div style="font-size:0.72rem;color:#888;font-weight:600;">{_wd_names[_wi]}</div>
  <div style="width:100%;background:#f0f0f0;border-radius:4px;height:70px;
              display:flex;align-items:flex-end;overflow:hidden;">
    <div style="width:100%;background:{_wc if _wv>0 else '#f0f0f0'};
                height:{max(_wp,2):.1f}%;border-radius:4px 4px 0 0;"></div>
  </div>
  <div style="font-size:0.7rem;color:#333;font-weight:700;">{_wv:,}</div>
</div>"""
                # So/Ne pokud mají data
                for _wi in range(5,7):
                    _wv = int(_by_wd.get(_wi, 0))
                    if _wv > 0:
                        _wp = _wv/_wd_max*100
                        _wc = f'rgb({int(40+_wp/100*195)},{int(112-_wp/100*72)},{int(240-_wp/100*197)})'
                        _wd_cells += f"""
<div style="display:flex;flex-direction:column;align-items:center;gap:3px;flex:1;opacity:.7;">
  <div style="font-size:0.72rem;color:#888;font-weight:600;">{_wd_names[_wi]}</div>
  <div style="width:100%;background:#f0f0f0;border-radius:4px;height:70px;
              display:flex;align-items:flex-end;overflow:hidden;">
    <div style="width:100%;background:{_wc};height:{max(_wp,2):.1f}%;border-radius:4px 4px 0 0;"></div>
  </div>
  <div style="font-size:0.7rem;color:#333;font-weight:700;">{_wv:,}</div>
</div>"""

                # ── HTML: barchart po hodinách ────────────────────────────────
                _hour_cells = ""
                if _has_hour:
                    for _hi in _hour_range:
                        _hv = int(_by_hour.get(_hi, 0))
                        _hpp = _hv/_hour_max*100
                        _hc = f'rgb({int(40+_hpp/100*195)},{int(112-_hpp/100*72)},{int(240-_hpp/100*197)})'
                        _hour_cells += f"""
<div style="display:flex;flex-direction:column;align-items:center;gap:2px;flex:1;min-width:0;">
  <div style="font-size:0.6rem;color:#888;">{_hi}</div>
  <div style="width:100%;background:#f0f0f0;border-radius:3px;height:60px;
              display:flex;align-items:flex-end;overflow:hidden;">
    <div style="width:100%;background:{_hc if _hv>0 else '#f0f0f0'};
                height:{max(_hpp,2):.1f}%;border-radius:3px 3px 0 0;"></div>
  </div>
  <div style="font-size:0.58rem;color:#444;">{_hv:,}</div>
</div>"""

                # ── HTML: heatmapa weekday × hour ─────────────────────────────
                _hm_html = ""
                if _has_hour:
                    _hm_header = "<tr><th style='padding:4px 6px;font-size:0.68rem;color:#888;'>Den↓ Hod→</th>"
                    for _hi in _hour_range:
                        _hm_header += f"<th style='padding:3px 4px;font-size:0.65rem;color:#999;text-align:center;'>{_hi}h</th>"
                    _hm_header += "</tr>"
                    _hm_rows = ""
                    for _wi in range(5):
                        _hm_rows += f"<tr><td style='padding:3px 8px;font-size:0.72rem;font-weight:600;color:#555;white-space:nowrap;'>{_wd_names[_wi]}</td>"
                        for _hi in _hour_range:
                            _hv2 = int(_hm.loc[_wi, _hi]) if _wi in _hm.index and _hi in _hm.columns else 0
                            _hm_rows += _hcell(_hv2, _hm_max)
                        _hm_rows += "</tr>"
                    _hm_html = f"""
<div style="margin-top:16px;">
  <div style="font-size:0.78rem;font-weight:700;color:#555;text-transform:uppercase;
              letter-spacing:.4px;margin-bottom:8px;">🗓️ Heatmapa den × hodina</div>
  <div style="overflow-x:auto;">
  <table style="border-collapse:collapse;width:100%;table-layout:fixed;">
    <thead style="background:#f8faff;">{_hm_header}</thead>
    <tbody>{_hm_rows}</tbody>
  </table>
  </div>
</div>"""

                # ── Join s parties_full → GPS + subsegment ────────────────────
                _geo_map_html = ""
                _subseg_html  = ""

                if parties_full is not None and not parties_full.empty:
                    _pf = parties_full.copy()
                    _pf.columns = [c.upper() for c in _pf.columns]

                    # ── Filtr přímo dle BRANCH_ID z parties_full ─────────────
                    _pf_bid = next((c for c in ['BRANCH_ID','DBS_HOME_BRANCH_CODE','BRANCH_CODE'] if c in _pf.columns), None)
                    if _pf_bid:
                        _pf_b = _pf[pd.to_numeric(_pf[_pf_bid], errors='coerce') == branch_code].copy()
                    else:
                        # fallback: join přes PT_UNIFIED_KEY z visits
                        if 'PT_UNIFIED_KEY' in _vb.columns and 'PT_UNIFIED_KEY' in _pf.columns:
                            _vb_keys = set(pd.to_numeric(_vb['PT_UNIFIED_KEY'], errors='coerce').dropna().astype(int))
                            _pf['PT_UNIFIED_KEY'] = pd.to_numeric(_pf['PT_UNIFIED_KEY'], errors='coerce')
                            _pf_b = _pf[_pf['PT_UNIFIED_KEY'].isin(_vb_keys)].copy()
                        else:
                            _pf_b = pd.DataFrame()

                    # ── GPS heatmapa z parties_full ───────────────────────────
                    if not _pf_b.empty and 'LOC_GEO_LATITUDE' in _pf_b.columns and 'LOC_GEO_LONGITUDE' in _pf_b.columns:
                        _geo = _pf_b[['LOC_GEO_LATITUDE','LOC_GEO_LONGITUDE']].copy()
                        _geo['_lat'] = pd.to_numeric(_geo['LOC_GEO_LATITUDE'], errors='coerce')
                        _geo['_lon'] = pd.to_numeric(_geo['LOC_GEO_LONGITUDE'], errors='coerce')
                        _geo = _geo.dropna(subset=['_lat','_lon'])
                        _geo = _geo[(_geo['_lat'].between(48.5, 51.2)) & (_geo['_lon'].between(12.0, 18.8))]

                        if len(_geo) > 0:
                            import plotly.graph_objects as _pgo

                            _map_clat = float(row.get('GPS_X') or _geo['_lat'].median())
                            _map_clon = float(row.get('GPS_Y') or _geo['_lon'].median())
                            _geo_cnt  = len(_geo)

                            # Agregace na mřížku 0.01° (~700m)
                            _geo['_glat'] = (_geo['_lat'] / 0.01).round() * 0.01
                            _geo['_glon'] = (_geo['_lon'] / 0.01).round() * 0.01
                            _grid = _geo.groupby(['_glat','_glon']).size().reset_index(name='cnt')

                            _fig = _pgo.Figure()
                            _fig.add_trace(_pgo.Densitymapbox(
                                lat=_grid['_glat'],
                                lon=_grid['_glon'],
                                z=_grid['cnt'],
                                radius=20,
                                colorscale=[
                                    [0.0,  'rgba(39,112,240,0)'],
                                    [0.2,  'rgba(39,112,240,0.5)'],
                                    [0.45, 'rgba(100,140,240,0.75)'],
                                    [0.7,  'rgba(220,100,60,0.88)'],
                                    [1.0,  'rgba(253,98,48,1.0)'],
                                ],
                                showscale=True,
                                colorbar=dict(
                                    title=dict(text='Klientů'),
                                    thickness=12, len=0.6,
                                    tickvals=[], ticktext=[],
                                ),
                                hovertemplate='Klientů v oblasti: %{z}<extra></extra>',
                            ))
                            _fig.add_trace(_pgo.Scattermapbox(
                                lat=[_map_clat], lon=[_map_clon],
                                mode='markers+text',
                                marker=dict(size=14, color='#2770f0'),
                                text=['Pobočka'], textposition='top right',
                                textfont=dict(size=11, color='#2770f0'),
                                hovertext='📍 Tato pobočka', hoverinfo='text',
                                showlegend=False,
                            ))
                            _fig.update_layout(
                                mapbox=_mapbox_layout(center_lat=_map_clat, center_lon=_map_clon, zoom=7),
                                height=420,
                                margin=dict(l=0, r=0, t=0, b=0),
                                paper_bgcolor='white',
                            )
                            _map_div = _fig.to_html(
                                full_html=False, include_plotlyjs='cdn',
                                div_id=f"geomap_{branch_code}",
                                config={'displayModeBar': False, 'scrollZoom': True},
                            )
                            _geo_map_html = f"""
<div style="margin-top:18px;">
  <div style="font-size:0.78rem;font-weight:700;color:#555;text-transform:uppercase;
              letter-spacing:.4px;margin-bottom:6px;">
    🗺️ Odkud přicházejí klienti — hustota adres
    <span style="font-size:0.7rem;font-weight:400;color:#aaa;margin-left:8px;">
      {_geo_cnt:,} klientů s GPS adresou
    </span>
  </div>
  <div style="border-radius:10px;overflow:hidden;border:1px solid #dde4f5;">
    {_map_div}
  </div>
</div>"""

                    # ── CP_SUBSEGMENT_ID z parties_full ───────────────────────
                    if not _pf_b.empty and 'CP_SUBSEGMENT_ID' in _pf_b.columns:
                        _SS_LABELS = {
                            'SK1': ('SK1 — Děti',               'klienti do 15 let'),
                            'SK2': ('SK2 — Mladí lidé',         'klienti 15–26 let'),
                            'SK3': ('SK3 — Vysokoškoláci',      'studenti VŠ'),
                            'SK4': ('SK4 — Zaměstnanci',        'zaměstnaní klienti'),
                            'SK5': ('SK5 — Svobodná povolání',  'OSVČ, svobodná povolání'),
                            'SK6': ('SK6 — Zaměstnanci FS ČS', 'zaměstnanci finanční skupiny ČS'),
                            'SK7': ('SK7 — Ostatní',            'ostatní segmenty'),
                            'SK8': ('SK8 — Důchodci',           'klienti v důchodu'),
                            'SK9': ('SK9 — Podnikatelé',        'podnikatelé a firmy'),
                        }
                        _SS_COLORS = {
                            'SK1': '#f472b6', 'SK2': '#3b82f6', 'SK3': '#8b5cf6',
                            'SK4': '#10b981', 'SK5': '#f59e0b', 'SK6': '#2770f0',
                            'SK7': '#6b7280', 'SK8': '#ef4444', 'SK9': '#d97706',
                        }
                        _ss_raw = _pf_b['CP_SUBSEGMENT_ID'].dropna().astype(str).str.strip().str.upper()
                        _ss_raw = _ss_raw[_ss_raw != '']
                        _by_subseg = _ss_raw.value_counts()
                        if len(_by_subseg) > 0:
                            _ss_total = _by_subseg.sum()
                            _ss_bars = ""
                            for _sskey, _ssc in _by_subseg.items():
                                _ssp  = _ssc / _ss_total * 100
                                if _sskey in _SS_LABELS:
                                    _sslbl, _ssdesc = _SS_LABELS[_sskey]
                                    _sscl = _SS_COLORS.get(_sskey, '#888')
                                else:
                                    _sslbl, _ssdesc = _sskey, ''
                                    _sscl = '#aaa'
                                _ss_bars += f"""
<div style="margin-bottom:10px;">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:3px;">
    <span style="display:inline-block;width:10px;height:10px;border-radius:50%;
                 background:{_sscl};flex-shrink:0;"></span>
    <span style="font-size:0.8rem;font-weight:700;color:#1a2a4a;flex:1;">{_sslbl}</span>
    <span style="font-size:0.73rem;font-weight:800;color:{_sscl};">{_ssp:.1f}%</span>
    <span style="font-size:0.7rem;color:#aaa;">{int(_ssc):,}×</span>
  </div>
  <div style="flex:1;background:#efefef;border-radius:4px;height:7px;overflow:hidden;">
    <div style="width:{_ssp:.1f}%;background:{_sscl};height:100%;border-radius:4px;"></div>
  </div>
  {"" if not _ssdesc else f'<div style="font-size:0.67rem;color:#aaa;margin-top:2px;padding-left:18px;font-style:italic;">{_ssdesc}</div>'}
</div>"""
                            _subseg_html = f"""
<div style="background:#fafafa;border-radius:8px;padding:12px 14px;
            border:1px solid #eee;margin-top:10px;">
  <div style="font-size:0.75rem;font-weight:700;color:#444;text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:12px;">🏷️ Subsegmenty klientů</div>
  {_ss_bars}
</div>"""

                visits_html = f"""
<div style="margin-top:28px;padding-top:20px;border-top:2px solid #e0e6f5;">
  <div class="section-header">🚶 Mapa návštěvnosti 2025</div>
  <div style="background:#fff;border-radius:12px;padding:18px 20px;
              box-shadow:0 2px 8px rgba(0,0,0,.07);margin-bottom:14px;">

    <!-- KPI řádek -->
    <div style="display:flex;gap:14px;flex-wrap:wrap;margin-bottom:18px;">
      <div style="background:#f0f4ff;border-radius:8px;padding:10px 16px;
                  border-left:4px solid #2770f0;min-width:110px;">
        <div style="font-size:0.7rem;color:#888;font-weight:600;text-transform:uppercase;">Návštěv celkem</div>
        <div style="font-size:1.6rem;font-weight:800;color:#2770f0;line-height:1.1;">{_total_v:,}</div>
      </div>
      <div style="background:#f0fff4;border-radius:8px;padding:10px 16px;
                  border-left:4px solid #27ae60;min-width:110px;">
        <div style="font-size:0.7rem;color:#888;font-weight:600;text-transform:uppercase;">Průměr / den</div>
        <div style="font-size:1.6rem;font-weight:800;color:#27ae60;line-height:1.1;">{_total_v/252:.1f}</div>
      </div>
      <div style="background:#fff8f0;border-radius:8px;padding:10px 16px;
                  border-left:4px solid #e07020;min-width:110px;">
        <div style="font-size:0.7rem;color:#888;font-weight:600;text-transform:uppercase;">Nejfrekv. měsíc</div>
        <div style="font-size:1.3rem;font-weight:800;color:#e07020;line-height:1.1;">{_month_names[int(_by_month.idxmax())-1]}</div>
        <div style="font-size:0.7rem;color:#aaa;">{int(_by_month.max()):,} návštěv</div>
      </div>
      <div style="background:#f5f0ff;border-radius:8px;padding:10px 16px;
                  border-left:4px solid #9b6bbf;min-width:110px;">
        <div style="font-size:0.7rem;color:#888;font-weight:600;text-transform:uppercase;">Nejfrekv. den</div>
        <div style="font-size:1.3rem;font-weight:800;color:#9b6bbf;line-height:1.1;">{_wd_names[int(_by_wd[:5].idxmax())]}</div>
        <div style="font-size:0.7rem;color:#aaa;">{int(_by_wd[:5].max()):,} návštěv</div>
      </div>
      {"" if not _has_hour else f'''<div style="background:#f0fbff;border-radius:8px;padding:10px 16px;border-left:4px solid #2aacbd;min-width:110px;"><div style="font-size:0.7rem;color:#888;font-weight:600;text-transform:uppercase;">Špička (hod.)</div><div style="font-size:1.3rem;font-weight:800;color:#2aacbd;line-height:1.1;">{int(_by_hour.idxmax())}:00</div><div style="font-size:0.7rem;color:#aaa;">{int(_by_hour.max()):,} návštěv</div></div>'''}
    </div>

    <!-- Grafy + mix -->
    <div style="display:grid;grid-template-columns:1fr 260px;gap:18px;align-items:start;">
      <div>
        <!-- Měsíce -->
        <div style="font-size:0.78rem;font-weight:700;color:#555;text-transform:uppercase;
                    letter-spacing:.4px;margin-bottom:8px;">📅 Návštěvnost po měsících</div>
        <div style="display:flex;gap:4px;align-items:flex-end;height:110px;margin-bottom:18px;">
          {_month_cells}
        </div>

        <!-- Dny v týdnu -->
        <div style="font-size:0.78rem;font-weight:700;color:#555;text-transform:uppercase;
                    letter-spacing:.4px;margin-bottom:8px;">📆 Návštěvnost dle dne v týdnu</div>
        <div style="display:flex;gap:8px;align-items:flex-end;height:95px;margin-bottom:18px;">
          {_wd_cells}
        </div>

        <!-- Hodiny -->
        {"" if not _has_hour else f'''
        <div style="font-size:0.78rem;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.4px;margin-bottom:8px;">⏰ Návštěvnost dle hodiny dne</div>
        <div style="display:flex;gap:3px;align-items:flex-end;height:80px;margin-bottom:6px;">{_hour_cells}</div>
        '''}
      </div>
      <div>
        {_seg_html}
        {_att_html}
        {_subseg_html}
      </div>
    </div>

    {_hm_html}
    {_geo_map_html}
  </div>
</div>"""

        # ── Bankomaty na pobočce ───────────────────────────────────────────────
        atm_html = ""
        _atm_df  = export_atm  if export_atm  is not None else None
        _kap_df  = kapacita_atm if kapacita_atm is not None else None

        if _atm_df is not None and not _atm_df.empty:
            # Detekce ID sloupce
            _aid = next((c for c in ['BRANCH_ID','BRANCH_CODE','POBOCKA'] if c in _atm_df.columns), None)
            if _aid:
                _atm_b = _atm_df[pd.to_numeric(_atm_df[_aid], errors='coerce') == branch_code].copy()
            else:
                _atm_b = pd.DataFrame()

            if not _atm_b.empty:
                # ── Kapacita ATM ─────────────────────────────────────────────
                _kap_html = ""
                if _kap_df is not None and not _kap_df.empty:
                    _kid = next((c for c in ['BRANCH_ID','BRANCH_CODE'] if c in _kap_df.columns), None)
                    if _kid:
                        _kap_b = _kap_df[pd.to_numeric(_kap_df[_kid], errors='coerce') == branch_code]
                        if not _kap_b.empty:
                            _kr = _kap_b.iloc[0]
                            def _kv(col): 
                                v = _kr.get(col)
                                try: return float(v) * 100
                                except: return None
                            _akt   = _kv('AKTUALNI_VYTIZENI_STROJU')
                            _prev  = _kv('VYTIZENI_VSECH_ATM_PO_PREVODU')
                            _volno = _kv('ZBYVAJICI_VOLNA_KAPACITA')

                            def _pocet(col):
                                v = _kr.get(col)
                                try: return round(float(v), 1)
                                except: return None
                            _p_vyberu = _pocet('POCET_ATM_VYBERU_PO_PREVODU_/_M')
                            _p_vkladu = _pocet('POCET_ATM_VKLADU_PO_PREVODU_/_M')

                            def _kap_bar(pct, color, label, sublabel=""):
                                pct_c = min(max(pct or 0, 0), 200)  # cap na 200% pro vizuál
                                bar_w = min(pct_c / 2, 100)  # bar jde do 100% šířky = 200% vytížení
                                warn  = pct_c > 100
                                return f"""
<div style="background:{'#fff5ef' if warn else '#f8f9ff'};border:1px solid {'#f5c9a8' if warn else '#e0e6f5'};
            border-radius:8px;padding:10px 14px;flex:1;min-width:140px;">
  <div style="font-size:0.68rem;color:#888;font-weight:600;text-transform:uppercase;">{label}</div>
  <div style="font-size:1.4rem;font-weight:800;color:{'#d62728' if warn else color};line-height:1.1;margin:3px 0;">
    {pct_c:.0f} %{'  ⚠️' if warn else ''}</div>
  <div style="background:#eee;border-radius:6px;height:7px;overflow:hidden;margin-top:4px;">
    <div style="width:{bar_w:.1f}%;background:{color};height:100%;border-radius:6px;transition:width .6s;"></div>
  </div>
  {f'<div style="font-size:0.68rem;color:#aaa;margin-top:3px;">{sublabel}</div>' if sublabel else ''}
</div>"""

                            def _pocet_card(val, label, icon, color):
                                if val is None: return ''
                                return f"""
<div style="background:#f8f9ff;border:1px solid #e0e6f5;border-radius:8px;
            padding:10px 14px;flex:1;min-width:140px;">
  <div style="font-size:0.68rem;color:#888;font-weight:600;text-transform:uppercase;">{label}</div>
  <div style="font-size:1.4rem;font-weight:800;color:{color};line-height:1.1;margin:3px 0;">
    {icon} {val:,.0f}</div>
  <div style="font-size:0.68rem;color:#aaa;margin-top:2px;">operací / měsíc</div>
</div>"""

                            _kap_html = f"""
<div style="margin-bottom:14px;">
  <div style="font-size:0.78rem;font-weight:700;color:#555;text-transform:uppercase;
              letter-spacing:.4px;margin-bottom:8px;">📊 Kapacita ATM</div>
  <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:8px;">
    {_kap_bar(_akt,  '#e07020', 'Aktuální vytížení', 'stávající stroje') if _akt  is not None else ''}
    {_kap_bar(_prev, '#2770f0', 'Vytížení po převodu', 'po přesunu hot. operací') if _prev is not None else ''}
    {_kap_bar(_volno,'#2ca02c', 'Volná kapacita', 'zbývá pro nové operace') if _volno is not None else ''}
  </div>
  <div style="display:flex;gap:10px;flex-wrap:wrap;">
    {_pocet_card(_p_vyberu, 'Výběry po převodu', '⬆️', '#e07020')}
    {_pocet_card(_p_vkladu, 'Vklady po převodu',  '⬇️', '#2770f0')}
  </div>
</div>"""

                # ── Tabulka bankomatů ─────────────────────────────────────────
                _atm_rows = ""
                for _ri, (_, _ar) in enumerate(_atm_b.iterrows()):
                    def _av(col, default='—'):
                        v = _ar.get(col)
                        return default if (v is None or (isinstance(v, float) and pd.isna(v)) or str(v) in ('nan','None','')) else str(v)
                    _bg = '#f9f9f9' if _ri % 2 == 0 else '#fff'
                    _loc_badge = (
                        '<span style="background:#e8f0ff;color:#2770f0;border:1px solid #b8ccf7;'
                        'border-radius:8px;padding:1px 8px;font-size:0.7rem;font-weight:600;">'
                        f'🏠 {_av("BRANCH_LOCATION")}</span>'
                    )
                    _avail = _av('AVAILABILITY_TO_CLIENTS')
                    _avail_badge = (
                        '<span style="background:#eaf7ec;color:#1a8a6a;border:1px solid #a8dfc8;'
                        f'border-radius:8px;padding:1px 8px;font-size:0.7rem;font-weight:600;">🕐 {_avail}</span>'
                    )
                    _lat = _av('LATITUDE'); _lon = _av('LONGITUDE')
                    _gps = f'<span style="font-size:0.68rem;color:#aaa;">{_lat}, {_lon}</span>' if _lat != '—' else ''
                    _selfserv = _av('SELFSERV_TYPE')
                    _selfserv_cell = (
                        f'<span style="background:#f5f0ff;color:#6b46c1;border:1px solid #d4b8f7;'
                        f'border-radius:8px;padding:1px 8px;font-size:0.7rem;font-weight:600;">⚙️ {_selfserv}</span>'
                        if _selfserv != '—' else '<span style="color:#ccc;">—</span>'
                    )
                    _atm_rows += (
                        f'<tr style="background:{_bg};">'
                        f'<td style="padding:7px 10px;font-size:0.8rem;font-weight:700;color:#1a3a5a;">{_av("ATM_ID")}</td>'
                        f'<td style="padding:7px 10px;font-size:0.78rem;color:#444;">{_av("ATM_TYPE")}</td>'
                        f'<td style="padding:7px 10px;">{_selfserv_cell}</td>'
                        f'<td style="padding:7px 10px;">{_loc_badge}</td>'
                        f'<td style="padding:7px 10px;">{_avail_badge}</td>'
                        f'<td style="padding:7px 10px;">{_gps}</td>'
                        f'</tr>'
                    )

                _n_atm = len(_atm_b)
                atm_html = f"""
<div style="margin-top:28px;padding-top:20px;border-top:2px solid #e0e6f5;">
  <div class="section-header">🏧 Bankomaty na pobočce ({_n_atm})</div>
  <div style="background:#fff;border-radius:10px;padding:16px;box-shadow:0 2px 8px rgba(0,0,0,.07);">
    {_kap_html}
    <table style="width:100%;border-collapse:collapse;font-size:0.82rem;">
      <thead>
        <tr style="background:#2770f0;color:#fff;">
          <th style="padding:7px 10px;text-align:left;border-radius:6px 0 0 0;">ATM ID</th>
          <th style="padding:7px 10px;text-align:left;">Typ HW</th>
          <th style="padding:7px 10px;text-align:left;">SelfServ typ</th>
          <th style="padding:7px 10px;text-align:left;">Umístění</th>
          <th style="padding:7px 10px;text-align:left;">Dostupnost</th>
          <th style="padding:7px 10px;text-align:left;border-radius:0 6px 0 0;">GPS</th>
        </tr>
      </thead>
      <tbody>{_atm_rows}</tbody>
    </table>
  </div>
</div>"""

        # ── Specialisté — přímo bez selectu ───────────────────────────────────
        spec_html = generate_specialiste_summary_table(
            [branch_code],
            title=f"👥 Obsazení pozic — {branch_name}",
            single_branch_code=branch_code
        )

        # ── Spádové pobočky ────────────────────────────────────────────────────
        spadovky_html = ""
        if spadovky is not None and not spadovky.empty:
            _sp = spadovky.copy()
            _sp.columns = [c.strip().upper() for c in _sp.columns]
            _sp['BRANCH_ID'] = pd.to_numeric(_sp['BRANCH_ID'], errors='coerce')
            _sp_row = _sp[_sp['BRANCH_ID'] == branch_code]
            if not _sp_row.empty:
                _sp_row = _sp_row.iloc[0]

                _sp_data = []
                for _rank, (_id_col, _nm_col, _vis_col) in enumerate(zip(
                    ['TOP_POBOCKA_ID_1',     'TOP_POBOCKA_ID_2',     'TOP_POBOCKA_ID_3'],
                    ['TOP_POBOCKA_NAZEV_1',  'TOP_POBOCKA_NAZEV_2',  'TOP_POBOCKA_NAZEV_3'],
                    ['TOP_POBOCKA_VISITS_1', 'TOP_POBOCKA_VISITS_2', 'TOP_POBOCKA_VISITS_3'],
                ), 1):
                    _vid = _sp_row.get(_id_col)
                    if _vid is None or (isinstance(_vid, float) and pd.isna(_vid)) or str(_vid).strip() in ('', 'nan'):
                        continue
                    _vname = _sp_row.get(_nm_col, '')
                    _vname = str(_vname) if pd.notna(_vname) else str(_vid)
                    _vis_raw = _sp_row.get(_vis_col)
                    _vis = int(float(_vis_raw)) if pd.notna(_vis_raw) and str(_vis_raw).strip() not in ('', 'nan') else None
                    _sp_data.append({'rank': _rank, 'name': _vname, 'bid': _vid, 'vis': _vis})

                if _sp_data:
                    _vis_total = sum(d['vis'] for d in _sp_data if d['vis'] is not None)
                    _rank_colors = {1: '#2770f0', 2: '#6b7fe8', 3: '#a5b4fc'}

                    _sp_rows = ""
                    for _d in _sp_data:
                        _rcol     = _rank_colors.get(_d['rank'], '#aaa')
                        _vis_str  = f"{_d['vis']:,}" if _d['vis'] is not None else "—"
                        _pct      = _d['vis'] / _vis_total * 100 if (_d['vis'] and _vis_total) else None
                        _pct_str  = f"{_pct:.1f} %" if _pct is not None else "—"
                        _bar_w    = f"{_pct:.1f}" if _pct is not None else "0"
                        _sp_rows += f"""
<tr>
  <td style="padding:7px 10px;border-bottom:1px solid #eef0f8;">
    <span style="display:inline-block;width:20px;height:20px;border-radius:50%;
                 background:{_rcol};color:#fff;font-size:0.7rem;font-weight:900;
                 text-align:center;line-height:20px;margin-right:6px;">{_d['rank']}</span>
    <span style="font-size:0.82rem;font-weight:700;color:#1a2a4a;">{_d['name']}</span>
    <div style="font-size:0.68rem;color:#bbb;padding-left:26px;">ID {_d['bid']}</div>
  </td>
  <td style="padding:7px 10px;border-bottom:1px solid #eef0f8;text-align:right;
             font-size:0.82rem;font-weight:700;color:#333;white-space:nowrap;">{_vis_str}</td>
  <td style="padding:7px 10px 7px 14px;border-bottom:1px solid #eef0f8;min-width:110px;">
    <div style="display:flex;align-items:center;gap:6px;">
      <div style="flex:1;background:#eef0f8;border-radius:4px;height:8px;overflow:hidden;">
        <div style="width:{_bar_w}%;background:{_rcol};height:100%;border-radius:4px;"></div>
      </div>
      <span style="font-size:0.75rem;font-weight:700;color:{_rcol};min-width:40px;text-align:right;">{_pct_str}</span>
    </div>
  </td>
</tr>"""

                    _vis_total_str = f"{_vis_total:,}" if _vis_total else "—"
                    spadovky_html = f"""
<div style="margin:18px 0 14px 0;">
  <div class="section-header">🔀 Spádové pobočky</div>
  <div style="background:#fff;border-radius:10px;padding:14px 16px;
              box-shadow:0 2px 8px rgba(0,0,0,.07);">
    <div style="font-size:0.72rem;color:#888;margin-bottom:10px;line-height:1.4;">
      Kam pravděpodobně přejdou klienti při uzavření pobočky — dle analýzy spádové oblasti:
    </div>
    <table style="width:100%;border-collapse:collapse;font-size:0.84rem;">
      <thead>
        <tr style="background:#f4f7ff;">
          <th style="padding:6px 10px;text-align:left;font-size:0.72rem;font-weight:700;
                     color:#555;text-transform:uppercase;letter-spacing:.4px;
                     border-bottom:2px solid #dde4f5;">Pobočka</th>
          <th style="padding:6px 10px;text-align:right;font-size:0.72rem;font-weight:700;
                     color:#555;text-transform:uppercase;letter-spacing:.4px;
                     border-bottom:2px solid #dde4f5;white-space:nowrap;">Návštěvy</th>
          <th style="padding:6px 14px;text-align:left;font-size:0.72rem;font-weight:700;
                     color:#555;text-transform:uppercase;letter-spacing:.4px;
                     border-bottom:2px solid #dde4f5;">Podíl</th>
        </tr>
      </thead>
      <tbody>{_sp_rows}</tbody>
      <tfoot>
        <tr style="background:#f4f7ff;">
          <td style="padding:7px 10px;font-weight:700;font-size:0.8rem;color:#444;">∑ Celkem</td>
          <td style="padding:7px 10px;text-align:right;font-weight:800;font-size:0.84rem;
                     color:#2770f0;">{_vis_total_str}</td>
          <td style="padding:7px 14px;font-size:0.72rem;color:#aaa;">100 %</td>
        </tr>
      </tfoot>
    </table>
  </div>
</div>"""

        # ── full_html ─────────────────────────────────────────────────────────
        full_html = f"""<!DOCTYPE html>
<html lang="cs">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{branch_name} ({branch_code}) — Report 2026</title>
  {bootstrap_css}
</head>
<body>
<div class="report-wrapper">

  <div style="display:flex;align-items:flex-start;justify-content:space-between;
              flex-wrap:wrap;gap:12px;margin-bottom:8px;">
    <div>
      <h1 style="color:#1a3a5a;border-left:8px solid #2770f0;padding-left:15px;
                 font-weight:bold;margin-bottom:4px;font-size:1.6rem;">{branch_name}</h1>
      <div style="padding-left:22px;color:#666;font-size:0.88rem;display:flex;flex-wrap:wrap;align-items:center;gap:6px;margin-top:4px;">
        📍 {region} &nbsp;·&nbsp; ID: <b>{branch_code}</b>
        &nbsp;&nbsp;
        {'<span style="background:#edf9f5;color:#1a8a6a;border:1px solid #a8dfc8;border-radius:12px;padding:2px 11px;font-size:0.82rem;font-weight:600;">🏦 Bezhotovostní</span>' if is_cashless_branch else '<span style="background:#fff5ef;color:#c05a10;border:1px solid #f5c9a8;border-radius:12px;padding:2px 11px;font-size:0.82rem;font-weight:600;">💵 Hotovostní</span>'}
        <span style="background:#f0f4ff;color:#2770f0;border:1px solid #b8ccf7;border-radius:12px;padding:2px 11px;font-size:0.82rem;font-weight:600;">📐 {format_p}</span>
        {(lambda close=_rok_str(_v(row,"PLAN_CLOSE_(ROK)")), inv=_inv_label, cl=_cl_str:
            f'<span style="background:#fdecea;color:#d62728;border:1px solid #f5b8b8;border-radius:12px;padding:2px 11px;font-size:0.82rem;font-weight:600;">🔴 Plán uzavření {close}</span>'
            if close else
            f'<span style="background:#fff8ec;color:#e07020;border:1px solid #f5d8a8;border-radius:12px;padding:2px 11px;font-size:0.82rem;font-weight:600;">🔨 Investice {inv}</span>'
            if inv else
            f'<span style="background:#edf9f5;color:#1a8a6a;border:1px solid #a8dfc8;border-radius:12px;padding:2px 11px;font-size:0.82rem;font-weight:600;">🏦 Cashless {cl}</span>'
            if cl else
            '<span style="background:#f5f5f5;color:#999;border:1px solid #ddd;border-radius:12px;padding:2px 11px;font-size:0.82rem;font-weight:500;">bez plánované akce</span>'
        )()}
      </div>
    </div>
    <div style="font-size:0.78rem;color:#aaa;text-align:right;padding-top:8px;">
      Report pobočky 2026 &nbsp;|&nbsp; {idx}/{total_branches}
    </div>
  </div>

  {scorecards_html}

  {rev_history_html}

  {cash_detail_html}

  <div class="section-header">🏢 Strategické informace</div>
  {strat_html}

  {make_collapsible(f"inv-all-{branch_code}",
    f"🔨 Historie investic ({_inv_count} záznamů)",
    investice_html, default_open=True)}

  {make_collapsible(f"bench-{branch_code}",
    f"📊 Benchmark — pobočka vs. region vs. síť vs. {fte_label}",
    benchmark_html, default_open=True)}

  <div class="section-header">🔍 Datově podobné pobočky</div>
  {similar_html}

  <div class="section-header">🛒 Obchody dle produktové skupiny — 2024 / 2025</div>
  <div class="age-section-card mb-4">
    {_prod_rev_html}
    {obchody_html_b}
  </div>

  <div class="section-header">👥 Věkové segmenty klientů &amp; Cizinci</div>
  <div class="age-section-card mb-4">{age_charts_html}</div>

  {metrics_html}

  <div style="display:flex;align-items:flex-start;gap:24px;flex-wrap:wrap;">
    <div style="flex:1;min-width:280px;">{spec_html if spec_html else ""}</div>
    <div style="flex:1;min-width:260px;">{spadovky_html}</div>
  </div>

  {visits_html}

  {atm_html}

</div>
</body>
</html>"""

        filename = os.path.join(output_dir, f"{slug}.html")
        with open(filename, "w", encoding="utf-8") as f:
            f.write(full_html)

        if idx % 20 == 0 or idx == total_branches:
            print(f"  ✅ {idx}/{total_branches} — {branch_name}")

    print(f"\n✅ Hotovo — {total_branches} reportů uloženo do '{output_dir}/'")



# =============================================================================
# 7. PŘÍPRAVA RATING_STATUS A GENEROVÁNÍ REPORTŮ
# =============================================================================

rating_status = prepare_rating_status(df)

# ── Přiřazení ORP dle GPS souřadnic ───────────────────────────────────────────
print("🗺️  Přiřazuji ORP k pobočkám dle GPS...")
try:
    import json as _json_orp
    from shapely.geometry import shape as _shp_shape, Point as _shp_Point
    from pyproj import Transformer as _Transformer

    # ── Debug: načtení GeoJSON ────────────────────────────────────
    _ORP_PATH = '../vypocet_ir_2026/zdroje/csu_geodb_sde_CISORP_byty_etl_20210326.geojson'
    print(f"  📂 Načítám GeoJSON: {_ORP_PATH}")
    with open(_ORP_PATH, encoding='utf-8') as _f:
        _orp_data = _json_orp.load(_f)
    print(f"  ✅ GeoJSON načten — {len(_orp_data['features'])} ORP polygonů")

    # Příprava polygonů (jednou, ne v každé iteraci)
    _orp_polygons = []
    for _feature in _orp_data['features']:
        _orp_polygons.append({
            'polygon':   _shp_shape(_feature['geometry']),
            'orp_kod':   _feature['properties']['kod'],
            'orp_nazev': _feature['properties']['nazev'],
        })

    # Transformer: WGS84 (lon, lat) → S-JTSK (EPSG:5514)
    _orp_transformer = _Transformer.from_crs("EPSG:4326", "EPSG:5514", always_xy=True)
    print("  ✅ Transformer WGS84 → S-JTSK připraven")

    def _find_orp(lat, lon):
        """Najde ORP pro danou GPS pozici. Vrátí (nazev, kod) nebo ('', '')."""
        try:
            if pd.isna(lat) or pd.isna(lon) or float(lat) == 0 or float(lon) == 0:
                return '', ''
            x_jtsk, y_jtsk = _orp_transformer.transform(float(lon), float(lat))
            point = _shp_Point(x_jtsk, y_jtsk)
            for _orp in _orp_polygons:
                if _orp['polygon'].contains(point):
                    return _orp['orp_nazev'], str(_orp['orp_kod'])
            return '', ''
        except Exception:
            return '', ''

    # ── GPS: přenes z df do rating_status pokud chybí ────────────
    for _gc in ['GPS_X', 'GPS_Y']:
        if _gc not in rating_status.columns and _gc in df.columns:
            print(f"  ℹ️  {_gc} chybí v rating_status — přenáším z df")
            _gps_map = df.set_index('BRANCH_CODE')[_gc].to_dict() if 'BRANCH_CODE' in df.columns else {}
            rating_status[_gc] = rating_status['BRANCH_CODE'].map(_gps_map)

    # ── Debug: stav GPS sloupců ───────────────────────────────────
    _gx_ok = 'GPS_X' in rating_status.columns
    _gy_ok = 'GPS_Y' in rating_status.columns
    print(f"  📍 GPS_X v rating_status: {_gx_ok}  |  GPS_Y: {_gy_ok}")
    if _gx_ok:
        _gx_series = pd.to_numeric(rating_status['GPS_X'], errors='coerce')
        rating_status['GPS_X'] = _gx_series
        print(f"     GPS_X nenulové: {(_gx_series > 0).sum()} / {len(rating_status)}"
              f"  — ukázka: {_gx_series.dropna().head(3).tolist()}")
    if _gy_ok:
        _gy_series = pd.to_numeric(rating_status['GPS_Y'], errors='coerce')
        rating_status['GPS_Y'] = _gy_series
        print(f"     GPS_Y nenulové: {(_gy_series > 0).sum()} / {len(rating_status)}"
              f"  — ukázka: {_gy_series.dropna().head(3).tolist()}")

    if _gx_ok and _gy_ok:
        print("  ⏳ Spouštím point-in-polygon pro všechny pobočky...")
        # Debug: ukázka skutečných hodnot GPS_X a GPS_Y
        _sample = rating_status[
            rating_status['GPS_X'].notna() & rating_status['GPS_Y'].notna() &
            (rating_status['GPS_X'] != 0) & (rating_status['GPS_Y'] != 0)
        ][['BRANCH_CODE', 'GPS_X', 'GPS_Y']].head(5)
        print(f"  🔍 Ukázka GPS hodnot:\n{_sample.to_string(index=False)}")

        # Zjisti rozsah — WGS84 lon je 12–19, lat je 48–51
        # S-JTSK X je -950000 až -400000
        _gx_med = float(rating_status['GPS_X'].dropna().median())
        _gy_med = float(rating_status['GPS_Y'].dropna().median())
        print(f"  📊 Medián GPS_X={_gx_med:.4f}, GPS_Y={_gy_med:.4f}")

        # GPS_X = lon (14.xx) nebo lat (50.xx)?
        # Pokud GPS_X ~ 14-19 → lon, GPS_Y ~ 48-51 → lat  (normální)
        # Pokud GPS_X ~ 48-51 → lat, GPS_Y ~ 14-19 → lon  (prohozeno)
        if 48 < _gx_med < 52:
            # GPS_X je latitude, GPS_Y je longitude — prohozeno!
            print("  🔄 GPS_X=latitude, GPS_Y=longitude — používám GPS_X jako lat, GPS_Y jako lon")
            _orp_results = rating_status.apply(
                lambda r: _find_orp(r['GPS_X'], r['GPS_Y']), axis=1
            )
        else:
            # GPS_X je longitude, GPS_Y je latitude — standardní
            print("  ✅ GPS_X=longitude, GPS_Y=latitude — standardní pořadí")
            _orp_results = rating_status.apply(
                lambda r: _find_orp(r['GPS_Y'], r['GPS_X']), axis=1
            )
        rating_status['ORP_NAZEV'] = _orp_results.apply(lambda x: x[0])
        rating_status['ORP_KOD']   = _orp_results.apply(lambda x: x[1])
        _n_found = (rating_status['ORP_NAZEV'] != '').sum()
        print(f"  ✅ ORP přiřazeno {_n_found} / {len(rating_status)} pobočkám")
        if _n_found > 0:
            print(f"     Ukázka: {rating_status[rating_status['ORP_NAZEV'] != ''][['BRANCH_NAME','ORP_NAZEV']].head(3).to_dict('records')}")
    else:
        print("  ⚠️  GPS_X/GPS_Y nejsou dostupné — ORP přeskočeno")
        rating_status['ORP_NAZEV'] = ''
        rating_status['ORP_KOD']   = ''

except FileNotFoundError as _fe:
    print(f"  ❌ GeoJSON nenalezen: {_fe}")
    import os as _os_orp
    print(f"     CWD: {_os_orp.getcwd()}")
    rating_status['ORP_NAZEV'] = ''
    rating_status['ORP_KOD']   = ''
    rating_status['ORP_NAZEV'] = ''
    rating_status['ORP_KOD']   = ''
except Exception as _orp_err:
    print(f"  ⚠️  ORP přiřazení selhalo: {_orp_err}")
    rating_status['ORP_NAZEV'] = ''
    rating_status['ORP_KOD']   = ''


# ── Přenos _Q sloupců z df do rating_status ───────────────────────────────────
# _safe_qcut loop zapsal _Q do df před vznikem rating_status — sync zpět.
_q_cols_to_sync = [c for c in df.columns if c.endswith("_Q") and c not in rating_status.columns]
if _q_cols_to_sync and "BRANCH_CODE" in df.columns and "BRANCH_CODE" in rating_status.columns:
    _q_sync = df[["BRANCH_CODE"] + _q_cols_to_sync].drop_duplicates("BRANCH_CODE")
    rating_status = rating_status.merge(_q_sync, on="BRANCH_CODE", how="left")

# Statický report (celkový přehled)
generate_report(rating_status, mode='static', output_prefix="report_rating_2026")

# Regionální reporty (jeden soubor per region)
generate_report(rating_status, mode='regional', output_prefix="report")

# Pobočkové reporty (jeden soubor per pobočka)
generate_branch_reports(rating_status, output_dir="report_pobocky", hotovostni_trn=hotovostni_trn, hotovostni_trn_detail=hotovostni_trn_detail, export_atm=export_atm, kapacita_atm=kapacita_atm, visits=visits, parties_full=parties_full, spadovky=spadovky)